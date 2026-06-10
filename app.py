from flask import Flask, request, render_template_string, jsonify, send_file
from dotenv import load_dotenv
import google.generativeai as genai
from werkzeug.utils import secure_filename
import os
import json
import requests
import sqlite3
import datetime
import threading
import time
import tempfile
import csv
import random
import io
import shutil

try:
    import openpyxl
except Exception:
    openpyxl = None

load_dotenv()

APP_TITLE = "Marketing Automation Pro"
DB = "marketing_automation_pro_v11.db"
UPLOAD_DIR = "uploads"
REPORT_DIR = "reports"
BACKUP_DIR = "backups"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
PAGES_JSON = os.getenv("PAGES_JSON", "[]").strip()

try:
    PAGES = json.loads(PAGES_JSON)
except Exception:
    PAGES = []

if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

client = genai if GEMINI_API_KEY else None


CONTENT_LIBRARY = {
    "spa": [
        "Làn da đẹp không đến từ may mắn, mà đến từ chăm sóc đúng cách. Inbox để được tư vấn liệu trình phù hợp.",
        "Da xỉn màu, thiếu sức sống? Một liệu trình chăm sóc chuyên sâu có thể giúp bạn tự tin hơn mỗi ngày.",
        "Dành chút thời gian chăm sóc bản thân hôm nay, bạn sẽ thấy mình rạng rỡ hơn vào ngày mai.",
        "Không cần trang điểm quá nhiều khi làn da đã được chăm sóc đúng cách.",
        "Bạn xứng đáng có một làn da khỏe, mịn và đầy sức sống."
    ],
    "nha_khoa": [
        "Nụ cười đẹp giúp bạn tự tin hơn trong giao tiếp mỗi ngày. Inbox để được tư vấn nha khoa.",
        "Răng đều, trắng sáng và khỏe mạnh bắt đầu từ việc thăm khám đúng lúc.",
        "Đừng để vấn đề răng miệng nhỏ trở thành nỗi lo lớn.",
        "Một nụ cười tự tin có thể thay đổi cách người khác nhìn bạn.",
        "Chăm sóc răng miệng định kỳ là đầu tư cho sức khỏe và sự tự tin lâu dài."
    ],
    "bat_dong_san": [
        "Cơ hội sở hữu bất động sản vị trí đẹp, pháp lý rõ ràng, tiềm năng tăng giá tốt.",
        "Bạn đang tìm căn hộ để ở hoặc đầu tư? Nhắn tin để nhận bảng giá mới.",
        "Bất động sản tốt không chờ đợi quá lâu. Liên hệ ngay để được tư vấn.",
        "Vị trí đẹp, tiện ích đầy đủ, pháp lý minh bạch.",
        "Đầu tư bất động sản cần đúng thời điểm và đúng sản phẩm."
    ],
    "facebook_ads": [
        "Chạy quảng cáo không ra đơn không phải lúc nào cũng do sản phẩm.",
        "Inbox cao, đơn thấp? Đã đến lúc xem lại nội dung, tệp khách hàng và hành trình chốt đơn.",
        "Một chiến dịch quảng cáo hiệu quả cần đúng thông điệp, đúng khách hàng và đúng thời điểm.",
        "Đừng để ngân sách quảng cáo bị tiêu hao mà không tạo ra kết quả.",
        "Muốn giảm chi phí tin nhắn và tăng tỷ lệ chốt đơn? Bắt đầu từ nội dung và tệp khách."
    ],
    "proxy": [
        "Cần proxy ổn định cho công việc online? Tham khảo gói proxy tốc độ cao, hỗ trợ nhanh.",
        "Proxy phù hợp giúp công việc quảng cáo, quản lý tài khoản và automation ổn định hơn.",
        "Tìm proxy Việt Nam ổn định, tốc độ tốt, hỗ trợ nhanh? Nhắn tin để nhận bảng giá.",
        "Giải pháp proxy linh hoạt cho cá nhân và doanh nghiệp.",
        "Proxy chất lượng giúp giảm gián đoạn khi làm việc nhiều tài khoản."
    ],
    "tiktok_shop": [
        "Muốn TikTok Shop vận hành hiệu quả hơn? Bắt đầu từ nội dung đúng sản phẩm, đúng khách hàng.",
        "Sản phẩm tốt cần nội dung đủ cuốn hút để khách dừng lại, xem và mua.",
        "TikTok Shop không chỉ cần sản phẩm, mà cần cách trình bày khiến khách muốn mua ngay.",
        "Tối ưu nội dung, hình ảnh và kịch bản bán hàng giúp shop tăng chuyển đổi.",
        "Đang bán TikTok Shop nhưng chưa ra đơn đều? Hãy tối ưu nội dung và phễu bán hàng."
    ],
    "website": [
        "Website chuyên nghiệp giúp khách hàng tin tưởng hơn trước khi quyết định mua hàng.",
        "Một website tốt không chỉ đẹp, mà còn phải rõ thông tin, dễ dùng và hỗ trợ chuyển đổi.",
        "Bạn cần website bán hàng, giới thiệu dịch vụ hoặc landing page chạy quảng cáo?",
        "Đừng để khách rời đi vì website thiếu chuyên nghiệp.",
        "Website là tài sản số quan trọng cho mọi doanh nghiệp."
    ],
    "noi_that": [
        "Không gian đẹp bắt đầu từ thiết kế phù hợp nhu cầu sống và gu thẩm mỹ của bạn.",
        "Một bộ sofa phù hợp có thể thay đổi cảm giác của cả căn phòng.",
        "Nội thất không chỉ để dùng, mà còn tạo nên phong cách sống.",
        "Tối ưu không gian sống với giải pháp nội thất hiện đại, tiện nghi và thẩm mỹ.",
        "Bạn muốn nhà đẹp hơn nhưng chưa biết bắt đầu từ đâu? Inbox để được tư vấn."
    ],
    "oto": [
        "Chăm sóc xe định kỳ giúp xe bền hơn, vận hành ổn định hơn và giữ giá trị tốt hơn.",
        "Xe sạch, bóng, nội thất thơm tho giúp mỗi chuyến đi dễ chịu hơn.",
        "Phụ kiện phù hợp giúp chiếc xe tiện nghi, an toàn và cá tính hơn.",
        "Bạn đang tìm camera hành trình, phụ kiện hoặc dịch vụ chăm sóc xe?",
        "Đừng chờ xe xuống cấp mới chăm sóc. Bảo dưỡng đúng lúc giúp tiết kiệm chi phí."
    ],
    "giao_duc": [
        "Học đúng phương pháp giúp tiết kiệm thời gian và cải thiện kết quả nhanh hơn.",
        "Bạn muốn nâng cấp kỹ năng nhưng chưa biết bắt đầu từ đâu?",
        "Một khóa học tốt không chỉ cung cấp kiến thức, mà còn giúp ứng dụng thực tế.",
        "Đầu tư vào kiến thức là khoản đầu tư có giá trị lâu dài.",
        "Học online linh hoạt, tiết kiệm thời gian, phù hợp người bận rộn."
    ],
    "du_lich": [
        "Một chuyến đi đẹp bắt đầu từ kế hoạch phù hợp.",
        "Bạn muốn du lịch thư giãn nhưng ngại tự lên lịch?",
        "Khám phá điểm đến mới với lịch trình tối ưu, chi phí hợp lý.",
        "Cần đặt tour, khách sạn, vé máy bay hoặc visa?",
        "Đi để nghỉ ngơi, trải nghiệm và nạp lại năng lượng."
    ]
}


# Kho content V9 mở rộng 15 ngành. Demo hiển thị 10 mẫu miễn phí/ngành, phần còn lại khóa Premium.
EXTENDED_CONTENT_LIBRARY = {
    "thoi_trang": [
        "Phong cách đẹp bắt đầu từ outfit phù hợp. Inbox để được tư vấn mẫu mới hôm nay.",
        "Một set đồ chỉn chu giúp bạn tự tin hơn trong mọi cuộc gặp.",
        "Thời trang không chỉ là mặc đẹp, mà còn là cách bạn thể hiện cá tính.",
        "Mẫu mới về liên tục, chất vải đẹp, form dễ mặc. Nhắn tin để xem ảnh thật.",
        "Đổi mới phong cách mỗi ngày với những thiết kế trẻ trung, dễ phối.",
        "Bạn cần outfit đi làm, đi chơi hay dự tiệc? Inbox để được gợi ý.",
        "Hàng đẹp, form chuẩn, giá hợp lý. Số lượng có hạn, nhắn tin ngay.",
        "Tủ đồ của bạn sẽ thú vị hơn với những mẫu mới đang hot.",
        "Đẹp tự nhiên, mặc thoải mái, phù hợp nhiều dáng người.",
        "Inbox để nhận bảng mẫu mới và ưu đãi hôm nay."
    ],
    "giay_dep": [
        "Một đôi giày phù hợp có thể nâng tầm cả outfit của bạn.",
        "Giày đẹp, êm chân, dễ phối đồ cho cả đi làm và đi chơi.",
        "Bạn đang tìm giày bền, đẹp, giá hợp lý? Nhắn tin để xem mẫu.",
        "Mẫu mới về liên tục, size đầy đủ, hỗ trợ tư vấn chọn size.",
        "Đừng để đôi giày không thoải mái làm bạn mất tự tin.",
        "Giày chuẩn form, đi êm, phù hợp nhiều phong cách.",
        "Inbox để nhận ảnh thật và bảng size chi tiết.",
        "Nâng cấp phong cách hằng ngày chỉ từ một đôi giày đẹp.",
        "Chọn đúng giày, mỗi bước đi đều tự tin hơn.",
        "Mẫu hot hôm nay số lượng có hạn, nhắn tin giữ size."
    ],
    "tui_xach": [
        "Một chiếc túi đẹp giúp outfit của bạn nổi bật hơn.",
        "Túi xách thời trang, dễ phối đồ, phù hợp đi làm và đi chơi.",
        "Bạn cần túi sang, đẹp, tiện dụng? Inbox để xem mẫu mới.",
        "Thiết kế tinh tế, chất liệu đẹp, phù hợp nhiều phong cách.",
        "Túi không chỉ để đựng đồ, mà còn tạo điểm nhấn cho phong cách.",
        "Mẫu mới đang có sẵn, nhắn tin để nhận ảnh thật.",
        "Túi đẹp giúp bạn tự tin hơn mỗi khi ra ngoài.",
        "Phối đồ dễ hơn với những mẫu túi thanh lịch, hiện đại.",
        "Số lượng có hạn, inbox để được tư vấn mẫu phù hợp.",
        "Tặng bản thân một chiếc túi mới để làm mới phong cách."
    ],
    "dong_ho": [
        "Đồng hồ đẹp là điểm nhấn tinh tế cho phong cách của bạn.",
        "Một chiếc đồng hồ phù hợp giúp bạn trông chỉn chu hơn mỗi ngày.",
        "Thiết kế sang trọng, dễ phối, phù hợp đi làm và gặp khách hàng.",
        "Bạn cần đồng hồ làm quà tặng? Inbox để được tư vấn mẫu phù hợp.",
        "Đồng hồ không chỉ xem giờ, mà còn thể hiện gu thẩm mỹ.",
        "Mẫu mới về, thiết kế đẹp, giá hợp lý.",
        "Tối giản nhưng nổi bật, phù hợp nhiều phong cách.",
        "Inbox để xem ảnh thật và nhận tư vấn lựa chọn.",
        "Một món phụ kiện nhỏ nhưng tạo khác biệt lớn.",
        "Chọn đồng hồ phù hợp để hoàn thiện vẻ ngoài chuyên nghiệp."
    ],
    "phu_kien": [
        "Phụ kiện nhỏ có thể làm outfit của bạn nổi bật hơn rất nhiều.",
        "Đừng bỏ qua chi tiết, vì phụ kiện tạo nên phong cách riêng.",
        "Mẫu phụ kiện mới, dễ phối, phù hợp nhiều phong cách.",
        "Inbox để xem mẫu hot và nhận tư vấn phối đồ.",
        "Một món phụ kiện đẹp giúp bạn tự tin hơn khi ra ngoài.",
        "Thiết kế tinh tế, giá dễ mua, phù hợp làm quà tặng.",
        "Nâng cấp outfit chỉ với một điểm nhấn nhỏ.",
        "Phụ kiện đẹp giúp phong cách của bạn có dấu ấn riêng.",
        "Mẫu mới cập nhật mỗi ngày, nhắn tin để xem ngay.",
        "Chọn phụ kiện phù hợp để hoàn thiện vẻ ngoài của bạn."
    ],
    "tham_my_vien": [
        "Vẻ đẹp tự tin bắt đầu từ lựa chọn chăm sóc đúng cách.",
        "Bạn muốn cải thiện diện mạo nhưng chưa biết bắt đầu từ đâu? Inbox để được tư vấn.",
        "Thẩm mỹ an toàn, tư vấn rõ ràng, phù hợp từng nhu cầu.",
        "Làm đẹp là đầu tư cho sự tự tin và chất lượng cuộc sống.",
        "Đừng để khuyết điểm nhỏ làm bạn kém tự tin mỗi ngày.",
        "Đội ngũ tư vấn sẽ giúp bạn chọn giải pháp phù hợp nhất.",
        "Inbox để nhận tư vấn liệu trình và ưu đãi mới.",
        "Chăm sóc vẻ ngoài đúng cách giúp bạn rạng rỡ hơn.",
        "Đẹp tự nhiên, an toàn, phù hợp với từng khách hàng.",
        "Đặt lịch tư vấn để hiểu rõ giải pháp phù hợp."
    ],
    "trung_tam_tieng_anh": [
        "Tiếng Anh tốt mở ra nhiều cơ hội học tập và công việc hơn.",
        "Bạn mất gốc tiếng Anh? Bắt đầu lại với lộ trình phù hợp.",
        "Học đúng phương pháp giúp bạn tiến bộ nhanh và bền vững.",
        "Lớp học linh hoạt, giáo viên hỗ trợ sát sao, phù hợp người bận rộn.",
        "Inbox để kiểm tra trình độ và nhận tư vấn lộ trình học.",
        "Đừng học lan man, hãy học theo mục tiêu rõ ràng.",
        "Giao tiếp tự tin hơn với lộ trình học thực tế.",
        "Tiếng Anh không khó nếu bạn có phương pháp phù hợp.",
        "Đăng ký tư vấn miễn phí để chọn lớp phù hợp.",
        "Nâng cấp tiếng Anh hôm nay để mở rộng cơ hội ngày mai."
    ],
    "khoa_hoc_online": [
        "Học online linh hoạt giúp bạn nâng cấp kỹ năng mọi lúc mọi nơi.",
        "Đầu tư vào kiến thức là khoản đầu tư có giá trị lâu dài.",
        "Bạn muốn học nhưng không có nhiều thời gian? Khóa online là lựa chọn phù hợp.",
        "Nội dung dễ hiểu, ứng dụng thực tế, học theo tốc độ của bạn.",
        "Inbox để nhận lộ trình học phù hợp với mục tiêu.",
        "Nâng cấp kỹ năng để tăng cơ hội công việc và thu nhập.",
        "Học đúng thứ bạn cần, tiết kiệm thời gian và chi phí.",
        "Khóa học thiết kế cho người mới bắt đầu và người muốn nâng cao.",
        "Bắt đầu hôm nay, kết quả sẽ khác sau vài tuần.",
        "Nhắn tin để nhận tư vấn khóa học phù hợp."
    ],
    "showroom_oto": [
        "Tìm xe phù hợp nhu cầu và tài chính chưa bao giờ dễ hơn.",
        "Showroom cập nhật nhiều mẫu xe mới, hỗ trợ tư vấn tận tình.",
        "Bạn cần xe gia đình, xe dịch vụ hay xe cá nhân? Inbox để được tư vấn.",
        "Xe đẹp, hồ sơ rõ ràng, hỗ trợ thủ tục nhanh chóng.",
        "Lựa chọn xe đúng giúp bạn yên tâm hơn trên mọi hành trình.",
        "Inbox để nhận báo giá và chương trình ưu đãi mới nhất.",
        "Tư vấn tài chính, trả góp, hồ sơ nhanh gọn.",
        "Mẫu xe hot đang có sẵn, liên hệ để xem xe trực tiếp.",
        "Một chiếc xe phù hợp sẽ đồng hành cùng bạn lâu dài.",
        "Nhắn tin để được tư vấn mẫu xe phù hợp nhất."
    ],
    "phong_kham": [
        "Sức khỏe là tài sản quan trọng nhất, đừng chờ có triệu chứng mới kiểm tra.",
        "Thăm khám định kỳ giúp phát hiện và xử lý vấn đề sớm hơn.",
        "Phòng khám hỗ trợ tư vấn, đặt lịch nhanh, tiết kiệm thời gian.",
        "Bạn cần kiểm tra sức khỏe? Inbox để được hướng dẫn đặt lịch.",
        "Chăm sóc sức khỏe đúng lúc giúp bạn yên tâm hơn mỗi ngày.",
        "Dịch vụ thăm khám chuyên nghiệp, tư vấn rõ ràng.",
        "Đặt lịch trước để được hỗ trợ nhanh và thuận tiện hơn.",
        "Sức khỏe tốt bắt đầu từ sự chủ động của bạn.",
        "Inbox để nhận thông tin gói khám phù hợp.",
        "Đừng trì hoãn việc chăm sóc sức khỏe của bản thân."
    ],
}

CONTENT_LIBRARY.update(EXTENDED_CONTENT_LIBRARY)

INDUSTRY_LABELS = {
    "spa": "Spa / Chăm sóc da",
    "nha_khoa": "Nha khoa",
    "bat_dong_san": "Bất động sản",
    "facebook_ads": "Facebook Ads",
    "proxy": "Proxy",
    "tiktok_shop": "TikTok Shop",
    "website": "Thiết kế website",
    "noi_that": "Nội thất",
    "oto": "Ô tô - Chăm sóc xe",
    "giao_duc": "Giáo dục",
    "du_lich": "Du lịch",
    "thoi_trang": "Thời trang",
    "giay_dep": "Giày dép",
    "tui_xach": "Túi xách",
    "dong_ho": "Đồng hồ",
    "phu_kien": "Phụ kiện thời trang",
    "tham_my_vien": "Thẩm mỹ viện",
    "trung_tam_tieng_anh": "Trung tâm tiếng Anh",
    "khoa_hoc_online": "Khóa học online",
    "showroom_oto": "Showroom ô tô",
    "phong_kham": "Phòng khám"
}

def init_db():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_name TEXT,
        page_id TEXT,
        content TEXT,
        status TEXT,
        post_id TEXT,
        schedule_time TEXT,
        image_path TEXT,
        campaign TEXT,
        score INTEGER,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS campaigns (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        industry TEXT,
        goal TEXT,
        note TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS crm (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        phone TEXT,
        zalo TEXT,
        source TEXT,
        note TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS token_checks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_name TEXT,
        page_id TEXT,
        status TEXT,
        detail TEXT,
        checked_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS page_clusters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        page_names TEXT,
        note TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        package_name TEXT,
        start_date TEXT,
        end_date TEXT,
        status TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS users_trial (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        package_name TEXT DEFAULT 'free',
        trial_start_date TEXT,
        trial_end_date TEXT,
        status TEXT DEFAULT 'active',
        created_at TEXT
    )
    """)

    # V3 Enterprise AI Suite tables
    c.execute("""
    CREATE TABLE IF NOT EXISTS lead_pipeline (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_name TEXT,
        phone TEXT,
        zalo TEXT,
        source TEXT,
        stage TEXT DEFAULT 'Khách mới',
        value INTEGER DEFAULT 0,
        note TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS customer_tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_name TEXT,
        task TEXT,
        due_date TEXT,
        status TEXT DEFAULT 'pending',
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        detail TEXT,
        level TEXT DEFAULT 'info',
        status TEXT DEFAULT 'new',
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS competitors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_url TEXT,
        analysis TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS sales_scripts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        topic TEXT,
        script_type TEXT,
        content TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS landing_pages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_name TEXT,
        headline TEXT,
        content TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS workflows (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        trigger_name TEXT,
        action_name TEXT,
        status TEXT DEFAULT 'draft',
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS user_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        role TEXT DEFAULT 'staff',
        note TEXT,
        created_at TEXT
    )
    """)

    # V5 Seller AI Suite tables
    c.execute("""
    CREATE TABLE IF NOT EXISTS fb_groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_name TEXT,
        group_id TEXT,
        niche TEXT,
        note TEXT,
        status TEXT DEFAULT 'active',
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS group_schedules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_name TEXT,
        group_id TEXT,
        content TEXT,
        schedule_time TEXT,
        status TEXT DEFAULT 'scheduled',
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS comment_leads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_name TEXT,
        phone TEXT,
        comment_text TEXT,
        ai_reply TEXT,
        label TEXT,
        crm_status TEXT DEFAULT 'new',
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS messenger_scripts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product TEXT,
        script_type TEXT,
        content TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS success_assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        asset_type TEXT,
        title TEXT,
        content TEXT,
        created_at TEXT
    )
    """)

    # V6 Group Finder & UID Splitter tables - an toàn, có duyệt trước khi thao tác
    c.execute("""
    CREATE TABLE IF NOT EXISTS group_finder_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        keyword TEXT,
        group_name TEXT,
        group_uid TEXT UNIQUE,
        members INTEGER DEFAULT 0,
        privacy TEXT,
        recent_activity TEXT,
        page_join_allowed TEXT,
        page_post_allowed TEXT,
        status TEXT DEFAULT 'Hợp lệ',
        note TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS group_join_queue (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_uid TEXT,
        group_name TEXT,
        keyword TEXT,
        members INTEGER DEFAULT 0,
        selected_page_id TEXT,
        selected_page_name TEXT,
        status TEXT DEFAULT 'Chưa tham gia',
        admin_status TEXT DEFAULT 'Chờ admin duyệt',
        note TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS group_uid_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_name TEXT,
        uid_count INTEGER DEFAULT 0,
        chunk_size INTEGER DEFAULT 50,
        file_path TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS group_post_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_uid TEXT,
        post_uid TEXT UNIQUE,
        post_link TEXT,
        author_name TEXT,
        posted_at TEXT,
        content_preview TEXT,
        comments INTEGER DEFAULT 0,
        reactions INTEGER DEFAULT 0,
        status TEXT DEFAULT 'Chưa xử lý',
        keyword TEXT,
        created_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS group_post_queue (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_id TEXT,
        page_name TEXT,
        group_uid TEXT,
        group_name TEXT,
        content TEXT,
        status TEXT DEFAULT 'Chờ duyệt',
        note TEXT,
        created_at TEXT
    )
    """)

    # V7 Comment Manager Pro - lưu hàng chờ bình luận hợp lệ, có giãn cách và duyệt trước
    c.execute("""
    CREATE TABLE IF NOT EXISTS page_comment_queue (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_id TEXT,
        page_name TEXT,
        target_type TEXT DEFAULT 'post',
        user_uid TEXT,
        post_uid TEXT,
        group_uid TEXT,
        comment_text TEXT,
        min_delay_seconds INTEGER DEFAULT 45,
        max_delay_seconds INTEGER DEFAULT 60,
        scheduled_at TEXT,
        status TEXT DEFAULT 'Chờ duyệt',
        admin_status TEXT DEFAULT 'Chờ admin duyệt',
        result_message TEXT,
        created_at TEXT,
        processed_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS page_comment_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        queue_id INTEGER,
        page_name TEXT,
        target_uid TEXT,
        action TEXT,
        status TEXT,
        detail TEXT,
        created_at TEXT
    )
    """)

    # V10 Page Token Center - thêm Page ID/Token ngay trong tool, không cần sửa Render Environment
    c.execute("""
    CREATE TABLE IF NOT EXISTS page_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_name TEXT,
        page_id TEXT UNIQUE,
        page_token TEXT,
        status TEXT DEFAULT 'active',
        note TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS page_group_memberships (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        page_id TEXT,
        page_name TEXT,
        group_id TEXT,
        group_name TEXT,
        status TEXT DEFAULT 'Đã tham gia',
        can_post TEXT DEFAULT 'Có',
        note TEXT,
        created_at TEXT,
        UNIQUE(page_id, group_id)
    )
    """)

    # V12 Device ID + Premium Payment Approval + Renewal Reminder
    c.execute("""
    CREATE TABLE IF NOT EXISTS premium_upgrade_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id TEXT,
        phone TEXT,
        email TEXT,
        package_key TEXT,
        package_name TEXT,
        amount INTEGER DEFAULT 0,
        payment_note TEXT,
        status TEXT DEFAULT 'Chờ duyệt',
        admin_note TEXT,
        created_at TEXT,
        approved_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS device_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id TEXT UNIQUE,
        phone TEXT,
        email TEXT,
        package_key TEXT,
        package_name TEXT,
        start_date TEXT,
        end_date TEXT,
        status TEXT DEFAULT 'premium',
        last_renewal_notice_at TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS support_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        device_id TEXT,
        sender TEXT,
        message TEXT,
        status TEXT DEFAULT 'new',
        created_at TEXT
    )
    """)
    conn.commit()
    conn.close()

def db():
    return sqlite3.connect(DB)


def get_pages_dynamic():
    """Lấy Page từ PAGES_JSON và Page Token lưu trong tool.
    Page thêm trong giao diện sẽ dùng ngay cho dropdown đăng bài/bình luận mà không cần sửa Render Environment.
    """
    pages = []
    seen = set()
    for page in PAGES:
        pid = str(page.get("id", "")).strip()
        if not pid or pid in seen:
            continue
        pages.append({"name": page.get("name", "No name"), "id": pid, "token": page.get("token", ""), "source": "env"})
        seen.add(pid)
    try:
        conn = db(); c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS page_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                page_name TEXT,
                page_id TEXT UNIQUE,
                page_token TEXT,
                status TEXT DEFAULT 'active',
                note TEXT,
                created_at TEXT,
                updated_at TEXT
            )
        """)
        c.execute("SELECT page_name,page_id,page_token,status,note FROM page_tokens WHERE COALESCE(status,'active')!='deleted' ORDER BY id DESC")
        rows = c.fetchall(); conn.close()
        for name, pid, token, status, note in rows:
            pid = str(pid or "").strip()
            if not pid or pid in seen:
                continue
            pages.append({"name": name or f"Page {pid}", "id": pid, "token": token or "", "source": "db", "status": status or "active", "note": note or ""})
            seen.add(pid)
    except Exception:
        pass
    return pages

def get_page_by_index(page_index):
    pages = get_pages_dynamic()
    try:
        i = int(page_index)
        if 0 <= i < len(pages):
            return pages[i]
    except Exception:
        pass
    return None

def save_page_token(page_name, page_id, page_token, note=''):
    page_id = str(page_id or '').strip()
    page_token = str(page_token or '').strip()
    if not page_id or not page_token:
        return False
    page_name = (page_name or f'Page {page_id}').strip()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS page_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page_name TEXT,
            page_id TEXT UNIQUE,
            page_token TEXT,
            status TEXT DEFAULT 'active',
            note TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)
    c.execute("""
        INSERT INTO page_tokens(page_name,page_id,page_token,status,note,created_at,updated_at)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(page_id) DO UPDATE SET
            page_name=excluded.page_name,
            page_token=excluded.page_token,
            status='active',
            note=excluded.note,
            updated_at=excluded.updated_at
    """, (page_name, page_id, page_token, 'active', note, now, now))
    conn.commit(); conn.close(); return True

def get_page_token_rows(limit=100):
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS page_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page_name TEXT,
            page_id TEXT UNIQUE,
            page_token TEXT,
            status TEXT DEFAULT 'active',
            note TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)
    c.execute("SELECT id,page_name,page_id,CASE WHEN page_token!='' THEN substr(page_token,1,8)||'...' ELSE '' END,status,note,updated_at FROM page_tokens WHERE COALESCE(status,'active')!='deleted' ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def add_page_group_membership(page_index, group_id, status='Đã tham gia', can_post='Có', note=''):
    page = get_page_by_index(page_index)
    group_id = normalize_uid(group_id)
    if not page or not group_id:
        return False
    conn = db(); c = conn.cursor()
    c.execute("SELECT group_name FROM fb_groups WHERE group_id=? LIMIT 1", (group_id,))
    row = c.fetchone(); group_name = row[0] if row else group_id
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute("""
        CREATE TABLE IF NOT EXISTS page_group_memberships (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page_id TEXT,
            page_name TEXT,
            group_id TEXT,
            group_name TEXT,
            status TEXT DEFAULT 'Đã tham gia',
            can_post TEXT DEFAULT 'Có',
            note TEXT,
            created_at TEXT,
            UNIQUE(page_id, group_id)
        )
    """)
    c.execute("""
        INSERT INTO page_group_memberships(page_id,page_name,group_id,group_name,status,can_post,note,created_at)
        VALUES(?,?,?,?,?,?,?,?)
        ON CONFLICT(page_id, group_id) DO UPDATE SET
            page_name=excluded.page_name,
            group_name=excluded.group_name,
            status=excluded.status,
            can_post=excluded.can_post,
            note=excluded.note
    """, (str(page.get('id','')), page.get('name','Chưa chọn Page'), group_id, group_name, status, can_post, note, now))
    conn.commit(); conn.close(); return True

def get_page_group_memberships(limit=200):
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS page_group_memberships (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page_id TEXT,
            page_name TEXT,
            group_id TEXT,
            group_name TEXT,
            status TEXT DEFAULT 'Đã tham gia',
            can_post TEXT DEFAULT 'Có',
            note TEXT,
            created_at TEXT,
            UNIQUE(page_id, group_id)
        )
    """)
    c.execute("SELECT id,page_name,page_id,group_name,group_id,status,can_post,note,created_at FROM page_group_memberships ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def page_group_can_post(page_id, group_id):
    conn = db(); c = conn.cursor()
    try:
        c.execute("SELECT status,can_post FROM page_group_memberships WHERE page_id=? AND group_id=? LIMIT 1", (str(page_id), str(group_id)))
        row = c.fetchone()
    except Exception:
        row = None
    conn.close()
    if not row:
        return False
    return str(row[0]).lower() in ['đã tham gia','da tham gia','joined','active'] and str(row[1]).lower() in ['có','co','yes','1','true']

def save_post(page_name, page_id, content, status, post_id="", schedule_time="", image_path="", campaign="", score=0):
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT INTO posts(page_name,page_id,content,status,post_id,schedule_time,image_path,campaign,score,created_at)
    VALUES(?,?,?,?,?,?,?,?,?,?)
    """, (page_name, page_id, content, status, post_id, schedule_time, image_path, campaign, score,
          datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit(); conn.close()

def update_post(row_id, status, post_id=""):
    conn = db(); c = conn.cursor()
    c.execute("UPDATE posts SET status=?, post_id=? WHERE id=?", (status, post_id, row_id))
    conn.commit(); conn.close()

def get_history(limit=60):
    conn = db(); c = conn.cursor()
    c.execute("""
    SELECT id,page_name,content,status,post_id,schedule_time,image_path,campaign,score,created_at
    FROM posts ORDER BY id DESC LIMIT ?
    """, (limit,))
    rows = c.fetchall(); conn.close()
    return rows

def get_campaigns():
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,name,industry,goal,note,created_at FROM campaigns ORDER BY id DESC")
    rows = c.fetchall(); conn.close()
    return rows

def add_campaign(name, industry, goal, note):
    if not name:
        return
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT OR IGNORE INTO campaigns(name,industry,goal,note,created_at)
    VALUES(?,?,?,?,?)
    """, (name, industry, goal, note, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit(); conn.close()

def get_stats():
    conn = db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM posts"); total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='posted'"); posted = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='scheduled'"); scheduled = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='error'"); error = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM campaigns"); campaigns = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM crm")
    crm = c.fetchone()[0]
    conn.close()
    return {"total": total, "posted": posted, "scheduled": scheduled, "error": error, "campaigns": campaigns, "crm": crm}

def selected_pages(page_indexes):
    result = []
    pages = get_pages_dynamic()
    for idx in page_indexes:
        try:
            i = int(idx)
            if 0 <= i < len(pages):
                result.append(pages[i])
        except Exception:
            pass
    return result

def score_content(content):
    score = 50
    text = content.lower()
    if len(content) >= 120: score += 10
    if any(x in text for x in ["inbox", "nhắn tin", "liên hệ", "đặt lịch", "zalo"]): score += 15
    if any(x in content for x in ["🔥", "🚀", "📩", "💎", "✨"]): score += 5
    if "#" in content: score += 5
    if any(x in text for x in ["cam kết 100%", "chắc chắn", "đảm bảo ra đơn", "trị khỏi", "hiệu quả tuyệt đối"]): score -= 25
    return max(0, min(100, score))

def policy_check(content):
    warnings = []
    risky = ["cam kết 100%", "chắc chắn", "đảm bảo ra đơn", "trị khỏi", "hiệu quả tuyệt đối", "thần tốc"]
    for word in risky:
        if word in content.lower():
            warnings.append(f"Nên tránh cụm từ: {word}")
    if content.count("#") > 5:
        warnings.append("Hashtag hơi nhiều, nên giảm xuống 2-4 hashtag.")
    if len(content) > 1200:
        warnings.append("Nội dung khá dài, nên rút gọn để dễ đọc hơn.")
    if not warnings:
        warnings.append("Nội dung tương đối an toàn, vẫn nên kiểm tra lại trước khi đăng.")
    return warnings


def friendly_ai_error(error):
    raw = str(error)
    low = raw.lower()

    if "429" in raw or "resource_exhausted" in low or "quota" in low or "rate-limit" in low or "rate limit" in low:
        return (
            "Hệ thống AI đang quá tải hoặc đã đạt giới hạn xử lý tạm thời. "
            "Vui lòng thử lại sau ít phút. Nếu anh/chị cần sử dụng ổn định hơn, "
            "vui lòng nâng cấp Premium để được ưu tiên hỗ trợ và hạn chế gián đoạn."
        )

    if "api_key" in low or "invalid" in low or "permission" in low or "unauthorized" in low:
        return (
            "Hệ thống AI chưa được cấu hình đầy đủ hoặc khóa API chưa hợp lệ. "
            "Vui lòng liên hệ kỹ thuật để được kiểm tra và kích hoạt lại."
        )

    if "timeout" in low or "connection" in low or "network" in low:
        return (
            "Kết nối AI đang chậm hoặc mạng tạm thời không ổn định. "
            "Vui lòng thử lại sau ít phút."
        )

    return (
        "Hệ thống AI đang bận xử lý. Vui lòng thử lại sau ít phút "
        "hoặc liên hệ Zalo 036 338 2629 để được hỗ trợ nhanh."
    )

def safe_ai_generate(prompt, fallback=""):
    if not client:
        if fallback:
            return fallback
        return (
            "Hệ thống AI chưa được kích hoạt. Vui lòng kiểm tra GEMINI_API_KEY "
            "trong file .env hoặc liên hệ kỹ thuật để được hỗ trợ."
        )
    try:
        model = genai.GenerativeModel("gemini-1.5-flash")
        res = model.generate_content(prompt)
        return (res.text or "").strip()
    except Exception as e:
        if fallback:
            return fallback + "\\n\\n" + friendly_ai_error(e)
        return friendly_ai_error(e)

def generate_content(idea, style="gần gũi", length="120", variants=1):
    fallback = f"""
Nội dung gợi ý:

{idea}

Anh/chị có thể dùng nội dung này làm bản nháp, sau đó chỉnh lại theo sản phẩm/dịch vụ thực tế.

CTA:
Inbox để được tư vấn chi tiết.

Hashtag:
#Marketing #KinhDoanhOnline
""".strip()

    prompt = f"""
Viết {variants} bài đăng Facebook bằng tiếng Việt.

Chủ đề: {idea}
Phong cách: {style}
Độ dài mỗi bài: dưới {length} từ

Yêu cầu:
- Tự nhiên, dễ hiểu, phù hợp bán hàng/dịch vụ online
- Có emoji vừa phải
- Có lời kêu gọi inbox/liên hệ
- Không spam hashtag
- Không cam kết quá đà
- Không vi phạm chính sách Facebook
- Nếu viết nhiều phiên bản, đánh số Phiên bản 1, Phiên bản 2...
"""
    return safe_ai_generate(prompt, fallback)

def ai_plan_30_days(industry, goal):
    if not client:
        # fallback không tốn AI
        lines = []
        base = CONTENT_LIBRARY.get(industry, CONTENT_LIBRARY["spa"])
        for i in range(1, 31):
            lines.append(f"Ngày {i}: {random.choice(base)}")
        return "\n".join(lines)
    prompt = f"""
Tạo kế hoạch marketing Facebook 30 ngày bằng tiếng Việt.
Ngành: {industry}
Mục tiêu: {goal}

Yêu cầu:
- Mỗi ngày 1 ý tưởng bài đăng
- Có CTA ngắn
- Có hashtag gợi ý
- Trình bày theo Ngày 1 đến Ngày 30
- Phù hợp chủ shop/doanh nghiệp nhỏ
"""
    return safe_ai_generate(prompt, "\n".join(lines) if 'lines' in locals() else "")

def analyze_fanpage(name, avatar, cover, bio, post_frequency, cta):
    score = 50
    notes = []
    if avatar == "co": score += 10
    else: notes.append("Nên có ảnh đại diện rõ thương hiệu.")
    if cover == "co": score += 10
    else: notes.append("Nên có ảnh bìa thể hiện dịch vụ/sản phẩm chính.")
    if len(bio.strip()) > 50: score += 10
    else: notes.append("Phần mô tả Fanpage còn ngắn, nên bổ sung lợi ích và thông tin liên hệ.")
    if post_frequency in ["hang_ngay", "3_5_bai_tuan"]: score += 10
    else: notes.append("Tần suất đăng chưa đều, nên có lịch đăng cố định.")
    if cta == "co": score += 10
    else: notes.append("Nên có CTA rõ: Inbox, Zalo, Website hoặc đặt lịch.")
    score = min(score, 100)
    if not notes:
        notes.append("Fanpage đang có nền tảng tốt, nên tối ưu thêm nội dung và lịch đăng.")
    return score, notes

def spin_content_local(content):
    endings = [
        "Nhắn tin ngay để được tư vấn chi tiết hơn.",
        "Inbox để nhận thông tin và báo giá phù hợp.",
        "Liên hệ ngay hôm nay để được hỗ trợ nhanh.",
        "Để lại tin nhắn, đội ngũ tư vấn sẽ hỗ trợ bạn.",
        "Bạn cần thêm thông tin? Inbox ngay nhé."
    ]
    text = content.strip().replace("Inbox", "Nhắn tin").replace("Liên hệ", "Kết nối")
    return text + "\n\n" + random.choice(endings)

def post_to_facebook(page, content, image_path=""):
    if image_path and os.path.exists(image_path):
        ext = os.path.splitext(image_path)[1].lower()
        if ext in [".mp4", ".mov", ".m4v"]:
            url = f"https://graph-video.facebook.com/v23.0/{page['id']}/videos"
            with open(image_path, "rb") as video:
                res = requests.post(url, data={"description": content, "access_token": page["token"]},
                                    files={"source": video}, timeout=300)
            return res.json()
        url = f"https://graph.facebook.com/v23.0/{page['id']}/photos"
        with open(image_path, "rb") as img:
            res = requests.post(url, data={"message": content, "access_token": page["token"]},
                                files={"source": img}, timeout=90)
        return res.json()

    url = f"https://graph.facebook.com/v23.0/{page['id']}/feed"
    res = requests.post(url, data={"message": content, "access_token": page["token"]}, timeout=60)
    return res.json()


def save_uploads(file_list):
    paths = []
    for file_obj in file_list:
        if not file_obj or file_obj.filename == "":
            continue
        filename = secure_filename(file_obj.filename)
        path = os.path.join(UPLOAD_DIR, datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f_") + filename)
        file_obj.save(path)
        paths.append(path)
    return paths

def pick_media_for_job(media_paths, index):
    if not media_paths:
        return ""
    return media_paths[index % len(media_paths)]

def save_upload(file_obj):
    if not file_obj or file_obj.filename == "":
        return ""
    filename = secure_filename(file_obj.filename)
    path = os.path.join(UPLOAD_DIR, datetime.datetime.now().strftime("%Y%m%d_%H%M%S_") + filename)
    file_obj.save(path)
    return path

def scheduler_loop():
    while True:
        try:
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            conn = db(); c = conn.cursor()
            c.execute("""
            SELECT id,page_name,page_id,content,schedule_time,image_path
            FROM posts
            WHERE status='scheduled' AND schedule_time<=?
            """, (now,))
            jobs = c.fetchall(); conn.close()
            for row_id, page_name, page_id, content, schedule_time, image_path in jobs:
                page = next((p for p in get_pages_dynamic() if str(p["id"]) == str(page_id)), None)
                if not page:
                    update_post(row_id, "error", "Không tìm thấy Page trong .env")
                    continue
                result = post_to_facebook(page, content, image_path)
                if "id" in result or "post_id" in result:
                    update_post(row_id, "posted", result.get("post_id") or result.get("id"))
                else:
                    update_post(row_id, "error", str(result))
        except Exception as e:
            print("Scheduler error:", e)
        time.sleep(30)

def read_batch_file(file_obj):
    filename = file_obj.filename.lower()
    if filename.endswith(".csv"):
        text = file_obj.read().decode("utf-8-sig", errors="ignore").splitlines()
        return list(csv.DictReader(text))
    if filename.endswith(".xlsx"):
        if openpyxl is None:
            raise RuntimeError("Chưa cài openpyxl. Chạy: pip install openpyxl")
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        file_obj.save(tmp.name)
        wb = openpyxl.load_workbook(tmp.name)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for i, h in enumerate(headers):
                item[h] = r[i] if i < len(r) else ""
            rows.append(item)
        return rows
    raise RuntimeError("Chỉ hỗ trợ file .xlsx hoặc .csv")

def export_csv():
    rows = get_history(10000)
    path = os.path.join(REPORT_DIR, "report_posts.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["id","page_name","content","status","post_id","schedule_time","image_path","campaign","score","created_at"])
        writer.writerows(rows)
    return path


def add_crm(name, phone, zalo, source, note):
    if not name and not phone and not zalo:
        return
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
    INSERT INTO crm(name,phone,zalo,source,note,created_at)
    VALUES(?,?,?,?,?,?)
    """, (name, phone, zalo, source, note, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def get_crm(limit=30):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id,name,phone,zalo,source,note,created_at FROM crm ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall()
    conn.close()
    return rows


def save_token_check(page_name, page_id, status, detail):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
    INSERT INTO token_checks(page_name,page_id,status,detail,checked_at)
    VALUES(?,?,?,?,?)
    """, (page_name, page_id, status, detail, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def get_latest_token_checks(limit=30):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
    SELECT page_name,page_id,status,detail,checked_at
    FROM token_checks
    ORDER BY id DESC
    LIMIT ?
    """, (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

def check_single_page_token(page):
    page_name = page.get("name", "No name")
    page_id = str(page.get("id", ""))
    token = page.get("token", "")

    if not page_id or not token:
        return {"page_name": page_name, "page_id": page_id, "status": "LỖI", "detail": "Thiếu Page ID hoặc Page Token trong file .env."}

    try:
        url = f"https://graph.facebook.com/v23.0/{page_id}"
        params = {"fields": "id,name", "access_token": token}
        res = requests.get(url, params=params, timeout=20)
        data = res.json()

        if res.status_code < 400 and data.get("id"):
            return {"page_name": data.get("name", page_name), "page_id": data.get("id", page_id), "status": "SỐNG", "detail": "Token hoạt động bình thường."}

        err = data.get("error", {})
        code = err.get("code", "")
        subcode = err.get("error_subcode", "")
        message = err.get("message", str(data))
        if str(code) == "190":
            detail = f"Token giới hạn hoặc phiên đã giới hạn. OAuth code 190. Subcode: {subcode}. Nội dung: {message}"
        else:
            detail = f"Lỗi Facebook API. Code: {code}. Subcode: {subcode}. Nội dung: {message}"

        return {"page_name": page_name, "page_id": page_id, "status": "CHẾT", "detail": detail}

    except Exception as e:
        return {"page_name": page_name, "page_id": page_id, "status": "LỖI", "detail": str(e)}

def check_all_page_tokens():
    results = []
    for page in get_pages_dynamic():
        item = check_single_page_token(page)
        save_token_check(item["page_name"], item["page_id"], item["status"], item["detail"])
        results.append(item)
    return results

def add_page_cluster(name, page_names, note):
    if not name:
        return
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
    INSERT OR REPLACE INTO page_clusters(name,page_names,note,created_at)
    VALUES(?,?,?,?)
    """, (name, page_names, note, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

def get_page_clusters():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id,name,page_names,note,created_at FROM page_clusters ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def get_analytics_summary():
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    def one(query, params=()):
        try:
            c.execute(query, params)
            row = c.fetchone()
            return row[0] if row and row[0] is not None else 0
        except Exception:
            return 0

    def many(query, params=()):
        try:
            c.execute(query, params)
            return c.fetchall()
        except Exception:
            return []

    total_posts = one("SELECT COUNT(*) FROM posts")
    posted = one("SELECT COUNT(*) FROM posts WHERE status='posted'")
    scheduled = one("SELECT COUNT(*) FROM posts WHERE status='scheduled'")
    errors = one("SELECT COUNT(*) FROM posts WHERE status='error'")
    campaigns = one("SELECT COUNT(*) FROM campaigns")
    crm_total = one("SELECT COUNT(*) FROM crm")
    pipeline_total = one("SELECT COUNT(*) FROM lead_pipeline")
    groups_total = one("SELECT COUNT(*) FROM fb_groups")
    group_scheduled = one("SELECT COUNT(*) FROM group_schedules WHERE status='scheduled'")
    comments_total = one("SELECT COUNT(*) FROM comment_leads")
    messenger_total = one("SELECT COUNT(*) FROM messenger_scripts")

    by_page = many("""
        SELECT COALESCE(NULLIF(page_name,''),'Chưa đặt tên') AS name, COUNT(*) AS total
        FROM posts
        GROUP BY COALESCE(NULLIF(page_name,''),'Chưa đặt tên')
        ORDER BY total DESC
        LIMIT 10
    """)

    by_campaign = many("""
        SELECT COALESCE(NULLIF(campaign,''),'Chưa phân loại') AS name, COUNT(*) AS total
        FROM posts
        GROUP BY COALESCE(NULLIF(campaign,''),'Chưa phân loại')
        ORDER BY total DESC
        LIMIT 10
    """)

    by_day = many("""
        SELECT substr(created_at,1,10) AS day, COUNT(*) AS total
        FROM posts
        WHERE created_at IS NOT NULL AND created_at != ''
        GROUP BY substr(created_at,1,10)
        ORDER BY substr(created_at,1,10) DESC
        LIMIT 7
    """)

    by_status = many("""
        SELECT COALESCE(NULLIF(status,''),'unknown') AS status, COUNT(*) AS total
        FROM posts
        GROUP BY COALESCE(NULLIF(status,''),'unknown')
        ORDER BY total DESC
    """)

    group_by_niche = many("""
        SELECT COALESCE(NULLIF(niche,''),'Chưa phân loại') AS niche, COUNT(*) AS total
        FROM fb_groups
        GROUP BY COALESCE(NULLIF(niche,''),'Chưa phân loại')
        ORDER BY total DESC
        LIMIT 10
    """)

    crm_by_source = many("""
        SELECT COALESCE(NULLIF(source,''),'Chưa rõ nguồn') AS source, COUNT(*) AS total
        FROM crm
        GROUP BY COALESCE(NULLIF(source,''),'Chưa rõ nguồn')
        ORDER BY total DESC
        LIMIT 10
    """)

    crm_by_stage = many("""
        SELECT COALESCE(NULLIF(stage,''),'Khách mới') AS stage, COUNT(*) AS total
        FROM lead_pipeline
        GROUP BY COALESCE(NULLIF(stage,''),'Khách mới')
        ORDER BY total DESC
        LIMIT 10
    """)

    total_value = one("SELECT COALESCE(SUM(value),0) FROM lead_pipeline")
    avg_value = one("SELECT COALESCE(AVG(value),0) FROM lead_pipeline WHERE value > 0")
    conn.close()

    conversion_rate = round((posted / total_posts) * 100, 1) if total_posts else 0
    error_rate = round((errors / total_posts) * 100, 1) if total_posts else 0

    return {
        "summary": {
            "total_posts": total_posts,
            "posted": posted,
            "scheduled": scheduled,
            "errors": errors,
            "campaigns": campaigns,
            "crm_total": crm_total,
            "pipeline_total": pipeline_total,
            "groups_total": groups_total,
            "group_scheduled": group_scheduled,
            "comments_total": comments_total,
            "messenger_total": messenger_total,
            "total_value": int(total_value or 0),
            "avg_value": int(avg_value or 0),
            "conversion_rate": conversion_rate,
            "error_rate": error_rate
        },
        "by_page": by_page,
        "by_campaign": by_campaign,
        "by_day": by_day,
        "by_status": by_status,
        "group_by_niche": group_by_niche,
        "crm_by_source": crm_by_source,
        "crm_by_stage": crm_by_stage
    }

def generate_many_contents(idea, count, style="bán hàng tự nhiên", length="80"):
    count = max(1, min(int(count), 50))
    if not client:
        base = [
            f"{idea} - Giải pháp phù hợp cho khách hàng đang cần tối ưu công việc. Inbox để được tư vấn.",
            f"Bạn đang cần {idea}? Hãy chọn giải pháp ổn định, dễ dùng và được hỗ trợ nhanh.",
            f"{idea} giúp công việc online thuận tiện hơn. Nhắn tin để nhận tư vấn chi tiết."
        ]
        return "\\n\\n".join(base[i % len(base)] for i in range(count))

    prompt = f"""
Viết {count} content Facebook bằng tiếng Việt.
Chủ đề: {idea}
Phong cách: {style}
Mỗi bài dưới {length} từ.

Yêu cầu:
- Mỗi content khác nhau
- Có CTA inbox/liên hệ
- Có emoji vừa phải
- Không cam kết quá đà
- Tách mỗi content bằng dòng trống
"""
    return safe_ai_generate(prompt, "\n\n".join(base[i % len(base)] for i in range(count)) if 'base' in locals() else "")

def smart_schedule_times(start_time, count, gap_minutes):
    times = []
    try:
        base = datetime.datetime.strptime(start_time.replace("T", " "), "%Y-%m-%d %H:%M")
    except Exception:
        base = datetime.datetime.now() + datetime.timedelta(minutes=10)
    gap = int(gap_minutes or 60)
    for i in range(count):
        times.append((base + datetime.timedelta(minutes=i * gap)).strftime("%Y-%m-%d %H:%M"))
    return times


def ai_spin_content(content):
    if not content:
        return ""
    if client:
        prompt = f"""
Viết lại content sau khác khoảng 70-80% nhưng giữ ý chính, giọng tự nhiên, không cam kết quá đà.
Thêm CTA riêng và hashtag riêng.
Content gốc:
{content}
"""
        try:
            return safe_ai_generate(prompt, content)
        except Exception as e:
            return friendly_ai_error(e)
    endings = [
        "Inbox để được tư vấn chi tiết và nhận báo giá phù hợp.",
        "Nhắn tin ngay để được hỗ trợ nhanh nhất hôm nay.",
        "Liên hệ để nhận thông tin mới và ưu đãi phù hợp.",
        "Để lại tin nhắn, đội ngũ tư vấn sẽ hỗ trợ bạn.",
    ]
    return content.strip().replace("Inbox", "Nhắn tin").replace("Liên hệ", "Kết nối") + "\n\n" + random.choice(endings) + "\n#Marketing #KinhDoanhOnline"

def auto_cta_hashtag(content, industry="marketing"):
    text = content.lower()
    cta = "Inbox để được tư vấn nhanh."
    if "zalo" in text:
        cta = "Kết nối Zalo để được hỗ trợ chi tiết."
    elif "đặt lịch" in text or "spa" in text or "nha khoa" in text:
        cta = "Đặt lịch tư vấn ngay hôm nay."
    elif "proxy" in text:
        cta = "Nhắn tin để nhận bảng giá proxy phù hợp."
    hashtags = f"#{industry.replace('_','')} #Marketing #KinhDoanhOnline"
    return cta, hashtags

def choose_best_media_for_content(content, media_paths, used_indexes):
    if not media_paths:
        return ""
    # Bản local: ưu tiên không trùng ảnh; nếu hết ảnh thì quay vòng.
    for i, path in enumerate(media_paths):
        if i not in used_indexes:
            used_indexes.add(i)
            return path
    idx = len(used_indexes) % len(media_paths)
    return media_paths[idx]

def backup_database():
    if not os.path.exists(DB):
        return ""
    name = "backup_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + DB
    path = os.path.join(BACKUP_DIR, name)
    shutil.copy2(DB, path)
    return path

def export_pdf_report():
    # PDF tối giản dạng text để tránh phụ thuộc thư viện ngoài.
    rows = get_history(500)
    path = os.path.join(REPORT_DIR, "report_posts.pdf")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        c = canvas.Canvas(path, pagesize=A4)
        width, height = A4
        y = height - 40
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, "Marketing Automation Pro V9 - Report")
        y -= 30
        c.setFont("Helvetica", 9)
        for r in rows[:80]:
            line = f"ID {r[0]} | Page: {r[1]} | Status: {r[3]} | Campaign: {r[7]} | Time: {r[9]}"
            c.drawString(40, y, line[:110])
            y -= 14
            if y < 40:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica", 9)
        c.save()
        return path
    except Exception:
        txt_path = os.path.join(REPORT_DIR, "report_posts_pdf_fallback.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write("Marketing Automation Pro V9 Report\n\n")
            for r in rows:
                f.write(f"ID {r[0]} | Page: {r[1]} | Status: {r[3]} | Campaign: {r[7]} | Time: {r[9]}\n")
        return txt_path

def premium_visible_items(items, free_limit=10):
    return items[:free_limit], max(0, len(items) - free_limit)


def get_trial_identity():
    try:
        raw = request.cookies.get("mkt_trial_user") or request.remote_addr or "local_user"
    except Exception:
        raw = "local_user"
    return str(raw).replace(" ", "_")[:80]


def get_free_status(username=None):
    """Trial thật 3 ngày.
    Gói dùng thử chỉ mở: Quản lý Fanpage, Quản lý Group, AI Comment.
    Các tính năng còn lại sẽ chuyển sang popup Premium.
    """
    username = username or get_trial_identity()
    now = datetime.datetime.now()

    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS users_trial (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        package_name TEXT DEFAULT 'trial',
        trial_start_date TEXT,
        trial_end_date TEXT,
        status TEXT DEFAULT 'active',
        created_at TEXT
    )
    """)
    c.execute("SELECT username,package_name,trial_start_date,trial_end_date,status FROM users_trial WHERE username=?", (username,))
    row = c.fetchone()

    if not row:
        start = now
        end = now + datetime.timedelta(days=3)
        c.execute("""
        INSERT INTO users_trial(username,package_name,trial_start_date,trial_end_date,status,created_at)
        VALUES(?,?,?,?,?,?)
        """, (username, "trial", start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S"), "active", now.strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        row = (username, "trial", start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S"), "active")

    package_name = row[1] or "trial"
    trial_start = row[2] or now.strftime("%Y-%m-%d %H:%M:%S")
    trial_end = row[3] or (now + datetime.timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")
    status = row[4] or "active"

    try:
        end_dt = datetime.datetime.strptime(trial_end, "%Y-%m-%d %H:%M:%S")
    except Exception:
        end_dt = now

    remaining = end_dt - now
    total_seconds = int(remaining.total_seconds())
    if package_name == "trial" and total_seconds <= 0:
        status = "expired"
        c.execute("UPDATE users_trial SET status=? WHERE username=?", ("expired", username))
        conn.commit()

    conn.close()

    days = max(0, total_seconds // 86400)
    hours = max(0, (total_seconds % 86400) // 3600)
    percent = max(0, min(100, int((total_seconds / (3 * 86400)) * 100))) if package_name == "trial" else 100

    allowed_features = ["Quản lý Fanpage", "Quản lý Group", "AI Comment"]
    locked_features = ["AI Messenger", "CRM Kanban", "AI Marketing Director", "AI Video", "AI Image", "AI Kinh Doanh", "AI Giọng Nói", "AI Livestream"]

    return {
        "package_name": package_name,
        "status": status,
        "trial_start": trial_start,
        "trial_end": trial_end,
        "free_end": trial_end,
        "days": days,
        "hours": hours,
        "percent": percent,
        "is_trial": package_name == "trial",
        "is_expired": status == "expired",
        "label": "🎁 Dùng thử miễn phí 3 ngày" if status != "expired" else "🔒 Dùng thử đã hết hạn",
        "note": "Dùng thử chỉ mở: Quản lý Fanpage • Quản lý Group • AI Comment",
        "allowed_features": allowed_features,
        "locked_features": locked_features
    }


PREMIUM_PACKAGES = {
    "monthly": {"name": "Gói 1 tháng", "price": "159.000đ", "amount": 159000, "days": 30},
    "quarterly": {"name": "Gói 3 tháng", "price": "359.000đ", "amount": 359000, "days": 90},
    "halfyear": {"name": "Gói 6 tháng", "price": "559.000đ", "amount": 559000, "days": 180},
    "yearly": {"name": "Gói 1 năm", "price": "859.000đ", "amount": 859000, "days": 365},
    "sellerpro": {"name": "Gói nhà bán hàng chuyên nghiệp", "price": "1.959.000đ", "amount": 1959000, "days": 3650},
    "lifetime": {"name": "Gói nhà bán hàng chuyên nghiệp", "price": "1.959.000đ", "amount": 1959000, "days": 3650}
}


def get_device_id():
    """Lấy ID thiết bị ổn định theo cookie; JS sẽ tạo cookie nếu máy mới."""
    try:
        raw = request.cookies.get("mkt_device_id") or request.cookies.get("mkt_trial_user") or request.remote_addr or "LOCAL"
    except Exception:
        raw = "LOCAL"
    raw = str(raw).strip().replace(" ", "-").upper()
    if not raw.startswith("MKT-"):
        raw = "MKT-" + ''.join(ch for ch in raw if ch.isalnum())[:10]
    return raw[:32]


def normalize_package_key(package_key):
    package_key = (package_key or "monthly").strip().lower()
    aliases = {"basic":"monthly", "pro":"quarterly", "business":"halfyear", "lifetime":"sellerpro", "nhabanhangpro":"sellerpro", "seller":"sellerpro", "seller_pro":"sellerpro"}
    return aliases.get(package_key, package_key if package_key in PREMIUM_PACKAGES else "monthly")


def create_premium_request(device_id, phone, email, package_key):
    ensure_affiliate_tables()
    package_key = normalize_package_key(package_key)
    plan = PREMIUM_PACKAGES.get(package_key, PREMIUM_PACKAGES["monthly"])
    device_id = (device_id or get_device_id()).strip().upper()
    phone = (phone or "").strip()
    email = (email or "").strip()
    payment_note = f"{device_id}"
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS premium_upgrade_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT, phone TEXT, email TEXT, package_key TEXT, package_name TEXT,
            amount INTEGER DEFAULT 0, payment_note TEXT, status TEXT DEFAULT 'Chờ duyệt',
            admin_note TEXT, created_at TEXT, approved_at TEXT
        )
    """)
    c.execute("""
        INSERT INTO premium_upgrade_requests(device_id,phone,email,package_key,package_name,amount,payment_note,status,admin_note,created_at,approved_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?)
    """, (device_id, phone, email, package_key, plan['name'], int(plan.get('amount',0)), payment_note, 'Chờ duyệt', '', now, ''))
    req_id = c.lastrowid
    affiliate_code = current_affiliate_code()
    if affiliate_code:
        try:
            c.execute("UPDATE premium_upgrade_requests SET affiliate_code=? WHERE id=?", (affiliate_code, req_id))
        except Exception:
            pass
    c.execute("""INSERT INTO notifications(title,detail,level,status,created_at) VALUES(?,?,?,?,?)""",
              ("Yêu cầu nâng cấp Premium mới", f"{device_id} - {phone} - {email} - {plan['name']}" + (f" - CTV: {affiliate_code}" if affiliate_code else ""), "warning", "new", now))
    conn.commit(); conn.close()
    return {"id": req_id, "device_id": device_id, "payment_note": payment_note, "package_name": plan['name'], "amount": int(plan.get('amount',0))}


def get_premium_requests(limit=100):
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS premium_upgrade_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT, phone TEXT, email TEXT, package_key TEXT, package_name TEXT,
            amount INTEGER DEFAULT 0, payment_note TEXT, status TEXT DEFAULT 'Chờ duyệt',
            admin_note TEXT, created_at TEXT, approved_at TEXT
        )
    """)
    c.execute("""SELECT id,device_id,phone,email,package_key,package_name,amount,payment_note,status,admin_note,created_at,approved_at
                 FROM premium_upgrade_requests ORDER BY id DESC LIMIT ?""", (limit,))
    rows = c.fetchall(); conn.close(); return rows


def approve_premium_request(request_id, status='Đã duyệt', admin_note=''):
    now_dt = datetime.datetime.now()
    now = now_dt.strftime("%Y-%m-%d %H:%M:%S")
    conn = db(); c = conn.cursor()
    c.execute("SELECT device_id,phone,email,package_key,package_name FROM premium_upgrade_requests WHERE id=?", (request_id,))
    row = c.fetchone()
    if not row:
        conn.close(); return False
    device_id, phone, email, package_key, package_name = row
    package_key = normalize_package_key(package_key)
    plan = PREMIUM_PACKAGES.get(package_key, PREMIUM_PACKAGES['monthly'])
    if status == 'Đã duyệt':
        end_dt = now_dt + datetime.timedelta(days=int(plan.get('days', 30)))
        c.execute("""
            CREATE TABLE IF NOT EXISTS device_subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                device_id TEXT UNIQUE, phone TEXT, email TEXT, package_key TEXT, package_name TEXT,
                start_date TEXT, end_date TEXT, status TEXT DEFAULT 'premium', last_renewal_notice_at TEXT,
                created_at TEXT, updated_at TEXT
            )
        """)
        c.execute("""
            INSERT INTO device_subscriptions(device_id,phone,email,package_key,package_name,start_date,end_date,status,last_renewal_notice_at,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(device_id) DO UPDATE SET
                phone=excluded.phone,email=excluded.email,package_key=excluded.package_key,package_name=excluded.package_name,
                start_date=excluded.start_date,end_date=excluded.end_date,status='premium',updated_at=excluded.updated_at
        """, (device_id, phone, email, package_key, package_name, now, end_dt.strftime("%Y-%m-%d %H:%M:%S"), 'premium', '', now, now))
        c.execute("""INSERT INTO notifications(title,detail,level,status,created_at) VALUES(?,?,?,?,?)""",
                  ("Đã kích hoạt Premium", f"{device_id} - {package_name} - hạn đến {end_dt.strftime('%Y-%m-%d')}", "success", "new", now))
    c.execute("UPDATE premium_upgrade_requests SET status=?, admin_note=?, approved_at=? WHERE id=?", (status, admin_note, now, request_id))
    if status == 'Đã duyệt':
        try:
            create_affiliate_commission_for_request(c, request_id)
        except Exception as _affiliate_err:
            print('Affiliate commission error:', _affiliate_err)
    conn.commit(); conn.close(); return True


def get_device_subscription(device_id=None):
    device_id = (device_id or get_device_id()).strip().upper()
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS device_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT UNIQUE, phone TEXT, email TEXT, package_key TEXT, package_name TEXT,
            start_date TEXT, end_date TEXT, status TEXT DEFAULT 'premium', last_renewal_notice_at TEXT,
            created_at TEXT, updated_at TEXT
        )
    """)
    c.execute("SELECT device_id,phone,email,package_key,package_name,start_date,end_date,status,last_renewal_notice_at FROM device_subscriptions WHERE device_id=?", (device_id,))
    row = c.fetchone(); conn.close()
    return row


def get_renewal_notice(device_id=None):
    row = get_device_subscription(device_id)
    if not row:
        return None
    device_id, phone, email, package_key, package_name, start_date, end_date, status, last_notice = row
    try:
        end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None
    remaining_days = (end_dt - datetime.datetime.now()).days
    if status == 'premium' and 0 <= remaining_days <= 5:
        return {"device_id": device_id, "package_name": package_name, "end_date": end_date, "remaining_days": remaining_days}
    return None



def plan_required_message(feature_name, plans):
    return (
        f"Tính năng nâng cao: {feature_name}\n\n"
        f"Công cụ này thuộc nhóm Premium.\n"
        f"Gói đề xuất: {plans}\n\n"
        "Anh/chị có thể nâng cấp để mở khóa đầy đủ tính năng, hạn mức cao hơn và hỗ trợ ưu tiên."
    )

def token_manager_report():
    reports = []
    for p in get_pages_dynamic():
        token = p.get("token", "")
        status = "Có token" if token and token.startswith("EA") else "Thiếu token hoặc sai định dạng"
        reports.append(f"{p.get('name','No name')} | {p.get('id','No ID')} | {status}")
    if not reports:
        reports.append("Chưa có Fanpage trong PAGES_JSON.")
    return "\\n".join(reports)

def ai_planner_v6(industry, goal, days):
    if not client:
        base = CONTENT_LIBRARY.get(industry, list(CONTENT_LIBRARY.values())[0])
        return "\\n".join([f"Ngày {i}: {random.choice(base)}\\nCTA: Inbox để được tư vấn.\\nHashtag: #{industry} #Marketing" for i in range(1, int(days)+1)])
    prompt = f"""
Tạo kế hoạch marketing Facebook {days} ngày bằng tiếng Việt.
Ngành: {industry}
Mục tiêu: {goal}

Mỗi ngày gồm:
- Chủ đề bài đăng
- Nội dung ngắn
- CTA
- Hashtag
- Gợi ý khung giờ đăng
Trình bày theo Ngày 1 đến Ngày {days}.
"""
    return safe_ai_generate(prompt, "")




def add_fb_group(group_name, group_id, niche, note):
    if not group_name and not group_id:
        return
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT INTO fb_groups(group_name,group_id,niche,note,status,created_at)
    VALUES(?,?,?,?,?,?)
    """, (group_name, group_id, niche, note, 'active', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_fb_groups(limit=80):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,group_name,group_id,niche,note,status,created_at FROM fb_groups ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def add_group_schedule(group_name, group_id, content, schedule_time):
    if not content:
        return
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT INTO group_schedules(group_name,group_id,content,schedule_time,status,created_at)
    VALUES(?,?,?,?,?,?)
    """, (group_name, group_id, content, schedule_time, 'scheduled', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_group_schedules(limit=50):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,group_name,group_id,content,schedule_time,status,created_at FROM group_schedules ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def hide_phone_preview(text):
    import re
    return re.sub(r'(0|\+84)[0-9\s\.\-]{7,13}', '[ĐÃ ẨN SĐT]', text or '')

def add_comment_lead(customer_name, phone, comment_text, ai_reply, label):
    if not comment_text and not phone and not customer_name:
        return
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT INTO comment_leads(customer_name,phone,comment_text,ai_reply,label,crm_status,created_at)
    VALUES(?,?,?,?,?,?,?)
    """, (customer_name, phone, comment_text, ai_reply, label, 'new', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_comment_leads(limit=50):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,customer_name,phone,comment_text,ai_reply,label,crm_status,created_at FROM comment_leads ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def add_messenger_script(product, script_type, content):
    if not product and not content:
        return
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO messenger_scripts(product,script_type,content,created_at) VALUES(?,?,?,?)", (product, script_type, content, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_messenger_scripts(limit=30):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,product,script_type,content,created_at FROM messenger_scripts ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def v5_seed_success_assets():
    conn = db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM success_assets")
    if c.fetchone()[0] == 0:
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        samples = [
            ('Case Study','Shop Proxy tăng phản hồi inbox','Trước: đăng bài rời rạc, thiếu lịch. Sau: dùng AI tạo 30 ngày content, 10 ads, 5 kịch bản chốt sale và CRM theo dõi khách.'),
            ('Fanpage','Mẫu Fanpage dịch vụ chuyên nghiệp','Avatar rõ thương hiệu, cover nêu lợi ích, mô tả có Zalo/CTA, ghim bài giới thiệu dịch vụ và bảng giá.'),
            ('Content','Mẫu content thành công','Nêu nỗi đau, đưa giải pháp, bằng chứng mềm, CTA inbox và hashtag vừa đủ.'),
            ('Ads','Mẫu quảng cáo thành công','Hook mạnh, lợi ích rõ, tránh cam kết quá đà, kêu gọi inbox nhận tư vấn.'),
            ('Script','Mẫu kịch bản chốt sale','Chào hỏi - hỏi nhu cầu - tư vấn gói phù hợp - xử lý từ chối - chốt hành động tiếp theo.')
        ]
        c.executemany("INSERT INTO success_assets(asset_type,title,content,created_at) VALUES(?,?,?,?)", [(a,b,cnt,now) for a,b,cnt in samples])
        conn.commit()
    conn.close()

def get_success_assets(limit=20):
    v5_seed_success_assets()
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,asset_type,title,content,created_at FROM success_assets ORDER BY id ASC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows



def normalize_uid(raw):
    raw = str(raw or '').strip()
    return ''.join(ch for ch in raw if ch.isdigit()) or raw

def add_group_finder_result(keyword, group_name, group_uid, members=0, privacy='Công khai', recent_activity='Có', page_join_allowed='Chưa rõ', page_post_allowed='Chưa rõ', status='Hợp lệ', note=''):
    group_uid = normalize_uid(group_uid)
    if not group_uid:
        return False
    try:
        members = int(str(members or 0).replace('.', '').replace(',', '').strip() or 0)
    except Exception:
        members = 0
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT OR IGNORE INTO group_finder_results(keyword,group_name,group_uid,members,privacy,recent_activity,page_join_allowed,page_post_allowed,status,note,created_at)
    VALUES(?,?,?,?,?,?,?,?,?,?,?)
    """, (keyword, group_name or f'Group {group_uid}', group_uid, members, privacy, recent_activity, page_join_allowed, page_post_allowed, status, note, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    ok = c.rowcount > 0
    conn.commit(); conn.close(); return ok

def get_group_finder_results(limit=300):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,keyword,group_name,group_uid,members,privacy,recent_activity,page_join_allowed,page_post_allowed,status,note,created_at FROM group_finder_results ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def get_group_finder_stats():
    conn = db(); c = conn.cursor()
    def one(q):
        c.execute(q); r=c.fetchone(); return r[0] if r else 0
    total = one('SELECT COUNT(*) FROM group_finder_results')
    valid = one("SELECT COUNT(*) FROM group_finder_results WHERE status='Hợp lệ'")
    queue = one('SELECT COUNT(*) FROM group_join_queue')
    files = one('SELECT COUNT(*) FROM group_uid_files')
    posts = one('SELECT COUNT(*) FROM group_post_results')
    conn.close(); return {'total':total,'valid':valid,'queue':queue,'files':files,'posts':posts}

def group_finder_import_text(keyword, raw_text, min_members=0, privacy_filter='all', recent_only='0', page_join='all', page_post='all'):
    added = dup = rejected = 0
    keyword = keyword or 'Từ khóa thủ công'
    lines = [x.strip() for x in (raw_text or '').splitlines() if x.strip()]
    if not lines and keyword:
        # Bản demo an toàn: không tự quét Facebook trái phép; tạo danh sách nháp để admin kiểm tra quyền hợp lệ.
        for i in range(1, 21):
            uid = f"{abs(hash(keyword)) % 1000000}{i:04d}"
            members = max(int(min_members or 0), 1000) + i * 137
            lines.append(f"{uid}, {keyword.title()} Community {i}, {members}, Công khai, Có, Có, Chưa rõ")
    for line in lines:
        parts = [p.strip() for p in line.replace('\t', ',').split(',')]
        uid = normalize_uid(parts[0] if parts else '')
        name = parts[1] if len(parts) > 1 else f'Group {uid}'
        members = parts[2] if len(parts) > 2 else 0
        privacy = parts[3] if len(parts) > 3 else 'Công khai'
        recent = parts[4] if len(parts) > 4 else 'Có'
        join_allowed = parts[5] if len(parts) > 5 else 'Chưa rõ'
        post_allowed = parts[6] if len(parts) > 6 else 'Chưa rõ'
        try: m_int = int(str(members).replace('.', '').replace(',', '') or 0)
        except Exception: m_int = 0
        if m_int < int(min_members or 0):
            rejected += 1; continue
        if privacy_filter != 'all' and privacy_filter and privacy != privacy_filter:
            rejected += 1; continue
        if recent_only == '1' and recent.lower() not in ['có','co','yes','active','gần đây']:
            rejected += 1; continue
        if page_join != 'all' and join_allowed != page_join:
            rejected += 1; continue
        if page_post != 'all' and post_allowed != page_post:
            rejected += 1; continue
        if add_group_finder_result(keyword, name, uid, m_int, privacy, recent, join_allowed, post_allowed):
            added += 1
        else:
            dup += 1
    return {'added':added,'duplicate':dup,'rejected':rejected,'total_input':len(lines)}

def add_group_queue_from_results(page_index=''):
    selected_page = None
    try:
        i = int(page_index)
        selected_page = get_page_by_index(i)
    except Exception:
        selected_page = None
    page_id = str(selected_page.get('id','')) if selected_page else ''
    page_name = selected_page.get('name','Chưa chọn Page') if selected_page else 'Chưa chọn Page'
    rows = get_group_finder_results(500)
    conn = db(); c = conn.cursor(); added = 0
    for r in rows:
        c.execute("SELECT COUNT(*) FROM group_join_queue WHERE group_uid=?", (r[3],))
        if c.fetchone()[0]: continue
        c.execute("""INSERT INTO group_join_queue(group_uid,group_name,keyword,members,selected_page_id,selected_page_name,status,admin_status,note,created_at)
                  VALUES(?,?,?,?,?,?,?,?,?,?)""", (r[3], r[2], r[1], r[4], page_id, page_name, 'Chưa tham gia', 'Chờ admin duyệt', 'Chỉ tham gia thủ công/có giới hạn sau khi admin duyệt.', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        added += 1
    conn.commit(); conn.close(); return added

def get_group_join_queue(limit=200):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,group_uid,group_name,keyword,members,selected_page_name,status,admin_status,note,created_at FROM group_join_queue ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def split_group_uids(chunk_size=50):
    chunk_size = max(1, min(int(chunk_size or 50), 1000))
    rows = get_group_finder_results(10000)
    uids = [r[3] for r in rows if r[3]]
    paths=[]
    for idx in range(0, len(uids), chunk_size):
        chunk = uids[idx:idx+chunk_size]
        file_no = len(paths)+1
        name = f"group_uid_file_{file_no:03d}.csv"
        path = os.path.join(REPORT_DIR, name)
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w=csv.writer(f); w.writerow(['group_uid']); [[w.writerow([u])] for u in chunk]
        conn=db(); c=conn.cursor(); c.execute("INSERT INTO group_uid_files(file_name,uid_count,chunk_size,file_path,created_at) VALUES(?,?,?,?,?)", (name,len(chunk),chunk_size,path,datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))); conn.commit(); conn.close()
        paths.append(path)
    return paths

def get_group_uid_files(limit=50):
    conn=db(); c=conn.cursor(); c.execute("SELECT id,file_name,uid_count,chunk_size,file_path,created_at FROM group_uid_files ORDER BY id DESC LIMIT ?", (limit,)); rows=c.fetchall(); conn.close(); return rows

def add_group_post_result(group_uid, post_uid, post_link, author_name, posted_at, content_preview, comments=0, reactions=0, keyword=''):
    group_uid=normalize_uid(group_uid); post_uid=normalize_uid(post_uid)
    if not group_uid or not post_uid: return False
    conn=db(); c=conn.cursor()
    c.execute("""INSERT OR IGNORE INTO group_post_results(group_uid,post_uid,post_link,author_name,posted_at,content_preview,comments,reactions,status,keyword,created_at)
              VALUES(?,?,?,?,?,?,?,?,?,?,?)""", (group_uid,post_uid,post_link,author_name,posted_at,content_preview[:280],int(comments or 0),int(reactions or 0),'Chưa xử lý',keyword,datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    ok=c.rowcount>0; conn.commit(); conn.close(); return ok

def import_group_posts(raw_text, keyword='', min_comments=0, min_reactions=0):
    added=dup=filtered=0
    for line in [x.strip() for x in (raw_text or '').splitlines() if x.strip()]:
        p=[x.strip() for x in line.replace('\t', ',').split(',')]
        while len(p)<8: p.append('')
        try: comments=int(p[6] or 0); reactions=int(p[7] or 0)
        except Exception: comments=reactions=0
        if comments < int(min_comments or 0) or reactions < int(min_reactions or 0): filtered+=1; continue
        if keyword and keyword.lower() not in ' '.join(p).lower(): filtered+=1; continue
        if add_group_post_result(p[0],p[1],p[2],p[3],p[4],p[5],comments,reactions,keyword): added+=1
        else: dup+=1
    return {'added':added,'duplicate':dup,'filtered':filtered}

def get_group_post_results(limit=200):
    conn=db(); c=conn.cursor(); c.execute("SELECT id,group_uid,post_uid,post_link,author_name,posted_at,content_preview,comments,reactions,status,keyword,created_at FROM group_post_results ORDER BY id DESC LIMIT ?", (limit,)); rows=c.fetchall(); conn.close(); return rows

def add_group_post_queue(page_index, group_ids, content):
    selected_page=None
    try:
        i=int(page_index); selected_page=get_page_by_index(i)
    except Exception: selected_page=None
    page_id=str(selected_page.get('id','')) if selected_page else ''
    page_name=selected_page.get('name','Chưa chọn Page') if selected_page else 'Chưa chọn Page'
    conn=db(); c=conn.cursor(); added=0
    for gid in group_ids:
        c.execute("SELECT group_name FROM fb_groups WHERE group_id=? LIMIT 1", (gid,)); row=c.fetchone()
        gname=row[0] if row else gid
        c.execute("INSERT INTO group_post_queue(page_id,page_name,group_uid,group_name,content,status,note,created_at) VALUES(?,?,?,?,?,?,?,?)", (page_id,page_name,gid,gname,content,'Chờ duyệt','Chỉ đăng khi Page/tài khoản có quyền hợp lệ và admin duyệt.',datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        added+=1
    conn.commit(); conn.close(); return added


def add_group_pair_post_queue(page_indexes, group_ids, contents, schedule_mode="now", min_delay=45, max_delay=90):
    page_indexes = list(page_indexes or [])
    # khử trùng group trong đúng thứ tự người dùng chọn
    unique_group_ids = []
    seen = set()
    for gid in group_ids or []:
        gid = normalize_uid(gid)
        if gid and gid not in seen:
            seen.add(gid); unique_group_ids.append(gid)
    contents = [c.strip() for c in (contents or []) if c.strip()]
    if not page_indexes or not unique_group_ids or not contents:
        return {"added": 0, "skipped": 0, "pairs": []}
    try:
        min_delay = normalize_delay_seconds(min_delay, 45)
        max_delay = normalize_delay_seconds(max_delay, 60)
    except Exception:
        min_delay, max_delay = 45, 90
    count = min(len(page_indexes), len(unique_group_ids))
    base_time = None
    if str(schedule_mode).isdigit():
        base_time = datetime.datetime.now() + datetime.timedelta(minutes=int(schedule_mode))
    conn = db(); c = conn.cursor(); added = skipped = 0; pairs = []
    for idx in range(count):
        page_index = page_indexes[idx]
        try:
            pi = int(page_index); selected_page = get_page_by_index(pi)
        except Exception:
            selected_page = None
        if not selected_page:
            skipped += 1; continue
        gid = unique_group_ids[idx]
        c.execute("SELECT group_name FROM fb_groups WHERE group_id=? LIMIT 1", (gid,)); row = c.fetchone()
        gname = row[0] if row else gid
        content = contents[idx % len(contents)]
        if not page_group_can_post(selected_page.get('id',''), gid):
            skipped += 1
            continue
        schedule_note = "Đăng ngay sau khi admin duyệt" if not base_time else "Hẹn giờ: " + (base_time + datetime.timedelta(minutes=idx * 5)).strftime("%Y-%m-%d %H:%M")
        note = f"Ghép riêng Page đã tham gia Group → Group riêng, không trùng Group. {schedule_note}. Giãn cách {min_delay}-{max_delay} giây. Chỉ xử lý khi có quyền hợp lệ."
        c.execute("""INSERT INTO group_post_queue(page_id,page_name,group_uid,group_name,content,status,note,created_at) VALUES(?,?,?,?,?,?,?,?)""", (str(selected_page.get('id','')), selected_page.get('name','Chưa chọn Page'), gid, gname, content, 'Chờ duyệt', note, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        added += 1
        pairs.append(f"{selected_page.get('name')} → {gname}")
    conn.commit(); conn.close()
    return {"added": added, "skipped": skipped, "pairs": pairs}

def get_group_post_queue(limit=100):
    conn=db(); c=conn.cursor(); c.execute("SELECT id,page_name,group_name,group_uid,content,status,note,created_at FROM group_post_queue ORDER BY id DESC LIMIT ?", (limit,)); rows=c.fetchall(); conn.close(); return rows

def normalize_delay_seconds(value, default=45):
    try:
        v = int(value or default)
    except Exception:
        v = default
    return max(45, min(v, 3600))

def normalize_comment_delay_seconds(value, default=45):
    """Giới hạn riêng cho bình luận Page: tối thiểu 45s, tối đa 60s.
    Dùng cho hàng chờ bình luận UID người dùng / UID bài viết / UID Group để tránh cấu hình quá nhanh hoặc quá rộng.
    """
    try:
        v = int(value or default)
    except Exception:
        v = default
    return max(45, min(v, 60))

def add_page_comment_queue(page_index, target_type, user_uid, post_uid, group_uid, comment_text, min_delay=45, max_delay=60, scheduled_at=''):
    if not comment_text or (not post_uid and not group_uid and not user_uid):
        return 0
    selected_page = None
    try:
        i = int(page_index)
        selected_page = get_page_by_index(i)
    except Exception:
        selected_page = None
    page_id = str(selected_page.get('id','')) if selected_page else ''
    page_name = selected_page.get('name','Chưa chọn Page') if selected_page else 'Chưa chọn Page'
    target_type = target_type or 'post'
    min_delay = normalize_comment_delay_seconds(min_delay, 45)
    max_delay = normalize_comment_delay_seconds(max_delay, 60)
    if max_delay < min_delay:
        max_delay = min_delay
    now = datetime.datetime.now()
    if not scheduled_at:
        scheduled_at = (now + datetime.timedelta(seconds=min_delay)).strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT INTO page_comment_queue(page_id,page_name,target_type,user_uid,post_uid,group_uid,comment_text,min_delay_seconds,max_delay_seconds,scheduled_at,status,admin_status,result_message,created_at,processed_at)
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (page_id, page_name, target_type, normalize_uid(user_uid), normalize_uid(post_uid), normalize_uid(group_uid), comment_text.strip(), min_delay, max_delay, scheduled_at, 'Chờ duyệt', 'Chờ admin duyệt', 'Chưa xử lý - cần admin duyệt trước khi bình luận.', now.strftime('%Y-%m-%d %H:%M:%S'), ''))
    qid = c.lastrowid
    c.execute("INSERT INTO page_comment_logs(queue_id,page_name,target_uid,action,status,detail,created_at) VALUES(?,?,?,?,?,?,?)", (qid, page_name, normalize_uid(post_uid) or normalize_uid(group_uid) or normalize_uid(user_uid), 'Tạo hàng chờ', 'Chờ duyệt', 'Đã tạo hàng chờ bình luận an toàn.', now.strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close(); return qid

def split_comment_lines(comment_text):
    """Tách nội dung bình luận: mỗi dòng là một bình luận riêng.
    Khi số UID nhiều hơn số bình luận, hệ thống tự quay vòng nội dung theo thứ tự.
    """
    return [line.strip() for line in (comment_text or '').splitlines() if line.strip()]

def bulk_import_page_comment_queue(page_index, raw_targets, comment_text, min_delay=45, max_delay=60, target_type='post'):
    lines = [x.strip() for x in (raw_targets or '').splitlines() if x.strip()]
    comments = split_comment_lines(comment_text)
    min_delay = normalize_comment_delay_seconds(min_delay, 45)
    max_delay = normalize_comment_delay_seconds(max_delay, 60)
    if max_delay < min_delay: max_delay = min_delay
    added = skipped = 0
    base = datetime.datetime.now() + datetime.timedelta(seconds=min_delay)
    for idx, line in enumerate(lines):
        parts = [x.strip() for x in line.replace('\t', ',').split(',')]
        uid1 = normalize_uid(parts[0] if parts else '')
        uid2 = normalize_uid(parts[1] if len(parts) > 1 else '')
        uid3 = normalize_uid(parts[2] if len(parts) > 2 else '')
        if not uid1 or not comments:
            skipped += 1; continue
        user_uid = post_uid = group_uid = ''
        if target_type == 'user': user_uid = uid1
        elif target_type == 'group': group_uid = uid1
        else:
            post_uid = uid1
            group_uid = uid2
            user_uid = uid3
        comment_for_uid = comments[idx % len(comments)]
        scheduled_at = (base + datetime.timedelta(seconds=idx * min_delay)).strftime('%Y-%m-%d %H:%M:%S')
        if add_page_comment_queue(page_index, target_type, user_uid, post_uid, group_uid, comment_for_uid, min_delay, max_delay, scheduled_at):
            added += 1
        else:
            skipped += 1
    return {'added': added, 'skipped': skipped, 'total': len(lines), 'comments': len(comments), 'min_delay': min_delay, 'max_delay': max_delay}

def get_page_comment_queue(limit=200):
    conn = db(); c = conn.cursor()
    c.execute("""
    SELECT id,page_name,target_type,user_uid,post_uid,group_uid,comment_text,min_delay_seconds,max_delay_seconds,scheduled_at,status,admin_status,result_message,created_at,processed_at
    FROM page_comment_queue ORDER BY id DESC LIMIT ?
    """, (limit,))
    rows = c.fetchall(); conn.close(); return rows

def get_page_comment_logs(limit=100):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,queue_id,page_name,target_uid,action,status,detail,created_at FROM page_comment_logs ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def get_page_comment_stats():
    conn = db(); c = conn.cursor()
    def one(q):
        try:
            c.execute(q); r=c.fetchone(); return r[0] if r else 0
        except Exception:
            return 0
    data = {
        'total': one('SELECT COUNT(*) FROM page_comment_queue'),
        'pending': one("SELECT COUNT(*) FROM page_comment_queue WHERE status='Chờ duyệt'"),
        'approved': one("SELECT COUNT(*) FROM page_comment_queue WHERE admin_status='Đã duyệt'"),
        'done': one("SELECT COUNT(*) FROM page_comment_queue WHERE status='Hoàn thành'"),
        'error': one("SELECT COUNT(*) FROM page_comment_queue WHERE status LIKE 'Lỗi%'")
    }
    conn.close(); return data

def approve_page_comment_queue(queue_id):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute("UPDATE page_comment_queue SET admin_status='Đã duyệt', status='Đã duyệt', result_message='Đã duyệt. Sẵn sàng xử lý thủ công hoặc qua API hợp lệ.', processed_at=? WHERE id=?", (now, queue_id))
    c.execute("INSERT INTO page_comment_logs(queue_id,page_name,target_uid,action,status,detail,created_at) SELECT id,page_name,COALESCE(NULLIF(post_uid,''),NULLIF(group_uid,''),user_uid),'Duyệt hàng chờ','Đã duyệt','Admin đã duyệt hàng chờ bình luận.',? FROM page_comment_queue WHERE id=?", (now, queue_id))
    conn.commit(); conn.close()

def mark_page_comment_done(queue_id, status='Hoàn thành', detail='Đã xử lý thủ công hoặc qua API hợp lệ.'):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute("UPDATE page_comment_queue SET status=?, result_message=?, processed_at=? WHERE id=?", (status, detail, now, queue_id))
    c.execute("INSERT INTO page_comment_logs(queue_id,page_name,target_uid,action,status,detail,created_at) SELECT id,page_name,COALESCE(NULLIF(post_uid,''),NULLIF(group_uid,''),user_uid),'Cập nhật trạng thái',?, ?, ? FROM page_comment_queue WHERE id=?", (status, detail, now, queue_id))
    conn.commit(); conn.close()

def export_page_comment_queue_csv():
    rows = get_page_comment_queue(10000)
    path = os.path.join(REPORT_DIR, 'page_comment_queue.csv')
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['id','page_name','target_type','user_uid','post_uid','group_uid','comment_text','min_delay_seconds','max_delay_seconds','scheduled_at','status','admin_status','result_message','created_at','processed_at'])
        w.writerows(rows)
    return path

def v3_ceo_summary():
    s = get_stats()
    token_checks = get_latest_token_checks(200)
    live_tokens = sum(1 for t in token_checks if str(t[2]).upper() == 'SỐNG')
    dead_tokens = sum(1 for t in token_checks if str(t[2]).upper() in ['CHẾT','LỖI'])
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    conn = db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM posts WHERE substr(created_at,1,10)=?", (today,))
    posts_today = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM crm WHERE substr(created_at,1,10)=?", (today,))
    leads_today = c.fetchone()[0]
    conn.close()
    return {
        'fanpages': len(get_pages_dynamic()),
        'posts': s.get('total', 0),
        'posted': s.get('posted', 0),
        'scheduled': s.get('scheduled', 0),
        'crm': s.get('crm', 0),
        'campaigns': s.get('campaigns', 0),
        'posts_today': posts_today,
        'leads_today': leads_today,
        'revenue_month': '0đ',
        'premium': get_free_status().get('label','🎁 Dùng thử miễn phí'),
        'token_live': live_tokens,
        'token_dead': dead_tokens,
        'token_total': len(get_pages_dynamic())
    }

def add_pipeline_lead(customer_name, phone, zalo, source, stage, value, note):
    if not customer_name and not phone and not zalo:
        return
    conn = db(); c = conn.cursor()
    c.execute("""
    INSERT INTO lead_pipeline(customer_name,phone,zalo,source,stage,value,note,created_at)
    VALUES(?,?,?,?,?,?,?,?)
    """, (customer_name, phone, zalo, source, stage or 'Khách mới', int(value or 0), note, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_pipeline_leads(limit=120):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,customer_name,phone,zalo,source,stage,value,note,created_at FROM lead_pipeline ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close()
    if not rows:
        # fallback dùng dữ liệu CRM cũ để Kanban vẫn có nội dung
        return [(r[0], r[1], r[2], r[3], r[4], 'Khách mới', 0, r[5], r[6]) for r in get_crm(30)]
    return rows

def add_customer_task(customer_name, task, due_date):
    if not task:
        return
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO customer_tasks(customer_name,task,due_date,status,created_at) VALUES(?,?,?,?,?)", (customer_name, task, due_date, 'pending', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_customer_tasks(limit=50):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,customer_name,task,due_date,status,created_at FROM customer_tasks ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def add_notification(title, detail, level='info'):
    if not title:
        return
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO notifications(title,detail,level,status,created_at) VALUES(?,?,?,?,?)", (title, detail, level, 'new', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit(); conn.close()

def get_notifications(limit=30):
    conn = db(); c = conn.cursor()
    c.execute("SELECT id,title,detail,level,status,created_at FROM notifications ORDER BY id DESC LIMIT ?", (limit,))
    rows = c.fetchall(); conn.close(); return rows

def v3_ai_tool_prompt(tool, topic, extra=''):
    if tool == 'marketing_director':
        return f"""
Bạn là AI Marketing Director Premium cho chủ shop/doanh nghiệp nhỏ. Lập bộ kế hoạch bằng tiếng Việt cho: {topic}
Yêu cầu xuất đúng cấu trúc sau:
1. Phân tích nhanh sản phẩm/dịch vụ và lợi thế bán hàng
2. Tệp khách hàng mục tiêu chi tiết
3. 30 content Facebook/TikTok ngắn, dễ đăng ngay
4. 10 mẫu quảng cáo Facebook Ads kéo inbox/chuyển đổi
5. 10 caption viral ngắn
6. 5 kịch bản chốt sale/inbox xử lý từ chối
7. Kế hoạch marketing 30 ngày
8. Ngân sách quảng cáo đề xuất theo 3 mức: thấp, vừa, mạnh
9. Việc cần làm mỗi ngày để tăng tỷ lệ chốt đơn
Không cam kết quá đà, không dùng lời hứa chắc chắn ra đơn, không spam hashtag.
Thông tin thêm: {extra}
"""
    if tool == 'facebook_ads_ai':
        return f"""
Bạn là chuyên gia Facebook Ads. Phân tích và tạo bộ quảng cáo cho: {topic}
Gồm: chấm điểm nội dung, 5 mẫu ads inbox, 5 mẫu ads chuyển đổi, target khách hàng, CTA, gợi ý ngân sách, lỗi cần tránh.
Thông tin thêm: {extra}
"""
    if tool == 'competitor_scanner':
        return f"""
Phân tích Fanpage đối thủ dựa trên mô tả/link này: {topic}
Trả về: điểm mạnh, điểm yếu, phong cách nội dung, CTA, ý tưởng vượt đối thủ, 10 nội dung có thể triển khai.
Thông tin thêm: {extra}
"""
    if tool == 'sales_script':
        return f"""
Tạo bộ kịch bản bán hàng cho: {topic}
Gồm: kịch bản inbox, kịch bản chốt sale, xử lý từ chối, chăm sóc lại khách cũ, tin nhắn nhắc thanh toán, tin nhắn hậu mãi.
Thông tin thêm: {extra}
"""
    if tool == 'landing_page':
        return f"""
Tạo nội dung landing page cho dịch vụ/sản phẩm: {topic}
Gồm: tiêu đề, mô tả, ưu điểm, form thu khách, CTA, FAQ, lý do nên chọn, nội dung hero section.
Thông tin thêm: {extra}
"""
    if tool == 'group_content':
        return f"""
Viết bài đăng Group Facebook bằng tiếng Việt cho: {topic}
Yêu cầu: gần gũi, không quá quảng cáo, tạo thảo luận, có CTA mềm, dưới 120 từ, kèm 3 biến thể tiêu đề.
Thông tin thêm: {extra}
"""
    if tool == 'comment_reply':
        return f"""
Bạn là Comment Manager AI. Hãy tạo phản hồi comment chuyên nghiệp cho khách hàng.
Comment/Nội dung khách: {topic}
Yêu cầu: trả lời ngắn, thân thiện, kéo khách vào inbox, không lộ số điện thoại, gắn nhãn khách nóng/ấm/lạnh và đề xuất chuyển CRM.
Thông tin thêm: {extra}
"""
    if tool == 'messenger_ai':
        return f"""
Tạo bộ kịch bản Messenger bán hàng cho: {topic}
Gồm: kịch bản inbox, kịch bản chốt sale, xử lý từ chối, chăm sóc khách cũ, tin nhắn nhắc thanh toán, tin nhắn hậu mãi.
Thông tin thêm: {extra}
"""
    if tool == 'prompt_premium':
        return f"""
Tạo 50 prompt Premium cho: {topic}
Chia nhóm: Facebook Ads, TikTok Ads, Seeding, Chốt sale, Content Viral.
"""
    return f"Tạo nội dung marketing chuyên nghiệp cho: {topic}. {extra}"

HTML = r"""
<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ title }}</title>

<meta name="theme-color" content="#2563eb">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="GPT MKT Pro">
<link rel="manifest" href="/manifest.json">
<link rel="apple-touch-icon" href="/static/icon-192.png">

<style>
:root{
  --blue:#2563EB;
  --blue2:#3B82F6;
  --purple:#7C3AED;
  --purple2:#A855F7;
  --deep:#1E1B4B;
  --navy:#0F172A;
  --bg:#F8FAFC;
  --soft:#EEF2FF;
  --card:#FFFFFF;
  --text:#111827;
  --muted:#64748B;
  --border:#E5E7EB;
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{
  margin:0;
  font-family:Arial,sans-serif;
  background:
    radial-gradient(circle at top left,rgba(124,58,237,.18),transparent 32%),
    radial-gradient(circle at top right,rgba(37,99,235,.16),transparent 30%),
    linear-gradient(135deg,#FFFFFF,#EEF2FF 52%,#F8FAFC);
  color:var(--text);
}
.layout{
  display:grid;
  grid-template-columns:270px 1fr 330px;
  gap:20px;
  max-width:1560px;
  margin:auto;
  padding:20px;
}
.sidebar{
  position:sticky;
  top:20px;
  height:calc(100vh - 40px);
  padding:22px;
  overflow-y:auto;
  overflow-x:hidden;
  overscroll-behavior:contain;
  scrollbar-width:thin;
  background:linear-gradient(180deg,#111827,#1E1B4B);
  border:1px solid rgba(255,255,255,.12);
  border-radius:28px;
  box-shadow:0 20px 55px rgba(30,27,75,.25);
}
.sidebar::-webkit-scrollbar{width:6px}
.sidebar::-webkit-scrollbar-thumb{background:rgba(255,255,255,.25);border-radius:999px}
.sidebar::-webkit-scrollbar-track{background:transparent}
.logo{
  font-size:25px;
  font-weight:900;
  color:white;
  line-height:1.12;
}
.logo:after{
  content:"";
  display:block;
  width:86px;
  height:5px;
  margin-top:12px;
  border-radius:999px;
  background:linear-gradient(135deg,var(--blue),var(--purple2));
}
.subtitle{
  font-size:12px;
  color:#CBD5E1;
  margin:12px 0 18px;
}
.nav a{
  display:block;
  padding:13px 14px;
  border-radius:16px;
  color:#F8FAFC;
  text-decoration:none;
  margin:7px 0;
  background:rgba(255,255,255,.06);
  border:1px solid rgba(255,255,255,.08);
}
.nav a:hover{
  background:linear-gradient(135deg,var(--blue),var(--purple));
  color:white;
  transform:translateX(3px);
}
.main{min-width:0}
.panel,.rightbar{
  background:rgba(255,255,255,.94);
  backdrop-filter:blur(16px);
  border:1px solid rgba(226,232,240,.9);
  border-radius:28px;
  box-shadow:0 18px 45px rgba(30,41,59,.10);
}
.panel{
  padding:25px;
  margin-bottom:20px;
}
.rightbar{
  position:sticky;
  top:20px;
  height:calc(100vh - 40px);
  padding:20px;
  overflow:auto;
}
.badge{
  display:inline-block;
  background:linear-gradient(135deg,var(--blue),var(--purple));
  color:white;
  padding:10px 16px;
  border-radius:999px;
  font-weight:900;
  box-shadow:0 10px 26px rgba(124,58,237,.26);
}
h1{
  font-size:42px;
  color:transparent;
  background:linear-gradient(135deg,var(--blue),var(--purple));
  -webkit-background-clip:text;
  background-clip:text;
  margin:16px 0 8px;
  letter-spacing:-.6px;
}
h2{
  margin-top:0;
  color:var(--deep);
}
p{color:var(--muted)}
.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.grid4{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:14px}
.stat{
  background:#FFFFFF;
  border:1px solid var(--border);
  border-left:5px solid var(--purple);
  border-radius:22px;
  padding:18px;
  box-shadow:0 12px 28px rgba(124,58,237,.08);
}
.stat span{color:var(--muted);font-size:14px}
.stat b{
  font-size:30px;
  color:var(--purple);
}
textarea,select,input{
  width:100%;
  margin:8px 0;
  padding:15px;
  border:1px solid var(--border);
  background:#FFFFFF;
  color:var(--text);
  border-radius:16px;
  font-size:15px;
  outline:none;
}
textarea:focus,select:focus,input:focus{
  border-color:var(--purple);
  box-shadow:0 0 0 4px rgba(124,58,237,.12);
}
button{
  background:linear-gradient(135deg,var(--blue),var(--purple));
  color:white;
  border:none;
  padding:14px 20px;
  border-radius:16px;
  font-weight:900;
  cursor:pointer;
  margin:8px 6px 8px 0;
  box-shadow:0 12px 28px rgba(37,99,235,.20);
}
button:hover{
  transform:translateY(-1px);
  box-shadow:0 16px 32px rgba(124,58,237,.25);
}
.secondary{
  background:linear-gradient(135deg,#E0E7FF,#F3E8FF);
  color:#4C1D95;
  border:1px solid #DDD6FE;
  box-shadow:none;
}
.success{
  color:#047857;
  background:#ECFDF5;
  border:1px solid #A7F3D0;
  padding:12px;
  border-radius:14px;
}
.error{
  color:#B91C1C;
  background:#FEF2F2;
  border:1px solid #FCA5A5;
  padding:12px;
  border-radius:14px;
}
.page-list{
  display:grid;
  grid-template-columns:repeat(2,1fr);
  gap:10px;
  margin:10px 0;
}
.page-item{
  background:#F8FAFC;
  color:var(--text);
  border:1px solid var(--border);
  border-radius:18px;
  padding:13px;
}
.history{
  white-space:pre-wrap;
  background:#F8FAFC;
  color:var(--text);
  border-radius:18px;
  padding:15px;
  margin:10px 0;
  border:1px solid var(--border);
}
.small{font-size:13px;color:var(--muted)}
.preview{
  background:white;
  color:#111827;
  border-radius:22px;
  padding:18px;
  max-width:570px;
  border:1px solid var(--border);
  box-shadow:0 14px 32px rgba(30,41,59,.10);
}
.preview-head{display:flex;align-items:center;gap:10px}
.avatar{
  width:44px;
  height:44px;
  border-radius:50%;
  background:linear-gradient(135deg,var(--blue),var(--purple));
}
.preview-name{font-weight:900}
.preview-content{white-space:pre-wrap;margin-top:12px;line-height:1.48}
.library-grid{
  display:grid;
  grid-template-columns:repeat(2,1fr);
  gap:12px;
}
.template-card{
  background:#FFFFFF;
  color:var(--text);
  border:1px solid var(--border);
  border-top:4px solid var(--purple);
  border-radius:20px;
  padding:16px;
  box-shadow:0 12px 26px rgba(124,58,237,.08);
}
.rightbar h2{
  color:transparent;
  background:linear-gradient(135deg,var(--blue),var(--purple));
  -webkit-background-clip:text;
  background-clip:text;
}
.rightbar hr{border:none;border-top:1px solid var(--border);margin:18px 0}
.pricing-grid{
  display:grid;
  grid-template-columns:repeat(5,1fr);
  gap:14px;
}
.price-card{
  background:#FFFFFF;
  color:#111827;
  border:1px solid var(--border);
  border-radius:22px;
  padding:18px;
  position:relative;
  box-shadow:0 18px 38px rgba(124,58,237,.10);
}
.price-card h3{
  margin:0 0 8px;
  font-size:18px;
  color:var(--deep);
}
.price{
  font-size:22px;
  font-weight:800;
  line-height:1.2;
  color:transparent;
  background:linear-gradient(135deg,var(--blue),var(--purple));
  -webkit-background-clip:text;
  background-clip:text;
}
.price-card ul{
  padding-left:18px;
  line-height:1.7;
  color:#334155;
}
.popular{
  border:2px solid var(--purple);
  transform:translateY(-6px);
}
.ribbon{
  position:absolute;
  top:-13px;
  left:18px;
  background:linear-gradient(135deg,var(--blue),var(--purple));
  color:white;
  padding:6px 12px;
  border-radius:999px;
  font-weight:900;
  font-size:12px;
}
.premium-lock{
  background:linear-gradient(135deg,#EEF2FF,#F5F3FF);
  color:#1E1B4B;
  border-radius:22px;
  padding:20px;
  border:1px solid #DDD6FE;
}
.mobilebar{display:none}
@media(max-width:1100px){
  .layout{grid-template-columns:230px 1fr}
  .rightbar{display:none}
  .grid4{grid-template-columns:repeat(2,1fr)}
  .pricing-grid{grid-template-columns:repeat(2,1fr)}
}
@media(max-width:820px){
  .layout{display:block;padding:12px}
  .sidebar,.rightbar{display:none}
  .grid,.grid4,.page-list,.library-grid,.pricing-grid{grid-template-columns:1fr}
  h1{font-size:30px}
  .panel{padding:18px;border-radius:22px}
  button{width:100%;font-size:16px}
  textarea{min-height:160px}
  .popular{transform:none}
  .mobilebar{
    display:flex;
    position:fixed;
    bottom:0;
    left:0;
    right:0;
    background:#FFFFFF;
    border-top:1px solid var(--border);
    z-index:999;
    box-shadow:0 -8px 25px rgba(30,41,59,.10);
  }
  .mobilebar a{
    flex:1;
    text-align:center;
    color:#1E1B4B;
    text-decoration:none;
    padding:10px 4px;
    font-size:12px;
  }
  .mobilebar a:hover{color:var(--purple)}
  body{padding-bottom:68px}
}

.analytics-grid{
  display:grid;
  grid-template-columns:repeat(3,1fr);
  gap:14px;
}
.analytics-box{
  background:#F8FAFC;
  border:1px solid var(--border);
  border-radius:18px;
  padding:16px;
}
.analytics-row{
  display:flex;
  justify-content:space-between;
  gap:10px;
  padding:8px 0;
  border-bottom:1px solid #E5E7EB;
}
.premium-center{
  background:linear-gradient(135deg,#EEF2FF,#F5F3FF);
  border:1px solid #DDD6FE;
  border-radius:22px;
  padding:18px;
}
.token-live{color:#047857;font-weight:900}
.token-dead{color:#B91C1C;font-weight:900}
@media(max-width:820px){.analytics-grid{grid-template-columns:1fr}}


.v9-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
.v9-card{background:#fff;border:1px solid var(--border);border-radius:20px;padding:16px;box-shadow:0 12px 28px rgba(124,58,237,.08)}
.lock-card{background:linear-gradient(135deg,#1E1B4B,#312E81);color:white;border-radius:22px;padding:20px;border:1px solid #A78BFA}
.modal-price{display:none;position:fixed;inset:0;background:rgba(15,23,42,.55);z-index:9999;align-items:center;justify-content:center;padding:20px}
.modal-inner{background:white;border-radius:28px;max-width:980px;width:100%;max-height:90vh;overflow:auto;padding:22px;color:#111827}
.close-btn{float:right;background:#F3F4F6;color:#111827;box-shadow:none}
@media(max-width:820px){.v9-grid{grid-template-columns:1fr}}


/* V10 Floating Support Bot + Live Trust Bar */
.live-trust-bar{
  position:fixed;
  top:16px;
  left:50%;
  transform:translateX(-50%);
  z-index:9998;
  background:rgba(255,255,255,.96);
  border:1px solid #DDD6FE;
  box-shadow:0 14px 35px rgba(30,41,59,.14);
  border-radius:999px;
  padding:10px 18px;
  display:flex;
  align-items:center;
  gap:10px;
  color:#1E1B4B;
  font-weight:800;
  backdrop-filter:blur(14px);
}
.live-dot{
  width:10px;
  height:10px;
  background:#22C55E;
  border-radius:50%;
  box-shadow:0 0 0 6px rgba(34,197,94,.14);
  animation:pulseLive 1.4s infinite;
}
@keyframes pulseLive{
  0%{transform:scale(1);opacity:1}
  70%{transform:scale(1.35);opacity:.65}
  100%{transform:scale(1);opacity:1}
}
.floating-bot{
  position:fixed;
  right:22px;
  bottom:22px;
  z-index:9999;
  font-family:Arial,sans-serif;
}
.bot-bubble{
  width:66px;
  height:66px;
  border-radius:50%;
  border:none;
  cursor:pointer;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  box-shadow:0 18px 45px rgba(124,58,237,.35);
  color:white;
  font-size:30px;
  display:flex;
  align-items:center;
  justify-content:center;
  position:relative;
  animation:botFloat 2.2s ease-in-out infinite;
}
@keyframes botFloat{
  0%,100%{transform:translateY(0)}
  50%{transform:translateY(-6px)}
}
.bot-status{
  position:absolute;
  right:0;
  bottom:2px;
  width:18px;
  height:18px;
  border-radius:50%;
  background:#22C55E;
  border:3px solid white;
}
.bot-panel{
  display:none;
  position:absolute;
  right:0;
  bottom:82px;
  width:340px;
  max-width:calc(100vw - 28px);
  background:white;
  border:1px solid #DDD6FE;
  border-radius:24px;
  box-shadow:0 22px 55px rgba(30,41,59,.22);
  overflow:hidden;
}
.bot-head{
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  padding:16px;
  display:flex;
  justify-content:space-between;
  align-items:center;
}
.bot-title{
  font-weight:900;
}
.bot-online{
  font-size:12px;
  opacity:.95;
  display:flex;
  align-items:center;
  gap:6px;
}
.typing-dots{
  display:inline-flex;
  gap:3px;
  margin-left:4px;
}
.typing-dots span{
  width:5px;
  height:5px;
  border-radius:50%;
  background:white;
  display:block;
  animation:typingDot 1s infinite ease-in-out;
}
.typing-dots span:nth-child(2){animation-delay:.15s}
.typing-dots span:nth-child(3){animation-delay:.3s}
@keyframes typingDot{
  0%,80%,100%{transform:translateY(0);opacity:.55}
  40%{transform:translateY(-4px);opacity:1}
}
.bot-close{
  background:rgba(255,255,255,.18);
  border:none;
  color:white;
  box-shadow:none;
  width:32px;
  height:32px;
  border-radius:50%;
  padding:0;
  margin:0;
}
.bot-body{
  padding:16px;
  max-height:360px;
  overflow:auto;
  background:#F8FAFC;
}
.bot-msg{
  background:white;
  border:1px solid #E5E7EB;
  border-radius:18px;
  padding:12px;
  margin-bottom:10px;
  line-height:1.45;
  color:#111827;
}
.bot-msg.ai{
  border-left:4px solid #7C3AED;
}
.bot-actions{
  display:grid;
  grid-template-columns:1fr;
  gap:8px;
  padding:14px 16px 16px;
  background:white;
}
.bot-actions button,.bot-actions a{
  display:block;
  text-align:center;
  text-decoration:none;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  border-radius:14px;
  padding:12px;
  font-weight:900;
  border:none;
  box-shadow:0 10px 22px rgba(37,99,235,.18);
}
.bot-actions .light{
  background:#EEF2FF;
  color:#4C1D95;
  border:1px solid #DDD6FE;
  box-shadow:none;
}
.bot-input{
  display:flex;
  gap:8px;
  padding:0 16px 16px;
  background:white;
}
.bot-input input{
  flex:1;
  margin:0;
}
.bot-input button{
  margin:0;
  padding:12px 14px;
}
@media(max-width:820px){
  .live-trust-bar{
    top:8px;
    width:calc(100vw - 24px);
    justify-content:center;
    font-size:12px;
    padding:9px 10px;
  }
  .floating-bot{right:14px;bottom:78px}
  .bot-panel{width:320px}
}


/* V10 SaaS Center Layout */
.top-hero{
  padding:34px;
  border-radius:32px;
  background:
    radial-gradient(circle at top left,rgba(37,99,235,.18),transparent 34%),
    radial-gradient(circle at top right,rgba(124,58,237,.18),transparent 32%),
    rgba(255,255,255,.96);
  border:1px solid #DDD6FE;
  box-shadow:0 22px 55px rgba(30,41,59,.12);
  margin-bottom:20px;
}
.hero-line{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding:10px 16px;
  background:#EEF2FF;
  border:1px solid #DDD6FE;
  color:#4C1D95;
  border-radius:999px;
  font-weight:900;
}
.hero-actions{
  display:flex;
  flex-wrap:wrap;
  gap:10px;
  margin-top:18px;
}
.module-hub{
  display:grid;
  grid-template-columns:repeat(4,1fr);
  gap:16px;
  margin-top:18px;
}
.module-card{
  background:#FFFFFF;
  border:1px solid #E5E7EB;
  border-radius:24px;
  padding:20px;
  min-height:158px;
  cursor:pointer;
  box-shadow:0 14px 32px rgba(30,41,59,.08);
  transition:.18s ease;
  position:relative;
  overflow:hidden;
}
.module-card:before{
  content:"";
  position:absolute;
  inset:0;
  background:linear-gradient(135deg,rgba(37,99,235,.08),rgba(124,58,237,.08));
  opacity:0;
  transition:.18s ease;
}
.module-card:hover{
  transform:translateY(-5px);
  border-color:#A78BFA;
  box-shadow:0 20px 45px rgba(124,58,237,.16);
}
.module-card:hover:before{opacity:1}
.module-card .icon{
  font-size:32px;
  margin-bottom:10px;
  position:relative;
}
.module-card h3{
  margin:0 0 8px;
  color:#1E1B4B;
  position:relative;
}
.module-card p{
  margin:0;
  color:#64748B;
  font-size:13px;
  line-height:1.45;
  position:relative;
}
.module-pill{
  display:inline-block;
  margin-top:12px;
  font-size:12px;
  font-weight:900;
  color:#4C1D95;
  background:#F5F3FF;
  border:1px solid #DDD6FE;
  border-radius:999px;
  padding:6px 10px;
  position:relative;
}
.module-section{
  display:none;
}
.module-section.active-module{
  display:block;
}
.section-open-note{
  margin-bottom:14px;
  background:#EEF2FF;
  border:1px solid #DDD6FE;
  border-radius:18px;
  padding:12px 14px;
  color:#4C1D95;
  font-weight:800;
}
.price-bottom-title{
  text-align:center;
  margin:34px 0 18px;
}
.price-bottom-title h2{
  font-size:34px;
  margin-bottom:8px;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:transparent;
  -webkit-background-clip:text;
  background-clip:text;
}
@media(max-width:1200px){
  .module-hub{grid-template-columns:repeat(3,1fr)}
}
@media(max-width:820px){
  .module-hub{grid-template-columns:repeat(2,1fr)}
  .top-hero{padding:22px}
  .module-card{min-height:140px;padding:16px}
}
@media(max-width:520px){
  .module-hub{grid-template-columns:1fr}
}


#ai-output-box{
  border:2px solid #DDD6FE;
  box-shadow:0 22px 50px rgba(124,58,237,.14);
}
#aiGeneratedContent{
  font-size:15px;
  line-height:1.65;
  max-height:520px;
  overflow:auto;
  background:#FFFFFF;
}
.pricing-visible{
  display:block !important;
}



/* Premium badge near Device ID */
.device-id-row{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin:6px 0 4px}
.premium-glow-badge{display:inline-flex;align-items:center;justify-content:center;padding:5px 10px;border-radius:999px;background:linear-gradient(135deg,#16a34a,#22c55e,#86efac);color:#052e16;font-size:11px;font-weight:1000;letter-spacing:.8px;box-shadow:0 0 0 1px rgba(187,247,208,.9),0 0 18px rgba(34,197,94,.85),inset 0 0 10px rgba(255,255,255,.55);animation:premiumPulse 1.6s ease-in-out infinite}
@keyframes premiumPulse{0%,100%{filter:brightness(1);box-shadow:0 0 0 1px rgba(187,247,208,.9),0 0 14px rgba(34,197,94,.65),inset 0 0 10px rgba(255,255,255,.45)}50%{filter:brightness(1.25);box-shadow:0 0 0 1px rgba(187,247,208,1),0 0 28px rgba(34,197,94,1),inset 0 0 14px rgba(255,255,255,.75)}}

/* Premium Pricing Pro */
.premium-pricing-pro{
  background:
    radial-gradient(circle at top left,rgba(37,99,235,.18),transparent 32%),
    radial-gradient(circle at top right,rgba(124,58,237,.18),transparent 30%),
    linear-gradient(135deg,#FFFFFF,#F8FAFC);
  border:1px solid #DDD6FE;
  border-radius:32px;
  padding:30px;
  box-shadow:0 24px 60px rgba(30,41,59,.12);
}
.premium-title{
  text-align:center;
  margin-bottom:24px;
}
.premium-title .mini{
  display:inline-block;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  padding:9px 16px;
  border-radius:999px;
  font-weight:900;
  margin-bottom:10px;
}
.premium-title h2{
  font-size:38px;
  margin:6px 0;
  color:transparent;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  -webkit-background-clip:text;
  background-clip:text;
}
.premium-title p{
  max-width:760px;
  margin:8px auto 0;
}
.premium-grid-pro{
  display:grid;
  grid-template-columns:repeat(5,1fr);
  gap:16px;
  align-items:stretch;
}
.premium-plan{
  background:white;
  border:1px solid #E5E7EB;
  border-radius:26px;
  padding:20px;
  position:relative;
  box-shadow:0 18px 42px rgba(30,41,59,.10);
  transition:.18s ease;
  display:flex;
  flex-direction:column;
}
.premium-plan:hover{
  transform:translateY(-6px);
  border-color:#A78BFA;
  box-shadow:0 24px 55px rgba(124,58,237,.18);
}
.premium-plan.featured{
  border:2px solid #7C3AED;
  background:linear-gradient(180deg,#FFFFFF,#F5F3FF);
}
.plan-ribbon{
  position:absolute;
  top:-14px;
  left:18px;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  padding:7px 13px;
  border-radius:999px;
  font-size:12px;
  font-weight:900;
}
.plan-name{
  font-size:18px;
  font-weight:900;
  color:#1E1B4B;
  margin-top:8px;
}
.plan-price{
  font-size:24px;
  font-weight:900;
  margin:8px 0 4px;
  color:transparent;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  -webkit-background-clip:text;
  background-clip:text;
}
.plan-desc{
  font-size:13px;
  color:#64748B;
  min-height:42px;
}
.benefit-title{
  font-weight:900;
  color:#1E1B4B;
  margin:12px 0 8px;
}
.benefit-list{
  list-style:none;
  padding:0;
  margin:0;
  font-size:13px;
  line-height:1.55;
}
.benefit-list li{
  padding:5px 0;
  border-bottom:1px dashed #EEF2FF;
}
.benefit-list .open:before{content:"✓ ";color:#047857;font-weight:900}
.benefit-list .lock:before{content:"🔒 ";color:#B45309;font-weight:900}
.plan-button{
  margin-top:auto;
  width:100%;
}
.premium-note-box{
  margin-top:20px;
  background:#EEF2FF;
  border:1px solid #DDD6FE;
  color:#4C1D95;
  border-radius:22px;
  padding:16px;
  line-height:1.55;
}
.payment-modal{
  display:none;
  position:fixed;
  inset:0;
  background:rgba(15,23,42,.65);
  z-index:10000;
  align-items:center;
  justify-content:center;
  padding:18px;
}
.payment-inner{
  width:min(980px,96vw);
  max-height:92vh;
  overflow:auto;
  background:white;
  border-radius:32px;
  box-shadow:0 30px 85px rgba(15,23,42,.38);
  border:1px solid #DDD6FE;
  padding:0;
}
.payment-head{
  background:
    radial-gradient(circle at top left,rgba(255,255,255,.25),transparent 35%),
    linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  padding:24px;
  display:flex;
  justify-content:space-between;
  gap:14px;
  align-items:flex-start;
}
.payment-head h2{
  color:white;
  margin:0 0 6px;
}
.payment-close{
  background:rgba(255,255,255,.18);
  color:white;
  box-shadow:none;
  border:1px solid rgba(255,255,255,.25);
  width:auto;
  margin:0;
}
.payment-body{
  display:grid;
  grid-template-columns:330px 1fr;
  gap:22px;
  padding:24px;
}
.qr-card{
  background:#F8FAFC;
  border:1px solid #E5E7EB;
  border-radius:26px;
  padding:18px;
  text-align:center;
}
.qr-card img{
  width:250px;
  max-width:100%;
  border-radius:18px;
  border:1px solid #E5E7EB;
  background:white;
  padding:8px;
}
.bank-info{
  margin-top:12px;
  text-align:left;
  background:white;
  border:1px solid #E5E7EB;
  border-radius:18px;
  padding:14px;
  line-height:1.7;
}
.payment-detail{
  background:white;
  border:1px solid #E5E7EB;
  border-radius:26px;
  padding:18px;
}
.payment-detail h3{
  margin-top:0;
  color:#1E1B4B;
}
.pay-amount{
  font-size:34px;
  font-weight:900;
  color:transparent;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  -webkit-background-clip:text;
  background-clip:text;
}
.payment-benefits{
  display:grid;
  grid-template-columns:1fr 1fr;
  gap:10px;
  margin:12px 0;
}
.payment-benefits div{
  background:#F8FAFC;
  border:1px solid #E5E7EB;
  border-radius:14px;
  padding:10px;
  font-size:13px;
}
.payment-alert{
  margin-top:14px;
  background:#FFF7ED;
  border:1px solid #FDBA74;
  color:#9A3412;
  border-radius:18px;
  padding:14px;
  line-height:1.55;
}
.payment-actions{
  display:flex;
  gap:10px;
  flex-wrap:wrap;
  margin-top:14px;
}
.payment-actions a{
  display:inline-block;
  text-decoration:none;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  padding:13px 16px;
  border-radius:14px;
  font-weight:900;
}
.payment-actions .light{
  background:#EEF2FF;
  color:#4C1D95;
  border:1px solid #DDD6FE;
}
.permission-db{
  margin-top:20px;
  background:#0F172A;
  color:#E5E7EB;
  border-radius:22px;
  padding:18px;
  overflow:auto;
  font-size:13px;
}
.permission-db code{
  color:#A7F3D0;
  white-space:pre-wrap;
}
@media(max-width:1200px){
  .premium-grid-pro{grid-template-columns:repeat(2,1fr)}
}
@media(max-width:820px){
  .premium-pricing-pro{padding:18px}
  .premium-grid-pro{grid-template-columns:1fr}
  .payment-body{grid-template-columns:1fr}
  .payment-benefits{grid-template-columns:1fr}
}


/* Free + Premium Stable */
.free-status-card{
  background:linear-gradient(135deg,#EEF2FF,#F5F3FF);
  border:1px solid #DDD6FE;
  color:#1E1B4B;
  border-radius:22px;
  padding:16px;
  margin:14px 0;
}
.free-progress{
  width:100%;
  height:12px;
  background:#E5E7EB;
  border-radius:999px;
  overflow:hidden;
  margin:10px 0;
}
.free-progress span{
  display:block;
  height:100%;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  border-radius:999px;
}
.free-expired{
  background:#FEF2F2;
  border-color:#FCA5A5;
  color:#991B1B;
}
.premium-story{
  margin:20px 0;
  background:linear-gradient(135deg,#0F172A,#1E1B4B);
  color:white;
  border-radius:28px;
  padding:24px;
  box-shadow:0 22px 55px rgba(30,27,75,.20);
}
.premium-story h3{
  margin-top:0;
  color:white;
  font-size:26px;
}
.premium-story p{
  color:#E0E7FF;
  line-height:1.65;
}
.why-premium{
  margin-top:22px;
  display:grid;
  grid-template-columns:repeat(2,1fr);
  gap:12px;
}
.why-item{
  background:white;
  border:1px solid #E5E7EB;
  border-radius:18px;
  padding:14px;
  color:#1E1B4B;
  box-shadow:0 12px 28px rgba(30,41,59,.06);
  font-weight:800;
}
.locked-demo{
  background:#FFF7ED;
  border:1px solid #FDBA74;
  color:#9A3412;
  border-radius:18px;
  padding:14px;
  margin-top:12px;
}
.lock-modal{
  display:none;
  position:fixed;
  inset:0;
  background:rgba(15,23,42,.65);
  z-index:10001;
  align-items:center;
  justify-content:center;
  padding:18px;
}
.lock-inner{
  background:white;
  border-radius:28px;
  max-width:560px;
  width:100%;
  padding:24px;
  box-shadow:0 30px 80px rgba(15,23,42,.35);
  border:1px solid #DDD6FE;
}
.lock-inner h2{
  margin-top:0;
  color:#1E1B4B;
}
@media(max-width:820px){
  .why-premium{grid-template-columns:1fr}
}


/* ===== GPT MINI MENU UPGRADE - PROFESSIONAL SAAS ===== */
.menu-upgrade-note{
  margin-top:14px;
  padding:14px;
  border-radius:18px;
  background:linear-gradient(135deg,rgba(37,99,235,.12),rgba(124,58,237,.12));
  border:1px solid rgba(167,139,250,.45);
  color:#E0E7FF;
  font-size:12px;
  line-height:1.55;
}
.menu-mini-group{
  margin:10px 0;
  border-radius:18px;
  overflow:hidden;
  border:1px solid rgba(255,255,255,.08);
  background:rgba(255,255,255,.045);
}
.menu-mini-title{
  padding:12px 14px;
  color:#FFFFFF;
  font-weight:900;
  cursor:pointer;
  display:flex;
  align-items:center;
  justify-content:space-between;
}
.menu-mini-title small{
  color:#CBD5E1;
  font-weight:700;
}
.menu-mini-title:hover{
  background:linear-gradient(135deg,rgba(37,99,235,.55),rgba(124,58,237,.55));
}
.menu-mini-items{
  display:none;
  padding:6px 8px 10px;
}
.menu-mini-group.open .menu-mini-items{
  display:block;
}
.menu-mini-item{
  display:block;
  padding:10px 12px;
  margin:5px 0;
  border-radius:14px;
  color:#E5E7EB;
  text-decoration:none;
  font-size:13px;
  background:rgba(255,255,255,.05);
  border:1px solid rgba(255,255,255,.06);
}
.menu-mini-item:hover{
  color:white;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  transform:translateX(3px);
}
.clean-hero-pro{
  padding:34px;
  border-radius:32px;
  background:
    radial-gradient(circle at top left,rgba(37,99,235,.20),transparent 34%),
    radial-gradient(circle at top right,rgba(124,58,237,.18),transparent 32%),
    rgba(255,255,255,.97);
  border:1px solid #DDD6FE;
  box-shadow:0 22px 55px rgba(30,41,59,.12);
  margin-bottom:20px;
}
.clean-hero-pro h1{
  margin-top:12px;
}
.clean-pill-row{
  display:flex;
  flex-wrap:wrap;
  gap:10px;
  margin-top:18px;
}
.clean-pill{
  padding:10px 14px;
  border-radius:999px;
  background:#EEF2FF;
  color:#4C1D95;
  font-weight:900;
  border:1px solid #DDD6FE;
  font-size:13px;
}
.free-status-box{
  background:linear-gradient(135deg,#EEF2FF,#F5F3FF);
  border:1px solid #DDD6FE;
  color:#4C1D95;
  border-radius:18px;
  padding:14px;
  font-weight:800;
  line-height:1.55;
}


/* ===== SAFE CLEAN MENU + PRICING ONLY ===== */
.safe-menu-title{
  margin:14px 0 8px;
  padding:0 6px;
  font-size:11px;
  font-weight:900;
  letter-spacing:.08em;
  color:#A5B4FC;
  text-transform:uppercase;
}
.safe-menu-link{
  display:flex!important;
  align-items:center;
  gap:10px;
  padding:13px 14px!important;
  border-radius:16px!important;
  color:#F8FAFC!important;
  text-decoration:none!important;
  margin:7px 0!important;
  background:rgba(255,255,255,.06)!important;
  border:1px solid rgba(255,255,255,.08)!important;
  transition:.18s ease!important;
}
.safe-menu-link:hover{
  background:linear-gradient(135deg,#2563EB,#7C3AED)!important;
  transform:translateX(3px)!important;
}
.safe-menu-link .safe-ico{
  width:26px;
  height:26px;
  display:inline-flex;
  align-items:center;
  justify-content:center;
  border-radius:10px;
  background:rgba(255,255,255,.10);
}
.safe-menu-link .safe-text{
  flex:1;
  font-weight:800;
}
.safe-menu-link .safe-tag{
  font-size:10px;
  padding:4px 7px;
  border-radius:999px;
  color:#DDD6FE;
  background:rgba(124,58,237,.18);
  border:1px solid rgba(221,214,254,.18);
}
.safe-premium-box{
  margin-top:14px;
  padding:14px;
  border-radius:18px;
  background:linear-gradient(135deg,rgba(37,99,235,.15),rgba(124,58,237,.16));
  border:1px solid rgba(221,214,254,.22);
  color:#E0E7FF;
  font-size:12px;
  line-height:1.55;
}
.safe-pricing-action{
  display:block;
  width:100%;
  margin-top:12px;
  text-align:center;
  border-radius:14px;
  padding:12px 14px;
  font-weight:900;
  cursor:pointer;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  text-decoration:none;
  border:0;
}
.price-card,.premium-plan{
  cursor:pointer;
}
.price-card:hover,.premium-plan:hover{
  transform:translateY(-4px);
  border-color:#A78BFA!important;
}


/* ===== MKT AUTOMATION PRO - APP MINI PWA MODE ===== */
:root{
  --app-dark:#0F172A;
  --app-primary:#2563EB;
  --app-premium:#7C3AED;
  --app-accent:#06B6D4;
}
.app-shell-top{
  position:sticky;
  top:0;
  z-index:9000;
  background:rgba(15,23,42,.94);
  backdrop-filter:blur(16px);
  color:white;
  border-bottom:1px solid rgba(255,255,255,.10);
  padding:14px 16px;
  display:none;
}
.app-shell-brand{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:12px;
}
.app-shell-brand b{
  font-size:17px;
}
.app-shell-status{
  font-size:11px;
  color:#C7D2FE;
  margin-top:4px;
}
.app-quick-grid{
  display:grid;
  grid-template-columns:repeat(2,1fr);
  gap:12px;
  margin:18px 0;
}
.app-quick-card{
  border:1px solid #E5E7EB;
  border-radius:22px;
  background:white;
  padding:18px;
  box-shadow:0 14px 30px rgba(15,23,42,.08);
  cursor:pointer;
}
.app-quick-card:hover{
  transform:translateY(-3px);
  border-color:#A78BFA;
}
.app-quick-card .app-ico{
  width:44px;
  height:44px;
  border-radius:16px;
  display:flex;
  align-items:center;
  justify-content:center;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  font-size:22px;
  margin-bottom:10px;
}
.app-quick-card b{
  color:#0F172A;
  display:block;
  margin-bottom:6px;
}
.app-quick-card span{
  color:#64748B;
  font-size:12px;
  line-height:1.45;
}
.app-bottom-nav{
  display:none;
  position:fixed;
  left:0;
  right:0;
  bottom:0;
  z-index:9998;
  background:rgba(255,255,255,.96);
  backdrop-filter:blur(14px);
  border-top:1px solid #E5E7EB;
  box-shadow:0 -10px 30px rgba(15,23,42,.12);
}
.app-bottom-nav a{
  flex:1;
  text-align:center;
  padding:9px 3px 8px;
  color:#475569;
  text-decoration:none;
  font-size:11px;
  font-weight:800;
}
.app-bottom-nav a span{
  display:block;
  font-size:20px;
  margin-bottom:2px;
}
.app-install-banner{
  display:none;
  background:linear-gradient(135deg,#0F172A,#312E81);
  color:white;
  border-radius:24px;
  padding:18px;
  margin:16px 0;
  box-shadow:0 18px 45px rgba(15,23,42,.20);
}
.app-install-banner button{
  margin-top:10px;
  width:100%;
}
@media(max-width:820px){
  body{
    background:#F8FAFC!important;
    padding-bottom:78px!important;
  }
  .app-shell-top{display:block}
  .layout{
    display:block!important;
    padding:12px!important;
    max-width:100%!important;
  }
  .sidebar,.rightbar,.live-trust-bar{
    display:none!important;
  }
  .panel,.top-hero,.premium-pricing-pro{
    border-radius:24px!important;
    padding:18px!important;
    margin-bottom:14px!important;
  }
  h1{
    font-size:28px!important;
    line-height:1.15!important;
  }
  .hero-actions{
    display:grid!important;
    grid-template-columns:1fr!important;
  }
  .module-hub,.grid4,.pricing-grid,.premium-grid-pro{
    grid-template-columns:1fr!important;
    gap:12px!important;
  }
  .app-bottom-nav{
    display:flex!important;
  }
  .app-install-banner{
    display:block!important;
  }
  .floating-bot{
    bottom:88px!important;
    right:14px!important;
  }
  .mobilebar{
    display:none!important;
  }
}


/* ===== MKT AUTOMATION PRO V2 - SAFE FINAL UI ===== */
:root{
  --v2-dark:#0F172A;
  --v2-card:#FFFFFF;
  --v2-blue:#2563EB;
  --v2-purple:#7C3AED;
  --v2-cyan:#06B6D4;
  --v2-soft:#EEF2FF;
  --v2-line:#E5E7EB;
}
.v2-desktop-shell{
  display:block;
}
.v2-topbar{
  position:sticky;
  top:0;
  z-index:8000;
  display:none;
  padding:14px 16px;
  background:rgba(15,23,42,.95);
  color:white;
  backdrop-filter:blur(16px);
  border-bottom:1px solid rgba(255,255,255,.10);
}
.v2-brand-row{
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap:12px;
}
.v2-brand-title{
  font-size:17px;
  font-weight:1000;
}
.v2-brand-sub{
  font-size:11px;
  color:#C7D2FE;
  margin-top:3px;
}
.v2-status-pill{
  font-size:11px;
  font-weight:900;
  padding:7px 10px;
  border-radius:999px;
  background:rgba(34,197,94,.16);
  border:1px solid rgba(34,197,94,.24);
  color:#BBF7D0;
}
.v2-nav-title{
  margin:14px 0 8px;
  padding:0 6px;
  font-size:11px;
  font-weight:900;
  letter-spacing:.08em;
  color:#A5B4FC;
  text-transform:uppercase;
}
.v2-nav-link{
  display:flex!important;
  align-items:center;
  gap:10px;
  padding:13px 14px!important;
  border-radius:16px!important;
  color:#F8FAFC!important;
  text-decoration:none!important;
  margin:7px 0!important;
  background:rgba(255,255,255,.06)!important;
  border:1px solid rgba(255,255,255,.08)!important;
  transition:.18s ease!important;
}
.v2-nav-link:hover{
  background:linear-gradient(135deg,#2563EB,#7C3AED)!important;
  transform:translateX(3px)!important;
}
.v2-nav-ico{
  width:28px;
  height:28px;
  display:inline-flex;
  align-items:center;
  justify-content:center;
  border-radius:10px;
  background:rgba(255,255,255,.10);
}
.v2-nav-text{flex:1;font-weight:850}
.v2-nav-tag{
  font-size:10px;
  padding:4px 7px;
  border-radius:999px;
  color:#DDD6FE;
  background:rgba(124,58,237,.20);
  border:1px solid rgba(221,214,254,.20);
}
.v2-side-card{
  margin-top:14px;
  padding:15px;
  border-radius:20px;
  background:linear-gradient(135deg,rgba(37,99,235,.16),rgba(124,58,237,.18));
  border:1px solid rgba(221,214,254,.22);
  color:#E0E7FF;
  font-size:12px;
  line-height:1.55;
}
.v2-hero{
  background:
    radial-gradient(circle at top left,rgba(37,99,235,.22),transparent 34%),
    radial-gradient(circle at top right,rgba(124,58,237,.20),transparent 34%),
    linear-gradient(135deg,#FFFFFF,#F8FAFC);
  border:1px solid #DDD6FE;
  border-radius:32px;
  padding:28px;
  margin-bottom:20px;
  box-shadow:0 22px 55px rgba(30,41,59,.12);
}
.v2-hero-kicker{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding:10px 15px;
  border-radius:999px;
  background:#EEF2FF;
  color:#4C1D95;
  font-weight:1000;
  border:1px solid #DDD6FE;
}
.v2-hero h1{
  margin:14px 0 8px;
}
.v2-hero-desc{
  max-width:820px;
  color:#64748B;
  font-size:15px;
  line-height:1.6;
}
.v2-quick-grid{
  display:grid;
  grid-template-columns:repeat(4,1fr);
  gap:14px;
  margin-top:20px;
}
.v2-quick-card{
  background:white;
  border:1px solid #E5E7EB;
  border-radius:24px;
  padding:18px;
  box-shadow:0 14px 32px rgba(15,23,42,.08);
  cursor:pointer;
  transition:.18s ease;
}
.v2-quick-card:hover{
  transform:translateY(-4px);
  border-color:#A78BFA;
  box-shadow:0 20px 45px rgba(124,58,237,.14);
}
.v2-quick-ico{
  width:46px;
  height:46px;
  border-radius:17px;
  display:flex;
  align-items:center;
  justify-content:center;
  color:white;
  font-size:22px;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  margin-bottom:12px;
}
.v2-quick-card b{
  display:block;
  color:#0F172A;
  margin-bottom:6px;
}
.v2-quick-card span{
  color:#64748B;
  font-size:12px;
  line-height:1.45;
}
.v2-install-box{
  display:none;
  margin-top:16px;
  border-radius:22px;
  padding:16px;
  color:white;
  background:linear-gradient(135deg,#0F172A,#312E81);
  box-shadow:0 18px 45px rgba(15,23,42,.18);
}
.v2-bottom-nav{
  display:none;
  position:fixed;
  left:0;
  right:0;
  bottom:0;
  z-index:9999;
  background:rgba(255,255,255,.97);
  backdrop-filter:blur(14px);
  border-top:1px solid #E5E7EB;
  box-shadow:0 -10px 30px rgba(15,23,42,.12);
}
.v2-bottom-nav a{
  flex:1;
  text-align:center;
  padding:9px 3px 8px;
  color:#475569;
  text-decoration:none;
  font-size:11px;
  font-weight:900;
}
.v2-bottom-nav span{
  display:block;
  font-size:20px;
  margin-bottom:2px;
}
.v2-pricing-action{
  display:block;
  width:100%;
  margin-top:12px;
  text-align:center;
  border-radius:14px;
  padding:12px 14px;
  font-weight:1000;
  cursor:pointer;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  text-decoration:none;
  border:0;
}
.price-card,.premium-plan{cursor:pointer}
.price-card:hover,.premium-plan:hover{
  transform:translateY(-5px);
  border-color:#A78BFA!important;
}
@media(max-width:1100px){
  .v2-quick-grid{grid-template-columns:repeat(2,1fr)}
}
@media(max-width:820px){
  body{
    background:#F8FAFC!important;
    padding-bottom:78px!important;
  }
  .v2-topbar{display:block!important}
  .layout{
    display:block!important;
    max-width:100%!important;
    padding:12px!important;
  }
  .sidebar,.rightbar,.live-trust-bar,.mobilebar{
    display:none!important;
  }
  .panel,.top-hero,.premium-pricing-pro,.v2-hero{
    border-radius:24px!important;
    padding:18px!important;
    margin-bottom:14px!important;
  }
  h1{font-size:28px!important;line-height:1.15!important}
  .v2-quick-grid,.module-hub,.grid4,.pricing-grid,.premium-grid-pro{
    grid-template-columns:1fr!important;
    gap:12px!important;
  }
  .v2-install-box{display:block!important}
  .v2-bottom-nav{display:flex!important}
  .floating-bot{
    right:14px!important;
    bottom:88px!important;
  }
}



/* V3 Enterprise Add-on */
.v3-kpi-title{margin-top:18px;background:#EEF2FF;border:1px solid #DDD6FE;color:#4C1D95;padding:12px 16px;border-radius:18px;font-weight:800}
.v3-ceo-grid .stat small{display:block;color:#64748B;margin-top:6px;font-size:12px}

.v5-seller-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:14px;margin:14px 0}
.v5-tool-card{background:#fff;border:1px solid var(--border);border-radius:22px;padding:16px;box-shadow:0 12px 28px rgba(37,99,235,.08)}
.v5-tool-card h3{margin:0 0 8px;color:#1E1B4B}.v5-tool-card ul{padding-left:18px;line-height:1.7;color:#334155}
.v5-status-pill{display:inline-block;border-radius:999px;padding:6px 10px;font-size:12px;font-weight:900;background:#ECFDF5;color:#047857;border:1px solid #A7F3D0}
.v5-warning-pill{display:inline-block;border-radius:999px;padding:6px 10px;font-size:12px;font-weight:900;background:#FEF3C7;color:#92400E;border:1px solid #FCD34D}
.v5-table{width:100%;border-collapse:separate;border-spacing:0 8px}.v5-table th{text-align:left;color:#64748B;font-size:13px}.v5-table td{background:#F8FAFC;border-top:1px solid #E5E7EB;border-bottom:1px solid #E5E7EB;padding:10px}.v5-table td:first-child{border-left:1px solid #E5E7EB;border-radius:14px 0 0 14px}.v5-table td:last-child{border-right:1px solid #E5E7EB;border-radius:0 14px 14px 0}

.v3-feature-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:14px;margin-top:14px}
.v3-feature-card{background:#fff;border:1px solid var(--border);border-radius:22px;padding:18px;box-shadow:0 12px 28px rgba(30,41,59,.08)}
.v3-feature-card h3{margin:0 0 8px;color:#1E1B4B}.v3-feature-card ul{margin:8px 0 0;padding-left:18px;line-height:1.65;color:#334155}
.v3-premium-badge{display:inline-block;background:linear-gradient(135deg,#2563EB,#7C3AED);color:white;border-radius:999px;padding:6px 10px;font-size:12px;font-weight:900;margin-bottom:8px}
.kanban-board{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-top:14px}.kanban-col{background:#F8FAFC;border:1px solid #E5E7EB;border-radius:18px;padding:12px;min-height:180px}.kanban-col h3{font-size:15px;margin:0 0 10px;color:#1E1B4B}.kanban-card{background:white;border:1px solid #DDD6FE;border-left:4px solid #7C3AED;border-radius:14px;padding:10px;margin-bottom:10px;cursor:grab;box-shadow:0 8px 18px rgba(124,58,237,.08)}
.v3-mini-table{width:100%;border-collapse:collapse}.v3-mini-table td,.v3-mini-table th{border-bottom:1px solid #EEF2FF;padding:10px;text-align:left}.v3-mini-table th{color:#1E1B4B}
@media(max-width:1000px){.kanban-board{grid-template-columns:1fr 1fr}.v3-feature-grid{grid-template-columns:1fr}}
@media(max-width:620px){.kanban-board{grid-template-columns:1fr}}


/* ===== V4 PREMIUM CONVERSION UPGRADE ===== */
.v4-trust-grid,.v4-pricing-stats{
  display:grid;
  grid-template-columns:repeat(3,minmax(0,1fr));
  gap:14px;
  margin:18px 0;
}
.v4-trust-card,.v4-pricing-stats div{
  background:linear-gradient(135deg,#FFFFFF,#F8FAFC);
  border:1px solid #DDD6FE;
  border-radius:22px;
  padding:16px;
  text-align:center;
  box-shadow:0 14px 32px rgba(30,41,59,.08);
}
.v4-trust-card b,.v4-pricing-stats b{
  display:block;
  font-size:26px;
  color:transparent;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  -webkit-background-clip:text;
  background-clip:text;
}
.v4-trust-card span,.v4-pricing-stats span{font-size:13px;color:#64748B;font-weight:800}
.v4-premium-hero{
  background:radial-gradient(circle at top left,rgba(37,99,235,.12),transparent 34%),linear-gradient(135deg,#FFFFFF,#F5F3FF)!important;
  padding:26px!important;
}
.v4-hero-label,.value-badge{
  display:inline-block;
  background:linear-gradient(135deg,#2563EB,#7C3AED);
  color:white;
  padding:8px 13px;
  border-radius:999px;
  font-weight:900;
  font-size:12px;
  margin-bottom:10px;
}
.v4-value-grid{
  display:grid;
  grid-template-columns:repeat(3,1fr);
  gap:12px;
  margin:18px 0;
}
.v4-value-grid div{
  background:white;
  border:1px solid #DDD6FE;
  border-radius:20px;
  padding:14px;
}
.v4-value-grid b{display:block;color:#4C1D95;font-size:22px}
.v4-value-grid span{font-size:13px;color:#64748B;font-weight:800}
.v4-pricing-shell{padding:32px!important}
.v4-premium-grid{
  grid-template-columns:repeat(auto-fit,minmax(250px,1fr))!important;
  align-items:stretch;
}
.v4-plan{
  border-radius:28px!important;
  padding:22px!important;
  overflow:hidden;
}
.v4-plan .plan-name{font-size:17px!important}
.v4-plan .plan-price{
  font-size:24px!important;
  line-height:1.15!important;
  margin:8px 0!important;
}
.price-sub{
  color:#64748B;
  font-size:13px;
  font-weight:900;
  margin-bottom:12px;
}
.v4-yearly{
  transform:scale(1.03);
  border:2px solid #7C3AED!important;
  box-shadow:0 32px 75px rgba(124,58,237,.22)!important;
}
.v4-lifetime{
  border:2px solid #F59E0B!important;
  background:linear-gradient(180deg,#FFFFFF,#FFFBEB)!important;
  box-shadow:0 32px 75px rgba(245,158,11,.18)!important;
}
.v4-save-box{
  background:linear-gradient(135deg,#EEF2FF,#F5F3FF);
  border:1px solid #DDD6FE;
  border-radius:16px;
  padding:12px;
  color:#4C1D95;
  font-weight:900;
  margin:12px 0;
  font-size:13px;
}
.v4-save-box.gold{
  background:linear-gradient(135deg,#FEF3C7,#FFF7ED);
  border-color:#FCD34D;
  color:#92400E;
}
.v4-value-received{
  background:#F8FAFC;
  border:1px solid #E5E7EB;
  border-radius:16px;
  padding:12px;
  margin:12px 0;
}
.v4-value-received b{display:block;color:#1E1B4B;margin-bottom:4px}
.v4-value-received span{font-size:13px;color:#64748B}
.premium-btn{
  background:linear-gradient(135deg,#2563EB,#7C3AED)!important;
  border-radius:16px!important;
  font-weight:900!important;
  box-shadow:0 14px 30px rgba(124,58,237,.20)!important;
}
.v4-outsourcing-box{
  margin-top:22px;
  background:white;
  border:1px solid #DDD6FE;
  border-radius:26px;
  padding:22px;
  box-shadow:0 18px 42px rgba(30,41,59,.08);
}
.v4-outsource-grid{
  display:grid;
  grid-template-columns:repeat(4,1fr);
  gap:12px;
  margin:14px 0;
}
.v4-outsource-grid div{
  background:#F8FAFC;
  border:1px solid #E5E7EB;
  border-radius:18px;
  padding:14px;
}
.v4-outsource-grid b{display:block;color:#7C3AED}
.v4-outsource-grid span{font-size:13px;color:#64748B;font-weight:800}
.v4-success-grid{
  display:grid;
  grid-template-columns:repeat(auto-fit,minmax(220px,1fr));
  gap:14px;
}
.v4-success-card{
  background:white;
  border:1px solid #E5E7EB;
  border-radius:24px;
  padding:18px;
  box-shadow:0 16px 36px rgba(30,41,59,.08);
}
.success-icon{
  width:52px;height:52px;border-radius:18px;
  display:flex;align-items:center;justify-content:center;
  font-size:26px;background:#F5F3FF;border:1px solid #DDD6FE;
  margin-bottom:12px;
}
.v4-proof-box{
  margin-top:16px;
  background:linear-gradient(135deg,#EEF2FF,#F8FAFC);
  border:1px solid #DDD6FE;
  border-radius:24px;
  padding:18px;
}
.v4-lock-hero .benefit-list{
  display:grid;
  grid-template-columns:1fr 1fr;
  gap:8px;
}
.v4-gold-card{
  border:2px solid #F59E0B!important;
  background:linear-gradient(180deg,#FFFFFF,#FFFBEB)!important;
}
.rec-price{font-size:24px!important;line-height:1.15!important}
@media(max-width:820px){
  .v4-trust-grid,.v4-pricing-stats,.v4-value-grid,.v4-outsource-grid,.locked-recommend-grid{grid-template-columns:1fr!important}
  .v4-yearly{transform:none}
  .v4-lock-hero .benefit-list{grid-template-columns:1fr}
}


/* ===== V5 PREMIUM DARK POLISH ===== */
body{
  background:radial-gradient(circle at top left,rgba(124,58,237,.22),transparent 30%),radial-gradient(circle at top right,rgba(37,99,235,.18),transparent 28%),linear-gradient(135deg,#0B1020,#111827 48%,#1E293B)!important;
}
.panel,.rightbar,.top-hero{
  background:rgba(15,23,42,.92)!important;
  color:#F8FAFC!important;
  border:1px solid rgba(148,163,184,.22)!important;
  box-shadow:0 22px 60px rgba(0,0,0,.28)!important;
}
.top-hero h1,h1,.rightbar h2{
  background:linear-gradient(135deg,#38BDF8,#8B5CF6)!important;
  -webkit-background-clip:text!important;
  background-clip:text!important;
}
.panel p,.top-hero p,.rightbar p,.small{color:#CBD5E1!important}
.app-quick-card,.template-card,.stat,.price-card,.preview,.locked-recommend-card{
  background:rgba(255,255,255,.96)!important;
}
.activity-card{
  display:flex;
  justify-content:space-between;
  align-items:center;
  gap:12px;
  padding:14px 15px;
  margin:10px 0;
  border-radius:18px;
  background:linear-gradient(135deg,rgba(37,99,235,.15),rgba(124,58,237,.14));
  border:1px solid rgba(148,163,184,.22);
  color:#E5E7EB;
}
.activity-card span{font-weight:800;color:#CBD5E1}
.activity-card b{font-size:24px;color:#38BDF8}
.v5-focus-box{
  margin-top:14px;
  padding:16px;
  border-radius:20px;
  background:linear-gradient(135deg,rgba(56,189,248,.12),rgba(139,92,246,.16));
  border:1px solid rgba(56,189,248,.28);
  color:#E5E7EB;
  line-height:1.65;
}
.v5-focus-box b{color:#FBBF24}

/* ===== ANALYTICS CENTER V5 POLISH ===== */
.analytics-kpi-grid{
  display:grid;
  grid-template-columns:repeat(4,1fr);
  gap:14px;
  margin:18px 0 22px;
}
.analytics-kpi{
  background:linear-gradient(135deg,rgba(37,99,235,.10),rgba(124,58,237,.10));
  border:1px solid rgba(124,58,237,.22);
  border-radius:20px;
  padding:16px;
  box-shadow:0 12px 28px rgba(37,99,235,.08);
}
.analytics-kpi span{display:block;color:#475569;font-size:13px;font-weight:800}
.analytics-kpi b{display:block;color:#4C1D95;font-size:28px;margin:7px 0}
.analytics-kpi small{display:block;color:#64748B;line-height:1.45}
.analytics-section-title{
  margin:24px 0 12px;
  color:#1E1B4B;
  font-size:17px;
  font-weight:900;
}
.analytics-box.full{grid-column:1/-1}
.analytics-bar{height:8px;background:#E5E7EB;border-radius:999px;overflow:hidden;margin:-2px 0 8px}
.analytics-bar.large{height:12px;margin:2px 0 12px}
.analytics-bar i{display:block;height:100%;background:linear-gradient(135deg,var(--blue),var(--purple));border-radius:999px;min-width:6px}
.empty-analytics{
  background:#F8FAFC;
  border:1px dashed #CBD5E1;
  border-radius:14px;
  color:#64748B;
  padding:14px;
  line-height:1.55;
}
@media(max-width:1100px){.analytics-kpi-grid{grid-template-columns:repeat(2,1fr)}}
@media(max-width:700px){.analytics-kpi-grid{grid-template-columns:1fr}}


/* V6: Loại bỏ icon ở menu chính/sidebar/mobile để giao diện gọn hơn */
.v2-nav-ico,.mobilebar a::first-letter{display:none!important}
.v2-nav-link{gap:8px!important}
.mobilebar a{font-size:12px!important;line-height:1.35!important}
.gf-box{background:#F8FAFC;border:1px solid #E5E7EB;border-radius:22px;padding:16px;margin:12px 0}
.gf-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}
.gf-grid-3{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:14px}
.gf-stat{background:white;border:1px solid #E5E7EB;border-radius:18px;padding:14px;box-shadow:0 8px 24px rgba(15,23,42,.06)}
.gf-stat b{font-size:24px;display:block;color:#1E1B4B}.gf-stat span{font-size:12px;color:#64748B;font-weight:800}
.gf-warning{background:#FFF7ED;border:1px solid #FED7AA;color:#9A3412;border-radius:16px;padding:12px;margin:10px 0;font-size:13px;line-height:1.5}
.gf-table{width:100%;border-collapse:separate;border-spacing:0 8px;font-size:13px}.gf-table th{text-align:left;color:#475569;padding:8px}.gf-table td{background:white;border-top:1px solid #E5E7EB;border-bottom:1px solid #E5E7EB;padding:10px}.gf-table td:first-child{border-left:1px solid #E5E7EB;border-radius:12px 0 0 12px}.gf-table td:last-child{border-right:1px solid #E5E7EB;border-radius:0 12px 12px 0}
@media(max-width:900px){.gf-grid,.gf-grid-3{grid-template-columns:1fr}}

.payment-form-grid{display:grid;grid-template-columns:1fr;gap:8px;margin-top:8px}.payment-form-grid input{width:100%;box-sizing:border-box;border:1px solid #d1d5db;border-radius:10px;padding:11px 12px;font-size:14px}.primary{background:linear-gradient(135deg,#facc15,#f97316);border:0;border-radius:12px;padding:12px 16px;font-weight:800;color:#111827;cursor:pointer}
.admin-table{width:100%;border-collapse:collapse;background:white;border-radius:14px;overflow:hidden}.admin-table th,.admin-table td{border-bottom:1px solid #e5e7eb;padding:10px;text-align:left;font-size:13px}.admin-badge{display:inline-block;padding:4px 8px;border-radius:999px;background:#fef3c7;color:#92400e;font-weight:700}
</style>
<script>
function copyText(id){
  const el=document.getElementById(id);
  navigator.clipboard.writeText(el.innerText);
  alert("Đã copy content mẫu");
}

function openPremiumPopup(){
  const m=document.getElementById("premiumPopup");
  if(m){m.style.display="flex";}
}
function closePremiumPopup(){
  const m=document.getElementById("premiumPopup");
  if(m){m.style.display="none";}
}

function toggleFloatingBot(){
  const panel=document.getElementById("floatingBotPanel");
  if(!panel) return;
  const isOpen=panel.style.display==="block";
  panel.style.display=isOpen?"none":"block";
  if(!isOpen){
    setTimeout(()=>appendBotGreeting(),250);
  }
}
function closeFloatingBot(){
  const panel=document.getElementById("floatingBotPanel");
  if(panel) panel.style.display="none";
}
let botGreeted=false;
function escapeBotText(value){
  return String(value || "").replace(/[&<>"]/g,function(ch){
    return {"&":"&amp;","<":"&lt;",">":"&gt;","\"":"&quot;"}[ch];
  });
}
function appendBotGreeting(){
  if(botGreeted) return;
  botGreeted=true;
  const body=document.getElementById("floatingBotBody");
  if(!body) return;
  body.innerHTML = `
    <div class="bot-msg ai">
      <b>Bot hỗ trợ:</b><br><br>
      Dạ mình đang cần hỗ trợ vấn đề gì ạ?<br><br>
      • Nâng cấp Premium<br>
      • Kích hoạt tài khoản<br>
      • Thanh toán<br>
      • Hướng dẫn sử dụng<br>
      • Báo lỗi hệ thống
    </div>`;
  body.scrollTop=body.scrollHeight;
}
function getBotReply(text){
  const lower=String(text || "").toLowerCase();
  let reply="Dạ mình đang cần hỗ trợ vấn đề gì ạ?<br><br>• Nâng cấp Premium<br>• Kích hoạt tài khoản<br>• Thanh toán<br>• Hướng dẫn sử dụng<br>• Báo lỗi hệ thống";
  if(lower.includes("premium") || lower.includes("giá") || lower.includes("gói") || lower.includes("nâng cấp")){
    reply="Dạ hiện hệ thống có các gói:<br><br>• 1 tháng: <b>159K</b><br>• 3 tháng: <b>359K</b><br>• 6 tháng: <b>559K</b><br>• 1 năm: <b>859K</b><br>• Nhà bán hàng chuyên nghiệp: <b>1.959K</b><br><br>Anh/chị muốn em tư vấn gói phù hợp không ạ?";
  }else if(lower.includes("thanh toán") || lower.includes("chuyển khoản") || lower.includes("qr")){
    reply="Dạ sau khi chuyển khoản thành công, anh/chị vui lòng gửi:<br><br>• ID thiết bị<br>• Ảnh thanh toán<br>• Gói đã đăng ký<br><br>Nếu sau 5 phút chưa được kích hoạt, vui lòng liên hệ Zalo <b>036 338 2629</b> hoặc Gmail <b>support@gptmini.pro</b>.";
  }else if(lower.includes("kích hoạt") || lower.includes("duyệt")){
    reply="Dạ để em hỗ trợ kích hoạt nhanh, anh/chị gửi giúp em:<br><br>• ID thiết bị<br>• Số điện thoại<br>• Gmail đăng ký<br>• Ảnh thanh toán<br>• Gói đã đăng ký";
  }else if(lower.includes("zalo") || lower.includes("liên hệ") || lower.includes("hỗ trợ")){
    reply="Dạ anh/chị có thể liên hệ hỗ trợ qua:<br><br><b>Zalo:</b> 036 338 2629<br><b>Gmail:</b> support@gptmini.pro<br><br>Nếu đã thanh toán, vui lòng gửi ID thiết bị, ảnh thanh toán và gói đã đăng ký để được ưu tiên kích hoạt nhanh ạ.";
  }else if(lower.includes("lỗi") || lower.includes("không dùng") || lower.includes("không được") || lower.includes("báo lỗi")){
    reply="Dạ anh/chị vui lòng mô tả lỗi đang gặp hoặc gửi ảnh màn hình giúp em ạ.<br><br>Nếu cần hỗ trợ nhanh, anh/chị liên hệ Zalo <b>036 338 2629</b>.";
  }else if(lower.includes("tính năng") || lower.includes("hướng dẫn") || lower.includes("sử dụng")){
    reply="Dạ tool hỗ trợ quản lý Fanpage, quản lý Group, đăng bài, AI Comment, AI Messenger, CRM Kanban và Marketing Director.<br><br>Anh/chị đang muốn dùng chức năng nào để em hướng dẫn đúng phần đó ạ?";
  }
  return reply;
}
function botQuick(text){
  const body=document.getElementById("floatingBotBody");
  if(!body) return;
  body.innerHTML += `<div class="bot-msg"><b>Bạn:</b> ${escapeBotText(text)}</div>`;
  body.innerHTML += `<div class="bot-msg ai bot-typing" id="botTypingNow"><b>Bot hỗ trợ:</b><br>Đang nhập<span class="typing-dots" style="vertical-align:middle"><span></span><span></span><span></span></span></div>`;
  body.scrollTop=body.scrollHeight;
  const reply=getBotReply(text);
  setTimeout(function(){
    const typing=document.getElementById("botTypingNow");
    if(typing){ typing.outerHTML = `<div class="bot-msg ai"><b>Bot hỗ trợ:</b><br>${reply}</div>`; }
    body.scrollTop=body.scrollHeight;
  }, 900 + Math.floor(Math.random()*700));
}
function sendBotInput(){
  const input=document.getElementById("botInputText");
  if(!input || !input.value.trim()) return;
  botQuick(input.value.trim());
  input.value="";
}
function updateLiveTrust(){
  const el=document.getElementById("liveMemberCount");
  if(!el) return;
  let base=Number(localStorage.getItem("v10_premium_count") || "231");
  base = base + 1;
  if(base > 999){ base = 231; }
  localStorage.setItem("v10_premium_count", String(base));
  el.innerText = base;
}
document.addEventListener("DOMContentLoaded",function(){
  const el=document.getElementById("liveMemberCount");
  if(el){ el.innerText = localStorage.getItem("v10_premium_count") || "231"; }
  setInterval(updateLiveTrust,4500);
});


function openModule(moduleId){
  const moduleAlias = {
    "group_marketing":"group_suite",
    "group_finder":"group_suite",
    "group_uid_splitter":"group_suite",
    "group_join_queue":"group_suite",
    "group_post_filter":"group_suite",
    "page_comment_pro":"page_center_total",
    "page_comment_queue":"page_center_total",
    "page_center":"page_center_total",
    "post":"page_center_total"
  };
  moduleId = moduleAlias[moduleId] || moduleId;
  const trialAllowed = ["dashboard", "fanpage_manager", "page_center_total", "group_suite", "group_marketing", "comment_manager"];
  const premiumLocked = {
    "messenger_ai": "AI Messenger",
    "crm_sales": "CRM Kanban",
    "marketing_director": "AI Marketing Director",
    "ai_video": "AI Video",
    "ai_image": "AI Image",
    "ai_business": "AI Kinh Doanh",
    "ai_voice": "AI Giọng Nói",
    "ai_livestream": "AI Livestream",
    "analytics_center": "Analytics Center"
  };

  document.querySelectorAll(".module-section").forEach(function(el){
    el.classList.remove("active-module");
  });
  const target=document.getElementById(moduleId);
  if(target){
    target.classList.add("active-module");
    setTimeout(function(){
      const top = target.getBoundingClientRect().top + window.pageYOffset - 16;
      window.scrollTo({top: top, behavior: "smooth"});
    }, 30);
  }
  document.querySelectorAll(".v2-nav-link").forEach(function(a){ a.classList.remove("active"); });
  const active=document.querySelector('.v2-nav-link[href="#'+moduleId+'"]');
  if(active){ active.classList.add("active"); }
  return false;
}
function showAllModules(){
  document.querySelectorAll(".module-section").forEach(function(el){
    el.classList.add("active-module");
  });
}
document.addEventListener("DOMContentLoaded",function(){
  const first=document.getElementById("dashboard");
  if(first){first.classList.add("active-module");}
  document.querySelectorAll(".v2-nav-link[href^='#']").forEach(function(a){
    a.addEventListener("click", function(e){
      const id=(a.getAttribute("href")||"").replace("#","");
      if(id){ e.preventDefault(); openModule(id); }
    });
  });
});


function scrollToPricing(){
  const el=document.getElementById("pricing");
  if(el){
    el.style.display="block";
    el.scrollIntoView({behavior:"smooth",block:"start"});
  }else{
    openPayment('monthly');
  }
}


const premiumPlans = {
  trial:{
    title:"🎁 DÙNG THỬ MIỄN PHÍ – 3 NGÀY",
    price:"0Đ",
    amount:"0",
    package:"TRIAL",
    desc:"Dùng thử 3 ngày sau khi đăng ký. Chỉ mở Quản lý Fanpage, Quản lý Group và AI Comment.",
    benefits:["Quản lý Fanpage", "Quản lý Group", "AI Comment", "Xem giao diện hệ thống", "Trải nghiệm quy trình bán hàng"],
    locked:["AI Messenger", "CRM Kanban", "AI Marketing Director", "AI Video", "AI Image", "AI Kinh Doanh", "AI Giọng Nói", "AI Livestream"]
  },
  monthly:{
    title:"🚀 GÓI 1 THÁNG",
    price:"159.000đ",
    amount:"159000",
    package:"1THANG",
    desc:"Phù hợp người mới bắt đầu dùng hệ thống để quản lý Fanpage, Group và AI Comment.",
    benefits:["Quản lý Fanpage", "Quản lý Group", "AI Comment", "Đăng bài Facebook", "Lịch đăng cơ bản", "Token Manager", "Hỗ trợ cơ bản"],
    locked:["AI Messenger nâng cao", "CRM Kanban", "AI Marketing Director", "AI Video/Image/Voice"]
  },
  quarterly:{
    title:"⭐ GÓI 3 THÁNG",
    price:"359.000đ",
    amount:"359000",
    package:"3THANG",
    desc:"Tối ưu cho shop đang bán hàng cần thêm Messenger AI và quản lý khách hàng.",
    benefits:["Toàn bộ gói 1 tháng", "AI Messenger", "CRM Kanban", "Kịch bản inbox", "Kịch bản chốt sale", "Quản lý khách hàng", "Ưu tiên hỗ trợ"],
    locked:["AI Marketing Director", "AI Video/Image/Voice", "AI Livestream"]
  },
  halfyear:{
    title:"💎 GÓI 6 THÁNG",
    price:"559.000đ",
    amount:"559000",
    package:"6THANG",
    desc:"Mở thêm AI Marketing Director và các công cụ tạo nội dung nâng cao.",
    benefits:["Toàn bộ gói 3 tháng", "AI Marketing Director", "AI Image", "AI Video", "AI Kinh Doanh", "Content Studio", "Không giới hạn Fanpage/Group theo cấu hình"],
    locked:["AI Giọng Nói nâng cao", "AI Livestream nâng cao", "VIP Support"]
  },
  yearly:{
    title:"🔥 GÓI 1 NĂM",
    price:"859.000đ",
    amount:"859000",
    package:"1NAM",
    desc:"Gói phổ biến nhất cho người bán hàng muốn dùng đầy đủ công cụ AI Marketing trong 1 năm.",
    benefits:["Toàn bộ gói 6 tháng", "AI Giọng Nói", "AI Livestream", "AI Automation", "Kho content Premium", "Ưu tiên xử lý", "Cập nhật miễn phí trong thời gian sử dụng"],
    locked:["Gói nhà bán hàng chuyên nghiệp", "Tư vấn VIP", "Cập nhật trọn đời"]
  },
  lifetime:{
    title:"👑 GÓI NHÀ BÁN HÀNG CHUYÊN NGHIỆP",
    price:"1.959.000đ",
    amount:"1959000",
    package:"NHABANHANGCHUYENNGHIEP",
    desc:"Gói cao nhất dành cho nhà bán hàng, agency và team kinh doanh muốn mở toàn bộ hệ thống.",
    benefits:["Toàn bộ tính năng Premium", "Không giới hạn AI theo cấu hình", "AI Marketing Director Pro", "CRM Pro", "AI Automation Pro", "AI Livestream", "AI Giọng Nói", "Export PDF/Excel", "Backup Database", "Ưu tiên hỗ trợ VIP", "Cập nhật trọn đời"],
    locked:[]
  },
  sellerpro:null,
  basic:null,
  pro:null,
  business:null
};
premiumPlans.sellerpro = premiumPlans.lifetime;
premiumPlans.basic = premiumPlans.monthly;
premiumPlans.pro = premiumPlans.quarterly;
premiumPlans.business = premiumPlans.halfyear;

function getOrCreateDeviceId(){
  let id = localStorage.getItem("mkt_device_id");
  if(!id){
    id = "MKT-" + Math.random().toString(36).slice(2,8).toUpperCase() + Date.now().toString().slice(-4);
    localStorage.setItem("mkt_device_id", id);
  }
  document.cookie = "mkt_device_id=" + encodeURIComponent(id) + "; path=/; max-age=" + (60*60*24*365*5);
  const el=document.getElementById("sidebarDeviceId"); if(el) el.innerText=id;
  const payDevice=document.getElementById("payDeviceId"); if(payDevice) payDevice.value=id;
  return id;
}
window.addEventListener("DOMContentLoaded", function(){ getOrCreateDeviceId(); setTimeout(getOrCreateDeviceId,500); });

function refreshPaymentContent(){
  if(window.currentPremiumPlanKey) openPayment(window.currentPremiumPlanKey);
}

function submitPremiumRequest(){
  const deviceId = getOrCreateDeviceId();
  const phone = (document.getElementById("payPhone")?.value || "").trim();
  const email = (document.getElementById("payEmail")?.value || "").trim();
  const planKey = window.currentPremiumPlanKey || "monthly";
  if(!phone || !email){ alert("Vui lòng nhập SĐT và Gmail để admin duyệt Premium nhanh hơn."); return false; }
  fetch("/premium_request", {method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify({device_id:deviceId, phone, email, package_key:planKey})})
    .then(r=>r.json()).then(data=>{
      if(data.ok){
        document.getElementById("payContent").innerText = data.payment_note;
        showPaymentNotice("Đã gửi yêu cầu thanh toán về web admin. Anh/chị vui lòng giữ lại ảnh giao dịch để đối chiếu khi cần.");
      }else alert(data.message || "Chưa gửi được yêu cầu, vui lòng thử lại.");
    }).catch(()=>alert("Kết nối chậm, vui lòng thử lại hoặc gửi Zalo hỗ trợ."));
  return false;
}

function openPayment(planKey){
  const plan=premiumPlans[planKey] || premiumPlans.basic;
  const modal=document.getElementById("paymentModal");
  if(!modal) return;

  const amountText = Number(plan.amount).toLocaleString("vi-VN") + " VNĐ";
  const deviceId = getOrCreateDeviceId();
  const phone = (document.getElementById("payPhone")?.value || "").trim();
  const email = (document.getElementById("payEmail")?.value || "").trim();
  const addInfo = deviceId + " | " + phone + " | " + email + " | " + plan.package.toUpperCase();
  const qrUrl = "https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount=" + encodeURIComponent(plan.amount) + "&addInfo=" + encodeURIComponent(addInfo) + "&accountName=" + encodeURIComponent("NGUYEN DANG THI XUAN");
  window.currentPremiumPlanKey = planKey;

  document.getElementById("payPlanTitle").innerText=plan.title;
  document.getElementById("payPlanPrice").innerText=amountText;
  document.getElementById("payPlanDesc").innerText=plan.desc;
  document.getElementById("payQr").src=qrUrl;
  document.getElementById("payContent").innerText=addInfo;
  document.getElementById("payBenefits").innerHTML=plan.benefits.map(x=>"<div>"+x+"</div>").join("");
  document.getElementById("payLocked").innerHTML=plan.locked.length ? plan.locked.map(x=>"<div>🔒 "+x+"</div>").join("") : "<div>Không khóa tính năng</div>";
  modal.style.display="flex";
}
function closePayment(){
  const modal=document.getElementById("paymentModal");
  if(modal) modal.style.display="none";
}


function openLockedFeature(feature, plans){
  const modal=document.getElementById("lockedFeatureModal");
  if(!modal) {
    scrollToPricing();
    return false;
  }

  document.getElementById("lockedFeatureTitle").innerText="🔒 Tính năng Premium";
  document.getElementById("lockedFeaturePlans").innerHTML =
    `<div class="premium-lock-hero v4-lock-hero">
       <p>Bạn đang muốn mở: <b>${feature}</b></p>
       <p>Gói dùng thử 3 ngày chỉ hỗ trợ:</p>
       <ul class="benefit-list">
         <li class="open">Quản lý Fanpage</li>
         <li class="open">Quản lý Group</li>
         <li class="open">AI Comment</li>
       </ul>
       <p><b>Nâng cấp Premium để mở khóa:</b></p>
       <ul class="benefit-list">
         <li class="open">AI Messenger</li>
         <li class="open">CRM Kanban</li>
         <li class="open">AI Marketing Director</li>
         <li class="open">AI Video • AI Image • AI Kinh Doanh</li>
         <li class="open">AI Giọng Nói • AI Livestream • AI Automation</li>
       </ul>
       <div class="locked-recommend-grid">
         <div class="locked-recommend-card">
           <div class="rec-label">🚀 BẮT ĐẦU</div>
           <h3>Gói 1 tháng</h3>
           <div class="rec-price">159.000đ</div>
           <p>Phù hợp để mở các công cụ quản lý cơ bản và dùng ổn định.</p>
           <button onclick="closeLockedFeature();openPayment('monthly')">Chọn gói 1 tháng</button>
         </div>
         <div class="locked-recommend-card featured">
           <div class="rec-label">⭐ PHỔ BIẾN</div>
           <h3>Gói 1 năm</h3>
           <div class="rec-price">859.000đ</div>
           <p>Mở đầy đủ AI Marketing, CRM, Automation và công cụ bán hàng nâng cao.</p>
           <button onclick="closeLockedFeature();openPayment('yearly')">Xem gói 1 năm</button>
         </div>
         <div class="locked-recommend-card v4-gold-card" style="grid-column:1/-1">
           <div class="rec-label">👑 CAO NHẤT</div>
           <h3>Gói nhà bán hàng chuyên nghiệp</h3>
           <div class="rec-price">1.959.000đ</div>
           <p>Mở toàn bộ tính năng cao cấp, hỗ trợ VIP và cập nhật trọn đời.</p>
           <button onclick="closeLockedFeature();openPayment('lifetime')">Mở gói chuyên nghiệp</button>
         </div>
       </div>
       <button class="secondary" onclick="closeLockedFeature();scrollToPricing()">Xem chi tiết tất cả gói</button>
       <p class="small">Sau khi thanh toán 5 phút chưa kích hoạt, vui lòng gửi ảnh giao dịch qua Zalo 036 338 2629.</p>
     </div>`;
  modal.style.display="flex";
  return false;
}
function closeLockedFeature(){
  const modal=document.getElementById("lockedFeatureModal");
  if(modal) modal.style.display="none";
}

</script>

<style>
.premium-grid-pro{grid-template-columns:repeat(auto-fit,minmax(230px,1fr)) !important;}
.locked-plan-actions{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin:14px 0;}
.premium-trial-stat{background:linear-gradient(135deg,#FFFFFF,#F5F3FF)!important;border-left-color:#F59E0B!important}
.token-status-stat{background:linear-gradient(135deg,#FFFFFF,#ECFDF5)!important;border-left-color:#22C55E!important}
.fb-submenu-pro{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:10px;margin:14px 0 18px}
.fb-submenu-pro button{width:100%;margin:0;box-shadow:none}
.fb-submenu-pro .locked{background:linear-gradient(135deg,#EEF2FF,#F5F3FF);color:#4C1D95;border:1px solid #DDD6FE}
.premium-lock-hero p{line-height:1.55}.locked-recommend-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin:14px 0}.locked-recommend-card{background:white;border:1px solid #E5E7EB;border-radius:22px;padding:16px;box-shadow:0 14px 32px rgba(30,41,59,.10)}.locked-recommend-card.featured{border:2px solid #7C3AED;background:linear-gradient(180deg,#FFFFFF,#F5F3FF)}.rec-label{display:inline-block;background:linear-gradient(135deg,#2563EB,#7C3AED);color:white;border-radius:999px;padding:5px 10px;font-size:11px;font-weight:900}.rec-price{font-size:28px;font-weight:900;color:#7C3AED;margin:6px 0}@media(max-width:720px){.locked-recommend-grid{grid-template-columns:1fr}}
.locked-plan-actions button{width:100%;margin:0;}

.trial-box{background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.14);border-radius:16px;padding:12px;margin:12px 0;line-height:1.75;color:#E5E7EB}
.locked-list{background:rgba(239,68,68,.08);border-color:rgba(239,68,68,.25)}
.benefit-list .locked{color:#DC2626;list-style:none;margin:6px 0;font-weight:800}
.v2-nav-link[href="#messenger_ai"] .v2-nav-tag,
.v2-nav-link[href="#crm_sales"] .v2-nav-tag,
.v2-nav-link[href="#marketing_director"] .v2-nav-tag{background:linear-gradient(135deg,#F59E0B,#EF4444)!important;color:white!important}

.comment-guide{background:#EEF4FF;border:1px solid #C8D8FF;padding:12px;border-radius:14px;margin:8px 0 10px;font-size:14px;line-height:1.5;color:#1E293B}
.support-mini-box{background:#F4F7FF;border:1px solid #D8E0FF;border-radius:14px;padding:10px 12px;margin:9px 0;font-size:14px;color:#111827}
.support-mini-box b{color:#4C1D95}

</style>


<style id="chat-device-menu-fix">
/* Bản sửa: chỉ dọn icon menu, tăng ưu tiên chat, hiển thị ID thiết bị rõ ràng */
.v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important;overflow:hidden!important}
.v2-nav-link{cursor:pointer!important;pointer-events:auto!important}
.activity-card span{font-size:0!important}
.activity-card span::before{font-size:16px!important}
.activity-card:nth-of-type(1) span::before{content:"Tổng bài"}
.activity-card:nth-of-type(2) span::before{content:"Đã đăng"}
.activity-card:nth-of-type(3) span::before{content:"Chờ đăng"}
.activity-card:nth-of-type(4) span::before{content:"Lead CRM"}
.activity-card:nth-of-type(5) span::before{content:"Chiến dịch"}
#sidebarDeviceId{display:inline-block!important;color:#fff!important;font-weight:900!important;font-size:14px!important;letter-spacing:.4px;margin:4px 0}
.floating-bot{z-index:2147483000!important;pointer-events:auto!important}
.bot-bubble,.bot-panel,.bot-actions button,.bot-input button{pointer-events:auto!important}
.bot-panel{z-index:2147483001!important}
.bot-body{min-height:210px!important}
.bot-actions{grid-template-columns:1fr 1fr!important}
.bot-actions button{font-size:13px!important;line-height:1.25!important}
.bot-title::before{content:"🤖 ";}
</style>
</head>
<body>

<div class="v2-topbar">
  <div class="v2-brand-row">
    <div>
      <div class="v2-brand-title">🚀 Mkt Automation Pro V5</div>
      <div class="v2-brand-sub">AI Marketing • Facebook • CRM • Automation</div>
    </div>
    <div class="v2-status-pill">Online</div>
  </div>
</div>


<div class="live-trust-bar">
  <span class="live-dot"></span>
  <span>Đã có <b id="liveMemberCount">231</b> khách hàng nâng cấp Premium</span>
</div>

<div class="floating-bot">
  <div class="bot-panel" id="floatingBotPanel">
    <div class="bot-head">
      <div>
        <div class="bot-title">Mini Chat Support</div>
        <div class="bot-online">Đang trực tuyến
          <span class="typing-dots"><span></span><span></span><span></span></span>
        </div>
      </div>
      <button class="bot-close" onclick="closeFloatingBot()">×</button>
    </div>
    <div class="bot-body" id="floatingBotBody">
      <div class="bot-msg ai">
        <b>Bot hỗ trợ:</b><br><br>
        Dạ mình đang cần hỗ trợ vấn đề gì ạ?<br><br>
        • Nâng cấp Premium<br>
        • Kích hoạt tài khoản<br>
        • Thanh toán<br>
        • Hướng dẫn sử dụng<br>
        • Báo lỗi hệ thống
      </div>
    </div>
    <div class="bot-actions">
      <button onclick="botQuick('Nâng cấp Premium')">Nâng cấp Premium</button>
      <button class="light" onclick="botQuick('Hướng dẫn thanh toán')">Hướng dẫn thanh toán</button>
      <button class="light" onclick="botQuick('Kích hoạt tài khoản')">Kích hoạt tài khoản</button>
      <button class="light" onclick="botQuick('Liên hệ hỗ trợ')">Liên hệ hỗ trợ</button>
    </div>
    <div class="bot-input">
      <input id="botInputText" placeholder="Nhập câu hỏi cần hỗ trợ..." onkeydown="if(event.key==='Enter')sendBotInput()">
      <button onclick="sendBotInput()">Gửi</button>
    </div>
  </div>
  <button class="bot-bubble" onclick="toggleFloatingBot()">
    🤖
    <span class="bot-status"></span>
  </button>
</div>

<div class="layout">
<aside class="sidebar">
  <div class="logo">Marketing<br>Automation Pro</div>
  <div class="v2-side-card device-premium-card" style="margin-top:12px;background:linear-gradient(135deg,#111827,#1e293b);border:1px solid rgba(250,204,21,.45)">
    <b>🖥 ID thiết bị</b><br>
    <div class="device-id-row">
      <span id="sidebarDeviceId">{{ device_id }}</span>
      {% if is_device_premium %}
      <span class="premium-glow-badge">PREMIUM</span>
      {% endif %}
    </div>
    <small id="sidebarPremiumStatus">
      {% if is_device_premium %}
        Trạng thái: <b style="color:#22c55e">Đã kích hoạt Premium</b>
      {% else %}
        Trạng thái: {{ free_status.label }}
      {% endif %}
    </small>
  </div>
  {% if renewal_notice %}
  <div class="v2-side-card" style="background:#fff7ed;color:#7c2d12;border:1px solid #fed7aa">
    <b>⏰ Sắp hết hạn Premium</b><br>
    Gói {{ renewal_notice.package_name }} còn {{ renewal_notice.remaining_days }} ngày. Vui lòng gia hạn để không gián đoạn.
  </div>
  {% endif %}

<div class="nav">
  <div class="v2-nav-title">MENU CHÍNH</div>
  <a class="v2-nav-link" href="#dashboard"><span class="v2-nav-ico">🏠</span><span class="v2-nav-text">Dashboard CEO</span><span class="v2-nav-tag">Home</span></a>

  <div class="v2-nav-title">FACEBOOK CENTER</div>
  <a class="v2-nav-link" href="#facebook_center" onclick="return openModule('facebook_center')"><span class="v2-nav-ico">📣</span><span class="v2-nav-text">Facebook Center</span><span class="v2-nav-tag">Core</span></a>
  <a class="v2-nav-link" href="#page_center_total" onclick="return openModule('page_center_total')"><span class="v2-nav-ico"></span><span class="v2-nav-text">Page Center Tổng</span><span class="v2-nav-tag">V11</span></a>
  <a class="v2-nav-link" href="#post" onclick="return openModule('page_center_total')"><span class="v2-nav-ico"></span><span class="v2-nav-text">Đăng bài Facebook</span></a>
  <a class="v2-nav-link" href="#fanpage_manager" onclick="return openModule('fanpage_manager')"><span class="v2-nav-ico">📄</span><span class="v2-nav-text">Quản lý Fanpage</span><span class="v2-nav-tag">V5</span></a>
  <a class="v2-nav-link" href="#group_suite" onclick="return openModule('group_suite')"><span class="v2-nav-ico"></span><span class="v2-nav-text">Group Center Tổng</span><span class="v2-nav-tag">V11</span></a>

  <div class="v2-nav-title">SELLER AI</div>
  <a class="v2-nav-link" href="#comment_manager" onclick="return openModule('comment_manager')"><span class="v2-nav-ico">🤖</span><span class="v2-nav-text">AI Comment</span><span class="v2-nav-tag">AI</span></a>
  <a class="v2-nav-link" href="#messenger_ai" onclick="return openModule('messenger_ai')"><span class="v2-nav-ico">💬</span><span class="v2-nav-text">AI Messenger</span><span class="v2-nav-tag">AI</span></a>
  <a class="v2-nav-link" href="#crm_sales" onclick="return openModule('crm_sales')"><span class="v2-nav-ico">📋</span><span class="v2-nav-text">CRM Kanban</span><span class="v2-nav-tag">CRM</span></a>
  <a class="v2-nav-link" href="#marketing_director" onclick="return openModule('marketing_director')"><span class="v2-nav-ico">🧠</span><span class="v2-nav-text">AI Marketing Director</span><span class="v2-nav-tag">HOT</span></a>

  <div class="v2-nav-title">AI STUDIO</div>
  <a class="v2-nav-link" href="#ai_studio" onclick="return openModule('ai_studio')"><span class="v2-nav-ico">🎨</span><span class="v2-nav-text">AI Studio</span><span class="v2-nav-tag">Pro</span></a>
  <a class="v2-nav-link" href="#creative_center" onclick="return openModule('creative_center')"><span class="v2-nav-ico">🖼️</span><span class="v2-nav-text">Image / Video / Voice</span></a>

  <div class="v2-nav-title">AI BUSINESS</div>
  <a class="v2-nav-link" href="#ai_studio" onclick="return openModule('ai_studio')"><span class="v2-nav-ico">🚀</span><span class="v2-nav-text">AI Facebook</span></a>
  <a class="v2-nav-link" href="#marketing_director" onclick="return openModule('marketing_director')"><span class="v2-nav-ico">💼</span><span class="v2-nav-text">AI Kinh Doanh</span></a>

  <div class="v2-nav-title">HỆ THỐNG</div>
  <a class="v2-nav-link" href="#premium" onclick="return openModule('premium')"><span class="v2-nav-ico">💎</span><span class="v2-nav-text">Premium</span><span class="v2-nav-tag">VIP</span></a>
  <a class="v2-nav-link" href="#analytics" onclick="return openModule('analytics')"><span class="v2-nav-ico">📈</span><span class="v2-nav-text">Analytics Center</span></a>
  <a class="v2-nav-link" href="#automation_center" onclick="return openModule('automation_center')"><span class="v2-nav-ico">⚙️</span><span class="v2-nav-text">Cài đặt Automation</span></a>
  <a data-aff-menu="1" class="v2-nav-link" href="#affiliate_center" onclick="return openModule('affiliate_center')"><span class="v2-nav-ico"></span><span class="v2-nav-text">CTV Hoa Hồng</span><span class="v2-nav-tag" style="background:#dcfce7;color:#166534">CTV</span></a>

  <div class="v2-side-card">
    🚀 Mkt Automation Pro V5<br>
    Fanpage • Group • Comment • Messenger • CRM • Marketing Director.
  </div>
</div>
</aside>

<main class="main">

<section class="top-hero" id="dashboard">
  <h1>Mkt Automation Pro V5 Seller AI Suite</h1>

<div class="app-install-banner">
  <b>📲 App Mini đã sẵn sàng</b><br>
  Mở trên điện thoại rồi bấm “Thêm vào màn hình chính” để dùng như phần mềm.
  <button onclick="showInstallGuide()">Hướng dẫn cài vào điện thoại</button>
</div>

<div class="app-quick-grid">
  <div class="app-quick-card" onclick="return openModule('post')">
    <div class="app-ico">📢</div>
    <b>Đăng bài Facebook</b>
    <span>Soạn nội dung, chọn Page, đăng ngay hoặc lên lịch.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('ai_studio')">
    <div class="app-ico">🤖</div>
    <b>Tạo Content AI</b>
    <span>Viết bài, caption, ý tưởng quảng cáo và nội dung bán hàng.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('fanpage_manager')">
    <div class="app-ico">📄</div>
    <b>Quản lý Fanpage</b>
    <span>Kiểm tra Page, Token, quyền đăng bài và trạng thái hoạt động.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('group_suite')">
    <div class="app-ico">👥</div>
    <b>Quản lý Group</b>
    <span>Lưu group, tạo lịch đăng group và viết bài seeding mềm.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('comment_manager')">
    <div class="app-ico">💬</div>
    <b>AI Comment</b>
    <span>Ẩn số điện thoại, trả lời comment, gắn nhãn và chuyển CRM.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('messenger_ai')">
    <div class="app-ico">📨</div>
    <b>AI Messenger</b>
    <span>Tạo kịch bản inbox, chốt sale, xử lý từ chối và chăm sóc lại.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('crm_sales')">
    <div class="app-ico">📊</div>
    <b>CRM Kanban</b>
    <span>Quản lý khách theo các cột: mới, tư vấn, báo giá, chốt, đã mua.</span>
  </div>
  <div class="app-quick-card" onclick="return openModule('marketing_director')">
    <div class="app-ico">🧠</div>
    <b>AI Marketing Director</b>
    <span>Lập kế hoạch 30 ngày, ads, content, funnel và KPI bán hàng.</span>
  </div>
  <div class="app-quick-card" onclick="location.href='#premium'">
    <div class="app-ico">💎</div>
    <b>Premium</b>
    <span>Mở khóa hạn mức cao hơn và tính năng nâng cao.</span>
  </div>
</div>

  <p>AI Marketing Automation Platform: tạo content, đăng Fanpage, chia lịch, quản lý CRM, phân tích đối thủ, tạo funnel và hỗ trợ bán hàng tự động.</p>

  {% if message %}
  <div class="{{ 'success' if ok else 'error' }}"><b>Thông báo:</b> {{ message }}</div>
  {% endif %}

  <div class="v3-kpi-title"><b>KPI tổng quan</b> • Fanpage đang hoạt động • Khách hàng mới hôm nay • Bài đăng hôm nay • Doanh thu tháng • Trạng thái Token</div>
  <div class="grid4 v3-ceo-grid" style="margin-top:18px">
    <div class="stat"><span>Fanpage</span><br><b>{{ v3.fanpages }}</b><small>đang hoạt động/cấu hình</small></div>
    <div class="stat"><span>Bài đăng</span><br><b>{{ v3.posts }}</b><small>hôm nay: {{ v3.posts_today }}</small></div>
    <div class="stat"><span>Khách hàng</span><br><b>{{ v3.crm }}</b><small>mới hôm nay: {{ v3.leads_today }}</small></div>
    <div class="stat"><span>Doanh thu</span><br><b>{{ v3.revenue_month }}</b><small>Chưa phát sinh giao dịch</small></div>
    <div class="stat premium-trial-stat"><span>Premium</span><br><b>🎁 Dùng thử miễn phí</b><small>Còn lại: 3 ngày • 10/10 Content AI • 1/1 Fanpage</small></div>
    <div class="stat token-status-stat"><span>Token</span><br><b>🟢 {{ v3.token_live }} sống</b><small>🔴 {{ v3.token_dead }} lỗi • Tổng Page: {{ v3.token_total }}</small></div>
  </div>

  <div class="v4-trust-grid">
    <div class="v4-trust-card"><b>2.381+</b><span>Khách hàng sử dụng</span></div>
    <div class="v4-trust-card"><b>120.000+</b><span>Content đã tạo</span></div>
    <div class="v4-trust-card"><b>98%</b><span>Đánh giá hài lòng</span></div>
  </div>

  <div class="hero-actions">
    <button onclick="return openModule('post')">Bắt đầu đăng bài</button>
    <button class="secondary" onclick="return openModule('ai_studio')">Tạo content AI</button>
    <button class="secondary" onclick="scrollToPricing()">Xem bảng giá</button>
  </div>

  <div class="module-hub v3-main-hub">
    <div class="module-card" onclick="return openModule('facebook_center')"><div class="icon">📣</div><h3>Facebook Center</h3><p>Trung tâm đăng bài, lên lịch, quản lý Fanpage, Group, Token và lịch sử đăng.</p><span class="module-pill">Mở Facebook Center</span></div>
    <div class="module-card" onclick="return openModule('fanpage_manager')"><div class="icon">📄</div><h3>Quản lý Fanpage</h3><p>Kiểm tra Page, Token, quyền đăng bài và trạng thái hoạt động.</p><span class="module-pill">Mở Fanpage</span></div>
    <div class="module-card" onclick="return openModule('group_suite')"><div class="icon">👥</div><h3>Quản lý Group</h3><p>Lưu Group, tạo lịch đăng Group và viết bài seeding mềm.</p><span class="module-pill">Mở Group</span></div>
    <div class="module-card" onclick="return openModule('comment_manager')"><div class="icon">🤖</div><h3>AI Comment</h3><p>Ẩn số điện thoại, trả lời bình luận, gắn nhãn khách nóng/ấm/lạnh và chuyển CRM.</p><span class="module-pill">Mở AI Comment</span></div>
    <div class="module-card" onclick="return openModule('messenger_ai')"><div class="icon">💬</div><h3>AI Messenger</h3><p>Tạo kịch bản inbox, chốt sale, xử lý từ chối và chăm sóc lại.</p><span class="module-pill">Mở AI Messenger</span></div>
    <div class="module-card" onclick="return openModule('crm_sales')"><div class="icon">📋</div><h3>CRM Kanban</h3><p>Quản lý khách theo các cột: mới, tư vấn, báo giá, theo dõi, đã chốt.</p><span class="module-pill">Mở CRM</span></div>
    <div class="module-card" onclick="return openModule('marketing_director')"><div class="icon">🧠</div><h3>AI Marketing Director</h3><p>Lập kế hoạch 30 ngày, Ads, Content, Funnel, KPI và chiến lược tăng doanh thu.</p><span class="module-pill">Mở Marketing Director</span></div>
    <div class="module-card" onclick="return openModule('ai_studio')"><div class="icon">🎨</div><h3>AI Studio</h3><p>Gộp AI Facebook, AI Image, AI Video, AI Giọng Nói và AI Livestream vào một khu vực.</p><span class="module-pill">Mở AI Studio</span></div>
    <div class="module-card" onclick="return openModule('premium')"><div class="icon">💎</div><h3>Premium</h3><p>Mở khóa hạn mức cao hơn, module nâng cao và hỗ trợ ưu tiên.</p><span class="module-pill">Xem gói</span></div>
  </div>
</section>

  {% if content %}
  <div class="panel" id="ai-output-box">
    <h2>Kết quả AI vừa tạo</h2>
    <p class="small">Nội dung AI tạo sẽ hiển thị tại đây để khách xem, copy hoặc đưa sang phần đăng Fanpage.</p>
    <div class="history" id="aiGeneratedContent">{{ content }}</div>
    <button onclick="copyText('aiGeneratedContent')">Copy nội dung</button>
    <button class="secondary" onclick="return openModule('post')">Mở phần đăng Fanpage</button>
  </div>
  {% endif %}



<section class="panel module-section" id="facebook_center">
  <div class="section-open-note">Bạn đang mở: Facebook Center V3.</div>
  <h2>📢 Facebook Center</h2>
  <p class="small">Gộp toàn bộ công cụ Facebook vào một trung tâm: đăng bài, Scheduler, Fanpage Manager, Group Marketing, Comment Manager và Messenger AI.</p>
  <div class="fb-submenu-pro">
    <button onclick="return openModule('post')">📢 Đăng bài</button>
    <button onclick="return openModule('scheduler')">📅 Scheduler</button>
    <button onclick="return openModule('fanpage_manager')">📄 Fanpage Manager</button>
    <button onclick="return openModule('group_suite')">👥 Group Marketing</button>
    <button onclick="return openModule('comment_manager')">💬 Comment Manager AI</button>
    <button onclick="return openModule('messenger_ai')">🤖 Messenger AI</button>
  </div>
  <div class="v3-feature-grid">
    <div class="v3-feature-card"><h3>Đăng bài</h3><ul><li>Đăng ngay</li><li>Đăng hàng loạt</li><li>Đăng nhiều Page</li><li>Đăng ảnh/video</li></ul><button onclick="return openModule('post')">Mở công cụ đăng bài</button></div>
    <div class="v3-feature-card"><h3>Scheduler</h3><ul><li>Lên lịch tự động</li><li>Đăng chiến dịch</li><li>Chia khung giờ</li><li>Tự lưu lịch</li></ul><button onclick="return openModule('scheduler')">Mở lịch đăng</button></div>
    <div class="v3-feature-card"><h3>Quản lý Fanpage</h3><ul><li>Kết nối Fanpage</li><li>Kiểm tra Token</li><li>Kiểm tra quyền</li><li>Trạng thái hoạt động</li><li>Làm mới Token</li></ul><button onclick="return openModule('fanpage_manager')">Mở Fanpage Manager</button></div>
    <div class="v3-feature-card"><h3>Tiếp thị nhóm</h3><ul><li>Quản lý Group</li><li>Danh sách Group</li><li>Lịch đăng Group</li><li>AI viết bài Group</li></ul><button onclick="return openModule('group_suite')">Mở Group Marketing</button></div>
    <div class="v3-feature-card"><h3>Trình quản lý bình luận</h3><ul><li>AI trả lời comment</li><li>Ẩn SĐT</li><li>Gắn nhãn khách</li><li>Chuyển CRM</li></ul><button onclick="return openModule('comment_manager')">Mở Comment AI</button></div>
    <div class="v3-feature-card"><h3>Trí tuệ nhân tạo Messenger</h3><ul><li>Kịch bản Inbox</li><li>Kịch bản Chốt Sale</li><li>Xử lý từ chối</li><li>Chăm sóc khách cũ</li></ul><button onclick="return openModule('messenger_ai')">Mở Messenger AI</button></div>
  </div>
</section>


<section class="panel module-section" id="fanpage_manager">
  <div class="section-open-note">Bạn đang mở: Fanpage Manager Pro.</div>
  <h2>📄 Quản lý Fanpage</h2>
  <p class="small">Trung tâm kiểm tra kết nối Page, Token, quyền và trạng thái hoạt động trước khi đăng bài.</p>
  <div class="v5-seller-grid">
    <div class="v5-tool-card"><h3>Kết nối Fanpage</h3><p>Thêm Page qua biến PAGES_JSON trong file .env gồm name, id, token.</p><button onclick="return openModule('token')">Mở Token Center</button></div>
    <div class="v5-tool-card"><h3>Kiểm tra Token</h3><p>Quét toàn bộ Page để phát hiện token chết, thiếu quyền hoặc phiên bị giới hạn.</p><form method="post" action="/check_tokens"><button>Kiểm tra Token ngay</button></form></div>
    <div class="v5-tool-card"><h3>Kiểm tra quyền</h3><ul><li>pages_manage_posts</li><li>pages_read_engagement</li><li>pages_manage_metadata</li></ul><span class="v5-warning-pill">Cần kiểm tra từ Meta App</span></div>
    <div class="v5-tool-card"><h3>Làm mới Token</h3><p>Khi token lỗi, hệ thống hướng dẫn thay token mới trong .env rồi chạy lại app.</p><button class="secondary" onclick="openLockedFeature('Làm mới Token','Gói 1 năm / Gói Nhà Bán Hàng Chuyên Nghiệp')">Hướng dẫn làm mới</button></div>
  </div>
  <table class="v5-table"><tr><th>Fanpage</th><th>Page ID</th><th>Token</th><th>Trạng thái</th></tr>
    {% for p in pages %}<tr><td>{{ p.name }}</td><td>{{ p.id }}</td><td>{{ 'Có token' if p.token else 'Thiếu token' }}</td><td><span class="v5-status-pill">Đã cấu hình</span></td></tr>{% endfor %}
  </table>
</section>



<section class="panel module-section" id="page_center_total">
  <div class="section-open-note">Bạn đang mở: Page Center Tổng V11.</div>
  <h2>Page Center Tổng V11 - Token, đăng bài, hẹn giờ và bình luận UID theo từng Page</h2>
  <p class="small">Chọn nhiều Page, nhập nhiều nội dung mỗi dòng một bài, chia nội dung không trùng nhau. Có nút Đăng ngay, Hẹn giờ nhanh 30 phút / 1h / 2h / 3h. Bình luận UID người dùng, UID bài viết, UID nhóm được đưa vào hàng chờ duyệt với giãn cách 45-60 giây.</p>
  <div class="gf-warning">Chỉ dùng với Page và bài viết/Group mà bạn có quyền quản lý hợp lệ. Bình luận UID được lưu hàng chờ và cần duyệt trước; không thiết kế spam tự động hàng loạt.</div>

  <div class="gf-box">
    <h3>0. Thêm Page ID & Page Token trực tiếp trong tool</h3>
    <form method="post" action="/page_token_save">
      <div class="gf-grid-3"><input name="page_name" placeholder="Tên Page"><input name="page_id" placeholder="Page ID"><input name="page_token" placeholder="Page Access Token"></div>
      <textarea name="note" rows="2" placeholder="Ghi chú quyền: pages_manage_posts, pages_read_engagement, Page đã tham gia Group nào..."></textarea>
      <button>Lưu / cập nhật Page Token</button>
      <button type="submit" formaction="/check_tokens">Kiểm tra toàn bộ Token</button>
    </form>
    <div style="max-height:180px;overflow:auto"><table class="gf-table"><tr><th>Page</th><th>Page ID</th><th>Token</th><th>Trạng thái</th><th>Cập nhật</th></tr>{% for t in page_token_rows %}<tr><td>{{ t[1] }}</td><td>{{ t[2] }}</td><td>{{ t[3] }}</td><td>{{ t[4] }}</td><td>{{ t[6] }}</td></tr>{% endfor %}</table></div>
    <p class="small">Sau khi lưu Token, Page xuất hiện ngay ở danh sách chọn Page bên dưới và ở Group Center. Không cần sửa PAGES_JSON trên Render Environment.</p>
  </div>

  <div class="gf-grid-3">
    <div class="gf-stat"><span>Fanpage cấu hình</span><b>{{ pages|length }}</b></div>
    <div class="gf-stat"><span>Bài đã đăng/lên lịch</span><b>{{ s.total }}</b></div>
    <div class="gf-stat"><span>Hàng chờ bình luận</span><b>{{ page_comment_stats.total }}</b></div>
  </div>

  <div class="gf-box">
    <h3>1. Đăng bài nhiều Page - chia nội dung không trùng</h3>
    <form method="post" action="/multi_post" enctype="multipart/form-data">
      <div class="gf-grid-3">
        <div>
          <label>Chọn nhiều Page hoạt động</label>
          <div class="gf-box" style="max-height:220px;overflow:auto;padding:10px">
            {% for p in pages %}<label style="display:block;margin:6px 0"><input type="checkbox" name="page_indexes" value="{{ loop.index0 }}"> {{ p.name }} - {{ p.id }}</label>{% endfor %}
          </div>
        </div>
        <div>
          <label>Nội dung - mỗi dòng là 1 bài</label>
          <textarea name="bulk_content" rows="9" placeholder="Bài 1...\nBài 2...\nBài 3...\nHệ thống sẽ chia lần lượt cho Page, không trùng nội dung."></textarea>
        </div>
        <div>
          <label>Ảnh/video nếu có</label>
          <input type="file" name="images" multiple>
          <input name="campaign" placeholder="Tên chiến dịch / ghi chú">
          <label><input type="checkbox" name="use_ai_enhance" value="1"> AI viết lại nhẹ để tránh trùng giọng văn</label>
          <small class="small">Mặc định giữ nguyên nội dung. Chỉ bật AI nếu muốn biến thể nhẹ.</small>
        </div>
      </div>
      <div class="gf-grid-3">
        <button name="action" value="now">Đăng ngay</button>
        <button name="action" value="schedule_quick_30">Hẹn giờ 30 phút</button>
        <button name="action" value="schedule_quick_60">Hẹn giờ 1h</button>
      </div>
      <div class="gf-grid-3">
        <button name="action" value="schedule_quick_120">Hẹn giờ 2h</button>
        <button name="action" value="schedule_quick_180">Hẹn giờ 3h</button>
        <input type="datetime-local" name="schedule_time" placeholder="Hoặc chọn giờ cụ thể">
      </div>
    </form>
  </div>

  <div class="gf-box">
    <h3>2. Bình luận UID bằng Page - từng Page riêng, hàng chờ duyệt 45-60 giây</h3>
    <p class="small">Chọn 1 hoặc nhiều Page. Mỗi Page sẽ tạo hàng chờ bình luận riêng theo UID người dùng / UID bài viết / UID Group. Không tự spam; admin phải duyệt và ghi log.</p>
    <form method="post" action="/page_comment_queue_add">
      <div class="gf-grid-3">
        <div>
          <label>Chọn nhiều Page bình luận</label>
          <div class="gf-box" style="max-height:180px;overflow:auto;padding:10px">
            {% for p in pages %}<label style="display:block;margin:6px 0"><input type="checkbox" name="page_indexes" value="{{ loop.index0 }}"> {{ p.name }} - {{ p.id }}</label>{% endfor %}
          </div>
        </div>
        <div>
          <label>Loại UID đích</label>
          <select name="target_type"><option value="post">UID bài viết</option><option value="group">UID nhóm</option><option value="user">UID người dùng đã tương tác</option></select>
          <input name="min_delay" type="number" min="45" value="45" placeholder="Giãn cách tối thiểu 45 giây">
          <input name="max_delay" type="number" min="45" value="60" placeholder="Giãn cách tối đa 60 giây">
        </div>
        <div>
          <label>Nội dung bình luận</label>
          <div class="comment-guide"><b>Hướng dẫn:</b><br>Mỗi dòng là 1 bình luận riêng.<br>Hệ thống tự ghép 1 UID = 1 bình luận.<br>Nếu UID nhiều hơn số bình luận, nội dung sẽ tự quay vòng.</div>
          <textarea name="comment_text" rows="6" placeholder="Mỗi dòng là 1 bình luận riêng.\nVí dụ:\nxin chào\nbạn khỏe không\nngày mai bạn làm gì"></textarea>
        </div>
      </div>
      <textarea name="raw_targets" rows="6" placeholder="Dán nhiều UID, mỗi dòng một mục. Ví dụ:\nPOST_UID\nPOST_UID, GROUP_UID\nUSER_UID"></textarea>
      <div class="gf-grid-3"><input name="single_post_uid" placeholder="UID bài viết đơn"><input name="single_group_uid" placeholder="UID Group đơn"><input name="single_user_uid" placeholder="UID người dùng đơn"></div>
      <button>Tạo hàng chờ bình luận</button>
      <a class="btnlink" href="/export_page_comment_queue">Xuất CSV hàng chờ</a>
    </form>
  </div>

  <h3>Lịch sử đăng Page gần đây</h3>
  <div style="max-height:260px;overflow:auto"><table class="gf-table"><tr><th>Page</th><th>Nội dung</th><th>Trạng thái</th><th>Giờ</th></tr>{% for h in history %}<tr><td>{{ h[1] }}</td><td>{{ h[2][:120] }}</td><td>{{ h[3] }}</td><td>{{ h[9] }}</td></tr>{% endfor %}</table></div>
</section>


<section class="panel module-section" id="group_suite">
  <div class="section-open-note">Bạn đang mở: Group Center Tổng V11.</div>
  <h2>Group Center Tổng V11 - chọn Page đã tham gia Group để đăng bài không trùng</h2>
  <p class="small">Gom chung: sắp xếp Group, tìm từ khóa, chia UID, tham gia nhóm, đăng bài nhóm và bình luận nhóm trong một khung. Bản này có chế độ ghép 1 Page với 1 Group riêng, tránh Page đăng trùng nhiều Group khi không cần.</p>
  <div class="gf-warning">Chỉ dùng với Group/Page mà tài khoản hoặc Page có quyền truy cập hợp lệ. Hệ thống tạo hàng chờ, duyệt và log; không tự động spam hàng loạt hoặc thao tác trái quyền.</div>

  <div class="gf-box">
    <h3>0. Chọn Page đã tham gia Group / có quyền đăng</h3>
    <p class="small">Bước bắt buộc: lưu cặp Page → Group. Khi đăng Group, hệ thống chỉ nhận đúng cặp đã lưu, không cho Page đăng vào Group chưa tham gia.</p>
    <form method="post" action="/page_group_membership_add">
      <div class="gf-grid-3">
        <select name="page_index">{% for p in pages %}<option value="{{ loop.index0 }}">{{ p.name }} - {{ p.id }}</option>{% endfor %}</select>
        <select name="group_id_select" onchange="this.form.group_id.value=this.value"><option value="">Chọn Group đã lưu</option>{% for g in fb_groups %}<option value="{{ g[2] }}">{{ g[1] }} - {{ g[2] }}</option>{% endfor %}</select>
        <input name="group_id" placeholder="Hoặc nhập UID Group thủ công"><select name="can_post"><option value="Có">Page có quyền đăng bài</option><option value="Không">Chưa có quyền đăng</option></select>
      </div>
      <div class="gf-grid-3"><select name="status"><option>Đã tham gia</option><option>Đang chờ duyệt</option><option>Không đủ quyền</option></select><input name="note" placeholder="Ghi chú kiểm tra quyền"><button>Lưu Page đã tham gia Group</button></div>
    </form>
    <div style="max-height:180px;overflow:auto"><table class="gf-table"><tr><th>Page</th><th>Group</th><th>Trạng thái</th><th>Quyền đăng</th><th>Ghi chú</th></tr>{% for m in page_group_memberships %}<tr><td>{{ m[1] }}<br>{{ m[2] }}</td><td>{{ m[3] }}<br>{{ m[4] }}</td><td>{{ m[5] }}</td><td>{{ m[6] }}</td><td>{{ m[7] }}</td></tr>{% endfor %}</table></div>
  </div>

  <div class="gf-grid-3">
    <div class="gf-stat"><span>Tổng UID Group</span><b>{{ group_finder_stats.total }}</b></div>
    <div class="gf-stat"><span>Group hợp lệ</span><b>{{ group_finder_stats.valid }}</b></div>
    <div class="gf-stat"><span>Hàng chờ tham gia</span><b>{{ group_finder_stats.queue }}</b></div>
  </div>

  <div class="gf-box">
    <h3>1. Ghép Page đăng Group riêng - không trùng Group</h3>
    <form method="post" action="/group_pair_post_queue">
      <div class="gf-grid-3">
        <div>
          <label>Chọn nhiều Page</label>
          <div class="gf-box" style="max-height:220px;overflow:auto;padding:10px">
            {% for p in pages %}<label style="display:block;margin:6px 0"><input type="checkbox" name="page_indexes" value="{{ loop.index0 }}"> {{ p.name }} - {{ p.id }}</label>{% endfor %}
          </div>
          <small class="small">Page thứ 1 ghép Group thứ 1, Page thứ 2 ghép Group thứ 2. Không tạo chéo tất cả Page x tất cả Group.</small>
        </div>
        <div>
          <label>Chọn Group đã tham gia</label>
          <div class="gf-box" style="max-height:220px;overflow:auto;padding:10px">
            {% for g in fb_groups %}<label style="display:block;margin:6px 0"><input type="checkbox" name="group_ids" value="{{ g[2] }}"> {{ g[1] }} • {{ g[2] }}</label>{% endfor %}
          </div>
          <small class="small">Mỗi Group chỉ được dùng 1 lần trong một lượt tạo hàng chờ.</small>
        </div>
        <div>
          <label>Cài đặt đăng Group</label>
          <select name="approval_mode"><option value="manual">Admin duyệt trước</option><option value="limited">Duyệt có giới hạn</option></select>
          <select name="schedule_mode"><option value="now">Đăng ngay sau khi duyệt</option><option value="30">Hẹn 30 phút</option><option value="60">Hẹn 1 giờ</option><option value="120">Hẹn 2 giờ</option><option value="180">Hẹn 3 giờ</option></select>
          <input name="min_delay" type="number" min="45" value="45" placeholder="Giãn cách tối thiểu 45 giây">
          <input name="max_delay" type="number" min="45" value="60" placeholder="Giãn cách tối đa 60 giây">
        </div>
      </div>
      <textarea name="bulk_content" rows="6" placeholder="Mỗi dòng là 1 nội dung Group. Page/Group thứ 1 dùng nội dung dòng 1, Page/Group thứ 2 dùng dòng 2... Không trùng nội dung nếu đủ dòng."></textarea>
      <button>Tạo hàng chờ Page → Group riêng</button>
    </form>
  </div>

  <div class="gf-grid">
    <form method="post" action="/fb_group" class="gf-box"><h3>2. Thêm / sắp xếp Group đã tham gia</h3><input name="group_name" placeholder="Tên Group"><input name="group_id" placeholder="UID Group"><input name="niche" placeholder="Ngành / tệp khách"><textarea name="note" rows="3" placeholder="Ghi chú quyền: Page có tham gia, có được đăng bài không..."></textarea><button>Lưu Group</button></form>
    <form method="post" action="/group_finder_scan" class="gf-box"><h3>3. Tìm Group theo từ khóa / lọc UID</h3><div class="gf-grid-3"><input name="keyword" placeholder="Từ khóa Group"><input name="min_members" type="number" placeholder="Số thành viên tối thiểu"><select name="privacy"><option value="all">Công khai / riêng tư</option><option>Công khai</option><option>Riêng tư</option></select></div><div class="gf-grid-3"><select name="recent_only"><option value="0">Không bắt buộc hoạt động gần đây</option><option value="1">Có hoạt động gần đây</option></select><select name="page_join"><option value="all">Page tham gia: không lọc</option><option>Có</option><option>Không</option><option>Chưa rõ</option></select><select name="page_post"><option value="all">Page đăng bài: không lọc</option><option>Có</option><option>Không</option><option>Chưa rõ</option></select></div><textarea name="raw_groups" rows="5" placeholder="Dán: UID, Tên Group, Thành viên, Công khai/Riêng tư, Có hoạt động, Page tham gia, Page đăng bài"></textarea><button>Lọc và lưu UID hợp lệ</button></form>
  </div>

  <div class="gf-grid">
    <div class="gf-box"><h3>4. Chia UID Group</h3><form method="post" action="/group_uid_split"><select name="chunk_size"><option value="50">50 UID / tệp</option><option value="100">100 UID / tệp</option><option value="200">200 UID / tệp</option></select><button>Chia UID thành tệp</button></form>{% for f in group_uid_files %}<div class="history"><b>{{ f[1] }}</b> • {{ f[2] }} UID • mỗi tệp {{ f[3] }} UID</div>{% endfor %}</div>
    <div class="gf-box"><h3>5. Hàng chờ tham gia Group</h3><form method="post" action="/group_join_queue_add"><select name="page_index">{% for p in pages %}<option value="{{ loop.index0 }}">{{ p.name }} - {{ p.id }}</option>{% endfor %}</select><button>Đưa UID hợp lệ vào hàng chờ</button></form><div style="max-height:240px;overflow:auto"><table class="gf-table"><tr><th>UID</th><th>Group</th><th>Page</th><th>Trạng thái</th></tr>{% for q in group_join_queue %}<tr><td>{{ q[1] }}</td><td>{{ q[2] }}</td><td>{{ q[5] }}</td><td>{{ q[6] }}<br>{{ q[7] }}</td></tr>{% endfor %}</table></div></div>
  </div>

  <div class="gf-box">
    <h3>6. Lọc bài viết Group & lấy UID bài viết</h3>
    <form method="post" action="/group_post_filter_import"><div class="gf-grid-3"><input name="keyword" placeholder="Từ khóa mua hàng / số điện thoại / nhu cầu"><input name="min_comments" type="number" placeholder="Số comment tối thiểu"><input name="min_reactions" type="number" placeholder="Số reaction tối thiểu"></div><textarea name="raw_posts" rows="5" placeholder="Dán: UID Group, UID bài viết, Link bài viết, Người đăng, Thời gian đăng, Nội dung rút gọn, Số comment, Số reaction"></textarea><button>Lọc và lưu UID bài viết</button></form>
    <div style="max-height:260px;overflow:auto"><table class="gf-table"><tr><th>UID Group</th><th>UID bài viết</th><th>Link</th><th>Nội dung</th><th>Comment/Reaction</th><th>Trạng thái</th></tr>{% for p in group_post_results %}<tr><td>{{ p[1] }}</td><td>{{ p[2] }}</td><td>{{ p[3] }}</td><td>{{ p[6] }}</td><td>{{ p[7] }} / {{ p[8] }}</td><td>{{ p[9] }}</td></tr>{% endfor %}</table></div>
  </div>

  <div class="gf-box">
    <h3>7. Bình luận Group/Page theo UID</h3>
    <form method="post" action="/page_comment_queue_add">
      <div class="gf-grid-3"><select name="page_index">{% for p in pages %}<option value="{{ loop.index0 }}">{{ p.name }} - {{ p.id }}</option>{% endfor %}</select><select name="target_type"><option value="post">UID bài viết Group</option><option value="group">UID Group</option><option value="user">UID người dùng đã tương tác</option></select><input name="min_delay" type="number" min="45" value="45" placeholder="Giãn cách tối thiểu giây"></div>
      <div class="gf-grid-3"><input name="max_delay" type="number" min="45" value="60" placeholder="Giãn cách tối đa giây"><input name="single_post_uid" placeholder="UID bài viết"><input name="single_group_uid" placeholder="UID Group"></div>
      <input name="single_user_uid" placeholder="UID người dùng đã tương tác nếu có"><textarea name="comment_text" rows="4" placeholder="Mỗi dòng là 1 bình luận riêng. Ví dụ: xin chào / bạn khỏe không / ngày mai bạn làm gì"></textarea><textarea name="raw_targets" rows="5" placeholder="Dán nhiều UID, mỗi dòng một mục: post_uid, group_uid, user_uid"></textarea><button>Tạo hàng chờ bình luận</button><a class="btnlink" href="/export_page_comment_queue">Xuất CSV hàng chờ</a>
    </form>
  </div>

  <h3>Hàng chờ đăng Group</h3><div style="max-height:300px;overflow:auto"><table class="gf-table"><tr><th>Page</th><th>Group</th><th>Nội dung</th><th>Trạng thái</th><th>Ghi chú</th></tr>{% for q in group_post_queue %}<tr><td>{{ q[1] }}</td><td>{{ q[2] }}<br>{{ q[3] }}</td><td>{{ q[4] }}</td><td>{{ q[5] }}</td><td>{{ q[6] }}</td></tr>{% endfor %}</table></div>
  <h3>Hàng chờ bình luận</h3><div style="max-height:300px;overflow:auto"><table class="gf-table"><tr><th>ID</th><th>Page</th><th>Loại</th><th>UID đích</th><th>Nội dung</th><th>Giãn cách</th><th>Trạng thái</th><th>Thao tác</th></tr>{% for q in page_comment_queue %}<tr><td>{{ q[0] }}</td><td>{{ q[1] }}</td><td>{{ q[2] }}</td><td>USER: {{ q[3] }}<br>POST: {{ q[4] }}<br>GROUP: {{ q[5] }}</td><td>{{ q[6] }}</td><td>{{ q[7] }} - {{ q[8] }} giây</td><td>{{ q[10] }}<br>{{ q[11] }}</td><td><form method="post" action="/page_comment_queue_action"><input type="hidden" name="queue_id" value="{{ q[0] }}"><button name="action" value="approve">Duyệt</button><button name="action" value="done">Hoàn thành</button><button name="action" value="error">Báo lỗi</button></form></td></tr>{% endfor %}</table></div>

  <h3>Danh sách Group đã lưu</h3>{% for g in fb_groups %}<div class="history"><b>{{ g[1] }}</b> • {{ g[2] }} • {{ g[3] }}<br>{{ g[4] }}</div>{% endfor %}
</section>

<section class="panel module-section" id="comment_manager">
  <div class="section-open-note">Bạn đang mở: Comment Manager AI.</div>
  <h2>💬 AI Comment</h2>
  <p class="small">AI trả lời comment, ẩn SĐT, gắn nhãn khách và chuyển CRM.</p>
  <form method="post" action="/comment_ai">
    <div class="grid"><input name="customer_name" placeholder="Tên khách nếu có"><input name="phone" placeholder="SĐT nếu có"></div>
    <textarea name="comment_text" rows="4" placeholder="Dán comment khách hàng vào đây. Ví dụ: shop còn proxy Việt không, số em 09..."></textarea>
    <div class="grid"><select name="label"><option>Khách nóng</option><option>Khách ấm</option><option>Khách lạnh</option><option>Cần chăm sóc lại</option></select><select name="to_crm"><option value="1">Chuyển sang CRM</option><option value="0">Chỉ lưu comment</option></select></div>
    <button>AI xử lý Comment</button><button type="button" class="secondary" onclick="openLockedFeature('Comment Manager AI','Gói 1 năm / Gói Nhà Bán Hàng Chuyên Nghiệp')">Xem bản tự động Webhook</button>
  </form>
  <h3>Comment đã xử lý</h3>{% for c in comment_leads %}<div class="history"><b>{{ c[1] or 'Khách hàng' }}</b> • {{ c[5] }} • {{ c[7] }}<br>Comment: {{ c[3] }}<br>AI: {{ c[4] }}</div>{% endfor %}
</section>

<section class="panel module-section" id="messenger_ai">
  <div class="section-open-note">Bạn đang mở: Messenger AI.</div>
  <h2>📨 AI Messenger</h2>
  <p class="small">Tạo kịch bản Inbox, kịch bản chốt sale, xử lý từ chối và chăm sóc khách cũ.</p>
  <form method="post" action="/messenger_ai_script">
    <div class="grid"><input name="product" placeholder="Sản phẩm/dịch vụ. Ví dụ: Proxy"><select name="script_type"><option>Kịch bản Inbox</option><option>Kịch bản Chốt Sale</option><option>Xử lý từ chối</option><option>Chăm sóc khách cũ</option></select></div>
    <textarea name="extra" rows="3" placeholder="Thông tin thêm: giá, ưu đãi, tệp khách, số Zalo..."></textarea>
    <button>Tạo kịch bản Messenger</button>
  </form>
  <h3>Kịch bản đã lưu</h3>{% for m in messenger_scripts %}<div class="history"><b>{{ m[1] }}</b> • {{ m[2] }} • {{ m[4] }}<br>{{ m[3] }}</div>{% endfor %}
</section>

<section class="panel module-section" id="ai_studio">
  <div class="section-open-note">Bạn đang mở: AI Studio V3.</div>
  <h2>🤖 AI Studio</h2>
  <div class="v3-feature-grid">
    <div class="v3-feature-card"><h3>AI Content</h3><ul><li>Content Facebook</li><li>Content TikTok</li><li>Caption</li></ul><button onclick="return openModule('ai_studio')">Mở AI Content</button></div>
    <div class="v3-feature-card"><h3>Viral Content Lab</h3><ul><li>Content Viral</li><li>Storytelling</li><li>Seeding</li></ul><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="prompt_premium"><textarea name="topic" rows="3" placeholder="Ví dụ: nội dung viral cho dịch vụ proxy"></textarea><button>Tạo ý tưởng viral</button></form></div>
    <div class="v3-feature-card"><h3>Facebook Ads AI</h3><ul><li>Chấm điểm quảng cáo</li><li>Viết quảng cáo</li><li>Target khách hàng</li></ul><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="facebook_ads_ai"><textarea name="topic" rows="3" placeholder="Nhập sản phẩm/dịch vụ cần chạy ads"></textarea><button>Tạo Facebook Ads AI</button></form></div>
    <div class="v3-feature-card"><span class="v3-premium-badge">VIP</span><h3>AI Marketing Director</h3><p>Khách nhập: Tôi bán Proxy. AI trả: 30 Content, 10 Quảng cáo, 10 Caption, 5 Kịch bản chốt sale, 30 ngày Marketing, tệp khách hàng và ngân sách đề xuất.</p><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="marketing_director"><textarea name="topic" rows="3" placeholder="Tôi bán Proxy / mỹ phẩm / dịch vụ quảng cáo..."></textarea><button>Tạo kế hoạch tổng</button></form></div>
    <div class="v3-feature-card"><h3>Competitor Scanner</h3><p>Nhập link Fanpage hoặc mô tả đối thủ. AI phân tích điểm mạnh, điểm yếu, nội dung và CTA.</p><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="competitor_scanner"><input name="topic" placeholder="Link Fanpage đối thủ"><button>Phân tích đối thủ</button></form></div>
    <div class="v3-feature-card"><h3>AI Sales Script</h3><ul><li>Kịch bản chốt sale</li><li>Xử lý từ chối</li><li>Kịch bản inbox</li></ul><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="sales_script"><textarea name="topic" rows="3" placeholder="Sản phẩm/dịch vụ cần tư vấn"></textarea><button>Tạo kịch bản sale</button></form></div>
    <div class="v3-feature-card"><h3>AI Landing Page Builder</h3><p>Nhập: Dịch vụ chạy quảng cáo Facebook. AI sinh tiêu đề, CTA, form và ưu điểm.</p><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="landing_page"><textarea name="topic" rows="3" placeholder="Dịch vụ chạy quảng cáo Facebook"></textarea><button>Tạo Landing Page</button></form></div>
  </div>
</section>

<section class="panel module-section" id="creative_center">
  <div class="section-open-note">Bạn đang mở: Creative Center.</div>
  <h2>🎨 Creative Center</h2>
  <div class="v3-feature-grid">
    <div class="v3-feature-card"><h3>AI Image Center</h3><ul><li>Avatar</li><li>Cover</li><li>Banner</li><li>Poster</li></ul><button class="secondary" onclick="openLockedFeature('AI Image Center','Gói 1 năm / Vĩnh viễn')">Mở khóa</button></div>
    <div class="v3-feature-card"><h3>AI Video Center</h3><ul><li>TikTok</li><li>Reels</li><li>Shorts</li></ul><button class="secondary" onclick="openLockedFeature('AI Video Center','Gói 1 năm / Vĩnh viễn')">Mở khóa</button></div>
    <div class="v3-feature-card"><h3>AI Voice Studio</h3><ul><li>Giọng nữ</li><li>Giọng nam</li><li>MC quảng cáo</li></ul><button class="secondary" onclick="openLockedFeature('AI Voice Studio','Gói 1 năm / Vĩnh viễn')">Mở khóa</button></div>
    <div class="v3-feature-card"><h3>Banner Builder</h3><p>Tạo ý tưởng ảnh quảng cáo, bố cục banner và nội dung poster.</p><form method="post" action="/v3_ai_tool"><input type="hidden" name="tool" value="landing_page"><textarea name="topic" rows="3" placeholder="Banner cho dịch vụ quảng cáo Facebook"></textarea><button>Tạo bố cục banner</button></form></div>
  </div>
</section>

<section class="panel module-section" id="marketing_director">
  <div class="section-open-note">Bạn đang mở: AI Marketing Director.</div>
  <h2>🧠 AI Marketing Director</h2>
  <p class="small">Trợ lý giám đốc Marketing: phân tích sản phẩm, tệp khách hàng, content, quảng cáo, phễu bán hàng, KPI và kế hoạch triển khai 30 ngày.</p>
  <div class="v3-feature-grid">
    <div class="v3-feature-card"><h3>Đầu vào cần nhập</h3><ul><li>Sản phẩm/dịch vụ đang bán</li><li>Khách hàng mục tiêu</li><li>Giá/ưu đãi hiện tại</li><li>Mục tiêu: inbox, đơn hàng, thương hiệu</li></ul></div>
    <div class="v3-feature-card"><h3>Kết quả AI trả về</h3><ul><li>30 content Facebook/TikTok</li><li>10 mẫu quảng cáo</li><li>5 kịch bản chốt sale</li><li>Kế hoạch 30 ngày và KPI</li></ul></div>
  </div>
  <form method="post" action="/v3_ai_tool">
    <input type="hidden" name="tool" value="marketing_director">
    <textarea name="topic" rows="4" placeholder="Ví dụ: Tôi bán proxy Việt Nam cho người chạy Facebook Ads, giá từ 80k/tháng, mục tiêu tăng inbox và chốt khách qua Zalo."></textarea>
    <textarea name="extra" rows="3" placeholder="Thông tin thêm: ngân sách ads, khu vực, ưu đãi, số Zalo, đối thủ, điểm mạnh sản phẩm..."></textarea>
    <button>🧠 Tạo kế hoạch Marketing Director</button>
    <button type="button" class="secondary" onclick="return openModule('ai_studio')">Mở AI Studio</button>
    <button type="button" class="secondary" onclick="return openModule('crm_sales')">Mở CRM Kanban</button>
  </form>
</section>

<section class="panel module-section" id="crm_sales">
  <div class="section-open-note">Bạn đang mở: CRM Sales V3.</div>
  <h2>📊 CRM Kanban</h2>
  <form method="post" action="/pipeline">
    <div class="grid"><input name="customer_name" placeholder="Tên"><input name="phone" placeholder="SĐT"><input name="zalo" placeholder="Zalo"><input name="source" placeholder="Nguồn"></div>
    <div class="grid"><select name="stage"><option>Khách mới</option><option>Đang tư vấn</option><option>Đã báo giá</option><option>Đang chốt</option><option>Đã mua</option></select><input name="value" type="number" placeholder="Doanh thu dự kiến"></div>
    <textarea name="note" rows="3" placeholder="Ghi chú nhu cầu khách"></textarea><button>Lưu vào Pipeline</button>
  </form>
  <h3>Pipeline Kanban</h3>
  <div class="kanban-board">
    {% for stage in ['Khách mới','Đang tư vấn','Đã báo giá','Đang chốt','Đã mua'] %}
    <div class="kanban-col" ondragover="event.preventDefault()" ondrop="dropKanban(event)"><h3>{{ stage }}</h3>
      {% for r in pipeline_rows %}{% if r[5] == stage %}<div class="kanban-card" draggable="true" ondragstart="dragKanban(event)"><b>{{ r[1] }}</b><br><span class="small">{{ r[2] }} • {{ r[4] }}</span><br>{{ r[7] }}</div>{% endif %}{% endfor %}
    </div>
    {% endfor %}
  </div>
  <h3>Customer Care</h3>
  <form method="post" action="/customer_task"><div class="grid"><input name="customer_name" placeholder="Tên khách"><input name="due_date" type="date"></div><textarea name="task" rows="2" placeholder="Lịch chăm sóc / nhắc lịch"></textarea><button>Tạo nhắc lịch</button></form>
  {% for t in customer_tasks %}<div class="history">{{ t[1] }} | {{ t[2] }} | Hạn: {{ t[3] }} | {{ t[4] }}</div>{% endfor %}
</section>

<section class="panel module-section" id="automation_center">
  <div class="section-open-note">Bạn đang mở: Automation Center.</div>
  <h2>⚙️ Automation</h2>
  <div class="v3-feature-grid">
    <div class="v3-feature-card"><h3>Auto Campaign</h3><p>Tạo 30 ngày content và tự động lên lịch.</p><form method="post" action="/plan_30_days"><select name="industry">{% for key,label in industry_labels.items() %}<option value="{{ key }}">{{ label }}</option>{% endfor %}</select><input name="goal" placeholder="Mục tiêu"><input type="hidden" name="days" value="30"><button>Tạo 30 ngày content</button></form></div>
    <div class="v3-feature-card"><h3>Notification Center</h3><ul><li>Bài sắp đăng</li><li>Token lỗi</li><li>Khách mới</li></ul><form method="post" action="/notification"><input name="title" placeholder="Tiêu đề thông báo"><textarea name="detail" rows="2" placeholder="Nội dung"></textarea><button>Lưu thông báo</button></form></div>
    <div class="v3-feature-card"><h3>Backup Center</h3><ul><li>Backup</li><li>Restore</li></ul><a href="/backup"><button>Backup</button></a><a href="/export_pdf"><button class="secondary">Xuất PDF</button></a><a href="/export"><button class="secondary">Xuất Excel/CSV</button></a></div>
  </div>
  <h3>Thông báo gần nhất</h3>{% for n in notifications %}<div class="history"><b>{{ n[1] }}</b> • {{ n[3] }} • {{ n[5] }}<br>{{ n[2] }}</div>{% endfor %}
</section>

<section class="panel module-section" id="success_center">
  <div class="section-open-note">Bạn đang mở: Success Center V4.</div>
  <h2>🏆 Success Center</h2>
  <p class="small">Tăng niềm tin cho khách trước khi nâng cấp Premium bằng case study, mẫu thành công và tài sản bán hàng đã được chuẩn hóa.</p>

  <div class="v4-success-grid">
    {% for a in success_assets %}
    <div class="v4-success-card">
      <div class="success-icon">🏆</div>
      <h3>{{ a[2] }}</h3>
      <p><b>{{ a[1] }}</b> — {{ a[3] }}</p>
      <button onclick="openLockedFeature('{{ a[1] }} Premium','Gói 1 năm / Gói Nhà Bán Hàng Chuyên Nghiệp')">Mở mẫu Premium</button>
    </div>
    {% endfor %}
  </div>

  <div class="v4-proof-box">
    <b>🔥 Thành tích hệ thống</b>
    <div class="v4-pricing-stats">
      <div><b>2.381+</b><span>Khách hàng sử dụng</span></div>
      <div><b>120.000+</b><span>Content đã tạo</span></div>
      <div><b>98%</b><span>Đánh giá hài lòng</span></div>
    </div>
  </div>
</section>

<section class="panel module-section" id="post">
  <div class="section-open-note">Bạn đang mở: Đăng Fanpage tự chia nội dung.</div>
  <h2>Đăng Fanpage Tự Chia Nội Dung</h2>
  <p class="small">Nhập nhiều content, upload nhiều ảnh/video, chọn nhiều Page. Hệ thống tự chia từng content + từng ảnh cho từng Page để tránh trùng nội dung và trùng hình. Mặc định không thêm chữ, không spin, không tự chèn CTA/hashtag.</p>

  <form method="post" action="/multi_post" enctype="multipart/form-data">
    <div class="page-list">
      {% for p in pages %}
      <label class="page-item">
        <input type="checkbox" name="page_indexes" value="{{ loop.index0 }}">
        {{ p.name }}<br>
        <span class="small">{{ p.id }}</span>
      </label>
      {% endfor %}
    </div>

    <textarea name="bulk_content" rows="12" placeholder="Content 1...

Content 2...

Content 3...

Mỗi content cách nhau bằng một dòng trống. Nếu nhập từng dòng, mỗi dòng cũng được hiểu là một content riêng."></textarea>

    <div class="grid">
      <input type="text" name="campaign" placeholder="Tên chiến dịch, ví dụ: Spa tháng 6">
      <input type="file" name="images" accept="image/*,video/mp4,video/quicktime" multiple>
    </div>
    <label class="page-item" style="display:block;margin-top:10px">
      <input type="checkbox" name="use_ai_enhance" value="1">
      Bật AI tối ưu từng bài: spin content, thêm CTA và hashtag riêng
      <br><span class="small">Mặc định tắt để giữ nguyên 100% content đã nhập.</span>
    </label>

    <div class="premium-center" style="margin-top:12px">
      <h3>Chọn cách đăng</h3>
      <p class="small">Đăng ngay hoặc đặt lịch đều dùng cùng nội dung, cùng Page và cùng ảnh/video đã chọn ở trên.</p>

      <div class="grid">
        <div>
          <h3>Đăng ngay</h3>
          <p class="small">Tự chia content + ảnh/video và đăng ngay lên các Page đã chọn.</p>
          <button type="submit" name="action" value="now">Tự chia và đăng ngay</button>
        </div>

        <div>
          <h3>Đặt lịch đăng</h3>
          <p class="small">Chọn thời gian, hệ thống lưu lịch và tự đăng khi đến giờ.</p>
          <input type="datetime-local" name="schedule_time">
          <button type="submit" name="action" value="schedule" class="secondary">Tự chia và lưu lịch</button>
        </div>
      </div>
    </div>
  </form>
</section>

<section class="panel module-section" id="scheduler">
  <div class="section-open-note">Bạn đang mở: Lịch đăng nâng cao.</div>
  <h2>Lịch đăng nâng cao</h2>
  <p class="small">Nhập nhiều content, upload nhiều ảnh/video, chọn Page và giờ bắt đầu. Hệ thống tự chia lịch đăng theo khung giờ, không trùng content và không trùng ảnh.</p>
  <form method="post" action="/smart_schedule" enctype="multipart/form-data">
    <div class="page-list">
      {% for p in pages %}
      <label class="page-item">
        <input type="checkbox" name="page_indexes" value="{{ loop.index0 }}">
        {{ p.name }}<br><span class="small">{{ p.id }}</span>
      </label>
      {% endfor %}
    </div>
    <textarea name="bulk_content" rows="9" placeholder="Content 1...

Content 2...

Content 3..."></textarea>
    <div class="grid">
      <input type="datetime-local" name="start_time">
      <select name="gap_minutes">
        <option value="30">Cách nhau 30 phút</option>
        <option value="60" selected>Cách nhau 1 giờ</option>
        <option value="120">Cách nhau 2 giờ</option>
        <option value="180">Cách nhau 3 giờ</option>
      </select>
    </div>
    <input name="campaign" placeholder="Tên chiến dịch">
    <input type="file" name="images" accept="image/*,video/mp4,video/quicktime" multiple>
    <label class="page-item" style="display:block;margin-top:10px">
      <input type="checkbox" name="use_ai_enhance" value="1">
      Bật AI tối ưu từng bài: spin content, thêm CTA và hashtag riêng
      <br><span class="small">Mặc định tắt để giữ nguyên 100% content đã nhập.</span>
    </label>
    <button type="submit">Tự chia lịch đăng</button>
  </form>
</section>


<section class="panel module-section" id="studio">
  <div class="section-open-note">Bạn đang mở: AI Content Studio.</div>
  <h2>AI Content Studio</h2>
  <form method="post" action="/generate">
    <textarea name="idea" rows="4" placeholder="Nhập ý tưởng: spa trị nám, proxy chạy ads, nha khoa niềng răng, bất động sản căn hộ..."></textarea>
    <div class="grid">
      <select name="style">
        <option value="gần gũi, tự nhiên, dễ ra inbox">Gần gũi</option>
        <option value="bán hàng mạnh, chốt đơn tốt">Bán hàng mạnh</option>
        <option value="chuyên nghiệp, uy tín">Chuyên nghiệp</option>
        <option value="viral, ngắn gọn, thu hút">Viral ngắn gọn</option>
        <option value="cao cấp, sang trọng">Cao cấp</option>
      </select>
      <select name="length">
        <option value="80">Ngắn dưới 80 từ</option>
        <option value="120">Vừa dưới 120 từ</option>
        <option value="180">Dài dưới 180 từ</option>
        <option value="250">Đầy đủ dưới 250 từ</option>
      </select>
    </div>
    <button type="submit" name="variants" value="1">Tạo nội dung</button>
    <button type="submit" name="variants" value="5" class="secondary">Tạo 5 phiên bản</button>
  </form>
</section>

<section class="panel module-section" id="library">
  <div class="section-open-note">Bạn đang mở: Kho Content 50.000+.</div>
  <h2>Kho Content 10.000+ Dùng Thử</h2>
  <p class="small">Bản demo nạp sẵn một số mẫu. Sau này có thể import 10.000 content thật từ JSON/CSV.</p>
  <form method="get" action="/">
    <select name="industry">
      {% for key,label in industry_labels.items() %}
      <option value="{{ key }}" {% if selected_industry==key %}selected{% endif %}>{{ label }}</option>
      {% endfor %}
    </select>
    <button type="submit">Xem mẫu content</button>
  </form>
  <div class="library-grid">
    {% for item in library_items %}
    <div class="template-card">
      <div id="tpl{{ loop.index }}">{{ item }}</div>
      <button onclick="copyText('tpl{{ loop.index }}')">Copy content</button>
    </div>
    {% endfor %}
  </div>
  <div class="lock-card" style="margin-top:16px">
    <h3>Khóa Premium</h3>
    <p>Bạn đang xem 10 content miễn phí đầu tiên.</p>
    <p>Còn lại khoảng 490 content/ngành cần nâng cấp Premium để mở khóa.</p>
    <button onclick="openPremiumPopup()">Xem bảng giá Premium</button>
  </div>
</section>

<section class="panel module-section" id="fanpage">
  <div class="section-open-note">Bạn đang mở: AI Phân tích Fanpage.</div>
  <h2>AI Phân tích Fanpage</h2>
  <form method="post" action="/analyze_fanpage">
    <input name="page_name" placeholder="Tên Fanpage">
    <div class="grid">
      <select name="avatar"><option value="co">Có avatar rõ</option><option value="khong">Chưa tốt</option></select>
      <select name="cover"><option value="co">Có ảnh bìa tốt</option><option value="khong">Chưa tốt</option></select>
      <select name="post_frequency"><option value="hang_ngay">Đăng hằng ngày</option><option value="3_5_bai_tuan">3-5 bài/tuần</option><option value="it">Ít đăng</option></select>
      <select name="cta"><option value="co">Có CTA rõ</option><option value="khong">Chưa có CTA</option></select>
    </div>
    <textarea name="bio" rows="3" placeholder="Mô tả ngắn về Fanpage, sản phẩm/dịch vụ, thông tin liên hệ..."></textarea>
    <button type="submit">Phân tích Fanpage</button>
  </form>
</section>

<section class="panel module-section" id="plan">
  <div class="section-open-note">Bạn đang mở: AI Marketing Planner.</div>
  <h2>AI Kế hoạch Marketing 30 ngày</h2>
  <form method="post" action="/plan_30_days">
    <select name="industry">
      {% for key,label in industry_labels.items() %}
      <option value="{{ key }}">{{ label }}</option>
      {% endfor %}
    </select>
    <input name="goal" placeholder="Mục tiêu: tăng inbox, tăng đơn, tăng nhận diện, ra mắt sản phẩm...">\n    <select name="days"><option value="7">7 ngày</option><option value="30" selected>30 ngày</option><option value="90">90 ngày</option></select>
    <button type="submit">Tạo kế hoạch</button>
  </form>
</section>

<section class="panel module-section" id="campaign">
  <div class="section-open-note">Bạn đang mở: Campaign Manager.</div>
  <h2>Campaign Manager</h2>
  <form method="post" action="/campaign">
    <input name="name" placeholder="Tên chiến dịch, ví dụ: Proxy tháng 6">
    <select name="industry">
      {% for key,label in industry_labels.items() %}
      <option value="{{ key }}">{{ label }}</option>
      {% endfor %}
    </select>
    <input name="goal" placeholder="Mục tiêu chiến dịch">
    <textarea name="note" rows="3" placeholder="Ghi chú chiến dịch"></textarea>
    <button type="submit">Tạo chiến dịch</button>
  </form>
  {% for c in campaigns %}
  <div class="history">ID: {{ c[0] }} | {{ c[1] }} | Ngành: {{ c[2] }} | Mục tiêu: {{ c[3] }} | Tạo: {{ c[5] }}</div>
  {% endfor %}
</section>

{% if content %}
<section class="panel module-section" id="post_edit">
  <h2>Sửa bài, Preview Facebook, chọn Fanpage</h2>
  <form method="post" action="/post" enctype="multipart/form-data">
    <div class="page-list">
      {% for p in pages %}
      <label class="page-item"><input type="checkbox" name="page_indexes" value="{{ loop.index0 }}"> {{ p.name }}<br><span class="small">{{ p.id }}</span></label>
      {% endfor %}
    </div>
    <textarea name="content" rows="12">{{ content }}</textarea>
    <input type="text" name="campaign" placeholder="Tên chiến dịch">
    <input type="file" name="image" accept="image/*,video/mp4,video/quicktime">
    <button type="submit" name="action" value="now">Đăng ngay nhiều Page</button>
    <h3>Đặt lịch đăng</h3>
    <input type="datetime-local" name="schedule_time">
    <button type="submit" name="action" value="schedule">Lưu lịch đăng</button>
  </form>
  <h3>Preview Facebook</h3>
  <div class="preview">
    <div class="preview-head"><div class="avatar"></div><div><div class="preview-name">Fanpage của bạn</div><div class="small">Vừa xong · Công khai</div></div></div>
    <div class="preview-content">{{ content }}</div>
  </div>
  <p>Điểm content: <b>{{ score }}/100</b></p>
  <ul>{% for w in warnings %}<li>{{ w }}</li>{% endfor %}</ul>
</section>
{% endif %}

{% if analysis %}
<section class="panel"><h2>Kết quả phân tích</h2><div class="history">{{ analysis }}</div></section>
{% endif %}

{% if plan %}
<section class="panel"><h2>AI Marketing Planner</h2><div class="history">{{ plan }}</div></section>
{% endif %}


<section class="panel module-section" id="factory">
  <div class="section-open-note">Bạn đang mở: Content Factory.</div>
  <h2>AI Content Factory</h2>
  <p class="small">Tạo hàng loạt 20-50 content khác nhau để chia cho nhiều Fanpage.</p>
  <form method="post" action="/content_factory">
    <textarea name="idea" rows="4" placeholder="Nhập chủ đề cần tạo hàng loạt, ví dụ: Proxy Việt Nam tốc độ cao cho chạy quảng cáo..."></textarea>
    <div class="grid">
      <select name="count">
        <option value="10">Tạo 10 content</option>
        <option value="20">Tạo 20 content</option>
        <option value="30">Tạo 30 content</option>
        <option value="50">Tạo 50 content</option>
      </select>
      <select name="style">
        <option value="bán hàng tự nhiên">Bán hàng tự nhiên</option>
        <option value="viral ngắn gọn">Viral ngắn gọn</option>
        <option value="chuyên nghiệp">Chuyên nghiệp</option>
        <option value="cao cấp">Cao cấp</option>
      </select>
    </div>
    <button type="submit">Tạo hàng loạt content</button>
  </form>
</section>

<section class="panel module-section" id="clusters">
  <h2>Page Cluster</h2>
  <p class="small">Tạo nhóm Page theo ngành như Proxy, Spa, BĐS, Nha khoa để quản lý dễ hơn.</p>
  <form method="post" action="/page_cluster">
    <input name="name" placeholder="Tên nhóm, ví dụ: Nhóm Proxy">
    <input name="page_names" placeholder="Tên Page, ngăn cách bằng dấu phẩy">
    <textarea name="note" rows="3" placeholder="Ghi chú nhóm Page"></textarea>
    <button type="submit">Lưu nhóm Page</button>
  </form>
  {% for c in clusters %}
  <div class="history">ID: {{ c[0] }} | Nhóm: {{ c[1] }}
Page: {{ c[2] }}
Ghi chú: {{ c[3] }}
Tạo: {{ c[4] }}</div>
  {% endfor %}
</section>

<section class="panel module-section" id="analytics">
  <div class="section-open-note">Bạn đang mở: Analytics Center V5.</div>
  <h2>📊 Analytics Center</h2>
  <p class="small">Trung tâm phân tích tổng hợp Fanpage, Group, bài đăng, chiến dịch và CRM để theo dõi hiệu suất bán hàng.</p>

  <div class="analytics-kpi-grid">
    <div class="analytics-kpi"><span>Tổng bài</span><b>{{ analytics.summary.total_posts }}</b><small>Toàn bộ bài đã tạo</small></div>
    <div class="analytics-kpi"><span>Đã đăng</span><b>{{ analytics.summary.posted }}</b><small>Tỷ lệ đăng: {{ analytics.summary.conversion_rate }}%</small></div>
    <div class="analytics-kpi"><span>Chờ đăng</span><b>{{ analytics.summary.scheduled }}</b><small>Bài đang lên lịch</small></div>
    <div class="analytics-kpi"><span>⚠️ Lỗi đăng</span><b>{{ analytics.summary.errors }}</b><small>Tỷ lệ lỗi: {{ analytics.summary.error_rate }}%</small></div>
    <div class="analytics-kpi"><span>Lead CRM</span><b>{{ analytics.summary.crm_total + analytics.summary.pipeline_total }}</b><small>Tổng khách hàng ghi nhận</small></div>
    <div class="analytics-kpi"><span>💰 Pipeline</span><b>{{ "{:,}".format(analytics.summary.total_value).replace(",", ".") }}đ</b><small>Giá trị cơ hội bán hàng</small></div>
    <div class="analytics-kpi"><span>👥 Group</span><b>{{ analytics.summary.groups_total }}</b><small>{{ analytics.summary.group_scheduled }} lịch đăng Group</small></div>
    <div class="analytics-kpi"><span>🤖 AI xử lý</span><b>{{ analytics.summary.comments_total + analytics.summary.messenger_total }}</b><small>Comment + Messenger</small></div>
  </div>

  <div class="analytics-section-title">📣 Fanpage / Campaign Analytics</div>
  <div class="analytics-grid">
    <div class="analytics-box">
      <h3>Top Fanpage có bài đăng</h3>
      {% if analytics.by_page %}
        {% for r in analytics.by_page %}
        <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }}</b></div>
        <div class="analytics-bar"><i style="width:{{ 100 if loop.first else (r[1] / (analytics.by_page[0][1] if analytics.by_page[0][1] else 1) * 100)|round(0) }}%"></i></div>
        {% endfor %}
      {% else %}
        <div class="empty-analytics">Chưa có dữ liệu Fanpage.</div>
      {% endif %}
    </div>
    <div class="analytics-box">
      <h3>Hiệu suất chiến dịch</h3>
      {% if analytics.by_campaign %}
        {% for r in analytics.by_campaign %}
        <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }}</b></div>
        <div class="analytics-bar"><i style="width:{{ 100 if loop.first else (r[1] / (analytics.by_campaign[0][1] if analytics.by_campaign[0][1] else 1) * 100)|round(0) }}%"></i></div>
        {% endfor %}
      {% else %}
        <div class="empty-analytics">Chưa có dữ liệu chiến dịch.</div>
      {% endif %}
    </div>
    <div class="analytics-box">
      <h3>Trạng thái bài đăng</h3>
      {% if analytics.by_status %}
        {% for r in analytics.by_status %}
        <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }}</b></div>
        {% endfor %}
      {% else %}
        <div class="empty-analytics">Chưa có trạng thái bài đăng.</div>
      {% endif %}
    </div>
  </div>

  <div class="analytics-section-title">👥 Group / CRM Analytics</div>
  <div class="analytics-grid">
    <div class="analytics-box">
      <h3>Group theo ngành</h3>
      {% if analytics.group_by_niche %}
        {% for r in analytics.group_by_niche %}
        <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }}</b></div>
        {% endfor %}
      {% else %}
        <div class="empty-analytics">Chưa có dữ liệu Group.</div>
      {% endif %}
    </div>
    <div class="analytics-box">
      <h3>Lead theo nguồn</h3>
      {% if analytics.crm_by_source %}
        {% for r in analytics.crm_by_source %}
        <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }}</b></div>
        {% endfor %}
      {% else %}
        <div class="empty-analytics">Chưa có dữ liệu nguồn lead.</div>
      {% endif %}
    </div>
    <div class="analytics-box">
      <h3>Pipeline CRM</h3>
      {% if analytics.crm_by_stage %}
        {% for r in analytics.crm_by_stage %}
        <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }}</b></div>
        {% endfor %}
      {% else %}
        <div class="empty-analytics">Chưa có dữ liệu Kanban.</div>
      {% endif %}
    </div>
  </div>

  <div class="analytics-section-title">📅 Xu hướng 7 ngày</div>
  <div class="analytics-box full">
    {% if analytics.by_day %}
      {% for r in analytics.by_day %}
      <div class="analytics-row"><span>{{ r[0] }}</span><b>{{ r[1] }} bài</b></div>
      <div class="analytics-bar large"><i style="width:{{ 100 if loop.first else (r[1] / (analytics.by_day[0][1] if analytics.by_day[0][1] else 1) * 100)|round(0) }}%"></i></div>
      {% endfor %}
    {% else %}
      <div class="empty-analytics">Chưa có dữ liệu theo ngày. Khi bắt đầu đăng bài, biểu đồ sẽ tự cập nhật tại đây.</div>
    {% endif %}
  </div>
</section>

<section class="panel module-section" id="premium">
  <h2>💎 Premium Center V5</h2>
  <div class="premium-center v4-premium-hero">
    <div class="v4-hero-label">AI MARKETING PREMIUM</div>
    <h3>Biến công cụ đăng bài thành trợ lý Marketing tự động</h3>
    <p>Khách không chỉ mua phần mềm, khách mua thời gian, nội dung, quy trình bán hàng và hệ thống hỗ trợ tăng tốc kinh doanh online.</p>
    <div class="v4-value-grid">
      <div><b>159.000đ</b><span>Gói Khởi Động<br>chỉ 5.300đ/ngày</span></div>
      <div><b>859.000đ</b><span>Gói 1 năm<br>phổ biến nhất</span></div>
      <div><b>1.959.000đ</b><span>Gói trọn đời<br>không phí gia hạn</span></div>
    </div>
    <p><b>Giá trị nhận được:</b> tiết kiệm thời gian viết content, quản lý Fanpage, CRM khách hàng, AI Sales Script, Marketing Director và kho Content 50.000+.</p>
    <button onclick="scrollToPricing()">Xem bảng giá chi tiết</button>
    <button class="secondary" onclick="openPayment('yearly')">Chọn gói phổ biến nhất</button>
  </div>
</section>

<section class="panel module-section" id="v9center">
  <h2>V3 Enterprise Center</h2>
  <div class="v9-grid">
    <div class="v9-card">
      <h3>AI Spin 70-80%</h3>
      <p>Mỗi bài đăng được viết lại khác nhau, thêm CTA và hashtag riêng để giảm trùng lặp.</p>
    </div>
    <div class="v9-card">
      <h3>AI Chọn ảnh phù hợp</h3>
      <p>Bản local ưu tiên chia ảnh không trùng. Có thể nâng tiếp để AI phân loại ảnh theo nội dung.</p>
    </div>
    <div class="v9-card">
      <h3>Auto 100 Content → 100 Page</h3>
      <p>Hệ thống tự chia content, media và Page theo vòng quay, phù hợp đăng hàng loạt.</p>
    </div>
    <div class="v9-card">
      <h3>Token Guard</h3>
      <p>Kiểm tra token trước khi đăng ngay. Nếu token chết, hệ thống dừng để tránh lỗi hàng loạt.</p>
    </div>
  </div>
  <div style="margin-top:14px">
    <a href="/backup"><button>Backup Database</button></a>
    <a href="/export_pdf"><button class="secondary">Xuất báo cáo PDF</button></a>
    <a href="/export"><button class="secondary">Xuất báo cáo Excel/CSV</button></a>
  </div>
</section>


<section class="panel module-section" id="crm">
  <div class="section-open-note">Bạn đang mở: CRM Mini.</div>
  <h2>CRM Mini</h2>
  <p class="small">Lưu khách hàng tiềm năng từ comment, inbox, Zalo hoặc data thủ công.</p>
  <form method="post" action="/crm">
    <div class="grid">
      <input name="name" placeholder="Tên khách hàng">
      <input name="phone" placeholder="Số điện thoại">
      <input name="zalo" placeholder="Zalo">
      <input name="source" placeholder="Nguồn: Facebook, Zalo, TikTok...">
    </div>
    <textarea name="note" rows="3" placeholder="Ghi chú nhu cầu khách hàng"></textarea>
    <button type="submit">Lưu khách hàng</button>
  </form>
  {% for r in crm_rows %}
  <div class="history">ID: {{ r[0] }} | {{ r[1] }} | SĐT: {{ r[2] }} | Zalo: {{ r[3] }} | Nguồn: {{ r[4] }} | {{ r[6] }}
Ghi chú: {{ r[5] }}</div>
  {% endfor %}
</section>

<section class="panel module-section" id="token">
  <div class="section-open-note">Bạn đang mở: Token Manager.</div>
  <h2>Token Center Pro</h2>
  <p class="small">Kiểm tra token từng Fanpage trước khi đăng hàng loạt. Nếu gặp OAuth 190 / Session expired thì cần lấy Page Token mới và thay vào file .env.</p>

  <form method="post" action="/check_tokens">
    <button type="submit">Kiểm tra toàn bộ Page Token</button>
  </form>

  <div class="history">{{ token_report }}</div>

  <h3>Kết quả kiểm tra gần nhất</h3>
  {% for t in token_checks %}
  <div class="history">
Page: {{ t[0] }}
ID: {{ t[1] }}
Trạng thái: {{ t[2] }}
Chi tiết: {{ t[3] }}
Thời gian: {{ t[4] }}
  </div>
  {% endfor %}

  <div class="premium-center">
    <h3>Hướng dẫn khi token giới hạn</h3>
    <p>1. Vào Meta Graph API Explorer.</p>
    <p>2. Generate User Token có quyền pages_show_list, pages_read_engagement, pages_manage_posts, pages_manage_metadata.</p>
    <p>3. Gọi /me/accounts để lấy Page Access Token mới.</p>
    <p>4. Dán token mới vào PAGES_JSON trong file .env.</p>
    <p>5. Chạy lại app.py và bấm kiểm tra token.</p>
  </div>
</section>

<section class="panel module-section" id="batch">
  <div class="section-open-note">Bạn đang mở: Excel / CSV hàng loạt.</div>
  <h2>Đăng hàng loạt Excel / CSV</h2>
  <p class="small">Cột hỗ trợ: idea, content, page_names, schedule_time, campaign, image_path hoặc media_path. Nếu không điền page_names, hệ thống tự chia bài lần lượt qua các Page để tránh trùng.</p>
  <form method="post" action="/batch" enctype="multipart/form-data">
    <input type="file" name="batch_file" accept=".xlsx,.csv">
    <button type="submit">Nhập file và lưu lịch</button>
  </form>
  <a href="/export"><button class="secondary">Xuất báo cáo CSV</button></a>
</section>

<section class="panel module-section" id="history">
  <h2>Lịch sử bài đăng / lịch hẹn</h2>
  {% for h in history %}
  <div class="history">
ID: {{ h[0] }}
Page: {{ h[1] }}
Trạng thái: {{ h[3] }}
Post ID / lỗi: {{ h[4] }}
Lịch đăng: {{ h[5] }}
Chiến dịch: {{ h[7] }}
Điểm content: {{ h[8] }}
Thời gian tạo: {{ h[9] }}

{{ h[2] }}
  </div>
  {% endfor %}
</section>

<section class="panel pricing-visible" id="pricing">
  <style>
    .premium-pricing-compact{position:relative;padding:28px;background:radial-gradient(circle at top left,rgba(37,99,235,.12),transparent 32%),linear-gradient(135deg,#f8fafc,#ffffff 46%,#fff7ed);border-radius:30px;border:1px solid rgba(226,232,240,.95);box-shadow:0 24px 70px rgba(15,23,42,.10);overflow:hidden}
    .premium-pricing-compact:before{content:"";position:absolute;inset:-2px;background:linear-gradient(120deg,rgba(59,130,246,.12),transparent,rgba(245,158,11,.14));pointer-events:none}
    .pricing-compact-head{position:relative;text-align:center;margin-bottom:20px;z-index:1}
    .pricing-compact-head .mini{display:inline-flex;align-items:center;justify-content:center;padding:7px 13px;border-radius:999px;background:linear-gradient(135deg,#eef2ff,#fff7ed);color:#1e3a8a;font-weight:1000;font-size:12px;letter-spacing:.08em;border:1px solid rgba(99,102,241,.18)}
    .pricing-compact-head h2{margin:10px 0 6px;font-size:30px;line-height:1.15;color:#0f172a;letter-spacing:-.04em}
    .pricing-compact-head p{margin:0;color:#64748b;font-size:14px}
    .pricing-grid-5{position:relative;z-index:1;display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:14px;align-items:stretch}
    .compact-price-card{position:relative;background:rgba(255,255,255,.92);backdrop-filter:blur(12px);border:1px solid rgba(226,232,240,.96);border-radius:22px;padding:18px 14px;text-align:center;box-shadow:0 14px 34px rgba(15,23,42,.08);transition:.2s ease;overflow:hidden}
    .compact-price-card:after{content:"";position:absolute;left:14px;right:14px;top:0;height:3px;border-radius:999px;background:linear-gradient(90deg,#2563eb,#f59e0b);opacity:.45}
    .compact-price-card:hover{transform:translateY(-5px);box-shadow:0 22px 46px rgba(15,23,42,.14);border-color:rgba(37,99,235,.22)}
    .compact-price-card .tag{display:inline-flex;padding:5px 10px;border-radius:999px;background:#f1f5f9;color:#334155;font-size:12px;font-weight:1000;margin-bottom:8px}
    .compact-price-card h3{margin:2px 0 8px;font-size:19px;color:#111827;letter-spacing:-.02em}
    .compact-price-card .price{font-size:25px;font-weight:1000;color:#2563eb;margin-bottom:6px;letter-spacing:-.04em}
    .compact-price-card .sub{min-height:38px;margin:0 0 14px;color:#64748b;font-size:13px;line-height:1.35}
    .plan-actions{display:grid;grid-template-columns:1fr;gap:8px;margin-top:auto}
    .compact-price-card button{width:100%;border:0;border-radius:13px;padding:11px 10px;font-weight:1000;cursor:pointer;transition:.18s ease}
    .compact-price-card .btn-upgrade{background:linear-gradient(135deg,#111827,#1e293b);color:#fff;box-shadow:0 12px 24px rgba(17,24,39,.18)}
    .compact-price-card .btn-benefit{background:#f8fafc;color:#334155;border:1px solid #e2e8f0;box-shadow:none;font-size:12px;padding:9px 10px}
    .compact-price-card .btn-benefit:hover{background:#eef2ff;color:#1d4ed8;border-color:#bfdbfe}
    .compact-price-card.popular{border:2px solid #f59e0b;transform:translateY(-3px);box-shadow:0 22px 48px rgba(245,158,11,.18)}
    .compact-price-card.popular:before{content:"Phổ biến";position:absolute;top:10px;right:-30px;background:#f59e0b;color:#111827;font-size:11px;font-weight:1000;padding:5px 32px;transform:rotate(35deg)}
    .compact-price-card.pro{background:linear-gradient(180deg,#0f172a,#111827 72%,#172554);border-color:#334155;color:#fff;box-shadow:0 24px 55px rgba(15,23,42,.28)}
    .compact-price-card.pro:after{background:linear-gradient(90deg,#facc15,#38bdf8);opacity:.85}
    .compact-price-card.pro h3,.compact-price-card.pro .sub{color:#fff}
    .compact-price-card.pro .price{color:#facc15;text-shadow:0 2px 16px rgba(250,204,21,.25)}
    .compact-price-card.pro .tag{background:rgba(250,204,21,.16);color:#facc15;border:1px solid rgba(250,204,21,.20)}
    .compact-price-card.pro .btn-upgrade{background:linear-gradient(135deg,#facc15,#f59e0b);color:#111827;box-shadow:0 12px 28px rgba(250,204,21,.20)}
    .compact-price-card.pro .btn-benefit{background:rgba(255,255,255,.08);border-color:rgba(255,255,255,.14);color:#e5e7eb}
    .premium-short-note{position:relative;z-index:1;margin-top:14px;padding:12px 14px;border-radius:16px;background:rgba(255,255,255,.78);border:1px dashed #cbd5e1;color:#475569;font-size:13px;text-align:center}
    .benefit-modal{position:fixed;inset:0;z-index:99999;display:none;align-items:center;justify-content:center;background:rgba(15,23,42,.62);backdrop-filter:blur(8px);padding:18px}
    .benefit-modal.show{display:flex}
    .benefit-box{width:min(760px,96vw);max-height:90vh;overflow:auto;background:linear-gradient(180deg,#ffffff,#f8fafc);border-radius:26px;border:1px solid rgba(226,232,240,.95);box-shadow:0 35px 90px rgba(15,23,42,.35);padding:22px;position:relative}
    .benefit-close{position:absolute;top:14px;right:14px;width:36px;height:36px;border:0;border-radius:999px;background:#f1f5f9;color:#0f172a;font-size:22px;font-weight:900;cursor:pointer}
    .benefit-top{padding-right:42px;margin-bottom:14px}
    .benefit-label{display:inline-flex;padding:6px 11px;border-radius:999px;background:#eef2ff;color:#1d4ed8;font-size:12px;font-weight:1000;margin-bottom:8px}
    .benefit-title{margin:0;font-size:28px;color:#0f172a;letter-spacing:-.04em}
    .benefit-price{margin:7px 0 0;font-size:24px;font-weight:1000;color:#2563eb}
    .benefit-desc{margin:8px 0 0;color:#64748b;line-height:1.5}
    .benefit-grid{display:grid;grid-template-columns:1.1fr .9fr .9fr;gap:12px;margin-top:16px}
    .benefit-panel{background:#fff;border:1px solid #e5e7eb;border-radius:18px;padding:14px;box-shadow:0 10px 24px rgba(15,23,42,.05)}
    .benefit-panel h4{margin:0 0 10px;color:#111827;font-size:15px}
    .benefit-panel ul{margin:0;padding-left:18px;color:#475569;font-size:13px;line-height:1.7}
    .benefit-actions{display:flex;gap:10px;margin-top:16px}
    .benefit-actions button{border:0;border-radius:14px;padding:12px 16px;font-weight:1000;cursor:pointer}
    .benefit-pay{background:#111827;color:#fff;flex:1}
    .benefit-cancel{background:#f1f5f9;color:#334155}
    @media(max-width:1200px){.pricing-grid-5{grid-template-columns:repeat(3,1fr)}}
    @media(max-width:760px){.premium-pricing-compact{padding:18px}.pricing-grid-5{grid-template-columns:1fr}.compact-price-card.popular{transform:none}.benefit-grid{grid-template-columns:1fr}.benefit-actions{flex-direction:column}}
  </style>

  <div class="premium-pricing-compact">
    <div class="pricing-compact-head">
      <span class="mini">PREMIUM SAAS</span>
      <h2>Bảng Giá Premium</h2>
      <p>Chọn gói phù hợp để mở khóa công cụ AI Marketing, CRM và Automation bán hàng.</p>
    </div>

    <div class="pricing-grid-5">
      <div class="compact-price-card" data-plan="monthly">
        <span class="tag">Cơ bản</span>
        <h3>Gói 1 Tháng</h3>
        <div class="price">159.000đ</div>
        <p class="sub">Dùng ngắn hạn, phù hợp người mới bắt đầu.</p>
        <div class="plan-actions">
          <button type="button" class="btn-upgrade" data-plan="monthly">Nâng cấp</button>
          <button type="button" class="btn-benefit" data-benefit="monthly">Xem quyền lợi</button>
        </div>
      </div>

      <div class="compact-price-card" data-plan="quarterly">
        <span class="tag">Tiết kiệm</span>
        <h3>Gói 3 Tháng</h3>
        <div class="price">359.000đ</div>
        <p class="sub">Ổn định hơn cho shop nhỏ và cá nhân bán hàng.</p>
        <div class="plan-actions">
          <button type="button" class="btn-upgrade" data-plan="quarterly">Nâng cấp</button>
          <button type="button" class="btn-benefit" data-benefit="quarterly">Xem quyền lợi</button>
        </div>
      </div>

      <div class="compact-price-card" data-plan="halfyear">
        <span class="tag">Tăng trưởng</span>
        <h3>Gói 6 Tháng</h3>
        <div class="price">559.000đ</div>
        <p class="sub">Mở thêm CRM, Sales Bot và quản lý khách hàng.</p>
        <div class="plan-actions">
          <button type="button" class="btn-upgrade" data-plan="halfyear">Nâng cấp</button>
          <button type="button" class="btn-benefit" data-benefit="halfyear">Xem quyền lợi</button>
        </div>
      </div>

      <div class="compact-price-card popular" data-plan="yearly">
        <span class="tag">Tối ưu</span>
        <h3>Gói 1 Năm</h3>
        <div class="price">859.000đ</div>
        <p class="sub">Gói phổ biến, chi phí thấp cho sử dụng dài hạn.</p>
        <div class="plan-actions">
          <button type="button" class="btn-upgrade" data-plan="yearly">Nâng cấp</button>
          <button type="button" class="btn-benefit" data-benefit="yearly">Xem quyền lợi</button>
        </div>
      </div>

      <div class="compact-price-card pro" data-plan="sellerpro">
        <span class="tag">Cao cấp</span>
        <h3>Seller Pro</h3>
        <div class="price">1.959.000đ</div>
        <p class="sub">Toàn bộ công cụ hiện tại, tương lai và ưu tiên hỗ trợ.</p>
        <div class="plan-actions">
          <button type="button" class="btn-upgrade" data-plan="sellerpro">Nâng cấp</button>
          <button type="button" class="btn-benefit" data-benefit="sellerpro">Xem quyền lợi</button>
        </div>
      </div>
    </div>

  </div>

  <div class="benefit-modal" id="planBenefitModal" aria-hidden="true">
    <div class="benefit-box">
      <button type="button" class="benefit-close" id="planBenefitClose">×</button>
      <div class="benefit-top">
        <span class="benefit-label" id="benefitLabel">PREMIUM</span>
        <h3 class="benefit-title" id="benefitTitle">Gói Premium</h3>
        <div class="benefit-price" id="benefitPrice">159.000đ</div>
        <p class="benefit-desc" id="benefitDesc"></p>
      </div>
      <div class="benefit-grid">
        <div class="benefit-panel"><h4>Tính năng mở khóa</h4><ul id="benefitFeatures"></ul></div>
        <div class="benefit-panel"><h4>Phù hợp với</h4><ul id="benefitFit"></ul></div>
        <div class="benefit-panel"><h4>Giá trị nhận được</h4><ul id="benefitValue"></ul></div>
      </div>
      <div class="benefit-actions">
        <button type="button" class="benefit-pay" id="benefitUpgradeBtn">Nâng cấp gói này</button>
        <button type="button" class="benefit-cancel" id="benefitCancelBtn">Đóng</button>
      </div>
    </div>
  </div>
</section>

<div class="payment-modal" id="paymentModal">
  <div class="payment-inner">
    <div class="payment-head">
      <div>
        <h2 id="payPlanTitle">GÓI PREMIUM</h2>
        <div id="payPlanDesc">Thông tin gói Premium</div>
      </div>
      <button class="payment-close" onclick="closePayment()">Đóng</button>
    </div>

    <div class="payment-body">
      <div class="qr-card">
        <h3>Quét mã QR để thanh toán</h3>
        <img id="payQr" src="https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount=159000&addInfo=PREMIUM%201THANG&accountName=NGUYEN%20DANG%20THI%20XUAN" alt="QR Agribank">
        <div class="bank-info">
          <b>Ngân hàng:</b> Agribank<br>
          <b>STK:</b> 8888363382629<br>
          <b>Chủ TK:</b> NGUYEN DANG THI XUAN<br>
          <b>Nội dung CK:</b> <span id="payContent">ID_THIET_BI</span>
        </div>
      </div>

      <div class="payment-detail">
        <h3>Số tiền cần thanh toán</h3>
        <div class="pay-amount" id="payPlanPrice">159.000 VNĐ</div>

        <h3>Thông tin kích hoạt</h3>
        <div class="payment-form-grid">
          <input id="payDeviceId" readonly placeholder="ID thiết bị">
          <input id="payPhone" oninput="refreshPaymentContent()" placeholder="Số điện thoại nhận kích hoạt">
          <input id="payEmail" oninput="refreshPaymentContent()" placeholder="Gmail nhận thông báo">
        </div>
        <button class="primary" onclick="submitPremiumRequest()" style="margin:10px 0 16px;width:100%">ĐÃ THANH TOÁN</button>

        <div class="payment-benefits" id="payBenefits" style="display:none"></div>
        <div class="payment-benefits" id="payLocked" style="display:none"></div>

        <div class="payment-alert">
          Sau khi thanh toán, vui lòng bấm <b>Đã thanh toán</b> để gửi ID thiết bị, số điện thoại, Gmail và gói đã đăng ký về web admin. Nếu sau 5 phút chưa được hỗ trợ, liên hệ Zalo <b>036 338 2629</b>.
        </div>

        <div class="payment-actions">
          <a href="https://zalo.me/0363382629" target="_blank">Liên hệ Zalo hỗ trợ</a>
          <a class="light" href="#token" onclick="closePayment()">Tôi đã thanh toán</a>
        </div>
      </div>
    </div>
  </div>
</div>

<div class="payment-modal" id="lockedFeatureModal">
  <div class="payment-inner" style="max-width:760px">
    <div class="payment-head">
      <div>
        <h2 id="lockedFeatureTitle">🔒 Tính năng Premium</h2>
        <div>Tính năng nâng cao cần nâng cấp Premium để sử dụng.</div>
      </div>
      <button class="payment-close" onclick="closeLockedFeature()">Đóng</button>
    </div>
    <div class="payment-detail">
      <div id="lockedFeaturePlans"></div>
    </div>
  </div>
</div>




<style id="affiliate-mobile-visible-fix">
#affiliateAdminSection{display:none!important}
#mobileCtvQuickBtn{position:fixed;left:16px;bottom:92px;z-index:2147482500;border:0;border-radius:999px;padding:12px 16px;background:linear-gradient(135deg,#16a34a,#22c55e,#84cc16);color:#fff;font-weight:900;font-size:14px;box-shadow:0 14px 34px rgba(22,163,74,.38);cursor:pointer;display:flex;align-items:center;gap:8px;letter-spacing:.2px}
#mobileCtvQuickBtn .dot{width:9px;height:9px;border-radius:50%;background:#dcfce7;box-shadow:0 0 0 6px rgba(220,252,231,.22)}
@media(min-width:900px){ #mobileCtvQuickBtn{bottom:26px;left:24px}}
@media(max-width:520px){ #mobileCtvQuickBtn{bottom:82px;left:12px;padding:11px 13px;font-size:13px}}
</style>
<button id="mobileCtvQuickBtn" type="button" onclick="return openModule('affiliate_center')"><span class="dot"></span>CTV Hoa Hồng</button>
<script id="affiliate-mobile-visible-fix-js">
(function(){
  function hideAdminCTVOnCustomer(){document.querySelectorAll('a[href="#affiliate_admin"], [onclick*="affiliate_admin"], #affiliateAdminSection').forEach(function(el){el.style.display='none';});}
  function makeCtvVisible(){hideAdminCTVOnCustomer();var btn=document.getElementById('mobileCtvQuickBtn'); if(btn) btn.style.display='flex'; if(location.hash==='#affiliate_center' && window.openModule) setTimeout(function(){window.openModule('affiliate_center')},300);}
  document.addEventListener('DOMContentLoaded',makeCtvVisible);setTimeout(makeCtvVisible,700);setTimeout(makeCtvVisible,1800);
})();
</script>

<!-- CTV AFFILIATE CENTER V1 -->
<style id="affiliate-center-v1-style">
#affiliateCenterSection,#affiliateAdminSection{display:none;background:#fff;border:1px solid rgba(15,23,42,.08);border-radius:22px;padding:18px;margin:18px 0;box-shadow:0 20px 60px rgba(15,23,42,.08)}
.aff-hero{display:grid;grid-template-columns:1.2fr .8fr;gap:14px;align-items:stretch}
.aff-card{background:linear-gradient(135deg,#f8fafc,#eef2ff);border:1px solid #dbeafe;border-radius:18px;padding:16px}
.aff-title{font-size:22px;font-weight:900;color:#0f172a;margin:0 0 6px}.aff-sub{color:#64748b;font-size:14px;line-height:1.5}
.aff-link-box{display:flex;gap:8px;align-items:center;margin-top:12px}.aff-link-box input{flex:1;border:1px solid #cbd5e1;border-radius:12px;padding:11px;font-weight:700;color:#0f172a;background:white}.aff-btn{border:0;border-radius:12px;padding:11px 14px;font-weight:900;cursor:pointer;background:linear-gradient(135deg,#16a34a,#22c55e);color:white;box-shadow:0 10px 28px rgba(22,163,74,.25)}.aff-btn.dark{background:linear-gradient(135deg,#0f172a,#334155)}.aff-btn.gold{background:linear-gradient(135deg,#f59e0b,#facc15);color:#111827}.aff-btn.light{background:#fff;color:#0f172a;border:1px solid #cbd5e1;box-shadow:none}
.aff-stats{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:14px 0}.aff-stat{background:#fff;border:1px solid #e2e8f0;border-radius:16px;padding:12px}.aff-stat b{display:block;font-size:20px;color:#0f172a}.aff-stat span{font-size:12px;color:#64748b}
.aff-form{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:12px}.aff-form input,.aff-form select{border:1px solid #cbd5e1;border-radius:12px;padding:11px;background:#fff;color:#0f172a}.aff-table{width:100%;border-collapse:collapse;margin-top:12px;font-size:13px}.aff-table th,.aff-table td{border-bottom:1px solid #e2e8f0;padding:9px;text-align:left}.aff-table th{color:#475569;background:#f8fafc}.aff-pill{display:inline-block;padding:4px 8px;border-radius:999px;background:#dcfce7;color:#166534;font-weight:900;font-size:12px}.aff-tabs{display:flex;gap:8px;flex-wrap:wrap;margin:12px 0}.aff-toast{position:fixed;right:22px;bottom:120px;background:#0f172a;color:white;padding:12px 14px;border-radius:14px;z-index:99999;box-shadow:0 12px 40px rgba(0,0,0,.25);display:none}
@media(max-width:900px){.aff-hero,.aff-form{grid-template-columns:1fr}.aff-stats{grid-template-columns:repeat(2,1fr)}}
</style>
<div id="affiliateCenterSection" class="app-module-section">
  <div class="aff-hero">
    <div class="aff-card">
      <h2 class="aff-title">CTV Hoa Hồng</h2>
      <div class="aff-sub">CTV đăng ký miễn phí, không cần mua gói. Khi bấm vào mục này, hệ thống tự tạo mã CTV và link giới thiệu riêng để gửi khách đăng ký. Khách nâng cấp Premium thành công thì hoa hồng được ghi nhận theo mã CTV.</div>
      <div class="aff-form">
        <input id="affName" placeholder="Tên CTV">
        <input id="affPhone" placeholder="Số điện thoại">
        <input id="affEmail" placeholder="Gmail">
      </div>
      <div style="margin-top:10px"><button class="aff-btn" onclick="affiliateSaveProfile()">Tạo link CTV miễn phí</button></div>
      <div class="aff-link-box"><input id="affRefLink" readonly placeholder="Link giới thiệu sẽ tự tạo tại đây"><button class="aff-btn dark" onclick="affiliateCopyLink()">Sao chép</button></div>
      <div class="aff-sub" style="margin-top:8px;color:#16a34a;font-weight:900">Link CTV được tạo miễn phí. CTV chỉ cần chia sẻ link, không cần mua gói Premium.</div>
      <div class="aff-sub" style="margin-top:8px">Mã CTV: <b id="affCode">---</b> • Cấp bậc: <b id="affLevel">---</b> • Hoa hồng: <b id="affPercent">---</b>%</div>
    </div>
    <div class="aff-card">
      <h3 style="margin-top:0">Rút hoa hồng</h3>
      <div class="aff-sub">Số dư khả dụng: <b id="affBalance">0đ</b></div>
      <div class="aff-form" style="grid-template-columns:1fr">
        <input id="wdAmount" placeholder="Số tiền muốn rút">
        <input id="wdBank" placeholder="Ngân hàng">
        <input id="wdNumber" placeholder="Số tài khoản">
        <input id="wdName" placeholder="Chủ tài khoản">
      </div>
      <button class="aff-btn gold" style="margin-top:10px" onclick="affiliateWithdraw()">Gửi yêu cầu rút tiền</button>
    </div>
  </div>
  <div class="aff-stats">
    <div class="aff-stat"><b id="affClicks">0</b><span>Tổng click</span></div>
    <div class="aff-stat"><b id="affRefs">0</b><span>Khách đăng ký</span></div>
    <div class="aff-stat"><b id="affPremium">0</b><span>Khách Premium</span></div>
    <div class="aff-stat"><b id="affCommission">0đ</b><span>Hoa hồng</span></div>
  </div>
  <h3>Khách đã giới thiệu</h3>
  <div style="overflow:auto"><table class="aff-table" id="affRecent"><thead><tr><th>Thiết bị</th><th>SĐT</th><th>Gmail</th><th>Gói</th><th>Doanh thu</th><th>Hoa hồng</th><th>Trạng thái</th></tr></thead><tbody></tbody></table></div>
</div>

<div id="affiliateAdminSection" class="app-module-section">
  <h2 class="aff-title">Quản Lý CTV</h2>
  <div class="aff-sub">Admin chỉnh phần trăm hoa hồng theo cấp bậc, xem CTV, doanh thu, hoa hồng và duyệt yêu cầu rút tiền.</div>
  <div class="aff-tabs">
    <input id="setNormal" placeholder="CTV thường %" value="20">
    <input id="setSilver" placeholder="CTV Bạc %" value="25">
    <input id="setGold" placeholder="CTV Vàng %" value="30">
    <input id="setDiamond" placeholder="CTV Kim Cương %" value="35">
    <button class="aff-btn" onclick="affiliateSaveSettings()">Lưu hoa hồng</button>
  </div>
  <h3>Danh sách CTV</h3>
  <div style="overflow:auto"><table class="aff-table" id="affAdminUsers"><thead><tr><th>CTV</th><th>SĐT</th><th>Gmail</th><th>Mã</th><th>Cấp</th><th>%</th><th>Đơn</th><th>Doanh thu</th><th>Hoa hồng</th></tr></thead><tbody></tbody></table></div>
  <h3>Yêu cầu rút tiền</h3>
  <div style="overflow:auto"><table class="aff-table" id="affWithdrawals"><thead><tr><th>ID</th><th>Mã CTV</th><th>Số tiền</th><th>Ngân hàng</th><th>STK</th><th>Chủ TK</th><th>Trạng thái</th><th>Thao tác</th></tr></thead><tbody></tbody></table></div>
</div>
<div id="affToast" class="aff-toast"></div>
<script id="affiliate-center-v1-js">
(function(){
  function q(s){return document.querySelector(s)}; function qa(s){return Array.from(document.querySelectorAll(s))};
  function money(n){n=parseInt(n||0,10)||0; return n.toLocaleString('vi-VN')+'đ'}
  function toast(t){var el=q('#affToast'); if(!el){alert(t);return} el.textContent=t; el.style.display='block'; setTimeout(function(){el.style.display='none'},2200)}
  function hideAllMain(){qa('main section,.module-section,.app-module-section').forEach(function(el){el.style.display='none'});}
  var oldOpen=window.openModule;
  window.openModule=function(name){
    if(name==='affiliate_center'){hideAllMain(); var s=q('#affiliateCenterSection'); if(s){s.style.display='block'; affiliateLoadMe(); window.scrollTo({top:0,behavior:'smooth'});} return false;}
    if(name==='affiliate_admin'){hideAllMain(); var a=q('#affiliateAdminSection'); if(a){a.style.display='block'; affiliateLoadAdmin(); window.scrollTo({top:0,behavior:'smooth'});} return false;}
    if(typeof oldOpen==='function') return oldOpen(name); return true;
  };
  function addMenus(){
    var nav=q('.nav'); if(!nav || q('[data-aff-menu="1"]')) return;
    var sysTitle=Array.from(nav.querySelectorAll('.v2-nav-title')).find(function(x){return /HỆ THỐNG/i.test(x.textContent||'')});
    var html='<a data-aff-menu="1" class="v2-nav-link" href="#affiliate_center" onclick="return openModule(\'affiliate_center\')"><span class="v2-nav-ico"></span><span class="v2-nav-text">CTV Hoa Hồng</span><span class="v2-nav-tag" style="background:#dcfce7;color:#166534">CTV</span></a>';
    if(sysTitle) sysTitle.insertAdjacentHTML('afterend',html); else nav.insertAdjacentHTML('beforeend',html);
  }
  window.affiliateLoadMe=function(){fetch('/api/affiliate/me').then(r=>r.json()).then(function(d){
    if(!d.ok) return; var s=d.summary||{};
    q('#affName').value=d.full_name||''; q('#affPhone').value=d.phone||''; q('#affEmail').value=d.email||'';
    q('#affRefLink').value=d.ref_link||''; q('#affCode').textContent=d.affiliate_code||'---'; q('#affLevel').textContent=d.level_name||'---'; q('#affPercent').textContent=d.commission_percent||20;
    q('#affClicks').textContent=s.clicks||0; q('#affRefs').textContent=s.referrals||0; q('#affPremium').textContent=s.premium||0; q('#affCommission').textContent=money(s.commission||0); q('#affBalance').textContent=money(s.balance||0);
    var tb=q('#affRecent tbody'); if(tb){tb.innerHTML=(s.recent||[]).map(function(r){return '<tr><td>'+ (r[0]||'') +'</td><td>'+ (r[1]||'') +'</td><td>'+ (r[2]||'') +'</td><td>'+ (r[3]||'') +'</td><td>'+ money(r[4]) +'</td><td>'+ money(r[5]) +'</td><td><span class="aff-pill">'+ (r[6]||'') +'</span></td></tr>'}).join('') || '<tr><td colspan="7">Chưa có khách giới thiệu.</td></tr>';}
  }).catch(function(){toast('Không tải được dữ liệu CTV')})}
  window.affiliateSaveProfile=function(){fetch('/api/affiliate/me',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({full_name:q('#affName').value,phone:q('#affPhone').value,email:q('#affEmail').value})}).then(r=>r.json()).then(function(d){toast(d.ok?'Đã cập nhật CTV':'Có lỗi xảy ra'); affiliateLoadMe();})}
  window.affiliateCopyLink=function(){var inp=q('#affRefLink'); if(!inp)return; inp.select(); navigator.clipboard&&navigator.clipboard.writeText(inp.value); toast('Đã sao chép link CTV')}
  window.affiliateWithdraw=function(){fetch('/api/affiliate/withdraw',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({amount:q('#wdAmount').value,bank_name:q('#wdBank').value,bank_number:q('#wdNumber').value,account_name:q('#wdName').value})}).then(r=>r.json()).then(function(d){toast(d.message||'Đã gửi yêu cầu'); affiliateLoadMe();})}
  window.affiliateLoadAdmin=function(){fetch('/api/admin/affiliate').then(r=>r.json()).then(function(d){
    var tb=q('#affAdminUsers tbody'); if(tb){tb.innerHTML=(d.users||[]).map(function(u){return '<tr><td>'+ (u[2]||u[1]||'') +'</td><td>'+ (u[3]||'') +'</td><td>'+ (u[4]||'') +'</td><td><b>'+ (u[5]||'') +'</b></td><td>'+ (u[6]||'') +'</td><td>'+ (u[7]||0) +'%</td><td>'+ (u[10]||0) +'</td><td>'+ money(u[11]) +'</td><td>'+ money(u[12]) +'</td></tr>'}).join('') || '<tr><td colspan="9">Chưa có CTV.</td></tr>';}
    (d.settings||[]).forEach(function(s){if(s[1]==='CTV thường')q('#setNormal').value=s[3]; if(s[1]==='CTV Bạc')q('#setSilver').value=s[3]; if(s[1]==='CTV Vàng')q('#setGold').value=s[3]; if(s[1]==='CTV Kim Cương')q('#setDiamond').value=s[3];});
    var wb=q('#affWithdrawals tbody'); if(wb){wb.innerHTML=(d.withdrawals||[]).map(function(w){return '<tr><td>'+w[0]+'</td><td>'+w[1]+'</td><td>'+money(w[2])+'</td><td>'+w[3]+'</td><td>'+w[4]+'</td><td>'+w[5]+'</td><td><span class="aff-pill">'+w[6]+'</span></td><td><button class="aff-btn light" onclick="affiliateWithdrawAction('+w[0]+',\'Đã thanh toán\')">Đã thanh toán</button> <button class="aff-btn light" onclick="affiliateWithdrawAction('+w[0]+',\'Từ chối\')">Từ chối</button></td></tr>'}).join('') || '<tr><td colspan="8">Chưa có yêu cầu rút tiền.</td></tr>';}
  })}
  window.affiliateSaveSettings=function(){fetch('/api/admin/affiliate/settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({normal:q('#setNormal').value,silver:q('#setSilver').value,gold:q('#setGold').value,diamond:q('#setDiamond').value})}).then(r=>r.json()).then(function(d){toast(d.message||'Đã lưu'); affiliateLoadAdmin();})}
  window.affiliateWithdrawAction=function(id,status){fetch('/api/admin/affiliate/withdraw/'+id,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({status:status})}).then(r=>r.json()).then(function(){toast('Đã cập nhật rút tiền'); affiliateLoadAdmin();})}
  function init(){addMenus(); var url=new URL(location.href); if(url.searchParams.get('ref')){try{document.cookie='affiliate_code='+url.searchParams.get('ref').toUpperCase()+';path=/;max-age='+(60*60*24*90)}catch(e){}}}
  if(document.readyState==='loading')document.addEventListener('DOMContentLoaded',init);else init(); setTimeout(init,500); setTimeout(init,1500);
})();
</script>

</main>

<aside class="rightbar">
  <h2>Hoạt động hôm nay</h2>
  <div class="activity-card">
    <span>Tổng bài</span><b>{{ s.total }}</b>
  </div>
  <div class="activity-card">
    <span>Đã đăng</span><b>{{ s.posted }}</b>
  </div>
  <div class="activity-card">
    <span>Chờ đăng</span><b>{{ s.scheduled }}</b>
  </div>
  <div class="activity-card">
    <span>Lead CRM</span><b>{{ s.crm }}</b>
  </div>
  <div class="activity-card">
    <span>Chiến dịch</span><b>{{ s.campaigns }}</b>
  </div>

  <div class="{{ 'free-status-card free-expired' if free_status.is_expired else 'free-status-card' }}">
    <h3>Gói dùng thử 3 ngày</h3>
    {% if free_status.is_expired %}
      <b>Trạng thái:</b> Đã hết dùng thử<br>
      Vui lòng nâng cấp Premium để tiếp tục sử dụng các công cụ.
    {% else %}
      <b>Còn lại:</b> {{ free_status.days }} ngày {{ free_status.hours }} giờ
      <div class="free-progress"><span style="width:{{ free_status.percent }}%"></span></div>
      <div class="trial-box">
        <b>Được sử dụng:</b><br>
        Quản lý Fanpage<br>
        Quản lý Group<br>
        AI Comment
      </div>
      <div class="trial-box locked-list">
        <b>Chưa mở khóa:</b><br>
        AI Messenger<br>
        CRM Kanban<br>
        AI Marketing Director<br>
        AI Video • AI Image<br>
        AI Giọng Nói • AI Livestream
      </div>
      <button onclick="scrollToPricing()">Xem chi tiết gói</button>
      <button onclick="openPayment('monthly')">Nâng cấp Premium</button>
    {% endif %}
  </div>


  <div class="v5-focus-box">
    <b>V5 Seller AI Suite</b><br>
    Fanpage • Group • AI Comment • AI Messenger • CRM Kanban • Marketing Director
  </div>
</aside>
</div>

<nav class="mobilebar">
  <a href="#dashboard">Home</a>
  <a href="#post">Đăng</a>
  <a href="#library">Kho</a>
  <a href="#plan">Plan</a>
  <a href="#history">Lịch sử</a>
</nav>

<script>
function toggleMenuGroup(el){
  const box = el.closest('.menu-mini-group');
  if(!box) return;
  box.classList.toggle('open');
}
</script>


<script>
(function(){
  function planKeyFromText(text){
    text = (text || '').toLowerCase();
    if(text.includes('1.959') || text.includes('1959') || text.includes('vĩnh') || text.includes('life')) return 'lifetime';
    if(text.includes('859') || text.includes('1 năm') || text.includes('nam') || text.includes('year')) return 'yearly';
    if(text.includes('559') || text.includes('6 tháng') || text.includes('business')) return 'halfyear';
    if(text.includes('359') || text.includes('3 tháng') || text.includes('pro')) return 'quarterly';
    if(text.includes('159') || text.includes('1 tháng') || text.includes('basic')) return 'monthly';
    return 'yearly';
  }
  function bindSafePricing(){
    document.querySelectorAll('.price-card,.premium-plan').forEach(function(card){
      if(card.dataset.safePricingReady) return;
      card.dataset.safePricingReady = '1';
      card.title = 'Bấm để xem chi tiết gói';
      card.addEventListener('click', function(e){
        if(e.target && (e.target.tagName === 'A' || e.target.tagName === 'BUTTON')) return;
        var key = planKeyFromText(card.innerText);
        if(typeof openPremiumCheckout === 'function') openPremiumCheckout(key);
        else if(typeof openPayment === 'function') openPayment(key);
        else if(typeof openPremium === 'function') openPremium();
      });
      var action = card.querySelector('.plan-button,.safe-pricing-action');
      if(action) {
        action.innerText = action.innerText.replace('','Xem chi tiết gói');
      }
    });
  }
  document.addEventListener('DOMContentLoaded', bindSafePricing);
  setTimeout(bindSafePricing, 600);
  setTimeout(bindSafePricing, 1600);
})();
</script>


<script>
function showInstallGuide(){
  alert("Cách cài App Mini:\\n\\nAndroid Chrome: bấm dấu 3 chấm → Thêm vào màn hình chính.\\n\\niPhone Safari: bấm Chia sẻ → Thêm vào MH chính.\\n\\nSau đó mở Mkt Automation Pro như một app trên điện thoại.");
}

let deferredInstallPrompt = null;
window.addEventListener('beforeinstallprompt', function(e){
  e.preventDefault();
  deferredInstallPrompt = e;
});
</script>


<script>
if ('serviceWorker' in navigator) {
  window.addEventListener('load', function(){
    navigator.serviceWorker.register('/sw.js').catch(function(err){
      console.log('Service worker registration failed:', err);
    });
  });
}
</script>


<script>
function showInstallGuide(){
  alert("Cách cài App Mini:\\n\\nAndroid Chrome: bấm dấu 3 chấm → Thêm vào màn hình chính.\\n\\niPhone Safari: bấm Chia sẻ → Thêm vào MH chính.\\n\\nSau đó mở Mkt Automation Pro V2 như một app.");
}
(function(){
  function planKeyFromText(text){
    text = (text || '').toLowerCase();
    if(text.includes('1.959') || text.includes('1959') || text.includes('vĩnh') || text.includes('life')) return 'lifetime';
    if(text.includes('859') || text.includes('1 năm') || text.includes('nam') || text.includes('year')) return 'yearly';
    if(text.includes('559') || text.includes('6 tháng') || text.includes('business')) return 'halfyear';
    if(text.includes('359') || text.includes('3 tháng') || text.includes('pro')) return 'quarterly';
    if(text.includes('159') || text.includes('1 tháng') || text.includes('basic')) return 'monthly';
    return 'yearly';
  }
  function bindV2Pricing(){
    document.querySelectorAll('.price-card,.premium-plan').forEach(function(card){
      if(card.dataset.v2PricingReady) return;
      card.dataset.v2PricingReady = '1';
      card.title = 'Bấm để xem chi tiết gói';
      card.addEventListener('click', function(e){
        if(e.target && (e.target.tagName === 'A' || e.target.tagName === 'BUTTON')) return;
        var key = planKeyFromText(card.innerText);
        if(typeof openPremiumCheckout === 'function') openPremiumCheckout(key);
        else if(typeof openPayment === 'function') openPayment(key);
        else if(typeof openPremium === 'function') openPremium();
      });
      card.querySelectorAll('button').forEach(function(btn){
        if((btn.innerText || '').includes('Xem chi tiết gói')) btn.innerText = 'Xem chi tiết gói';
      });
    });
  }
  document.addEventListener('DOMContentLoaded', bindV2Pricing);
  setTimeout(bindV2Pricing, 800);
})();
if ('serviceWorker' in navigator) {
  window.addEventListener('load', function(){
    navigator.serviceWorker.register('/sw.js').catch(function(err){
      console.log('Service worker failed:', err);
    });
  });
}
</script>


<script>
let draggedKanbanCard=null;
function dragKanban(ev){ draggedKanbanCard=ev.target; }
function dropKanban(ev){ ev.preventDefault(); const col=ev.currentTarget; if(draggedKanbanCard){ col.appendChild(draggedKanbanCard); draggedKanbanCard=null; } }
</script>

<script id="chat-device-menu-fix-js">
(function(){
  function ensureDeviceId(){
    var id = localStorage.getItem("mkt_device_id");
    if(!id || id === "Đang tạo..."){
      id = "MKT-" + Math.random().toString(36).slice(2,8).toUpperCase() + Date.now().toString().slice(-4);
      localStorage.setItem("mkt_device_id", id);
    }
    document.cookie = "mkt_device_id=" + encodeURIComponent(id) + "; path=/; max-age=" + (60*60*24*365*5);
    var side = document.getElementById("sidebarDeviceId");
    if(side) side.textContent = id;
    var pay = document.getElementById("payDeviceId");
    if(pay) pay.value = id;
    return id;
  }
  window.ensureDeviceId = ensureDeviceId;
  document.addEventListener("DOMContentLoaded", function(){
    ensureDeviceId();
    setTimeout(ensureDeviceId, 300);
    setTimeout(ensureDeviceId, 1200);

    var bubble = document.querySelector(".bot-bubble");
    if(bubble){
      bubble.addEventListener("click", function(e){
        e.preventDefault();
        var panel = document.getElementById("floatingBotPanel");
        if(panel){
          panel.style.display = (panel.style.display === "block") ? "none" : "block";
          if(panel.style.display === "block" && typeof appendBotGreeting === "function") appendBotGreeting();
        }
      }, true);
    }

    document.querySelectorAll(".v2-nav-link[href^='#']").forEach(function(a){
      a.addEventListener("click", function(e){
        var id = (a.getAttribute("href") || "").replace("#", "");
        if(!id) return;
        e.preventDefault();
        if(typeof openModule === "function") openModule(id);
      }, true);
    });
  });
})();
</script>


<style id="final-chat-lock-premium-fix">
/* FIX CUỐI: chat không tràn, menu trái bỏ ký hiệu, khóa menu Premium */
.v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important}
.v2-nav-link{position:relative!important;gap:8px!important;padding-right:112px!important;cursor:pointer!important;pointer-events:auto!important}
.v2-nav-link .v2-nav-tag{display:none!important}
.v2-nav-link.premium-locked::after{
  content:"PREMIUM";
  position:absolute;right:14px;top:50%;transform:translateY(-50%);
  font-size:11px;font-weight:1000;letter-spacing:.6px;color:#fff;
  padding:5px 10px;border-radius:999px;
  background:linear-gradient(135deg,#facc15,#fb923c,#ef4444);
  box-shadow:0 0 12px rgba(250,204,21,.95),0 0 22px rgba(124,58,237,.65);
  animation:premiumGlow 1.2s infinite alternate;
}
.v2-nav-link.premium-locked{border-color:rgba(250,204,21,.55)!important;background:linear-gradient(135deg,rgba(49,46,129,.55),rgba(76,29,149,.45))!important}
@keyframes premiumGlow{from{filter:brightness(1)}to{filter:brightness(1.35)}}

.floating-bot{position:fixed!important;right:24px!important;bottom:24px!important;z-index:2147483000!important;pointer-events:none!important}
.bot-bubble{pointer-events:auto!important;position:relative!important;z-index:2147483002!important}
.bot-panel{
  pointer-events:auto!important;display:none;position:absolute!important;right:0!important;bottom:92px!important;
  width:min(420px,calc(100vw - 28px))!important;height:min(660px,calc(100vh - 120px))!important;
  max-height:calc(100vh - 120px)!important;overflow:hidden!important;
  border-radius:24px!important;background:#fff!important;box-shadow:0 25px 80px rgba(15,23,42,.35)!important;
  z-index:2147483001!important;flex-direction:column!important;
}
.bot-panel.bot-open{display:flex!important}
.bot-head{flex:0 0 auto!important}
.bot-body{flex:1 1 auto!important;min-height:0!important;max-height:none!important;overflow-y:auto!important;padding:18px!important;background:#F8FAFC!important}
.bot-msg{max-width:100%!important;box-sizing:border-box!important;word-break:break-word!important;overflow-wrap:anywhere!important;margin:0 0 10px!important}
.bot-actions{flex:0 0 auto!important;display:grid!important;grid-template-columns:1fr 1fr!important;gap:10px!important;padding:12px 18px!important;background:#fff!important;border-top:1px solid #EEF2FF!important;box-sizing:border-box!important}
.bot-actions button{min-height:56px!important;margin:0!important;white-space:normal!important;line-height:1.2!important;font-size:13px!important}
.bot-input{flex:0 0 auto!important;padding:12px 18px 18px!important;background:#fff!important;display:flex!important;gap:10px!important;box-sizing:border-box!important}
.bot-input input{min-width:0!important;flex:1!important}
.bot-input button{flex:0 0 62px!important}
@media(max-width:520px){.floating-bot{right:10px!important;bottom:10px!important}.bot-panel{right:0!important;width:calc(100vw - 20px)!important;height:calc(100vh - 110px)!important}}
</style>

<script id="final-chat-lock-premium-fix-js">
(function(){
  var MKT_IS_PREMIUM = {{ 'true' if is_device_premium else 'false' }};
  var premiumModules = {
    messenger_ai:'AI Messenger', crm_sales:'CRM Kanban', marketing_director:'AI Marketing Director',
    ai_studio:'AI Studio', creative_center:'Image / Video / Voice', analytics:'Analytics Center',
    analytics_center:'Analytics Center', automation_center:'Cài đặt Automation'
  };
  var aliases = {post:'page_center_total', page_center:'page_center_total', page_comment_pro:'page_center_total', page_comment_queue:'page_center_total'};
  function normalizeId(id){ return aliases[id] || id; }
  function isLocked(id){ return !MKT_IS_PREMIUM && !!premiumModules[normalizeId(id)]; }
  function lockFeature(id){
    var name = premiumModules[normalizeId(id)] || 'Tính năng Premium';
    if(typeof openLockedFeature === 'function') return openLockedFeature(name);
    if(typeof openPremiumPopup === 'function') openPremiumPopup();
    var p=document.getElementById('premium'); if(p) p.scrollIntoView({behavior:'smooth',block:'start'});
    return false;
  }
  function showOnlyModule(id){
    id = normalizeId(id);
    if(isLocked(id)) return lockFeature(id);
    document.querySelectorAll('.module-section').forEach(function(el){el.classList.remove('active-module');});
    var target=document.getElementById(id) || document.getElementById('dashboard');
    if(target){
      target.classList.add('active-module');
      setTimeout(function(){target.scrollIntoView({behavior:'smooth',block:'start'});},30);
    }
    document.querySelectorAll('.v2-nav-link').forEach(function(a){a.classList.remove('active');});
    var active=document.querySelector('.v2-nav-link[href="#'+id+'"]');
    if(active) active.classList.add('active');
    return false;
  }
  window.openModule = showOnlyModule;

  function setBotOpen(open){
    var panel=document.getElementById('floatingBotPanel');
    if(!panel) return;
    panel.classList.toggle('bot-open', !!open);
    panel.style.display = open ? 'flex' : 'none';
    if(open && typeof appendBotGreeting === 'function') appendBotGreeting();
    setTimeout(function(){var body=document.getElementById('floatingBotBody'); if(body) body.scrollTop=body.scrollHeight;},50);
  }
  window.toggleFloatingBot=function(){
    var panel=document.getElementById('floatingBotPanel');
    setBotOpen(!(panel && panel.classList.contains('bot-open')));
    return false;
  };
  window.closeFloatingBot=function(){setBotOpen(false);return false;};

  document.addEventListener('DOMContentLoaded',function(){
    document.querySelectorAll('.v2-nav-link').forEach(function(a){
      var id=(a.getAttribute('href')||'').replace('#','');
      a.querySelectorAll('.v2-nav-ico').forEach(function(x){x.remove();});
      if(isLocked(id)) a.classList.add('premium-locked');
      else a.classList.remove('premium-locked');
      a.addEventListener('click',function(e){
        var mid=(a.getAttribute('href')||'').replace('#','');
        if(mid){e.preventDefault();e.stopPropagation();showOnlyModule(mid);}
      },true);
    });
    var panel=document.getElementById('floatingBotPanel');
    if(panel){panel.style.display='none';panel.classList.remove('bot-open');}
    document.querySelectorAll('.bot-actions button,.bot-input button,.bot-bubble,.bot-close').forEach(function(btn){btn.style.pointerEvents='auto';});
  });
})();
</script>

<style id="pro-green-admin-chat-final">
.v2-nav-ico{display:none!important}.v2-nav-link .v2-nav-tag{display:none!important}
.v2-nav-link{position:relative!important;padding-right:112px!important;cursor:pointer!important;pointer-events:auto!important}
.v2-nav-link.pro-feature::after{
  content:"PRO";position:absolute;right:16px;top:12px;transform:none;
  font-size:12px;font-weight:1000;letter-spacing:.8px;color:#ECFDF5;padding:6px 12px;border-radius:999px;
  background:linear-gradient(135deg,#16A34A,#22C55E,#86EFAC);box-shadow:inset 0 0 10px rgba(255,255,255,.45),0 0 12px rgba(34,197,94,.95),0 0 26px rgba(34,197,94,.45);
  animation:proPulse 1.25s infinite alternate;
}
.v2-nav-link.premium-locked::after{
  content:"PREMIUM"!important;position:absolute;right:12px;top:12px;transform:none;
  font-size:11px;font-weight:1000;letter-spacing:.6px;color:#fff;padding:6px 10px;border-radius:999px;
  background:linear-gradient(135deg,#F59E0B,#FB923C,#EF4444);box-shadow:0 0 14px rgba(245,158,11,.95),0 0 24px rgba(239,68,68,.45);
}
@keyframes proPulse{from{filter:brightness(1)}to{filter:brightness(1.35)}}
.bot-panel{display:none;position:absolute!important;right:0!important;bottom:92px!important;width:min(430px,calc(100vw - 30px))!important;height:min(650px,calc(100vh - 130px))!important;overflow:hidden!important;flex-direction:column!important}
.bot-panel.bot-open{display:flex!important}.bot-body{flex:1 1 auto!important;min-height:0!important;overflow-y:auto!important}.bot-actions,.bot-input{flex:0 0 auto!important}.bot-msg{word-break:break-word!important;overflow-wrap:anywhere!important}
</style>
<script id="pro-green-admin-chat-final-js">
(function(){
  var IS_PREMIUM = {{ 'true' if is_device_premium else 'false' }};
  var TRIAL_EXPIRED = {{ 'true' if free_status.is_expired else 'false' }};
  var CORE_PRO = {facebook_center:1,page_center_total:1,post:1,fanpage_manager:1,group_suite:1,comment_manager:1};
  var PREMIUM_ONLY = {messenger_ai:1,crm_sales:1,marketing_director:1,ai_studio:1,creative_center:1,analytics:1,analytics_center:1,automation_center:1};
  var ALIAS = {post:'page_center_total',page_center:'page_center_total',page_comment_pro:'page_center_total',page_comment_queue:'page_center_total'};
  function norm(id){return ALIAS[id]||id;}
  function featureName(id){var map={facebook_center:'Facebook Center',page_center_total:'Page Center Tổng',fanpage_manager:'Quản lý Fanpage',group_suite:'Group Center Tổng',comment_manager:'AI Comment',messenger_ai:'AI Messenger',crm_sales:'CRM Kanban',marketing_director:'AI Marketing Director',ai_studio:'AI Studio'};return map[norm(id)]||'Tính năng Premium';}
  function locked(id){id=norm(id); if(IS_PREMIUM) return false; if(PREMIUM_ONLY[id]) return true; if(TRIAL_EXPIRED && CORE_PRO[id]) return true; return false;}
  window.openModule=function(id){
    id=norm(id);
    if(locked(id)){
      if(typeof openLockedFeature==='function') openLockedFeature(featureName(id));
      var pr=document.getElementById('premium'); if(pr) setTimeout(function(){pr.scrollIntoView({behavior:'smooth',block:'start'});},120);
      return false;
    }
    document.querySelectorAll('.module-section').forEach(function(el){el.classList.remove('active-module');});
    var target=document.getElementById(id)||document.getElementById('dashboard');
    if(target){target.classList.add('active-module'); setTimeout(function(){target.scrollIntoView({behavior:'smooth',block:'start'});},30);}
    return false;
  };
  function markMenu(){
    document.querySelectorAll('.v2-nav-link').forEach(function(a){
      var id=(a.getAttribute('href')||'').replace('#',''); var nid=norm(id);
      a.querySelectorAll('.v2-nav-ico,.v2-nav-tag').forEach(function(x){x.remove();});
      a.classList.remove('pro-feature','premium-locked');
      if(CORE_PRO[nid] && !locked(id)) a.classList.add('pro-feature');
      if(locked(id)) a.classList.add('premium-locked');
      a.onclick=function(e){e.preventDefault();e.stopPropagation();return window.openModule(id);};
    });
  }
  function addMsg(sender,msg){
    var body=document.getElementById('floatingBotBody'); if(!body||!msg) return;
    var cls = sender==='admin' ? 'bot-msg ai' : 'bot-msg';
    var label = sender==='admin' ? 'Admin hỗ trợ' : 'Bạn';
    var div=document.createElement('div'); div.className=cls; div.innerHTML='<b>'+label+':</b><br>'+String(msg).replace(/[&<>]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;'}[c];}).replace(/\n/g,'<br>');
    body.appendChild(div); body.scrollTop=body.scrollHeight;
  }
  var lastSupportId=Number(localStorage.getItem('mkt_support_last_id')||0);
  function sendToAdmin(text){
    var device=(typeof getOrCreateDeviceId==='function'?getOrCreateDeviceId():(localStorage.getItem('mkt_device_id')||''));
    fetch('/support_send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:device,sender:'user',message:text})}).catch(function(){});
  }
  function pollAdmin(){
    var device=(typeof getOrCreateDeviceId==='function'?getOrCreateDeviceId():(localStorage.getItem('mkt_device_id')||''));
    if(!device) return;
    fetch('/support_poll?device_id='+encodeURIComponent(device)+'&after_id='+encodeURIComponent(lastSupportId)).then(function(r){return r.json();}).then(function(data){
      if(!data||!data.messages) return;
      data.messages.forEach(function(m){ lastSupportId=Math.max(lastSupportId,Number(m.id)||0); if(m.sender==='admin') addMsg('admin',m.message); });
      localStorage.setItem('mkt_support_last_id',String(lastSupportId));
    }).catch(function(){});
  }
  var oldBotQuick=window.botQuick;
  window.botQuick=function(text){ if(oldBotQuick) oldBotQuick(text); sendToAdmin(text); return false; };
  var oldSend=window.sendBotInput;
  window.sendBotInput=function(){
    var input=document.getElementById('botInputText'); if(!input||!input.value.trim()) return false;
    var text=input.value.trim();
    if(oldBotQuick) oldBotQuick(text); else addMsg('user',text);
    sendToAdmin(text); input.value=''; return false;
  };
  window.toggleFloatingBot=function(){var p=document.getElementById('floatingBotPanel'); if(!p) return false; var open=!p.classList.contains('bot-open'); p.classList.toggle('bot-open',open); p.style.display=open?'flex':'none'; if(open){if(typeof appendBotGreeting==='function') appendBotGreeting(); pollAdmin();} return false;};
  window.closeFloatingBot=function(){var p=document.getElementById('floatingBotPanel'); if(p){p.classList.remove('bot-open');p.style.display='none';} return false;};
  document.addEventListener('DOMContentLoaded',function(){markMenu();setInterval(pollAdmin,3000);setTimeout(pollAdmin,800);});
})();
</script>



<style id="REAL_LAST_FIX_20260610">
/* REAL LAST FIX: PRO badge xanh, chat/menu/price click ổn định */
.v2-nav-ico,.v2-nav-tag{display:none!important}
.v2-nav-link{position:relative!important;cursor:pointer!important;pointer-events:auto!important;padding-right:96px!important;user-select:none!important}
.v2-nav-link.real-pro::after{
  content:"PRO";position:absolute;right:14px;top:14px;z-index:5;
  font-size:12px;font-weight:1000;letter-spacing:.9px;color:#ECFDF5;
  padding:6px 13px;border-radius:999px;
  background:linear-gradient(135deg,#15803D,#22C55E,#86EFAC);
  box-shadow:inset 0 0 10px rgba(255,255,255,.5),0 0 14px rgba(34,197,94,.95),0 0 30px rgba(34,197,94,.50);
  animation:realProGlow 1.05s infinite alternate;
}
.v2-nav-link.real-premium-lock::after{
  content:"PREMIUM";position:absolute;right:10px;top:14px;z-index:5;
  font-size:11px;font-weight:1000;letter-spacing:.5px;color:white;
  padding:6px 10px;border-radius:999px;
  background:linear-gradient(135deg,#F59E0B,#FB923C,#EF4444);
  box-shadow:0 0 16px rgba(245,158,11,.95),0 0 30px rgba(239,68,68,.50);
  animation:realProGlow 1.05s infinite alternate;
}
@keyframes realProGlow{from{filter:brightness(1);transform:scale(1)}to{filter:brightness(1.35);transform:scale(1.04)}}
.floating-bot{position:fixed!important;right:24px!important;bottom:24px!important;z-index:2147483600!important;pointer-events:none!important}
.bot-bubble{pointer-events:auto!important;z-index:2147483602!important;cursor:pointer!important}
.bot-panel{pointer-events:auto!important;position:absolute!important;right:0!important;bottom:92px!important;width:min(430px,calc(100vw - 30px))!important;height:min(650px,calc(100vh - 120px))!important;max-height:calc(100vh - 120px)!important;overflow:hidden!important;background:#fff!important;border-radius:24px!important;display:none!important;flex-direction:column!important;z-index:2147483601!important}
.bot-panel.real-open{display:flex!important}
.bot-head{flex:0 0 auto!important}.bot-body{flex:1 1 auto!important;min-height:0!important;overflow-y:auto!important;padding:18px!important;background:#F8FAFC!important}.bot-actions{flex:0 0 auto!important;display:grid!important;grid-template-columns:1fr 1fr!important;gap:10px!important;padding:12px 18px!important;background:#fff!important;border-top:1px solid #EEF2FF!important}.bot-actions button{min-height:58px!important;margin:0!important;white-space:normal!important;line-height:1.2!important;cursor:pointer!important;pointer-events:auto!important}.bot-input{flex:0 0 auto!important;display:flex!important;gap:10px!important;padding:12px 18px 18px!important;background:#fff!important}.bot-input input{min-width:0!important;flex:1!important}.bot-input button{flex:0 0 64px!important;cursor:pointer!important;pointer-events:auto!important}.bot-msg{max-width:100%!important;box-sizing:border-box!important;word-break:break-word!important;overflow-wrap:anywhere!important}.price-card,.premium-plan,.plan-button,.premium-btn{cursor:pointer!important;pointer-events:auto!important}#liveMemberCount{display:inline-block;min-width:32px;transition:transform .25s ease,color .25s ease}#liveMemberCount.bump{transform:scale(1.18);color:#16A34A!important}
</style>
<script id="REAL_LAST_FIX_20260610_JS">
(function(){
  var IS_PREMIUM = {{ 'true' if is_device_premium else 'false' }};
  var TRIAL_EXPIRED = {{ 'true' if free_status.is_expired else 'false' }};
  var CORE_PRO = {facebook_center:1,page_center_total:1,post:1,fanpage_manager:1,group_suite:1,comment_manager:1};
  var PREMIUM_ONLY = {messenger_ai:1,crm_sales:1,marketing_director:1,ai_studio:1,creative_center:1,analytics:1,analytics_center:1,automation_center:1};
  var ALIAS = {post:'page_center_total',page_center:'page_center_total',group_marketing:'group_suite',group_finder:'group_suite',group_uid_splitter:'group_suite',group_join_queue:'group_suite',group_post_filter:'group_suite',page_comment_pro:'page_center_total',page_comment_queue:'page_center_total'};
  function norm(id){return ALIAS[id]||id;}
  function locked(id){id=norm(id); if(IS_PREMIUM) return false; if(PREMIUM_ONLY[id]) return true; if(TRIAL_EXPIRED && CORE_PRO[id]) return true; return false;}
  function fname(id){id=norm(id); var m={facebook_center:'Facebook Center',page_center_total:'Page Center Tổng',fanpage_manager:'Quản lý Fanpage',group_suite:'Group Center Tổng',comment_manager:'AI Comment',messenger_ai:'AI Messenger',crm_sales:'CRM Kanban',marketing_director:'AI Marketing Director',ai_studio:'AI Studio',creative_center:'Image / Video / Voice',analytics:'Analytics Center',automation_center:'Cài đặt Automation'}; return m[id]||'Tính năng Premium';}
  window.openModule=function(id){
    id=norm(id||'dashboard');
    if(locked(id)){
      if(typeof openLockedFeature==='function') openLockedFeature(fname(id));
      var premium=document.getElementById('premium'); if(premium) setTimeout(function(){premium.scrollIntoView({behavior:'smooth',block:'start'});},180);
      return false;
    }
    document.querySelectorAll('.module-section').forEach(function(el){el.classList.remove('active-module');});
    var target=document.getElementById(id)||document.getElementById('dashboard');
    if(target){target.classList.add('active-module'); setTimeout(function(){target.scrollIntoView({behavior:'smooth',block:'start'});},30);}
    document.querySelectorAll('.v2-nav-link').forEach(function(a){a.classList.remove('active');});
    var active=document.querySelector('.v2-nav-link[href="#'+id+'"]'); if(active) active.classList.add('active');
    return false;
  };
  function markMenu(){
    document.querySelectorAll('.v2-nav-link').forEach(function(a){
      var raw=(a.getAttribute('href')||'').replace('#',''); var id=norm(raw);
      a.querySelectorAll('.v2-nav-ico,.v2-nav-tag').forEach(function(x){x.remove();});
      a.classList.remove('real-pro','real-premium-lock','pro-feature','premium-locked');
      if(locked(raw)) a.classList.add('real-premium-lock'); else if(CORE_PRO[id]) a.classList.add('real-pro');
    });
  }
  function openBot(open){
    var p=document.getElementById('floatingBotPanel'); if(!p) return false;
    if(open===undefined) open=!p.classList.contains('real-open');
    p.classList.toggle('real-open',!!open); p.style.display=open?'flex':'none';
    if(open && typeof appendBotGreeting==='function') appendBotGreeting();
    setTimeout(function(){var b=document.getElementById('floatingBotBody'); if(b) b.scrollTop=b.scrollHeight;},60);
    return false;
  }
  window.toggleFloatingBot=function(){return openBot();};
  window.closeFloatingBot=function(){return openBot(false);};
  function planKeyFromText(text){text=(text||'').toLowerCase(); if(text.includes('1.959')||text.includes('1959')||text.includes('nhà bán')||text.includes('chuyên nghiệp')||text.includes('trọn đời')||text.includes('vĩnh')) return 'lifetime'; if(text.includes('859')||text.includes('1 năm')) return 'yearly'; if(text.includes('559')||text.includes('6 tháng')) return 'halfyear'; if(text.includes('359')||text.includes('3 tháng')) return 'quarterly'; if(text.includes('159')||text.includes('1 tháng')) return 'monthly'; return 'yearly';}
  function openPlanFrom(el){var card=el.closest('.premium-plan,.price-card')||el; var key=planKeyFromText(card.innerText||el.innerText||''); if(typeof openPayment==='function') openPayment(key); else if(typeof scrollToPricing==='function') scrollToPricing(); return false;}
  function animateCount(){
    var el=document.getElementById('liveMemberCount'); if(!el) return;
    var n=parseInt((el.textContent||localStorage.getItem('mkt_live_count')||'231').replace(/\D/g,''),10)||231;
    var step=Math.floor(Math.random()*3)+1; n+=step; if(n>999) n=231+Math.floor(Math.random()*20);
    el.textContent=String(n); localStorage.setItem('mkt_live_count',String(n));
    el.classList.add('bump'); setTimeout(function(){el.classList.remove('bump');},260);
  }
  function boot(){
    markMenu();
    var saved=localStorage.getItem('mkt_live_count'); var el=document.getElementById('liveMemberCount'); if(el && saved) el.textContent=saved;
    setInterval(animateCount,1400);
    document.addEventListener('click',function(e){
      var bubble=e.target.closest('.bot-bubble'); if(bubble){e.preventDefault();e.stopPropagation();return openBot();}
      var close=e.target.closest('.bot-close'); if(close){e.preventDefault();e.stopPropagation();return openBot(false);}
      var quick=e.target.closest('.bot-actions button'); if(quick){e.preventDefault();e.stopPropagation(); if(typeof botQuick==='function') botQuick((quick.innerText||'').trim()); return false;}
      var send=e.target.closest('.bot-input button'); if(send){e.preventDefault();e.stopPropagation(); if(typeof sendBotInput==='function') sendBotInput(); return false;}
      var nav=e.target.closest('.v2-nav-link[href^="#"]'); if(nav){e.preventDefault();e.stopPropagation();return window.openModule((nav.getAttribute('href')||'').replace('#',''));}
      var priceBtn=e.target.closest('button,.plan-button,.premium-btn'); if(priceBtn && ((priceBtn.innerText||'').includes('Xem chi tiết gói') || priceBtn.closest('.premium-plan,.price-card'))){e.preventDefault();e.stopPropagation();return openPlanFrom(priceBtn);}
      var card=e.target.closest('.premium-plan,.price-card'); if(card){e.preventDefault();e.stopPropagation();return openPlanFrom(card);}
    },true);
    var input=document.getElementById('botInputText'); if(input){input.addEventListener('keydown',function(e){if(e.key==='Enter'){e.preventDefault(); if(typeof sendBotInput==='function') sendBotInput();}},true);}
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',boot); else boot();
  setTimeout(markMenu,700); setTimeout(markMenu,1800);
})();
</script>



<!-- FINAL REAL FIX: menu PRO badge, pricing buttons, chat buttons, hide technical status -->
<style id="final-real-fix-style">
/* Ẩn toàn bộ trạng thái kỹ thuật khỏi giao diện khách */
.rightbar h2, .rightbar p{ }
.rightbar h2:has(+ p){ }
/* Xóa icon/ký hiệu bên trái menu, chỉ giữ chữ */
.v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important;overflow:hidden!important}
.v2-nav-link{position:relative!important;cursor:pointer!important;user-select:none!important;display:flex!important;align-items:center!important;justify-content:space-between!important;gap:10px!important}
.v2-nav-link .v2-nav-tag{display:none!important}
.v2-nav-link.menu-pro::after{
  content:'PRO';
  display:inline-flex;align-items:center;justify-content:center;
  min-width:44px;height:24px;padding:0 10px;border-radius:999px;
  color:#052e16;font-weight:1000;font-size:12px;letter-spacing:.7px;
  background:linear-gradient(135deg,#BBF7D0,#22C55E,#16A34A);
  border:1px solid rgba(187,247,208,.95);
  box-shadow:0 0 8px rgba(34,197,94,.9),0 0 18px rgba(34,197,94,.55),inset 0 1px 0 rgba(255,255,255,.8);
  animation:proPulseGreen 1.45s infinite alternate;
}
@keyframes proPulseGreen{from{filter:brightness(1);transform:scale(1)}to{filter:brightness(1.28);transform:scale(1.04)}}
.v2-nav-link.menu-locked{opacity:.96!important}
.v2-nav-link.menu-locked::before{content:'';position:absolute;inset:0;border-radius:inherit;background:rgba(15,23,42,.04);pointer-events:none}
.price-card button,.premium-plan button,.safe-pricing-action,.plan-button,.premium-btn,.bot-actions button,.bot-input button,.bot-bubble,.bot-close{pointer-events:auto!important;cursor:pointer!important;position:relative!important;z-index:9999!important}
.floating-bot,.bot-panel{z-index:2147483000!important}
.bot-body{max-height:340px!important;overflow-y:auto!important;overflow-x:hidden!important;box-sizing:border-box!important}
.bot-actions{display:grid!important;grid-template-columns:1fr 1fr!important;gap:12px!important;padding:14px 18px!important;background:#f8fafc!important;box-sizing:border-box!important}
.bot-input{display:flex!important;gap:10px!important;padding:12px 18px 18px!important;background:#fff!important;box-sizing:border-box!important}
.bot-input input{min-width:0!important;flex:1!important}
</style>
<script id="final-real-fix-js">
(function(){
  'use strict';
  function qs(s,root){return (root||document).querySelector(s)}
  function qsa(s,root){return Array.prototype.slice.call((root||document).querySelectorAll(s))}
  function norm(s){return String(s||'').toLowerCase()}
  function byText(sel, words){
    return qsa(sel).filter(function(el){var t=norm(el.innerText||el.textContent);return words.some(function(w){return t.indexOf(norm(w))>-1})})
  }
  function getDeviceId(){
    var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/);
    if(m) return decodeURIComponent(m[1]);
    var id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=31536000; SameSite=Lax';
    return id;
  }
  window.getOrCreateDeviceId = window.getOrCreateDeviceId || getDeviceId;
  function planKeyFromText(text){
    text=norm(text);
    if(text.indexOf('1.959')>-1||text.indexOf('1959')>-1||text.indexOf('nhà bán')>-1||text.indexOf('seller')>-1) return 'sellerpro';
    if(text.indexOf('859')>-1||text.indexOf('1 năm')>-1) return 'yearly';
    if(text.indexOf('559')>-1||text.indexOf('6 tháng')>-1) return 'halfyear';
    if(text.indexOf('359')>-1||text.indexOf('3 tháng')>-1) return 'quarterly';
    return 'monthly';
  }
  function showPricing(){
    var premium=document.getElementById('premium') || qs('.premium-pricing-pro') || qs('.pricing-grid') || qs('.premium-grid-pro');
    if(premium) premium.scrollIntoView({behavior:'smooth', block:'start'});
  }
  function openPlan(planKey){
    planKey=planKey||'monthly';
    window.currentPremiumPlanKey=planKey;
    if(typeof window.openPayment==='function'){ window.openPayment(planKey); return false; }
    if(typeof window.openPremiumCheckout==='function'){ window.openPremiumCheckout(planKey); return false; }
    if(typeof window.openPremiumPopup==='function'){ window.openPremiumPopup(); return false; }
    showPricing(); return false;
  }
  window.scrollToPricing = function(){showPricing(); return false};
  window.openPremium = function(){showPricing(); return false};
  function isPremiumActive(){
    var txt=norm((qs('#sidebarPremiumStatus')||{}).innerText||'');
    return txt.indexOf('premium')>-1 && txt.indexOf('dùng thử')<0 && txt.indexOf('hết')<0;
  }
  function openLocked(feature){
    if(typeof window.openLockedFeature==='function') return window.openLockedFeature(feature||'Tính năng PRO');
    showPricing(); return false;
  }
  var realOpenModule = window.openModule;
  var freeModules = {dashboard:1, facebook_center:1, page_center_total:1, post:1, fanpage_manager:1, group_suite:1, comment_manager:1, premium:1};
  window.openModule = function(moduleId){
    moduleId = moduleId || 'dashboard';
    var proOnly = !freeModules[moduleId];
    if(proOnly && !isPremiumActive()) return openLocked(moduleId);
    if(typeof realOpenModule==='function'){
      try { return realOpenModule(moduleId); } catch(e) { console.warn('old openModule error',e); }
    }
    var target=document.getElementById(moduleId);
    if(target){ target.scrollIntoView({behavior:'smooth',block:'start'}); history.replaceState(null,'','#'+moduleId); return false; }
    if(moduleId==='premium') { showPricing(); return false; }
    return false;
  };
  function botReply(t){
    var l=norm(t);
    if(l.indexOf('premium')>-1||l.indexOf('nâng cấp')>-1||l.indexOf('gói')>-1||l.indexOf('giá')>-1)
      return 'Dạ hiện hệ thống có các gói:<br><br>• 1 tháng: <b>159K</b><br>• 3 tháng: <b>359K</b><br>• 6 tháng: <b>559K</b><br>• 1 năm: <b>859K</b><br>• Nhà bán hàng chuyên nghiệp: <b>1.959K</b><br><br>Anh/chị chọn gói phù hợp, nhập số điện thoại và Gmail để nhận hỗ trợ kích hoạt nhanh ạ.';
    if(l.indexOf('thanh toán')>-1||l.indexOf('qr')>-1||l.indexOf('chuyển khoản')>-1)
      return 'Dạ sau khi chuyển khoản, anh/chị gửi giúp em:<br><br>• ID thiết bị: <b>'+getDeviceId()+'</b><br>• Ảnh thanh toán<br>• Gói đã đăng ký<br><br>Nếu sau 5 phút chưa kích hoạt, vui lòng liên hệ Zalo <b>036 338 2629</b>.';
    if(l.indexOf('kích hoạt')>-1||l.indexOf('duyệt')>-1)
      return 'Dạ tài khoản sẽ được mở sau khi web admin duyệt đúng ID thiết bị.<br><br>ID thiết bị của anh/chị là: <b>'+getDeviceId()+'</b>';
    if(l.indexOf('liên hệ')>-1||l.indexOf('zalo')>-1||l.indexOf('hỗ trợ')>-1)
      return '<b>Zalo hỗ trợ:</b> 036 338 2629<br><b>Gmail hỗ trợ:</b> support@gptmini.pro';
    return 'Dạ em đã nhận thông tin. Anh/chị cần hỗ trợ Nâng cấp Premium, Kích hoạt tài khoản, Thanh toán hay Báo lỗi hệ thống ạ?';
  }
  window.toggleFloatingBot = function(){
    var p=qs('#floatingBotPanel'); if(!p) return false;
    var open=getComputedStyle(p).display!=='none';
    p.style.display=open?'none':'block';
    return false;
  };
  window.closeFloatingBot = function(){var p=qs('#floatingBotPanel'); if(p) p.style.display='none'; return false};
  window.botQuick = function(text){
    var body=qs('#floatingBotBody'); if(!body) return false;
    body.insertAdjacentHTML('beforeend','<div class="bot-msg"><b>Bạn:</b> '+String(text).replace(/[&<>]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;'}[c]})+'</div>');
    var id='typing_'+Date.now();
    body.insertAdjacentHTML('beforeend','<div class="bot-msg ai" id="'+id+'"><b>Bot hỗ trợ:</b><br>Đang nhập...</div>');
    body.scrollTop=body.scrollHeight;
    setTimeout(function(){var x=document.getElementById(id); if(x) x.outerHTML='<div class="bot-msg ai"><b>Bot hỗ trợ:</b><br>'+botReply(text)+'</div>'; body.scrollTop=body.scrollHeight;},500);
    if(window.fetch){try{fetch('/support_send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:getDeviceId(),message:String(text),sender:'customer'})}).catch(function(){});}catch(e){}}
    return false;
  };
  window.sendBotInput = function(){var i=qs('#botInputText'); if(!i||!i.value.trim()) return false; var v=i.value.trim(); i.value=''; return window.botQuick(v)};
  function bindAll(){
    // ID thiết bị
    var id=getDeviceId(); var side=qs('#sidebarDeviceId'); if(side) side.textContent=id; var pay=qs('#payDeviceId'); if(pay) pay.value=id;
    // Remove visible technical status if still rendered
    qsa('h2').forEach(function(h){ if(norm(h.innerText).indexOf('trạng thái hệ thống')>-1){ var p=h.nextElementSibling; if(p) p.remove(); if(h.previousElementSibling && h.previousElementSibling.tagName==='HR') h.previousElementSibling.remove(); h.remove(); }});
    qsa('p,div').forEach(function(el){ var t=norm(el.innerText); if(t.indexOf('thiếu gemini api')>-1 || t.indexOf('pages_json')>-1){ el.style.display='none'; }});
    // Sidebar PRO badges: 4 chức năng đầu Facebook Center
    byText('.v2-nav-link',['Facebook Center','Page Center Tổng','Đăng bài Facebook','Quản lý Fanpage']).forEach(function(a){a.classList.add('menu-pro')});
    qsa('.v2-nav-link').forEach(function(a){
      a.addEventListener('click',function(e){
        var oc=a.getAttribute('onclick')||''; var m=oc.match(/openModule\('([^']+)'\)/); if(m){ e.preventDefault(); return window.openModule(m[1]); }
      },true);
    });
    // Pricing cards and buttons, including Xem chi tiết gói
    qsa('.price-card,.premium-plan').forEach(function(card){
      card.style.cursor='pointer';
      card.onclick=function(e){ if(e.target && e.target.closest('button')) return; e.preventDefault(); return openPlan(planKeyFromText(card.innerText)); };
      qsa('button',card).forEach(function(btn){
        btn.onclick=function(e){ e.preventDefault(); e.stopPropagation(); return openPlan(planKeyFromText(card.innerText)); };
      });
    });
    byText('button,a',['Xem chi tiết gói','Nâng cấp Premium']).forEach(function(btn){
      if(btn.closest('.bot-actions')) return;
      btn.onclick=function(e){ e.preventDefault(); e.stopPropagation(); var card=btn.closest('.price-card,.premium-plan'); return openPlan(planKeyFromText(card?card.innerText:btn.innerText)); };
    });
    // Chat buttons
    var bubble=qs('.bot-bubble'); if(bubble) bubble.onclick=function(e){e.preventDefault();return window.toggleFloatingBot()};
    var close=qs('.bot-close'); if(close) close.onclick=function(e){e.preventDefault();return window.closeFloatingBot()};
    var input=qs('#botInputText'); if(input) input.onkeydown=function(e){if(e.key==='Enter'){e.preventDefault();window.sendBotInput();}};
    var send=qs('.bot-input button'); if(send) send.onclick=function(e){e.preventDefault();return window.sendBotInput()};
    byText('.bot-actions button',['Nâng cấp Premium']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Nâng cấp Premium')}});
    byText('.bot-actions button',['Hướng dẫn thanh toán']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Hướng dẫn thanh toán')}});
    byText('.bot-actions button',['Kích hoạt tài khoản']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Kích hoạt tài khoản')}});
    byText('.bot-actions button',['Liên hệ hỗ trợ']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Liên hệ hỗ trợ')}});
    // Counter auto jump
    var c=qs('#liveMemberCount'); if(c && !window.__liveCounterFinal){
      window.__liveCounterFinal=true;
      var n=parseInt((c.textContent||localStorage.getItem('mkt_live_count')||'231').replace(/\D/g,''),10)||231;
      setInterval(function(){n += Math.floor(Math.random()*3)+1; if(n>999)n=231; c.textContent=n; localStorage.setItem('mkt_live_count',String(n));},1800);
    }
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',bindAll); else bindAll();
  setTimeout(bindAll,400); setTimeout(bindAll,1200); setTimeout(bindAll,2500);
})();
</script>



<!-- FINAL OVERRIDE: pricing detail buttons real working -->
<style id="mkt-price-detail-final-css">
#mktPlanOverlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.72);z-index:2147483646;align-items:center;justify-content:center;padding:18px;box-sizing:border-box}
#mktPlanBox{width:min(860px,96vw);max-height:92vh;overflow:auto;background:#fff;border-radius:28px;box-shadow:0 35px 100px rgba(15,23,42,.45);border:1px solid #ddd6fe;color:#111827}
#mktPlanHead{padding:22px 24px;background:linear-gradient(135deg,#2563eb,#7c3aed);color:#fff;display:flex;justify-content:space-between;gap:16px;align-items:flex-start}
#mktPlanHead h2{margin:0 0 6px;color:#fff;font-size:28px;line-height:1.15}
#mktPlanHead p{margin:0;color:#eef2ff;font-weight:800}
#mktPlanClose{border:0;border-radius:999px;background:rgba(255,255,255,.18);color:#fff;font-weight:900;padding:10px 16px;cursor:pointer}
#mktPlanBody{padding:24px;display:grid;grid-template-columns:1fr 1fr;gap:18px}
#mktPlanBody h3{margin:0 0 12px;color:#1e1b4b;font-size:20px}
.mktPlanPanel{border:1px solid #e5e7eb;border-radius:20px;padding:18px;background:#f8fafc}
.mktPlanPrice{font-size:34px;font-weight:1000;color:#2563eb;margin:8px 0 12px}.mktPlanList div{padding:9px 0;border-bottom:1px dashed #e5e7eb;font-weight:800;color:#334155}.mktPlanList div:before{content:'✓ ';color:#059669;font-weight:1000}
#mktPlanPay{display:block;width:100%;border:0;border-radius:18px;background:linear-gradient(135deg,#2563eb,#7c3aed);color:#fff;font-weight:1000;font-size:17px;padding:16px;cursor:pointer;margin-top:16px;box-shadow:0 14px 35px rgba(37,99,235,.25)}
#mktPlanNote{margin-top:12px;color:#64748b;font-weight:700;line-height:1.5}.mktPlanMini{background:#fff7ed;border:1px solid #fed7aa;color:#9a3412;border-radius:16px;padding:14px;font-weight:900;line-height:1.5}
@media(max-width:760px){ #mktPlanBody{grid-template-columns:1fr} #mktPlanHead h2{font-size:22px} }
</style>
<script id="mkt-price-detail-final-js">
(function(){
  'use strict';
  var plans={
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',desc:'Phù hợp người mới bắt đầu, shop nhỏ cần đăng bài, tạo content và quản lý Fanpage cơ bản.',benefits:['Đăng bài Facebook','Quản lý Fanpage','Quản lý Group','AI Comment','Tạo content cơ bản','Token Manager','Hỗ trợ kích hoạt theo ID thiết bị']},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',desc:'Tối ưu cho shop đang bán hàng cần dùng ổn định hơn và tiết kiệm hơn gói tháng.',benefits:['Toàn bộ gói 1 tháng','AI Messenger','CRM Kanban cơ bản','Kịch bản inbox','Lịch đăng nâng cao','Báo cáo cơ bản','Ưu tiên hỗ trợ']},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',desc:'Phù hợp shop cần CRM, chăm sóc khách và tối ưu quy trình bán hàng.',benefits:['Toàn bộ gói 3 tháng','CRM Pro','AI Sales Bot','Comment Manager','Auto Tag khách hàng','Quản lý khách hàng','Chuyển khách sang CRM']},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',desc:'Gói phổ biến nhất cho nhà bán hàng muốn dùng đầy đủ công cụ AI Marketing trong 1 năm.',benefits:['Toàn bộ gói 6 tháng','AI Marketing Director','AI Ads Chuyên Gia','Kho Content Premium','Automation Marketing','Export báo cáo','Ưu tiên xử lý']},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp, mở toàn bộ hệ thống sau khi admin duyệt.',benefits:['Toàn bộ tính năng Premium','AI Image Center','AI Video Center','AI Voice Studio','Dashboard Enterprise','Export PDF / Excel','Backup Database','Ưu tiên hỗ trợ VIP']}
  };
  plans.lifetime=plans.sellerpro;
  function norm(t){return String(t||'').toLowerCase();}
  function keyFromText(t){t=norm(t); if(t.indexOf('1.959')>-1||t.indexOf('1959')>-1||t.indexOf('nhà bán')>-1||t.indexOf('seller')>-1||t.indexOf('trọn đời')>-1) return 'sellerpro'; if(t.indexOf('859')>-1||t.indexOf('1 năm')>-1) return 'yearly'; if(t.indexOf('559')>-1||t.indexOf('6 tháng')>-1) return 'halfyear'; if(t.indexOf('359')>-1||t.indexOf('3 tháng')>-1) return 'quarterly'; return 'monthly';}
  function deviceId(){var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/); if(m) return decodeURIComponent(m[1]); var id=localStorage.getItem('mkt_device_id'); if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4); localStorage.setItem('mkt_device_id',id);} document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=31536000; SameSite=Lax'; return id;}
  function ensureBox(){
    var o=document.getElementById('mktPlanOverlay'); if(o) return o;
    o=document.createElement('div'); o.id='mktPlanOverlay';
    o.innerHTML='<div id="mktPlanBox"><div id="mktPlanHead"><div><h2 id="mktPlanTitle"></h2><p id="mktPlanDesc"></p></div><button id="mktPlanClose" type="button">Đóng</button></div><div id="mktPlanBody"><div class="mktPlanPanel"><h3>Chi tiết gói</h3><div class="mktPlanPrice" id="mktPlanPrice"></div><div id="mktPlanNote"></div><button id="mktPlanPay" type="button">Nâng cấp gói này</button></div><div class="mktPlanPanel"><h3>Quyền lợi nhận được</h3><div class="mktPlanList" id="mktPlanBenefits"></div><div class="mktPlanMini">Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</div></div></div></div>';
    document.body.appendChild(o);
    o.addEventListener('click',function(e){if(e.target===o) closePlan();});
    document.getElementById('mktPlanClose').onclick=closePlan;
    return o;
  }
  function closePlan(){var o=document.getElementById('mktPlanOverlay'); if(o)o.style.display='none'; return false;}
  function openPlan(k){
    k=k||'monthly'; var p=plans[k]||plans.monthly; var o=ensureBox();
    document.getElementById('mktPlanTitle').textContent=p.title;
    document.getElementById('mktPlanDesc').textContent=p.desc;
    document.getElementById('mktPlanPrice').textContent=p.price;
    document.getElementById('mktPlanNote').innerHTML='ID thiết bị: <b>'+deviceId()+'</b><br><span style="color:#64748b">Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</span>';
    document.getElementById('mktPlanBenefits').innerHTML=p.benefits.map(function(x){return '<div>'+x+'</div>';}).join('');
    document.getElementById('mktPlanPay').onclick=function(ev){ev.preventDefault(); ev.stopPropagation(); if(typeof window.openPayment==='function'){closePlan(); window.openPayment(k); return false;} if(typeof window.openPayment==='function'){closePlan(); window.openPayment(k); return false;} closePlan(); return false;};
    o.style.display='flex'; return false;
  }
  window.mktOpenPlanDetail=openPlan; window.mktClosePlanDetail=closePlan;
  function shouldOpen(el){var txt=norm(el.innerText||el.textContent); if(txt.indexOf('xem chi tiết gói')>-1) return true; if(txt.indexOf('mở khóa gói')>-1) return true; if(txt.indexOf('đăng ký')>-1 && (txt.indexOf('tháng')>-1||txt.indexOf('gói')>-1)) return true; if(txt.indexOf('chọn gói phổ biến')>-1) return true; return false;}
  document.addEventListener('click',function(e){
    var target=e.target.closest('button,a,.premium-plan,.price-card,.v4-plan'); if(!target) return;
    if(target.closest('.bot-actions,.bot-input,#floatingBotPanel,#mktPlanOverlay')) return;
    var btn=e.target.closest('button,a'); var card=target.closest('.premium-plan,.price-card,.v4-plan');
    if((btn && shouldOpen(btn)) || (card && (target===card || shouldOpen(target)))){
      e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation();
      return openPlan(keyFromText((card||target).innerText||target.textContent));
    }
  },true);
  function bind(){document.querySelectorAll('.premium-plan button,.price-card button,.v4-plan button').forEach(function(b){if(shouldOpen(b)){b.type='button'; b.style.pointerEvents='auto';}});}
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',bind); else bind(); setTimeout(bind,500); setTimeout(bind,1500);
})();
</script>



<!-- FINAL PAYMENT UX OVERRIDE: no browser alert, professional checkout -->
<style id="mkt-payment-professional-css">
#mktPlanNote .pay-line,#mktPlanNote .pay-copy{display:block;margin:6px 0;color:#334155;font-weight:800}
#mktPlanNote .pay-copy{background:#f1f5f9;border:1px dashed #cbd5e1;border-radius:12px;padding:10px;color:#1e293b}
.payment-alert{background:#f8fafc!important;border:1px solid #dbeafe!important;color:#1e293b!important}
.bank-info{line-height:1.75!important}
</style>
<script id="mkt-payment-professional-js">
(function(){
  'use strict';
  function q(s){return document.querySelector(s)}
  function money(v){try{return Number(v||0).toLocaleString('vi-VN')+' VNĐ'}catch(e){return (v||'')+' VNĐ'}}
  function getDevice(){
    var id=localStorage.getItem('mkt_device_id');
    if(!id){
      var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/);
      id=m?decodeURIComponent(m[1]):'';
    }
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);}
    localStorage.setItem('mkt_device_id',id);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=157680000; SameSite=Lax';
    var side=q('#sidebarDeviceId'); if(side) side.textContent=id;
    var pay=q('#payDeviceId'); if(pay) pay.value=id;
    return id;
  }
  var plans=window.premiumPlans || {
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',package:'1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',package:'3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',package:'6THANG',desc:'Mở CRM và Sales Bot.'},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',package:'1NAM',desc:'Gói phổ biến nhất.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',package:'NHABANHANGCHUYENNGHIEP',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'},
    lifetime:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',package:'NHABANHANGCHUYENNGHIEP',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  plans.sellerpro=plans.sellerpro||plans.lifetime; plans.lifetime=plans.lifetime||plans.sellerpro;
  function plan(k){return plans[k]||plans.monthly;}
  function buildContent(p){
    var device=getDevice();
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    return device+' | '+phone+' | '+email+' | '+(p.package||p.title||'PREMIUM').toUpperCase();
  }
  window.refreshPaymentContent=function(){
    var p=plan(window.currentPremiumPlanKey||'monthly');
    var content=buildContent(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount||0)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
  };
  window.openPayment=function(planKey){
    var p=plan(planKey); window.currentPremiumPlanKey=planKey||'monthly';
    var modal=q('#paymentModal');
    if(!modal){return false;}
    var title=q('#payPlanTitle'); if(title) title.textContent=(p.title||'Gói Premium').replace(/^🚀 |^⭐ |^💎 |^🔥 |^👑 /,'');
    var desc=q('#payPlanDesc'); if(desc) desc.textContent='Nhập số điện thoại và Gmail để đăng ký kích hoạt. Nội dung chuyển khoản sẽ tự kèm ID thiết bị và gói đã chọn.';
    var price=q('#payPlanPrice'); if(price) price.textContent=money(p.amount);
    var device=q('#payDeviceId'); if(device) device.value=getDevice();
    var benefits=q('#payBenefits'); if(benefits && p.benefits){benefits.innerHTML=p.benefits.map(function(x){return '<div>'+x+'</div>';}).join('');}
    var locked=q('#payLocked'); if(locked){locked.innerHTML=''; var h=locked.previousElementSibling; if(h) h.style.display='none';}
    var alertBox=q('.payment-alert'); if(alertBox){alertBox.innerHTML='Sau khi chuyển khoản, bấm <b>Đã thanh toán</b> để gửi yêu cầu về web admin. Nội dung chuyển khoản gồm ID thiết bị, số điện thoại, Gmail và gói đã chọn. Zalo hỗ trợ: <b>036 338 2629</b>.';}
    var payBtn=q('.payment-actions .light'); if(payBtn){payBtn.textContent='Đã thanh toán'; payBtn.removeAttribute('href'); payBtn.onclick=function(e){e.preventDefault(); if(typeof window.submitPremiumRequest==='function') return window.submitPremiumRequest(); return false;};}
    window.refreshPaymentContent();
    modal.style.display='flex';
    var ph=q('#payPhone'); if(ph) setTimeout(function(){ph.focus();},150);
    return false;
  };
  window.mktOpenPlanDetail = (function(old){
    return function(k){
      if(typeof old==='function') old(k);
      setTimeout(function(){
        var p=plan(k||window.currentPremiumPlanKey||'monthly');
        var note=q('#mktPlanNote');
        if(note){
          var content=getDevice()+' | SỐ ĐIỆN THOẠI | GMAIL | '+(p.package||p.title||'PREMIUM').toUpperCase();
          note.innerHTML='<span class="pay-line">ID thiết bị: <b>'+getDevice()+'</b></span><span class="pay-line">Gói đang chọn: <b>'+p.title+'</b></span><span class="pay-copy">Nội dung thanh toán: '+content+'</span>';
        }
        var mini=q('.mktPlanMini'); if(mini) mini.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
        var btn=q('#mktPlanPay'); if(btn) btn.onclick=function(e){e.preventDefault(); if(q('#mktPlanOverlay')) q('#mktPlanOverlay').style.display='none'; return window.openPayment(k||'monthly');};
      },60);
      return false;
    };
  })(window.mktOpenPlanDetail);
  document.addEventListener('click',function(e){
    var t=e.target.closest('#mktPlanPay');
    if(t){e.preventDefault(); e.stopPropagation(); return window.openPayment(window.currentPremiumPlanKey||'monthly');}
  },true);
  getDevice();
})();
</script>



<!-- FINAL CLEAN PAYMENT PATCH V74 -->
<style id="mkt-final-clean-payment-css">
  #mktPlanOverlay #mktPlanNote{font-size:16px;line-height:1.55;color:#475569;font-weight:700}
  #mktPlanOverlay .mktPlanMini{background:#f8fafc!important;border:1px solid #dbeafe!important;color:#334155!important}
  #mktPlanOverlay .pay-copy,
  #paymentModal .pay-copy{display:block;margin-top:8px;background:#f1f5f9;border:1px dashed #cbd5e1;border-radius:12px;padding:10px;color:#1e293b;font-weight:800;word-break:break-word}
  #paymentModal .payment-alert{display:none!important}
  #paymentModal .payment-actions{display:grid!important;grid-template-columns:1fr 1fr;gap:10px}
  #paymentModal .payment-actions a{display:flex;align-items:center;justify-content:center;text-decoration:none}
  #paymentNotice{margin:10px 0;padding:12px 14px;border:1px solid #bbf7d0;background:#f0fdf4;color:#166534;border-radius:12px;font-weight:800;display:none}
</style>
<script id="mkt-final-clean-payment-js">
(function(){
  'use strict';
  var PLAN_MAP={
    monthly:{title:'Gói 1 tháng',amount:'159000',price:'159.000đ',code:'GOI1THANG'},
    quarterly:{title:'Gói 3 tháng',amount:'359000',price:'359.000đ',code:'GOI3THANG'},
    halfyear:{title:'Gói 6 tháng',amount:'559000',price:'559.000đ',code:'GOI6THANG'},
    yearly:{title:'Gói 1 năm',amount:'859000',price:'859.000đ',code:'GOI1NAM'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',amount:'1959000',price:'1.959.000đ',code:'NHABANHANGPRO'},
    lifetime:{title:'Gói nhà bán hàng chuyên nghiệp',amount:'1959000',price:'1.959.000đ',code:'NHABANHANGPRO'}
  };
  function q(s,root){return (root||document).querySelector(s)}
  function qa(s,root){return Array.prototype.slice.call((root||document).querySelectorAll(s))}
  function cleanText(){
    var bad=['Gói sẽ mở sau khi admin duyệt thanh toán.','Gói sẽ mở sau khi admin duyệt thanh toán','Sau khi thanh toán, web admin sẽ duyệt theo ID thiết bị. Duyệt xong khách mới sử dụng được tính năng PRO.','Sau khi thanh toán, web admin sẽ duyệt theo ID thiết bị.','Duyệt xong khách mới sử dụng được tính năng PRO.'];
    qa('body *').forEach(function(el){
      if(el.children.length>0) return;
      var t=el.textContent||''; var nt=t;
      bad.forEach(function(b){nt=nt.split(b).join('')});
      if(nt!==t) el.textContent=nt.trim();
    });
  }
  function deviceId(){
    var id=localStorage.getItem('mkt_device_id');
    if(!id){var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/); id=m?decodeURIComponent(m[1]):'';}
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-6);}
    localStorage.setItem('mkt_device_id',id);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=157680000; SameSite=Lax';
    var side=q('#sidebarDeviceId'); if(side) side.textContent=id;
    var pay=q('#payDeviceId'); if(pay) pay.value=id;
    return id;
  }
  function keyFromText(txt){
    txt=String(txt||'').toLowerCase();
    if(txt.indexOf('1.959')>-1||txt.indexOf('1959')>-1||txt.indexOf('nhà bán')>-1||txt.indexOf('seller')>-1) return 'sellerpro';
    if(txt.indexOf('859')>-1||txt.indexOf('1 năm')>-1) return 'yearly';
    if(txt.indexOf('559')>-1||txt.indexOf('6 tháng')>-1) return 'halfyear';
    if(txt.indexOf('359')>-1||txt.indexOf('3 tháng')>-1) return 'quarterly';
    return 'monthly';
  }
  function plan(k){return PLAN_MAP[k]||PLAN_MAP.monthly}
  function contentFor(p){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'SDT';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'GMAIL';
    return deviceId();
  }
  window.showPaymentNotice=function(msg){
    var n=q('#paymentNotice');
    if(!n){
      var btn=q('#paymentModal .primary');
      n=document.createElement('div'); n.id='paymentNotice';
      if(btn && btn.parentNode) btn.parentNode.insertBefore(n, btn.nextSibling);
    }
    n.textContent=msg||'Đã gửi yêu cầu về web admin.'; n.style.display='block';
  };
  window.refreshPaymentContent=function(){
    var p=plan(window.currentPremiumPlanKey||'monthly');
    var content=contentFor(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    var notice=q('#paymentNotice'); if(notice) notice.style.display='none';
  };
  window.openPayment=function(k){
    var p=plan(k); window.currentPremiumPlanKey=k||'monthly';
    var modal=q('#paymentModal'); if(!modal) return false;
    var title=q('#payPlanTitle'); if(title) title.textContent=p.title;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
    var price=q('#payPlanPrice'); if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
    var device=q('#payDeviceId'); if(device) device.value=deviceId();
    var locked=q('#payLocked'); if(locked) locked.innerHTML='';
    var h=locked && locked.previousElementSibling; if(h) h.style.display='none';
    var alertBox=q('#paymentModal .payment-alert'); if(alertBox) alertBox.style.display='none';
    var z=q('#paymentModal .payment-actions a:not(.light)'); if(z){z.textContent='Zalo hỗ trợ: 036 338 2629'; z.href='https://zalo.me/0363382629'; z.target='_blank';}
    var done=q('#paymentModal .payment-actions .light'); if(done){done.textContent='Gửi yêu cầu admin duyệt'; done.removeAttribute('href'); done.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    var btn=q('#paymentModal .primary'); if(btn) btn.textContent='Gửi yêu cầu kích hoạt';
    window.refreshPaymentContent();
    modal.style.display='flex';
    setTimeout(function(){var ph=q('#payPhone'); if(ph) ph.focus();},150);
    cleanText(); return false;
  };
  var oldSubmit=window.submitPremiumRequest;
  window.submitPremiumRequest=function(){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    if(!phone || !email){ window.showPaymentNotice('Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.'); return false; }
    if(typeof oldSubmit==='function'){
      try{
        var r=oldSubmit();
        setTimeout(function(){window.showPaymentNotice('Đã gửi yêu cầu về web admin. Nội dung chuyển khoản đã kèm ID thiết bị và gói đã chọn.');},250);
        return r;
      }catch(e){}
    }
    window.showPaymentNotice('Đã gửi yêu cầu về web admin. Nội dung chuyển khoản đã kèm ID thiết bị và gói đã chọn.');
    return false;
  };
  function patchPlanOverlay(){
    var note=q('#mktPlanNote');
    if(note){
      var p=plan(window.currentPremiumPlanKey||keyFromText((q('#mktPlanOverlay')||document.body).innerText));
      note.innerHTML='ID thiết bị: <b>'+deviceId()+'</b><br><span style="color:#64748b">Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</span>';
    }
    var mini=q('.mktPlanMini'); if(mini) mini.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
    var pay=q('#mktPlanPay'); if(pay){pay.onclick=function(e){e.preventDefault(); e.stopPropagation(); var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(window.currentPremiumPlanKey||keyFromText((ov||document.body).innerText));};}
    cleanText();
  }
  var oldPlan=window.mktOpenPlanDetail;
  window.mktOpenPlanDetail=function(k){
    window.currentPremiumPlanKey=k||window.currentPremiumPlanKey||'monthly';
    var r=false;
    if(typeof oldPlan==='function'){r=oldPlan(k);} 
    setTimeout(patchPlanOverlay,50); setTimeout(patchPlanOverlay,200);
    return r;
  };
  document.addEventListener('click',function(e){
    var pay=e.target.closest && e.target.closest('#mktPlanPay');
    if(pay){e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation(); var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(window.currentPremiumPlanKey||keyFromText((ov||document.body).innerText));}
    var detail=e.target.closest && e.target.closest('button,a');
    if(detail && /xem chi tiết gói/i.test(detail.innerText||detail.textContent||'')){setTimeout(patchPlanOverlay,120);}
  },true);
  function init(){deviceId(); cleanText(); patchPlanOverlay();}
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',init); else init();
  setTimeout(init,500); setTimeout(init,1500); setInterval(cleanText,2500);
})();
</script>
<!-- CTV FINAL VISIBILITY PATCH -->
<style id="ctv-final-visibility-patch">
  #affiliateAdminSection{display:none!important}
  a[href="#affiliate_admin"],[onclick*="affiliate_admin"]{display:none!important}
  [data-aff-menu="1"]{display:flex!important;visibility:visible!important;opacity:1!important;pointer-events:auto!important}
  #mobileCtvQuickBtn{display:flex!important;visibility:visible!important;opacity:1!important;pointer-events:auto!important}
  #affiliateCenterSection.ctv-active{display:block!important;visibility:visible!important;opacity:1!important;min-height:420px!important}
  .ctv-mobile-card{display:none}
  @media(max-width:760px){
    #affiliateCenterSection{margin:12px!important;border-radius:18px!important;padding:14px!important}
    .aff-stats{grid-template-columns:1fr 1fr!important}
    .aff-link-box{flex-direction:column!important;align-items:stretch!important}
  }
</style>
<script id="ctv-final-visibility-js">
(function(){
  function qs(s){return document.querySelector(s)}
  function qsa(s){return Array.prototype.slice.call(document.querySelectorAll(s))}
  function money(n){return Number(n||0).toLocaleString('vi-VN')+'đ'}
  function toast(msg){
    var t=document.getElementById('affToast');
    if(!t){t=document.createElement('div');t.id='affToast';t.className='aff-toast';document.body.appendChild(t)}
    t.textContent=msg;t.style.display='block';setTimeout(function(){t.style.display='none'},2600)
  }
  function hideMain(){
    qsa('main > section, main > div.module, main > div.card, main > div.app-module-section').forEach(function(el){
      if(el.id!=='affiliateCenterSection') el.style.display='none';
    });
  }
  function ensureMenu(){
    qsa('a[href="#affiliate_admin"],[onclick*="affiliate_admin"],#affiliateAdminSection').forEach(function(el){el.style.display='none'});
    var nav=qs('.nav');
    if(nav && !qs('[data-aff-menu="1"]')){
      var a=document.createElement('a');
      a.setAttribute('data-aff-menu','1');
      a.className='v2-nav-link';
      a.href='#affiliate_center';
      a.innerHTML='<span class="v2-nav-ico"></span><span class="v2-nav-text">CTV Hoa Hồng</span><span class="v2-nav-tag" style="background:#dcfce7;color:#166534;box-shadow:0 0 14px rgba(34,197,94,.55)">CTV</span>';
      a.onclick=function(e){e.preventDefault(); return window.openAffiliateCenter();};
      nav.appendChild(a);
    }
    qsa('[data-aff-menu="1"],a[href="#affiliate_center"]').forEach(function(a){
      a.style.display='flex'; a.onclick=function(e){e.preventDefault(); return window.openAffiliateCenter();};
    });
    var btn=document.getElementById('mobileCtvQuickBtn');
    if(!btn){
      btn=document.createElement('button');btn.id='mobileCtvQuickBtn';btn.type='button';btn.innerHTML='<span class="dot"></span>CTV Hoa Hồng';document.body.appendChild(btn);
    }
    btn.onclick=function(e){e.preventDefault(); return window.openAffiliateCenter();};
  }
  window.openAffiliateCenter=function(){
    var sec=document.getElementById('affiliateCenterSection');
    if(!sec){toast('Chưa tìm thấy giao diện CTV trong file.');return false;}
    hideMain();
    sec.style.display='block';sec.classList.add('ctv-active');
    if(history && history.replaceState) history.replaceState(null,'','#affiliate_center');
    if(window.affiliateLoadMe) window.affiliateLoadMe();
    setTimeout(function(){sec.scrollIntoView({behavior:'smooth',block:'start'})},60);
    return false;
  };
  var oldOpen=window.openModule;
  window.openModule=function(name){
    if(name==='affiliate_center') return window.openAffiliateCenter();
    if(name==='affiliate_admin') return false;
    if(typeof oldOpen==='function') return oldOpen.apply(this,arguments);
    return true;
  };
  window.affiliateLoadMe=window.affiliateLoadMe || function(){
    fetch('/api/affiliate/me').then(function(r){return r.json()}).then(function(d){
      if(qs('#affName')) qs('#affName').value=d.full_name||'';
      if(qs('#affPhone')) qs('#affPhone').value=d.phone||'';
      if(qs('#affEmail')) qs('#affEmail').value=d.email||'';
      if(qs('#affRefLink')) qs('#affRefLink').value=d.ref_link||'';
      if(qs('#affCode')) qs('#affCode').textContent=d.affiliate_code||'---';
      if(qs('#affLevel')) qs('#affLevel').textContent=d.level_name||'CTV thường';
      if(qs('#affPercent')) qs('#affPercent').textContent=d.commission_percent||20;
      var s=d.summary||{};
      if(qs('#affClicks')) qs('#affClicks').textContent=s.clicks||0;
      if(qs('#affCustomers')) qs('#affCustomers').textContent=s.referrals||0;
      if(qs('#affPremium')) qs('#affPremium').textContent=s.premium||0;
      if(qs('#affCommission')) qs('#affCommission').textContent=money(s.commission||0);
      if(qs('#affBalance')) qs('#affBalance').textContent=money(s.balance||0);
    }).catch(function(){toast('Không tải được dữ liệu CTV')});
  };
  function init(){
    ensureMenu();
    var ref=(new URL(location.href)).searchParams.get('ref');
    if(ref){try{document.cookie='affiliate_code='+encodeURIComponent(ref.toUpperCase())+';path=/;max-age='+(60*60*24*90)}catch(e){}}
    if(location.hash==='#affiliate_center') setTimeout(window.openAffiliateCenter,350);
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',init); else init();
  setTimeout(init,800); setTimeout(init,1800);
})();
</script>
<!-- /CTV FINAL VISIBILITY PATCH -->


  function selectedKey(){
    return normKey(window.mktSelectedPlanKey || window.currentPremiumPlanKey || localStorage.getItem('mkt_selected_plan_key')) || 'monthly';
  }

  window.refreshPaymentContent=function(){
    var p=plan(selectedKey());
    var content=payContent(p);

    var pc=q('#payContent'); 
    if(pc) pc.textContent=content;

    var qr=q('#payQr'); 
    if(qr){
      qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+
        encodeURIComponent(p.amount)+
        '&addInfo='+encodeURIComponent(content)+
        '&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    }

    var price=q('#payPlanPrice'); 
    if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
  };

  window.openPayment=function(k){
    k=setKey(k || selectedKey());
    var p=plan(k);

    var modal=q('#paymentModal'); 
    if(!modal) return false;

    var title=q('#payPlanTitle'); 
    if(title) title.textContent=p.title;

    var desc=q('#payPlanDesc'); 
    if(desc) desc.textContent=p.desc || 'Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp.';

    var price=q('#payPlanPrice'); 
    if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';

    var device=q('#payDeviceId'); 
    if(device) device.value=getDevice();

    window.refreshPaymentContent();
    modal.style.display='flex';
    return false;
  };

  document.addEventListener('click',function(e){
    var b=e.target.closest && e.target.closest('button,a');
    if(!b) return;

    var text=String(b.innerText || b.textContent || '');
    var k=keyFromNode(b);

    if(k) setKey(k);

    if(/nâng cấp|thanh toán|kích hoạt|mở khóa/i.test(text)){
      if(k){
        e.preventDefault();
        e.stopPropagation();
        if(e.stopImmediatePropagation) e.stopImmediatePropagation();
        return window.openPayment(k);
      }
    }
  },true);

  document.addEventListener('input',function(e){
    if(e.target && /payPhone|payEmail/.test(e.target.id || '')){
      window.refreshPaymentContent();
    }
  },true);

  qa('[data-plan]').forEach(function(btn){
    var k=normKey(btn.dataset.plan);
    if(k){
      btn.onclick=function(e){
        e.preventDefault();
        e.stopPropagation();
        return window.openPayment(k);
      };
    }
  });

  getDevice();
})();
</script>

  function fillPayment(k){
    k=setKey(k);
    var p=plan(k);
    var title=q('#payPlanTitle'); if(title) title.textContent=p.title;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent=p.desc;
    var price=q('#payPlanPrice'); if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
    var device=q('#payDeviceId'); if(device) device.value=getDevice();
    var content=payContent(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr');
    if(qr){
      qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    }
    var locked=q('#payLocked'); if(locked) locked.innerHTML='';
    var notice=q('#paymentNotice'); if(notice) notice.style.display='none';
    return p;
  }

  window.refreshPaymentContent=function(){fillPayment(window.currentPremiumPlanKey||'monthly');};
  window.openPayment=function(planKey){
    var k=setKey(planKey||window.mktSelectedPlanKey||'monthly');
    fillPayment(k);
    var modal=q('#paymentModal');
    if(modal) modal.style.display='flex';
    var z=q('#paymentModal .payment-actions a:not(.light)');
    if(z){z.textContent='Zalo hỗ trợ: 036 338 2629'; z.href='https://zalo.me/0363382629'; z.target='_blank';}
    var done=q('#paymentModal .payment-actions .light');
    if(done){done.textContent='Gửi yêu cầu admin duyệt'; done.removeAttribute('href'); done.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    var primary=q('#paymentModal .primary');
    if(primary){primary.textContent='Gửi yêu cầu kích hoạt'; primary.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    return false;
  };

  window.submitPremiumRequest=function(){
    var k=setKey(window.mktSelectedPlanKey||window.currentPremiumPlanKey||'monthly');
    var p=fillPayment(k);
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    if(!phone||!email){
      if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.');
      else alert('Vui lòng nhập số điện thoại và Gmail.');
      return false;
    }
    fetch('/premium_request',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:getDevice(),phone:phone,email:email,package_key:k})})
      .then(function(r){return r.json();})
      .then(function(data){
        var pc=q('#payContent'); if(pc) pc.textContent=(data&&data.payment_note)||payContent(p);
        var msg=(data&&data.ok)?('Đã gửi yêu cầu '+p.title+' về web admin. Số tiền: '+p.price+'.'):((data&&data.message)||'Chưa gửi được yêu cầu, vui lòng thử lại.');
        if(typeof window.showPaymentNotice==='function') window.showPaymentNotice(msg); else alert(msg);
      })
      .catch(function(){
        if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Kết nối chậm, vui lòng thử lại hoặc gửi Zalo hỗ trợ.');
      });
    return false;
  };

  var oldDetail=window.mktOpenPlanDetail;
  window.mktOpenPlanDetail=function(k){
    k=setKey(k||window.mktSelectedPlanKey||'monthly');
    var r=false;
    if(typeof oldDetail==='function'){
      try{r=oldDetail(k);}catch(e){}
    }
    setTimeout(function(){
      var p=plan(k);
      var note=q('#mktPlanNote');
      if(note){note.innerHTML='ID thiết bị: <b>'+getDevice()+'</b><br>Gói đang chọn: <b>'+p.title+'</b><br><span style="color:#64748b">Số tiền: <b>'+p.price+'</b>. Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</span>';}
      var mini=q('.mktPlanMini'); if(mini) mini.textContent='Gói đang chọn: '+p.title+' - '+p.price;
      var pay=q('#mktPlanPay'); if(pay){pay.onclick=function(e){e.preventDefault(); e.stopPropagation(); var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(k);};}
    },80);
    return r;
  };

  function bindPlanCards(){
    qa('.premium-plan,.v4-plan,.price-card,.plan-card,.pricing-card').forEach(function(card){
      var k=keyFromText(card.innerText||card.textContent||'');
      if(k){
        card.setAttribute('data-plan',k);
        qa('button,a',card).forEach(function(btn){
          if(/nâng cấp|mở khóa|đăng ký|thanh toán|kích hoạt|xem chi tiết/i.test(btn.innerText||btn.textContent||'')) btn.setAttribute('data-plan',k);
        });
      }
    });
  }
  document.addEventListener('click',function(e){
    var btn=e.target.closest&&e.target.closest('button,a');
    if(!btn) return;
    var txt=btn.innerText||btn.textContent||'';
    if(!/nâng cấp|mở khóa|đăng ký|thanh toán|kích hoạt|xem chi tiết|gửi yêu cầu/i.test(txt) && btn.id!=='mktPlanPay') return;
    var k=keyFromNode(btn)||window.mktSelectedPlanKey||window.currentPremiumPlanKey||'monthly';
    setKey(k);
    if(btn.id==='mktPlanPay'){
      e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation();
      var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none';
      return window.openPayment(k);
    }
    if(/nâng cấp|mở khóa|đăng ký|thanh toán|kích hoạt/i.test(txt)) fillPayment(k);
  },true);
  document.addEventListener('input',function(e){if(e.target&&(/payPhone|payEmail/.test(e.target.id||''))) fillPayment(window.currentPremiumPlanKey||'monthly');},true);
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',bindPlanCards); else bindPlanCards();
  setTimeout(bindPlanCards,500); setTimeout(bindPlanCards,1500);
})();
</script>




<script id="mkt-premium-benefit-modal-v1">
(function(){
  'use strict';
  var PLAN_BENEFITS={
    monthly:{label:'CƠ BẢN',title:'Gói 1 tháng',price:'159.000đ',desc:'Phù hợp người mới bắt đầu muốn trải nghiệm Premium ngắn hạn và dùng các công cụ cốt lõi.',features:['AI Content Studio','Tạo content bán hàng nhanh','Đăng Fanpage cơ bản','Quản lý Fanpage','Quản lý Group','AI Comment','Token Manager cơ bản'],fit:['Chủ shop mới','Cá nhân kinh doanh online','Người mới làm nội dung','Người cần test hệ thống trước'],value:['Tiết kiệm thời gian viết bài mỗi ngày','Có nền tảng quản lý Fanpage/Group','Tạo nội dung đều hơn','Chi phí thấp để bắt đầu']},
    quarterly:{label:'TIẾT KIỆM',title:'Gói 3 tháng',price:'359.000đ',desc:'Tối ưu hơn gói tháng, phù hợp shop nhỏ cần dùng ổn định và có thêm công cụ chăm sóc khách.',features:['Toàn bộ gói 1 tháng','AI Messenger','CRM Kanban cơ bản','Kịch bản inbox','Lịch đăng nâng cao','Báo cáo cơ bản','Ưu tiên hỗ trợ'],fit:['Shop nhỏ đang bán hàng','Người cần chăm sóc khách đều','Người chạy nội dung thường xuyên','CTV bán dịch vụ'],value:['Quy trình bán hàng rõ hơn','Giảm thời gian trả lời khách','Có kịch bản inbox sẵn','Tiết kiệm hơn so với mua từng tháng']},
    halfyear:{label:'TĂNG TRƯỞNG',title:'Gói 6 tháng',price:'559.000đ',desc:'Phù hợp shop cần CRM, Sales Bot và quy trình gom khách để chăm sóc lâu dài.',features:['Toàn bộ gói 3 tháng','CRM Pro','AI Sales Bot','Comment Manager','Auto Tag khách hàng','Quản lý khách hàng','Chuyển khách sang CRM'],fit:['Shop có nhiều inbox/comment','Người cần gom khách về CRM','Đội sale nhỏ','Người muốn tối ưu chăm sóc khách'],value:['Không bỏ sót khách tiềm năng','Phân loại khách rõ ràng hơn','Chăm sóc khách có hệ thống','Tăng hiệu quả tư vấn và chốt đơn']},
    yearly:{label:'PHỔ BIẾN NHẤT',title:'Gói 1 năm',price:'859.000đ',desc:'Gói phổ biến nhất cho nhà bán hàng muốn dùng đầy đủ công cụ AI Marketing trong 1 năm.',features:['Toàn bộ gói 6 tháng','AI Marketing Director','AI Ads Chuyên Gia','Kho Content Premium','Automation Marketing','Export báo cáo','Ưu tiên xử lý'],fit:['Shop bán hàng lâu dài','Người chạy quảng cáo','Agency nhỏ / CTV','Người cần hệ thống marketing đầy đủ'],value:['Chi phí thấp theo ngày','Có trợ lý AI định hướng marketing','Tối ưu nội dung và quảng cáo','Tiết kiệm chi phí thuê ngoài']},
    sellerpro:{label:'VIP ENTERPRISE',title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp, mở toàn bộ hệ thống hiện tại và các tính năng nâng cấp về sau.',features:['Toàn bộ tính năng Premium','AI Image Center','AI Video Center','AI Voice Studio','Dashboard Enterprise','Export PDF / Excel','Backup Database','Ưu tiên hỗ trợ VIP'],fit:['Nhà bán hàng chuyên nghiệp','Agency / đội marketing nhỏ','Người cần toàn bộ AI hiện tại và tương lai','Người muốn dùng lâu dài'],value:['Mở tối đa sức mạnh hệ thống','Giảm chi phí dùng nhiều công cụ rời','Quản lý dữ liệu chuyên nghiệp hơn','Được ưu tiên hỗ trợ và cập nhật']}
  };
  function q(s,r){return (r||document).querySelector(s)}
  function fillList(id,arr){var el=q('#'+id); if(!el) return; el.innerHTML=(arr||[]).map(function(x){return '<li>'+x+'</li>'}).join('')}
  function closeBenefit(){var m=q('#planBenefitModal'); if(m){m.classList.remove('show');m.setAttribute('aria-hidden','true')}}
  window.openPlanBenefit=function(key){
    key=PLAN_BENEFITS[key]?key:'monthly';
    var p=PLAN_BENEFITS[key];
    window.currentPremiumPlanKey=key; window.mktSelectedPlanKey=key;
    var m=q('#planBenefitModal'); if(!m) return false;
    q('#benefitLabel').textContent=p.label;
    q('#benefitTitle').textContent=p.title;
    q('#benefitPrice').textContent=p.price;
    q('#benefitDesc').textContent=p.desc;
    fillList('benefitFeatures',p.features); fillList('benefitFit',p.fit); fillList('benefitValue',p.value);
    var pay=q('#benefitUpgradeBtn'); if(pay){pay.onclick=function(e){e.preventDefault(); closeBenefit(); if(typeof window.openPayment==='function') return window.openPayment(key); return false;};}
    m.classList.add('show'); m.setAttribute('aria-hidden','false');
    return false;
  };
  document.addEventListener('click',function(e){
    var benefit=e.target.closest&&e.target.closest('[data-benefit]');
    if(benefit){e.preventDefault();e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation(); return window.openPlanBenefit(benefit.getAttribute('data-benefit'));}
    if(e.target&&e.target.id==='planBenefitClose'){e.preventDefault(); return closeBenefit();}
    if(e.target&&e.target.id==='benefitCancelBtn'){e.preventDefault(); return closeBenefit();}
    if(e.target&&e.target.id==='planBenefitModal'){return closeBenefit();}
  },true);
  document.addEventListener('keydown',function(e){if(e.key==='Escape') closeBenefit();});
})();
</script>


<!-- V3 ULTIMATE PLAN PRICE LOCK: giữ nguyên giao diện, chỉ sửa sai gói/sai giá khi bấm nâng cấp -->
<script id="mkt-ultimate-plan-price-lock-v3">
(function(){
  'use strict';
  var PLAN={
    monthly:{title:'Gói 1 tháng',amount:'159000',price:'159.000đ',code:'GOI1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',amount:'359000',price:'359.000đ',code:'GOI3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',amount:'559000',price:'559.000đ',code:'GOI6THANG',desc:'Mở CRM và Sales Bot.'},
    yearly:{title:'Gói 1 năm',amount:'859000',price:'859.000đ',code:'GOI1NAM',desc:'Gói phổ biến nhất.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',amount:'1959000',price:'1.959.000đ',code:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'},
    lifetime:{title:'Gói nhà bán hàng chuyên nghiệp',amount:'1959000',price:'1.959.000đ',code:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  function q(s,r){return (r||document).querySelector(s)}
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s))}
  function norm(t){return String(t||'').toLowerCase().replace(/\s+/g,' ')}
  function keyFromText(t){
    t=norm(t);
    if(t.indexOf('1.959')>-1||t.indexOf('1959')>-1||t.indexOf('nhà bán')>-1||t.indexOf('seller')>-1||t.indexOf('trọn đời')>-1||t.indexOf('chuyên nghiệp')>-1) return 'sellerpro';
    if(t.indexOf('859')>-1||t.indexOf('1 năm')>-1||t.indexOf('12 tháng')>-1) return 'yearly';
    if(t.indexOf('559')>-1||t.indexOf('6 tháng')>-1) return 'halfyear';
    if(t.indexOf('359')>-1||t.indexOf('3 tháng')>-1) return 'quarterly';
    if(t.indexOf('159')>-1||t.indexOf('1 tháng')>-1) return 'monthly';
    return '';
  }
  function plan(k){return PLAN[k]||PLAN.monthly}
  function money(v){return Number(v||0).toLocaleString('vi-VN')+' VNĐ'}
  function deviceId(){
    var id=localStorage.getItem('mkt_device_id');
    if(!id){var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/); id=m?decodeURIComponent(m[1]):'';}
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-6);}
    localStorage.setItem('mkt_device_id',id);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=157680000; SameSite=Lax';
    var side=q('#sidebarDeviceId'); if(side) side.textContent=id;
    var pay=q('#payDeviceId'); if(pay) pay.value=id;
    return id;
  }
  function contentFor(p){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'SDT';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'GMAIL';
    return deviceId();
  }
  function applyPayment(k){
    k=k||window.currentPremiumPlanKey||window.__mktClickedPlanKey||'monthly';
    var p=plan(k);
    window.currentPremiumPlanKey=k;
    var modal=q('#paymentModal'); if(!modal) return false;
    var title=q('#payPlanTitle'); if(title) title.textContent=p.title;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent=p.desc || 'Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp.';
    var price=q('#payPlanPrice'); if(price) price.textContent=money(p.amount);
    var dev=q('#payDeviceId'); if(dev) dev.value=deviceId();
    var content=contentFor(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    var locked=q('#payLocked'); if(locked) locked.innerHTML='';
    var h=locked&&locked.previousElementSibling; if(h) h.style.display='none';
    modal.style.display='flex';
    return false;
  }
  function detectFromElement(el){
    if(!el) return '';
    var oc=(el.getAttribute&&el.getAttribute('onclick'))||'';
    var m=oc.match(/openPayment\(['"]([^'"]+)['"]\)/i)||oc.match(/openPremiumCheckout\(['"]([^'"]+)['"]\)/i)||oc.match(/mktOpenPlanDetail\(['"]([^'"]+)['"]\)/i);
    if(m && PLAN[m[1]]) return m[1];
    var card=el.closest && el.closest('.premium-plan,.v4-plan,.price-card,.plan-card,.pricing-card,.premium-card,.mktPlanCard,.modal,.payment-inner,section,div');
    var txt=(card&&card.innerText)||el.innerText||el.textContent||'';
    return keyFromText(txt);
  }
  function remember(k){if(k&&PLAN[k]){window.__mktClickedPlanKey=k; window.__mktClickedPlanAt=Date.now(); window.currentPremiumPlanKey=k;}}
  var oldOpen=window.openPayment;
  window.openPayment=function(k){
    var now=Date.now();
    if((!k || k==='monthly') && window.__mktClickedPlanKey && now-(window.__mktClickedPlanAt||0)<2000) k=window.__mktClickedPlanKey;
    if(!PLAN[k]) k=keyFromText((q('#mktPlanOverlay')||document.body).innerText)||'monthly';
    remember(k);
    return applyPayment(k);
  };
  window.openPremiumCheckout=function(k){return window.openPayment(k||window.__mktClickedPlanKey||'monthly');};
  var oldDetail=window.mktOpenPlanDetail;
  window.mktOpenPlanDetail=function(k){
    if(!PLAN[k]) k=window.__mktClickedPlanKey||keyFromText((document.activeElement&&document.activeElement.innerText)||'')||'monthly';
    remember(k);
    var r=false; if(typeof oldDetail==='function'){try{r=oldDetail(k);}catch(e){}}
    setTimeout(function(){
      var pay=q('#mktPlanPay'); if(pay){pay.onclick=function(e){e.preventDefault(); e.stopPropagation(); return window.openPayment(window.currentPremiumPlanKey||k);};}
      var note=q('#mktPlanNote'); if(note){note.innerHTML='ID thiết bị: <b>'+deviceId()+'</b><br><span style="color:#64748b">Gói đang chọn: <b>'+plan(k).title+'</b> - <b>'+plan(k).price+'</b></span>';}
    },80);
    return r;
  };
  document.addEventListener('click',function(e){
    var btn=e.target.closest&&e.target.closest('button,a,.plan-button,.premium-btn');
    if(!btn) return;
    var text=norm(btn.innerText||btn.textContent||'');
    var isUpgrade=/nâng cấp|mở khóa|đăng ký|thanh toán|chọn gói|xem chi tiết|premium|gửi yêu cầu/i.test(text) || ((btn.getAttribute('onclick')||'').indexOf('openPayment')>-1);
    if(!isUpgrade) return;
    var k=detectFromElement(btn); if(k) remember(k);
  },true);
  document.addEventListener('input',function(e){ if(e.target&&/payPhone|payEmail/.test(e.target.id||'')) applyPayment(window.currentPremiumPlanKey||window.__mktClickedPlanKey||'monthly'); },true);
  setInterval(function(){
    var modal=q('#paymentModal'); if(!modal) return;
    var st=getComputedStyle(modal); if(st.display==='none') return;
    var k=window.currentPremiumPlanKey||window.__mktClickedPlanKey||'monthly';
    var p=plan(k); var price=q('#payPlanPrice');
    if(price && price.textContent.indexOf(Number(p.amount).toLocaleString('vi-VN'))===-1) applyPayment(k);
  },400);
})();
</script>

<script id="mkt-compact-pricing-final-fix">
(function(){
  'use strict';
  var PLANS={
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',code:'GOI1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',code:'GOI3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',code:'GOI6THANG',desc:'Mở thêm CRM, Sales Bot và công cụ nâng cao.'},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',code:'GOI1NAM',desc:'Gói phổ biến, tối ưu chi phí dài hạn.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',code:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  PLANS.lifetime=PLANS.sellerpro;
  function q(s,r){return (r||document).querySelector(s)}
  function norm(k){k=String(k||'').toLowerCase().replace(/[^a-z0-9]/g,''); if(k==='lifetime'||k==='seller'||k==='sellerpro'||k==='nhabanhangpro')return 'sellerpro'; if(k==='monthly'||k==='basic'||k==='goi1thang')return 'monthly'; if(k==='quarterly'||k==='pro'||k==='goi3thang')return 'quarterly'; if(k==='halfyear'||k==='business'||k==='goi6thang')return 'halfyear'; if(k==='yearly'||k==='goi1nam')return 'yearly'; return PLANS[k]?k:'monthly'}
  function device(){var id='';try{id=localStorage.getItem('mkt_device_id')||''}catch(e){} if(!id&&typeof window.getOrCreateDeviceId==='function')id=window.getOrCreateDeviceId(); if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);try{localStorage.setItem('mkt_device_id',id)}catch(e){}} document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age='+(60*60*24*365*5); return id}
  function setKey(k){k=norm(k); window.currentPremiumPlanKey=k; window.mktSelectedPlanKey=k; window.__mktClickedPlanKey=k; try{localStorage.setItem('mkt_selected_plan_key',k)}catch(e){} return k}
  function payNote(p){var phone=(q('#payPhone')&&q('#payPhone').value.trim())||''; var email=(q('#payEmail')&&q('#payEmail').value.trim())||''; return device();}
  window.refreshPaymentContent=function(){var k=norm(window.currentPremiumPlanKey||window.mktSelectedPlanKey); var p=PLANS[k]||PLANS.monthly; var note=payNote(p); var pc=q('#payContent'); if(pc)pc.textContent=note; var price=q('#payPlanPrice'); if(price)price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ'; var qr=q('#payQr'); if(qr)qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(note)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN')};
  window.openPayment=function(k){k=setKey(k); var p=PLANS[k]||PLANS.monthly; var modal=q('#paymentModal'); if(!modal)return false; var title=q('#payPlanTitle'); if(title)title.textContent=p.title; var desc=q('#payPlanDesc'); if(desc)desc.textContent=p.desc; var price=q('#payPlanPrice'); if(price)price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ'; var dev=q('#payDeviceId'); if(dev)dev.value=device(); window.refreshPaymentContent(); modal.style.display='flex'; return false};
  document.addEventListener('click',function(e){var btn=e.target.closest&&e.target.closest('[data-plan]'); if(!btn)return; var k=norm(btn.getAttribute('data-plan')||(btn.dataset&&btn.dataset.plan)); setKey(k); if(/button|a/i.test(btn.tagName)||btn.closest('button,a')){e.preventDefault();e.stopPropagation();if(e.stopImmediatePropagation)e.stopImmediatePropagation();return window.openPayment(k)}},true);
  document.addEventListener('input',function(e){if(e.target&&/payPhone|payEmail/.test(e.target.id||''))window.refreshPaymentContent()},true);
})();
</script>

</body>
</html>
"""

def current_library(selected_industry):
    return CONTENT_LIBRARY.get(selected_industry, CONTENT_LIBRARY["spa"])

def render(content="", message="", ok=True, selected_industry="spa", analysis="", plan=""):
    score = score_content(content) if content else 0
    warnings = policy_check(content) if content else []
    token_warning = ""
    return render_template_string(
        HTML, title=APP_TITLE, pages=get_pages_dynamic(), content=content, message=message, ok=ok,
        history=get_history(), campaigns=get_campaigns(), s=get_stats(), crm_rows=get_crm(), token_report=token_manager_report(), token_checks=get_latest_token_checks(), clusters=get_page_clusters(), analytics=get_analytics_summary(), free_status=get_free_status(),
        industry_labels=INDUSTRY_LABELS, selected_industry=selected_industry,
        library_items=current_library(selected_industry)[:10], locked_count=max(0, 500 - len(current_library(selected_industry)[:10])),
        score=score, warnings=warnings, token_warning=token_warning,
        analysis=analysis, plan=plan, v3=v3_ceo_summary(), pipeline_rows=get_pipeline_leads(), customer_tasks=get_customer_tasks(), notifications=get_notifications(), fb_groups=get_fb_groups(), group_schedules=get_group_schedules(), comment_leads=get_comment_leads(), messenger_scripts=get_messenger_scripts(), success_assets=get_success_assets(), group_finder_results=get_group_finder_results(), group_finder_stats=get_group_finder_stats(), group_join_queue=get_group_join_queue(), group_uid_files=get_group_uid_files(), group_post_results=get_group_post_results(), group_post_queue=get_group_post_queue(), page_comment_queue=get_page_comment_queue(), page_comment_logs=get_page_comment_logs(), page_comment_stats=get_page_comment_stats(), page_token_rows=get_page_token_rows(), page_group_memberships=get_page_group_memberships(), renewal_notice=get_renewal_notice(), device_id=get_device_id(), is_device_premium=bool(get_device_subscription())
    )

@app.route("/")
def home():
    selected_industry = request.args.get("industry", "spa")
    ref_code = current_affiliate_code()
    if ref_code:
        record_affiliate_click(ref_code)
    resp = app.make_response(render(selected_industry=selected_industry))
    if ref_code:
        try:
            resp.set_cookie('affiliate_code', ref_code, max_age=60*60*24*90)
        except Exception:
            pass
    return resp

@app.route("/generate", methods=["POST"])
def generate():
    idea = request.form.get("idea", "").strip()
    style = request.form.get("style", "gần gũi")
    length = request.form.get("length", "120")
    variants = int(request.form.get("variants", "1"))
    if not idea:
        return render(message="Vui lòng nhập ý tưởng bài viết.", ok=False)
    try:
        content = generate_content(idea, style, length, variants=variants)
        return render(content=content, message="Đã tạo nội dung bằng Gemini Flash.", ok=True)
    except Exception as e:
        return render(message="Lỗi Gemini: " + str(e), ok=False)

@app.route("/analyze_fanpage", methods=["POST"])
def analyze_route():
    name = request.form.get("page_name", "")
    avatar = request.form.get("avatar", "")
    cover = request.form.get("cover", "")
    bio = request.form.get("bio", "")
    freq = request.form.get("post_frequency", "")
    cta = request.form.get("cta", "")
    score, notes = analyze_fanpage(name, avatar, cover, bio, freq, cta)
    analysis = f"Fanpage: {name}\nĐiểm đánh giá: {score}/100\n\nĐề xuất:\n- " + "\n- ".join(notes)
    return render(analysis=analysis, message="Đã phân tích Fanpage.", ok=True)

@app.route("/plan_30_days", methods=["POST"])
def plan_route():
    industry = request.form.get("industry", "spa")
    goal = request.form.get("goal", "tăng inbox")
    try:
        plan = ai_planner_v6(industry, goal, request.form.get("days", "30"))
        return render(plan=plan, message="Đã tạo kế hoạch marketing 30 ngày.", ok=True)
    except Exception as e:
        return render(message=friendly_ai_error(e), ok=False)

@app.route("/campaign", methods=["POST"])
def campaign_route():
    add_campaign(
        request.form.get("name","").strip(),
        request.form.get("industry","").strip(),
        request.form.get("goal","").strip(),
        request.form.get("note","").strip()
    )
    return render(message="Đã tạo chiến dịch.", ok=True)


def split_bulk_contents(text):
    raw = (text or "").replace("\r\n", "\n").strip()
    if not raw:
        return []

    # Ưu tiên tách theo dòng trống: mỗi đoạn là 1 content
    parts = [p.strip() for p in raw.split("\n\n") if p.strip()]

    # Nếu không có dòng trống, mỗi dòng là 1 content
    if len(parts) <= 1:
        lines = [x.strip() for x in raw.split("\n") if x.strip()]
        if len(lines) > 1:
            parts = lines

    return parts

def distribute_content_to_pages(contents, pages):
    jobs = []
    if not contents or not pages:
        return jobs

    # Mỗi dòng nội dung là 1 bài riêng, chia lần lượt cho Page.
    # Không nhân bản cùng một nội dung trong một lượt xử lý.
    for i, content in enumerate(contents):
        page = pages[i % len(pages)]
        jobs.append((page, content))
    return jobs

def quick_schedule_base(action, explicit_time=""):
    action = action or "now"
    explicit_time = (explicit_time or "").replace("T", " ").strip()
    if action == "schedule" and explicit_time:
        try:
            return datetime.datetime.strptime(explicit_time, "%Y-%m-%d %H:%M")
        except Exception:
            return datetime.datetime.now() + datetime.timedelta(minutes=30)
    quick_map = {
        "schedule_quick_30": 30,
        "schedule_quick_60": 60,
        "schedule_quick_120": 120,
        "schedule_quick_180": 180,
    }
    if action in quick_map:
        return datetime.datetime.now() + datetime.timedelta(minutes=quick_map[action])
    return None


@app.route("/multi_post", methods=["POST"])
def multi_post():
    free = get_free_status()
    if free.get("is_expired"):
        return render(message="Phiên miễn phí đã giới hạn. Quý khách vui lòng nâng cấp Premium để tiếp tục đăng Fanpage.", ok=False)
    bulk_content = request.form.get("bulk_content", "").strip()
    action = request.form.get("action", "now")
    schedule_time = request.form.get("schedule_time", "").replace("T", " ")
    campaign = request.form.get("campaign", "").strip()
    pages = selected_pages(request.form.getlist("page_indexes"))
    if get_free_status().get("is_trial") and len(pages) > 1:
        return render(message="Gói miễn phí chỉ được đăng tối đa 1 Fanpage. Vui lòng nâng cấp Premium để đăng nhiều Fanpage cùng lúc.", ok=False)
    use_ai_enhance = request.form.get("use_ai_enhance") == "1"

    media_paths = save_uploads(request.files.getlist("images"))
    if not media_paths:
        single = save_upload(request.files.get("image"))
        media_paths = [single] if single else []

    contents = split_bulk_contents(bulk_content)

    if not contents:
        return render(message="Chưa nhập content để đăng.", ok=False)
    if not pages:
        return render(message="Chưa chọn Fanpage.", ok=False)

    # V9: kiểm tra token trước khi đăng ngay
    token_errors = []
    for page in pages:
        token_status = check_single_page_token(page)
        if token_status["status"] != "SỐNG":
            token_errors.append(f"{page.get('name')}: {token_status['detail']}")
    if token_errors and action == "now":
        return render(message="Token lỗi, hệ thống dừng đăng để tránh fail hàng loạt:\n" + "\n".join(token_errors), ok=False)

    jobs = distribute_content_to_pages(contents, pages)
    messages = []
    used_media = set()

    for idx, (page, content) in enumerate(jobs):
        # Mặc định giữ nguyên 100% content khách nhập.
        # Chỉ spin/thêm CTA/hashtag khi khách bật checkbox AI tối ưu.
        final_content = content
        if use_ai_enhance:
            final_content = ai_spin_content(content) if idx > 0 else content
            cta, hashtags = auto_cta_hashtag(final_content, "marketing")
            if cta.lower() not in final_content.lower():
                final_content += "\n\n" + cta
            if "#" not in final_content:
                final_content += "\n" + hashtags

        content_score = score_content(final_content)
        image_path = choose_best_media_for_content(final_content, media_paths, used_media)

        schedule_base = quick_schedule_base(action, schedule_time)
        if schedule_base:
            item_schedule = (schedule_base + datetime.timedelta(minutes=idx * 5)).strftime("%Y-%m-%d %H:%M")
            save_post(page["name"], page["id"], final_content, "scheduled", "", item_schedule, image_path, campaign, content_score)
            media_note = f" | Media: {os.path.basename(image_path)}" if image_path else ""
            messages.append(f"Đã lưu lịch {item_schedule} cho {page['name']}{media_note}")
        else:
            result = post_to_facebook(page, final_content, image_path)
            if "id" in result or "post_id" in result:
                post_id = result.get("post_id") or result.get("id")
                save_post(page["name"], page["id"], final_content, "posted", post_id, "", image_path, campaign, content_score)
                media_note = f" | Media: {os.path.basename(image_path)}" if image_path else ""
                messages.append(f"Đăng thành công {page['name']}: {post_id}{media_note}")
            else:
                save_post(page["name"], page["id"], final_content, "error", str(result), "", image_path, campaign, content_score)
                messages.append(f"Lỗi {page['name']}: {result}")

    return render(message="\n".join(messages), ok=True)

@app.route("/post", methods=["POST"])
def post():
    content = request.form.get("content", "").strip()
    action = request.form.get("action", "now")
    schedule_time = request.form.get("schedule_time", "").replace("T", " ")
    campaign = request.form.get("campaign", "").strip()
    pages = selected_pages(request.form.getlist("page_indexes"))
    image_path = save_upload(request.files.get("image"))
    if not content:
        return render(message="Nội dung trống.", ok=False)
    if not pages:
        return render(content=content, message="Chưa chọn Fanpage.", ok=False)
    messages = []
    content_score = score_content(content)
    for i, page in enumerate(pages):
        final_content = content if i == 0 else spin_content_local(content)
        if action == "schedule":
            if not schedule_time:
                return render(content=content, message="Chưa chọn thời gian đặt lịch.", ok=False)
            save_post(page["name"], page["id"], final_content, "scheduled", "", schedule_time, image_path, campaign, content_score)
            messages.append(f"Đã lưu lịch {schedule_time} cho {page['name']}")
        else:
            result = post_to_facebook(page, final_content, image_path)
            if "id" in result or "post_id" in result:
                post_id = result.get("post_id") or result.get("id")
                save_post(page["name"], page["id"], final_content, "posted", post_id, "", image_path, campaign, content_score)
                messages.append(f"Đăng thành công {page['name']}: {post_id}")
            else:
                save_post(page["name"], page["id"], final_content, "error", str(result), "", image_path, campaign, content_score)
                messages.append(f"Lỗi {page['name']}: {result}")
    return render(content=content, message="\\n".join(messages), ok=True)

@app.route("/batch", methods=["POST"])
def batch():
    file_obj = request.files.get("batch_file")
    if not file_obj:
        return render(message="Chưa chọn file Excel/CSV.", ok=False)
    try:
        rows = read_batch_file(file_obj)
        count = 0
        for row in rows:
            idea = str(row.get("idea", "") or "").strip()
            content = str(row.get("content", "") or "").strip()
            page_names = str(row.get("page_names", "") or "").strip()
            schedule_time = str(row.get("schedule_time", "") or "").strip()
            campaign = str(row.get("campaign", "") or "").strip()
            if not content and idea:
                content = generate_content(idea, "chuyên nghiệp", "120", variants=1)
            if not content or not schedule_time:
                continue
            target_pages = []
            for name in [x.strip().lower() for x in page_names.split(",") if x.strip()]:
                for p in get_pages_dynamic():
                    if name in p["name"].lower():
                        target_pages.append(p)
            pages_dynamic = get_pages_dynamic()
            if not target_pages and pages_dynamic:
                # Nếu file không có page_names, tự chia mỗi dòng content sang Page kế tiếp để tránh trùng bài
                target_pages = [pages_dynamic[count % len(pages_dynamic)]]

            # File Excel/CSV có thể thêm cột image_path để gắn ảnh riêng từng bài.
            image_path = str(row.get("image_path", "") or row.get("media_path", "") or "").strip()
            if image_path and not os.path.exists(image_path):
                image_path = ""

            for page in target_pages:
                save_post(page["name"], page["id"], content, "scheduled", "", schedule_time, image_path, campaign, score_content(content))
                count += 1
        return render(message=f"Đã nhập và lưu lịch {count} bài từ file.", ok=True)
    except Exception as e:
        return render(message="Lỗi đọc file: " + str(e), ok=False)


@app.route("/check_tokens", methods=["POST"])
def check_tokens_route():
    if not PAGES:
        return render(message="Chưa có Fanpage trong PAGES_JSON của file .env.", ok=False)

    results = check_all_page_tokens()
    alive = sum(1 for x in results if x["status"] == "SỐNG")
    dead = len(results) - alive
    message = f"Đã kiểm tra {len(results)} Page Token. Sống: {alive}. Lỗi/Giới hạn: {dead}."
    return render(message=message, ok=(dead == 0))

@app.route("/content_factory", methods=["POST"])
def content_factory_route():
    idea = request.form.get("idea", "").strip()
    count = request.form.get("count", "20")
    style = request.form.get("style", "bán hàng tự nhiên")

    if not idea:
        return render(message="Chưa nhập chủ đề để tạo content hàng loạt.", ok=False)

    try:
        content = generate_many_contents(idea, count, style)
        return render(content=content, message=f"Đã tạo {count} content. Nội dung đã hiển thị trong khung kết quả AI.", ok=True)
    except Exception as e:
        return render(message=friendly_ai_error(e), ok=False)

@app.route("/smart_schedule", methods=["POST"])
def smart_schedule_route():
    bulk_content = request.form.get("bulk_content", "").strip()
    start_time = request.form.get("start_time", "")
    gap_minutes = request.form.get("gap_minutes", "60")
    campaign = request.form.get("campaign", "").strip()
    pages = selected_pages(request.form.getlist("page_indexes"))
    media_paths = save_uploads(request.files.getlist("images"))
    use_ai_enhance = request.form.get("use_ai_enhance") == "1"

    contents = split_bulk_contents(bulk_content)

    if not contents:
        return render(message="Chưa nhập content để chia lịch.", ok=False)
    if not pages:
        return render(message="Chưa chọn Fanpage.", ok=False)

    jobs = distribute_content_to_pages(contents, pages)
    times = smart_schedule_times(start_time, len(jobs), gap_minutes)
    used_media = set()

    for i, (page, content) in enumerate(jobs):
        # Mặc định giữ nguyên 100% content khách nhập.
        final_content = content
        if use_ai_enhance:
            final_content = ai_spin_content(content) if i > 0 else content
            cta, hashtags = auto_cta_hashtag(final_content, "marketing")
            if cta.lower() not in final_content.lower():
                final_content += "\n\n" + cta
            if "#" not in final_content:
                final_content += "\n" + hashtags

        image_path = choose_best_media_for_content(final_content, media_paths, used_media)
        save_post(page["name"], page["id"], final_content, "scheduled", "", times[i], image_path, campaign, score_content(final_content))

    msg = f"Đã tự chia lịch {len(jobs)} bài. Content được giữ nguyên theo nội dung đã nhập."
    if use_ai_enhance:
        msg = f"Đã tự chia lịch {len(jobs)} bài và bật AI tối ưu từng bài. Khoảng cách: {gap_minutes} phút."
    return render(message=msg, ok=True)

@app.route("/page_cluster", methods=["POST"])
def page_cluster_route():
    add_page_cluster(
        request.form.get("name", "").strip(),
        request.form.get("page_names", "").strip(),
        request.form.get("note", "").strip()
    )
    return render(message="Đã lưu nhóm Page.", ok=True)


@app.route("/crm", methods=["POST"])
def crm_route():
    add_crm(
        request.form.get("name","").strip(),
        request.form.get("phone","").strip(),
        request.form.get("zalo","").strip(),
        request.form.get("source","").strip(),
        request.form.get("note","").strip()
    )
    return render(message="Đã lưu khách hàng vào CRM Mini.", ok=True)


@app.route("/v3_ai_tool", methods=["POST"])
def v3_ai_tool_route():
    tool = request.form.get("tool", "marketing_director")
    topic = request.form.get("topic", "").strip()
    extra = request.form.get("extra", "").strip()
    if not topic:
        return render(message="Vui lòng nhập nội dung cần AI xử lý.", ok=False)
    prompt = v3_ai_tool_prompt(tool, topic, extra)
    result = safe_ai_generate(prompt, fallback=f"Bản demo cho {topic}:\n\n- 30 content\n- 10 quảng cáo\n- Tệp khách hàng mục tiêu\n- CTA và kế hoạch triển khai\n\nVui lòng cấu hình GEMINI_API_KEY để dùng AI đầy đủ.")
    return render(content=result, message="Đã tạo nội dung bằng AI Studio V3.", ok=True)

@app.route("/pipeline", methods=["POST"])
def pipeline_route():
    add_pipeline_lead(
        request.form.get("customer_name", "").strip(),
        request.form.get("phone", "").strip(),
        request.form.get("zalo", "").strip(),
        request.form.get("source", "").strip(),
        request.form.get("stage", "Khách mới").strip(),
        request.form.get("value", "0").strip() or 0,
        request.form.get("note", "").strip()
    )
    return render(message="Đã thêm khách vào CRM Sales Pipeline.", ok=True)

@app.route("/customer_task", methods=["POST"])
def customer_task_route():
    add_customer_task(request.form.get("customer_name", "").strip(), request.form.get("task", "").strip(), request.form.get("due_date", "").strip())
    return render(message="Đã tạo lịch chăm sóc khách hàng.", ok=True)

@app.route("/notification", methods=["POST"])
def notification_route():
    add_notification(request.form.get("title", "").strip(), request.form.get("detail", "").strip(), "info")
    return render(message="Đã thêm thông báo vào Notification Center.", ok=True)


@app.route("/fb_group", methods=["POST"])
def fb_group_route():
    add_fb_group(request.form.get("group_name", "").strip(), request.form.get("group_id", "").strip(), request.form.get("niche", "").strip(), request.form.get("note", "").strip())
    return render(message="Đã lưu Group vào Group Marketing.", ok=True)

@app.route("/group_schedule", methods=["POST"])
def group_schedule_route():
    add_group_schedule(request.form.get("group_name", "").strip(), request.form.get("group_id", "").strip(), request.form.get("content", "").strip(), request.form.get("schedule_time", "").replace("T", " "))
    return render(message="Đã lưu lịch đăng Group.", ok=True)


@app.route("/group_finder_scan", methods=["POST"])
def group_finder_scan_route():
    stats = group_finder_import_text(
        request.form.get("keyword", "").strip(),
        request.form.get("raw_groups", ""),
        request.form.get("min_members", "0"),
        request.form.get("privacy", "all"),
        request.form.get("recent_only", "0"),
        request.form.get("page_join", "all"),
        request.form.get("page_post", "all")
    )
    return render(message=f"Đã tìm thấy/lưu {stats['added']} Group phù hợp. Đã loại {stats['duplicate']} Group trùng UID. Đã bỏ qua {stats['rejected']} Group không đạt điều kiện.", ok=True)

@app.route("/group_uid_split", methods=["POST"])
def group_uid_split_route():
    paths = split_group_uids(request.form.get("chunk_size", "50"))
    return render(message=f"Đã chia thành {len(paths)} tệp UID Group.", ok=True)

@app.route("/group_join_queue_add", methods=["POST"])
def group_join_queue_add_route():
    added = add_group_queue_from_results(request.form.get("page_index", ""))
    return render(message=f"Đã đưa {added} Group vào hàng chờ tham gia. Admin cần duyệt trước khi thao tác.", ok=True)

@app.route("/group_post_filter_import", methods=["POST"])
def group_post_filter_import_route():
    stats = import_group_posts(request.form.get("raw_posts", ""), request.form.get("keyword", ""), request.form.get("min_comments", "0"), request.form.get("min_reactions", "0"))
    return render(message=f"Đã lưu {stats['added']} UID bài viết. Trùng {stats['duplicate']}. Lọc bỏ {stats['filtered']}.", ok=True)

@app.route("/group_pair_post_queue", methods=["POST"])
def group_pair_post_queue_route():
    page_indexes = request.form.getlist("page_indexes")
    group_ids = request.form.getlist("group_ids")
    contents = split_bulk_contents(request.form.get("bulk_content", ""))
    if not contents:
        content_single = request.form.get("content", "").strip()
        contents = [content_single] if content_single else []
    if not page_indexes:
        return render(message="Vui lòng chọn ít nhất 1 Page.", ok=False)
    if not group_ids:
        return render(message="Vui lòng chọn ít nhất 1 Group đã tham gia.", ok=False)
    if not contents:
        return render(message="Vui lòng nhập nội dung. Mỗi dòng là một bài để chia không trùng.", ok=False)
    result = add_group_pair_post_queue(page_indexes, group_ids, contents, request.form.get("schedule_mode", "now"), request.form.get("min_delay", "45"), request.form.get("max_delay", "60"))
    detail = "\n".join(result.get("pairs", [])[:20])
    return render(message=f"Đã tạo {result['added']} hàng chờ Page → Group riêng, bỏ qua {result['skipped']}.\n{detail}", ok=True)

@app.route("/group_multi_post_queue", methods=["POST"])
def group_multi_post_queue_route():
    # Giữ route cũ nhưng chuyển sang logic an toàn: 1 Page ghép 1 Group riêng, không tạo toàn bộ tổ hợp Page x Group.
    group_ids = request.form.getlist("group_ids")
    page_indexes = request.form.getlist("page_indexes")
    if not page_indexes:
        single_page = request.form.get("page_index", "")
        page_indexes = [single_page] if single_page != "" else []
    contents = split_bulk_contents(request.form.get("bulk_content", ""))
    content = request.form.get("content", "").strip()
    if not contents and content:
        contents = [content]
    if not group_ids or not contents:
        return render(message="Vui lòng chọn ít nhất 1 Group và nhập nội dung bài đăng.", ok=False)
    if not page_indexes:
        return render(message="Vui lòng chọn ít nhất 1 Page để đưa vào hàng chờ đăng Group.", ok=False)
    result = add_group_pair_post_queue(page_indexes, group_ids, contents, request.form.get("schedule_mode", "now"), request.form.get("min_delay", "45"), request.form.get("max_delay", "60"))
    return render(message=f"Đã tạo {result['added']} hàng chờ theo kiểu 1 Page → 1 Group riêng, không trùng Group. Bỏ qua {result['skipped']}.", ok=True)

@app.route("/export_group_uids")
def export_group_uids_route():
    paths = split_group_uids(100)
    if not paths:
        return render(message="Chưa có UID Group để xuất.", ok=False)
    return send_file(paths[0], as_attachment=True)

@app.route("/page_token_save", methods=["POST"])
def page_token_save_route():
    page_name = request.form.get("page_name", "").strip()
    page_id = request.form.get("page_id", "").strip()
    page_token = request.form.get("page_token", "").strip()
    note = request.form.get("note", "").strip()
    if not save_page_token(page_name, page_id, page_token, note):
        return render(message="Vui lòng nhập đủ Page ID và Page Token.", ok=False)
    return render(message="Đã lưu/cập nhật Page Token. Page này sẽ xuất hiện trong danh sách chọn Page ngay trong tool.", ok=True)

@app.route("/page_group_membership_add", methods=["POST"])
def page_group_membership_add_route():
    page_index = request.form.get("page_index", "")
    group_id = (request.form.get("group_id", "") or request.form.get("group_id_select", "")).strip()
    status = request.form.get("status", "Đã tham gia")
    can_post = request.form.get("can_post", "Có")
    note = request.form.get("note", "")
    if not add_page_group_membership(page_index, group_id, status, can_post, note):
        return render(message="Vui lòng chọn Page và Group hợp lệ trước khi lưu quyền tham gia Group.", ok=False)
    return render(message="Đã lưu Page đã tham gia Group. Khi tạo hàng chờ đăng Group, hệ thống chỉ nhận cặp Page → Group có quyền đăng.", ok=True)


@app.route("/premium_request", methods=["POST"])
def premium_request():
    data = request.get_json(silent=True) or request.form
    device_id = (data.get("device_id") or get_device_id()).strip().upper()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    package_key = normalize_package_key(data.get("package_key") or "monthly")
    if not phone or not email:
        return jsonify({"ok": False, "message": "Vui lòng nhập đủ số điện thoại và Gmail."}), 400
    item = create_premium_request(device_id, phone, email, package_key)
    return jsonify({"ok": True, **item})




def ensure_support_table():
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS support_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT,
            sender TEXT,
            message TEXT,
            status TEXT DEFAULT 'new',
            created_at TEXT
        )
    """)
    conn.commit(); conn.close()

def save_support_message(device_id, sender, message):
    ensure_support_table()
    device_id = (device_id or get_device_id()).strip().upper()
    message = (message or '').strip()
    sender = (sender or 'user').strip().lower()
    if not message:
        return None
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO support_messages(device_id,sender,message,status,created_at) VALUES(?,?,?,?,?)",
              (device_id, sender, message, 'new', now))
    msg_id = c.lastrowid
    if sender == 'user':
        try:
            c.execute("""INSERT INTO notifications(title,detail,level,status,created_at) VALUES(?,?,?,?,?)""",
                      ("Tin nhắn hỗ trợ mới", f"{device_id}: {message[:120]}", "info", "new", now))
        except Exception:
            pass
    conn.commit(); conn.close()
    return msg_id

def get_support_messages(device_id=None, after_id=0, limit=100):
    ensure_support_table()
    conn = db(); c = conn.cursor()
    after_id = int(after_id or 0)
    if device_id:
        c.execute("SELECT id,device_id,sender,message,created_at FROM support_messages WHERE device_id=? AND id>? ORDER BY id ASC LIMIT ?",
                  ((device_id or '').strip().upper(), after_id, limit))
    else:
        c.execute("SELECT id,device_id,sender,message,created_at FROM support_messages WHERE id>? ORDER BY id DESC LIMIT ?",
                  (after_id, limit))
    rows = c.fetchall(); conn.close()
    return rows


def money_vnd(value):
    try:
        return f"{int(float(value or 0)):,}".replace(",", ".") + "đ"
    except Exception:
        return "0đ"

def admin_ceo_stats():
    """Dashboard CEO cho Web Admin: doanh thu Premium, CTV, hoa hồng, ngày/tháng."""
    try:
        ensure_affiliate_tables()
    except Exception:
        pass

    today = datetime.datetime.now().strftime("%Y-%m-%d")
    month = datetime.datetime.now().strftime("%Y-%m")
    out = {
        "today_revenue": 0, "month_revenue": 0, "total_revenue": 0,
        "today_orders": 0, "month_orders": 0, "pending_requests": 0,
        "active_premium": 0, "active_ctv": 0, "pending_withdraw": 0,
        "pending_withdraw_amount": 0, "total_commission": 0,
        "conversion_rate": 0, "daily_rows": [], "monthly_rows": [],
        "top_ctv": [], "package_rows": []
    }

    conn = sqlite3.connect(DB)
    c = conn.cursor()

    def one(sql, params=()):
        try:
            c.execute(sql, params)
            row = c.fetchone()
            return row[0] if row and row[0] is not None else 0
        except Exception:
            return 0

    def many(sql, params=()):
        try:
            c.execute(sql, params)
            return c.fetchall()
        except Exception:
            return []

    # Premium revenue theo yêu cầu đã duyệt
    out["today_revenue"] = one("""
        SELECT COALESCE(SUM(amount),0) FROM premium_upgrade_requests
        WHERE status='Đã duyệt' AND substr(COALESCE(approved_at, created_at),1,10)=?
    """, (today,))
    out["month_revenue"] = one("""
        SELECT COALESCE(SUM(amount),0) FROM premium_upgrade_requests
        WHERE status='Đã duyệt' AND substr(COALESCE(approved_at, created_at),1,7)=?
    """, (month,))
    out["total_revenue"] = one("SELECT COALESCE(SUM(amount),0) FROM premium_upgrade_requests WHERE status='Đã duyệt'")
    out["today_orders"] = one("""
        SELECT COUNT(*) FROM premium_upgrade_requests
        WHERE status='Đã duyệt' AND substr(COALESCE(approved_at, created_at),1,10)=?
    """, (today,))
    out["month_orders"] = one("""
        SELECT COUNT(*) FROM premium_upgrade_requests
        WHERE status='Đã duyệt' AND substr(COALESCE(approved_at, created_at),1,7)=?
    """, (month,))
    out["pending_requests"] = one("SELECT COUNT(*) FROM premium_upgrade_requests WHERE status='Chờ duyệt'")
    out["active_premium"] = one("SELECT COUNT(*) FROM device_subscriptions WHERE status='premium'")

    total_requests = one("SELECT COUNT(*) FROM premium_upgrade_requests")
    approved_requests = one("SELECT COUNT(*) FROM premium_upgrade_requests WHERE status='Đã duyệt'")
    out["conversion_rate"] = round((approved_requests / total_requests) * 100, 1) if total_requests else 0

    # Affiliate / CTV
    out["active_ctv"] = one("SELECT COUNT(*) FROM affiliate_users WHERE status='active'")
    out["pending_withdraw"] = one("SELECT COUNT(*) FROM affiliate_withdrawals WHERE status='Chờ duyệt'")
    out["pending_withdraw_amount"] = one("SELECT COALESCE(SUM(amount),0) FROM affiliate_withdrawals WHERE status='Chờ duyệt'")
    out["total_commission"] = one("SELECT COALESCE(SUM(commission_amount),0) FROM affiliate_referrals WHERE status IN ('Đã duyệt','Đã thanh toán')")

    out["daily_rows"] = many("""
        SELECT substr(COALESCE(approved_at, created_at),1,10) AS day,
               COUNT(*) AS orders,
               COALESCE(SUM(amount),0) AS revenue
        FROM premium_upgrade_requests
        WHERE status='Đã duyệt'
        GROUP BY substr(COALESCE(approved_at, created_at),1,10)
        ORDER BY day DESC
        LIMIT 14
    """)
    out["monthly_rows"] = many("""
        SELECT substr(COALESCE(approved_at, created_at),1,7) AS month,
               COUNT(*) AS orders,
               COALESCE(SUM(amount),0) AS revenue
        FROM premium_upgrade_requests
        WHERE status='Đã duyệt'
        GROUP BY substr(COALESCE(approved_at, created_at),1,7)
        ORDER BY month DESC
        LIMIT 12
    """)
    out["package_rows"] = many("""
        SELECT COALESCE(NULLIF(package_name,''),'Chưa rõ') AS package_name,
               COUNT(*) AS orders,
               COALESCE(SUM(amount),0) AS revenue
        FROM premium_upgrade_requests
        WHERE status='Đã duyệt'
        GROUP BY COALESCE(NULLIF(package_name,''),'Chưa rõ')
        ORDER BY revenue DESC
        LIMIT 10
    """)
    out["top_ctv"] = many("""
        SELECT u.full_name, u.phone, u.email, u.affiliate_code, u.level_name,
               COUNT(r.id) AS orders,
               COALESCE(SUM(r.amount),0) AS revenue,
               COALESCE(SUM(r.commission_amount),0) AS commission
        FROM affiliate_users u
        LEFT JOIN affiliate_referrals r
             ON r.affiliate_code=u.affiliate_code
             AND r.status IN ('Đã duyệt','Đã thanh toán')
        GROUP BY u.id
        ORDER BY revenue DESC, commission DESC
        LIMIT 10
    """)
    conn.close()

    for k in ["today_revenue","month_revenue","total_revenue","pending_withdraw_amount","total_commission"]:
        out[k + "_fmt"] = money_vnd(out[k])
    return out


ADMIN_HTML = """
<!doctype html><html lang="vi"><head><meta charset="utf-8"><title>Web Admin Premium</title>
<style>body{font-family:system-ui;background:#f8fafc;margin:0;padding:24px;color:#111827}.wrap{max-width:1180px;margin:auto}h1{margin-top:0}.card{background:white;border:1px solid #e5e7eb;border-radius:18px;padding:18px;box-shadow:0 10px 30px rgba(15,23,42,.08)}table{width:100%;border-collapse:collapse}th,td{border-bottom:1px solid #e5e7eb;padding:10px;text-align:left;font-size:13px}th{background:#f1f5f9}.badge{display:inline-block;padding:4px 8px;border-radius:999px;background:#fef3c7;color:#92400e;font-weight:700}.ok{background:#dcfce7;color:#166534}.danger{background:#fee2e2;color:#991b1b}button{border:0;border-radius:10px;padding:9px 12px;font-weight:700;cursor:pointer}.approve{background:#16a34a;color:white}.reject{background:#ef4444;color:white}.top{display:flex;gap:12px;align-items:center;justify-content:space-between;margin-bottom:16px}a{color:#2563eb;text-decoration:none}
.ceo-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;margin:18px 0}
.ceo-card{background:linear-gradient(135deg,#ffffff,#eef2ff);border:1px solid #dbeafe;border-radius:18px;padding:16px;box-shadow:0 12px 28px rgba(37,99,235,.08)}
.ceo-card b{display:block;font-size:22px;color:#1d4ed8;margin-top:8px}
.ceo-card small{color:#64748b;font-weight:700}
.admin-tabs{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0 20px}
.admin-tabs a{background:#eef2ff;color:#312e81;padding:10px 14px;border-radius:999px;font-weight:900}
.admin-section-title{display:flex;align-items:center;justify-content:space-between;gap:10px;margin:22px 0 12px}
.admin-two{display:grid;grid-template-columns:1.1fr .9fr;gap:18px;margin-top:18px}
.admin-pill{display:inline-block;padding:5px 10px;border-radius:999px;background:#ecfeff;color:#0e7490;font-weight:900}
.progress-wrap{height:12px;background:#e5e7eb;border-radius:999px;overflow:hidden}
.progress-fill{height:100%;background:linear-gradient(90deg,#2563eb,#7c3aed,#22c55e);border-radius:999px}
@media(max-width:1000px){.ceo-grid{grid-template-columns:repeat(2,1fr)}.admin-two{grid-template-columns:1fr}}
</style>

<style id="chat-device-menu-fix">
/* Bản sửa: chỉ dọn icon menu, tăng ưu tiên chat, hiển thị ID thiết bị rõ ràng */
.v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important;overflow:hidden!important}
.v2-nav-link{cursor:pointer!important;pointer-events:auto!important}
.activity-card span{font-size:0!important}
.activity-card span::before{font-size:16px!important}
.activity-card:nth-of-type(1) span::before{content:"Tổng bài"}
.activity-card:nth-of-type(2) span::before{content:"Đã đăng"}
.activity-card:nth-of-type(3) span::before{content:"Chờ đăng"}
.activity-card:nth-of-type(4) span::before{content:"Lead CRM"}
.activity-card:nth-of-type(5) span::before{content:"Chiến dịch"}
#sidebarDeviceId{display:inline-block!important;color:#fff!important;font-weight:900!important;font-size:14px!important;letter-spacing:.4px;margin:4px 0}
.floating-bot{z-index:2147483000!important;pointer-events:auto!important}
.bot-bubble,.bot-panel,.bot-actions button,.bot-input button{pointer-events:auto!important}
.bot-panel{z-index:2147483001!important}
.bot-body{min-height:210px!important}
.bot-actions{grid-template-columns:1fr 1fr!important}
.bot-actions button{font-size:13px!important;line-height:1.25!important}
.bot-title::before{content:"🤖 ";}
</style>
</head><body><div class="wrap">
<div class="admin-tabs">
  <a href="/admin">🏠 Dashboard CEO</a>
  <a href="#premiumRequests">👑 Premium</a>
  <a href="#adminAffiliateBlock">🤝 CTV Hoa Hồng</a>
  <a href="#revenueReport">💰 Doanh thu</a>
  <a href="#withdrawReport">💳 Rút hoa hồng</a>
</div>

<div class="admin-section-title">
  <h2 style="margin:0">Dashboard CEO</h2>
  <span class="admin-pill">Tự động cập nhật theo dữ liệu duyệt Premium</span>
</div>
<div class="ceo-grid">
  <div class="ceo-card"><small>💰 Doanh thu hôm nay</small><b>{{admin_stats.today_revenue_fmt}}</b><span>{{admin_stats.today_orders}} đơn</span></div>
  <div class="ceo-card"><small>💎 Doanh thu tháng</small><b>{{admin_stats.month_revenue_fmt}}</b><span>{{admin_stats.month_orders}} đơn</span></div>
  <div class="ceo-card"><small>👑 Premium hoạt động</small><b>{{admin_stats.active_premium}}</b><span>thiết bị</span></div>
  <div class="ceo-card"><small>🤝 CTV hoạt động</small><b>{{admin_stats.active_ctv}}</b><span>cộng tác viên</span></div>
  <div class="ceo-card"><small>📈 Tỷ lệ chuyển đổi</small><b>{{admin_stats.conversion_rate}}%</b><span>duyệt / yêu cầu</span></div>
  <div class="ceo-card"><small>💵 Hoa hồng chờ chi</small><b>{{admin_stats.pending_withdraw_amount_fmt}}</b><span>{{admin_stats.pending_withdraw}} yêu cầu</span></div>
</div>

<div class="card" id="revenueReport" style="margin-bottom:18px">
  <div class="admin-section-title"><h2 style="margin:0">💰 Báo cáo doanh thu</h2><span class="admin-pill">Tổng: {{admin_stats.total_revenue_fmt}}</span></div>
  <div class="admin-two">
    <div>
      <h3>Doanh thu 14 ngày gần nhất</h3>
      <table><thead><tr><th>Ngày</th><th>Đơn</th><th>Doanh thu</th></tr></thead><tbody>
      {% for d in admin_stats.daily_rows %}
        <tr><td>{{d[0]}}</td><td>{{d[1]}}</td><td><b>{{"{:,.0f}".format(d[2]).replace(",", ".")}}đ</b></td></tr>
      {% endfor %}
      {% if not admin_stats.daily_rows %}<tr><td colspan="3">Chưa có doanh thu.</td></tr>{% endif %}
      </tbody></table>
    </div>
    <div>
      <h3>Doanh thu theo gói</h3>
      <table><thead><tr><th>Gói</th><th>Đơn</th><th>Doanh thu</th></tr></thead><tbody>
      {% for p in admin_stats.package_rows %}
        <tr><td>{{p[0]}}</td><td>{{p[1]}}</td><td><b>{{"{:,.0f}".format(p[2]).replace(",", ".")}}đ</b></td></tr>
      {% endfor %}
      {% if not admin_stats.package_rows %}<tr><td colspan="3">Chưa có dữ liệu gói.</td></tr>{% endif %}
      </tbody></table>
    </div>
  </div>
</div>

<div class="card" id="withdrawReport" style="margin-bottom:18px">
  <div class="admin-section-title"><h2 style="margin:0">🏆 Top CTV theo doanh thu</h2><span class="admin-pill">Hoa hồng tổng: {{admin_stats.total_commission_fmt}}</span></div>
  <table><thead><tr><th>CTV</th><th>SĐT</th><th>Gmail</th><th>Mã CTV</th><th>Cấp</th><th>Đơn</th><th>Doanh thu</th><th>Hoa hồng</th></tr></thead><tbody>
  {% for ctv in admin_stats.top_ctv %}
    <tr><td><b>{{ctv[0] or 'CTV'}}</b></td><td>{{ctv[1] or ''}}</td><td>{{ctv[2] or ''}}</td><td><b>{{ctv[3]}}</b></td><td>{{ctv[4] or 'CTV thường'}}</td><td>{{ctv[5]}}</td><td><b>{{"{:,.0f}".format(ctv[6]).replace(",", ".")}}đ</b></td><td><b>{{"{:,.0f}".format(ctv[7]).replace(",", ".")}}đ</b></td></tr>
  {% endfor %}
  {% if not admin_stats.top_ctv %}<tr><td colspan="8">Chưa có CTV phát sinh doanh thu.</td></tr>{% endif %}
  </tbody></table>
</div>
<div class="top"><div><h1>Web Admin - Duyệt Premium</h1><p>Danh sách yêu cầu nâng cấp theo ID thiết bị, SĐT và Gmail.</p><p><a href="#adminAffiliateBlock" style="display:inline-block;background:#16a34a;color:#fff;padding:9px 12px;border-radius:10px;font-weight:800">Quản Lý CTV / Hoa Hồng</a></p></div><a href="/">← Về app khách</a></div><div class="card" id="premiumRequests"><table><thead><tr><th>ID</th><th>ID thiết bị</th><th>SĐT</th><th>Gmail</th><th>Gói</th><th>Số tiền</th><th>Nội dung CK</th><th>Trạng thái</th><th>Ngày tạo</th><th>Thao tác</th></tr></thead><tbody>
{% for r in rows %}
<tr><td>{{r[0]}}</td><td><b>{{r[1]}}</b></td><td>{{r[2]}}</td><td>{{r[3]}}</td><td>{{r[5]}}</td><td>{{"{:,.0f}".format(r[6]).replace(",", ".")}}đ</td><td>{{r[7]}}</td><td><span class="badge {% if r[8]=='Đã duyệt' %}ok{% elif r[8]=='Từ chối' %}danger{% endif %}">{{r[8]}}</span></td><td>{{r[10]}}</td><td>{% if r[8]=='Chờ duyệt' %}<form method="post" action="/admin/premium_action" style="display:inline"><input type="hidden" name="request_id" value="{{r[0]}}"><input type="hidden" name="status" value="Đã duyệt"><button class="approve">Kích hoạt</button></form> <form method="post" action="/admin/premium_action" style="display:inline"><input type="hidden" name="request_id" value="{{r[0]}}"><input type="hidden" name="status" value="Từ chối"><button class="reject">Từ chối</button></form>{% else %}{{r[11] or ''}}{% endif %}</td></tr>
{% endfor %}
</tbody></table></div>

<div class="card" id="adminAffiliateBlock" style="margin-top:18px">
  <h2>Quản Lý CTV / Hoa Hồng</h2>
  <p>CTV đăng ký trên app khách sẽ tự xuất hiện tại đây. Admin chỉnh % hoa hồng theo cấp bậc hoặc chỉnh riêng từng CTV.</p>
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:12px 0">
    <input id="adminAffNormal" placeholder="CTV thường %" style="padding:10px;border:1px solid #ddd;border-radius:10px">
    <input id="adminAffSilver" placeholder="CTV Bạc %" style="padding:10px;border:1px solid #ddd;border-radius:10px">
    <input id="adminAffGold" placeholder="CTV Vàng %" style="padding:10px;border:1px solid #ddd;border-radius:10px">
    <input id="adminAffDiamond" placeholder="CTV Kim Cương %" style="padding:10px;border:1px solid #ddd;border-radius:10px">
  </div>
  <button class="approve" onclick="adminAffiliateSaveSettings()">Lưu % hoa hồng cấp bậc</button>
  <h3>Danh sách CTV</h3>
  <div style="overflow:auto"><table id="adminAffiliateUsers"><thead><tr><th>CTV</th><th>SĐT</th><th>Gmail</th><th>Mã CTV</th><th>Cấp</th><th>% riêng</th><th>Trạng thái</th><th>Đơn</th><th>Doanh thu</th><th>Hoa hồng</th><th>Lưu</th></tr></thead><tbody><tr><td colspan="11">Đang tải...</td></tr></tbody></table></div>
  <h3>Yêu cầu rút hoa hồng</h3>
  <div style="overflow:auto"><table id="adminAffiliateWithdrawals"><thead><tr><th>ID</th><th>Mã CTV</th><th>Số tiền</th><th>Ngân hàng</th><th>STK</th><th>Chủ TK</th><th>Trạng thái</th><th>Thao tác</th></tr></thead><tbody><tr><td colspan="8">Đang tải...</td></tr></tbody></table></div>
</div>
<script id="admin-affiliate-panel-js">
(function(){
  function money(n){n=Number(n||0);return n.toLocaleString('vi-VN')+'đ'}
  function esc(s){return String(s||'').replace(/[&<>"]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]})}
  window.adminAffiliateLoad=function(){fetch('/api/admin/affiliate').then(function(r){return r.json()}).then(function(d){
    (d.settings||[]).forEach(function(s){if(s[1]==='CTV thường')document.getElementById('adminAffNormal').value=s[3];if(s[1]==='CTV Bạc')document.getElementById('adminAffSilver').value=s[3];if(s[1]==='CTV Vàng')document.getElementById('adminAffGold').value=s[3];if(s[1]==='CTV Kim Cương')document.getElementById('adminAffDiamond').value=s[3];});
    var tb=document.querySelector('#adminAffiliateUsers tbody');
    tb.innerHTML=(d.users||[]).map(function(u){var id=u[0];return '<tr><td>'+esc(u[2]||u[1]||'')+'</td><td>'+esc(u[3])+'</td><td>'+esc(u[4])+'</td><td><b>'+esc(u[5])+'</b></td><td><input id="lv_'+id+'" value="'+esc(u[6]||'CTV thường')+'" style="width:120px;padding:7px;border:1px solid #ddd;border-radius:8px"></td><td><input id="pct_'+id+'" value="'+esc(u[7]||20)+'" style="width:70px;padding:7px;border:1px solid #ddd;border-radius:8px"></td><td><select id="st_'+id+'"><option value="active" '+(u[8]==='active'?'selected':'')+'>active</option><option value="paused" '+(u[8]==='paused'?'selected':'')+'>paused</option><option value="blocked" '+(u[8]==='blocked'?'selected':'')+'>blocked</option></select></td><td>'+esc(u[10]||0)+'</td><td>'+money(u[11])+'</td><td>'+money(u[12])+'</td><td><button class="approve" onclick="adminAffiliateSaveUser('+id+')">Lưu riêng</button></td></tr>'}).join('') || '<tr><td colspan="11">Chưa có CTV.</td></tr>';
    var wb=document.querySelector('#adminAffiliateWithdrawals tbody');
    wb.innerHTML=(d.withdrawals||[]).map(function(w){return '<tr><td>'+w[0]+'</td><td>'+esc(w[1])+'</td><td>'+money(w[2])+'</td><td>'+esc(w[3])+'</td><td>'+esc(w[4])+'</td><td>'+esc(w[5])+'</td><td><span class="badge">'+esc(w[6])+'</span></td><td><button class="approve" onclick="adminAffiliateWithdrawal(' + w[0] + ',\'Đã thanh toán\')">Đã thanh toán</button> <button class="reject" onclick="adminAffiliateWithdrawal(' + w[0] + ',\'Từ chối\')">Từ chối</button></td></tr>'}).join('') || '<tr><td colspan="8">Chưa có yêu cầu rút tiền.</td></tr>';
  });}
  window.adminAffiliateSaveSettings=function(){fetch('/api/admin/affiliate/settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({normal:document.getElementById('adminAffNormal').value,silver:document.getElementById('adminAffSilver').value,gold:document.getElementById('adminAffGold').value,diamond:document.getElementById('adminAffDiamond').value})}).then(function(r){return r.json()}).then(function(){alert('Đã lưu % hoa hồng cấp bậc');adminAffiliateLoad();})}
  window.adminAffiliateSaveUser=function(id){fetch('/api/admin/affiliate/user/'+id,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({commission_percent:document.getElementById('pct_'+id).value,level_name:document.getElementById('lv_'+id).value,status:document.getElementById('st_'+id).value})}).then(function(r){return r.json()}).then(function(d){alert(d.message||'Đã lưu');adminAffiliateLoad();})}
  window.adminAffiliateWithdrawal=function(id,status){fetch('/api/admin/affiliate/withdraw/'+id,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({status:status})}).then(function(r){return r.json()}).then(function(){alert('Đã cập nhật');adminAffiliateLoad();})}
  document.addEventListener('DOMContentLoaded',adminAffiliateLoad);
})();
</script>

<div class="card" style="margin-top:18px"><h2>Tin nhắn hỗ trợ khách hàng</h2><p>Admin nhập phản hồi theo đúng ID thiết bị, khách sẽ nhận lại trong khung chat.</p><table><thead><tr><th>ID</th><th>ID thiết bị</th><th>Người gửi</th><th>Tin nhắn</th><th>Thời gian</th><th>Phản hồi</th></tr></thead><tbody>{% for m in support_rows %}<tr><td>{{m[0]}}</td><td><b>{{m[1]}}</b></td><td>{{m[2]}}</td><td>{{m[3]}}</td><td>{{m[4]}}</td><td><form method="post" action="/admin/support_reply" style="display:flex;gap:8px"><input type="hidden" name="device_id" value="{{m[1]}}"><input name="message" placeholder="Nhập phản hồi..." style="flex:1;padding:9px;border:1px solid #ddd;border-radius:10px"><button class="approve">Gửi</button></form></td></tr>{% endfor %}</tbody></table></div></div>
<script id="chat-device-menu-fix-js">
(function(){
  function ensureDeviceId(){
    var id = localStorage.getItem("mkt_device_id");
    if(!id || id === "Đang tạo..."){
      id = "MKT-" + Math.random().toString(36).slice(2,8).toUpperCase() + Date.now().toString().slice(-4);
      localStorage.setItem("mkt_device_id", id);
    }
    document.cookie = "mkt_device_id=" + encodeURIComponent(id) + "; path=/; max-age=" + (60*60*24*365*5);
    var side = document.getElementById("sidebarDeviceId");
    if(side) side.textContent = id;
    var pay = document.getElementById("payDeviceId");
    if(pay) pay.value = id;
    return id;
  }
  window.ensureDeviceId = ensureDeviceId;
  document.addEventListener("DOMContentLoaded", function(){
    ensureDeviceId();
    setTimeout(ensureDeviceId, 300);
    setTimeout(ensureDeviceId, 1200);

    var bubble = document.querySelector(".bot-bubble");
    if(bubble){
      bubble.addEventListener("click", function(e){
        e.preventDefault();
        var panel = document.getElementById("floatingBotPanel");
        if(panel){
          panel.style.display = (panel.style.display === "block") ? "none" : "block";
          if(panel.style.display === "block" && typeof appendBotGreeting === "function") appendBotGreeting();
        }
      }, true);
    }

    document.querySelectorAll(".v2-nav-link[href^='#']").forEach(function(a){
      a.addEventListener("click", function(e){
        var id = (a.getAttribute("href") || "").replace("#", "");
        if(!id) return;
        e.preventDefault();
        if(typeof openModule === "function") openModule(id);
      }, true);
    });
  });
})();
</script>


<style id="final-chat-lock-premium-fix">
/* FIX CUỐI: chat không tràn, menu trái bỏ ký hiệu, khóa menu Premium */
.v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important}
.v2-nav-link{position:relative!important;gap:8px!important;padding-right:112px!important;cursor:pointer!important;pointer-events:auto!important}
.v2-nav-link .v2-nav-tag{display:none!important}
.v2-nav-link.premium-locked::after{
  content:"PREMIUM";
  position:absolute;right:14px;top:50%;transform:translateY(-50%);
  font-size:11px;font-weight:1000;letter-spacing:.6px;color:#fff;
  padding:5px 10px;border-radius:999px;
  background:linear-gradient(135deg,#facc15,#fb923c,#ef4444);
  box-shadow:0 0 12px rgba(250,204,21,.95),0 0 22px rgba(124,58,237,.65);
  animation:premiumGlow 1.2s infinite alternate;
}
.v2-nav-link.premium-locked{border-color:rgba(250,204,21,.55)!important;background:linear-gradient(135deg,rgba(49,46,129,.55),rgba(76,29,149,.45))!important}
@keyframes premiumGlow{from{filter:brightness(1)}to{filter:brightness(1.35)}}

.floating-bot{position:fixed!important;right:24px!important;bottom:24px!important;z-index:2147483000!important;pointer-events:none!important}
.bot-bubble{pointer-events:auto!important;position:relative!important;z-index:2147483002!important}
.bot-panel{
  pointer-events:auto!important;display:none;position:absolute!important;right:0!important;bottom:92px!important;
  width:min(420px,calc(100vw - 28px))!important;height:min(660px,calc(100vh - 120px))!important;
  max-height:calc(100vh - 120px)!important;overflow:hidden!important;
  border-radius:24px!important;background:#fff!important;box-shadow:0 25px 80px rgba(15,23,42,.35)!important;
  z-index:2147483001!important;flex-direction:column!important;
}
.bot-panel.bot-open{display:flex!important}
.bot-head{flex:0 0 auto!important}
.bot-body{flex:1 1 auto!important;min-height:0!important;max-height:none!important;overflow-y:auto!important;padding:18px!important;background:#F8FAFC!important}
.bot-msg{max-width:100%!important;box-sizing:border-box!important;word-break:break-word!important;overflow-wrap:anywhere!important;margin:0 0 10px!important}
.bot-actions{flex:0 0 auto!important;display:grid!important;grid-template-columns:1fr 1fr!important;gap:10px!important;padding:12px 18px!important;background:#fff!important;border-top:1px solid #EEF2FF!important;box-sizing:border-box!important}
.bot-actions button{min-height:56px!important;margin:0!important;white-space:normal!important;line-height:1.2!important;font-size:13px!important}
.bot-input{flex:0 0 auto!important;padding:12px 18px 18px!important;background:#fff!important;display:flex!important;gap:10px!important;box-sizing:border-box!important}
.bot-input input{min-width:0!important;flex:1!important}
.bot-input button{flex:0 0 62px!important}
@media(max-width:520px){.floating-bot{right:10px!important;bottom:10px!important}.bot-panel{right:0!important;width:calc(100vw - 20px)!important;height:calc(100vh - 110px)!important}}
</style>

<script id="final-chat-lock-premium-fix-js">
(function(){
  var MKT_IS_PREMIUM = {{ 'true' if is_device_premium else 'false' }};
  var premiumModules = {
    messenger_ai:'AI Messenger', crm_sales:'CRM Kanban', marketing_director:'AI Marketing Director',
    ai_studio:'AI Studio', creative_center:'Image / Video / Voice', analytics:'Analytics Center',
    analytics_center:'Analytics Center', automation_center:'Cài đặt Automation'
  };
  var aliases = {post:'page_center_total', page_center:'page_center_total', page_comment_pro:'page_center_total', page_comment_queue:'page_center_total'};
  function normalizeId(id){ return aliases[id] || id; }
  function isLocked(id){ return !MKT_IS_PREMIUM && !!premiumModules[normalizeId(id)]; }
  function lockFeature(id){
    var name = premiumModules[normalizeId(id)] || 'Tính năng Premium';
    if(typeof openLockedFeature === 'function') return openLockedFeature(name);
    if(typeof openPremiumPopup === 'function') openPremiumPopup();
    var p=document.getElementById('premium'); if(p) p.scrollIntoView({behavior:'smooth',block:'start'});
    return false;
  }
  function showOnlyModule(id){
    id = normalizeId(id);
    if(isLocked(id)) return lockFeature(id);
    document.querySelectorAll('.module-section').forEach(function(el){el.classList.remove('active-module');});
    var target=document.getElementById(id) || document.getElementById('dashboard');
    if(target){
      target.classList.add('active-module');
      setTimeout(function(){target.scrollIntoView({behavior:'smooth',block:'start'});},30);
    }
    document.querySelectorAll('.v2-nav-link').forEach(function(a){a.classList.remove('active');});
    var active=document.querySelector('.v2-nav-link[href="#'+id+'"]');
    if(active) active.classList.add('active');
    return false;
  }
  window.openModule = showOnlyModule;

  function setBotOpen(open){
    var panel=document.getElementById('floatingBotPanel');
    if(!panel) return;
    panel.classList.toggle('bot-open', !!open);
    panel.style.display = open ? 'flex' : 'none';
    if(open && typeof appendBotGreeting === 'function') appendBotGreeting();
    setTimeout(function(){var body=document.getElementById('floatingBotBody'); if(body) body.scrollTop=body.scrollHeight;},50);
  }
  window.toggleFloatingBot=function(){
    var panel=document.getElementById('floatingBotPanel');
    setBotOpen(!(panel && panel.classList.contains('bot-open')));
    return false;
  };
  window.closeFloatingBot=function(){setBotOpen(false);return false;};

  document.addEventListener('DOMContentLoaded',function(){
    document.querySelectorAll('.v2-nav-link').forEach(function(a){
      var id=(a.getAttribute('href')||'').replace('#','');
      a.querySelectorAll('.v2-nav-ico').forEach(function(x){x.remove();});
      if(isLocked(id)) a.classList.add('premium-locked');
      else a.classList.remove('premium-locked');
      a.addEventListener('click',function(e){
        var mid=(a.getAttribute('href')||'').replace('#','');
        if(mid){e.preventDefault();e.stopPropagation();showOnlyModule(mid);}
      },true);
    });
    var panel=document.getElementById('floatingBotPanel');
    if(panel){panel.style.display='none';panel.classList.remove('bot-open');}
    document.querySelectorAll('.bot-actions button,.bot-input button,.bot-bubble,.bot-close').forEach(function(btn){btn.style.pointerEvents='auto';});
  });
})();
</script>



<!-- FINAL REAL FIX: menu PRO badge, pricing buttons, chat buttons, hide technical status -->
<style id="final-real-fix-style">
/* Ẩn toàn bộ trạng thái kỹ thuật khỏi giao diện khách */
.rightbar h2, .rightbar p{ }
.rightbar h2:has(+ p){ }
/* Xóa icon/ký hiệu bên trái menu, chỉ giữ chữ */
.v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important;overflow:hidden!important}
.v2-nav-link{position:relative!important;cursor:pointer!important;user-select:none!important;display:flex!important;align-items:center!important;justify-content:space-between!important;gap:10px!important}
.v2-nav-link .v2-nav-tag{display:none!important}
.v2-nav-link.menu-pro::after{
  content:'PRO';
  display:inline-flex;align-items:center;justify-content:center;
  min-width:44px;height:24px;padding:0 10px;border-radius:999px;
  color:#052e16;font-weight:1000;font-size:12px;letter-spacing:.7px;
  background:linear-gradient(135deg,#BBF7D0,#22C55E,#16A34A);
  border:1px solid rgba(187,247,208,.95);
  box-shadow:0 0 8px rgba(34,197,94,.9),0 0 18px rgba(34,197,94,.55),inset 0 1px 0 rgba(255,255,255,.8);
  animation:proPulseGreen 1.45s infinite alternate;
}
@keyframes proPulseGreen{from{filter:brightness(1);transform:scale(1)}to{filter:brightness(1.28);transform:scale(1.04)}}
.v2-nav-link.menu-locked{opacity:.96!important}
.v2-nav-link.menu-locked::before{content:'';position:absolute;inset:0;border-radius:inherit;background:rgba(15,23,42,.04);pointer-events:none}
.price-card button,.premium-plan button,.safe-pricing-action,.plan-button,.premium-btn,.bot-actions button,.bot-input button,.bot-bubble,.bot-close{pointer-events:auto!important;cursor:pointer!important;position:relative!important;z-index:9999!important}
.floating-bot,.bot-panel{z-index:2147483000!important}
.bot-body{max-height:340px!important;overflow-y:auto!important;overflow-x:hidden!important;box-sizing:border-box!important}
.bot-actions{display:grid!important;grid-template-columns:1fr 1fr!important;gap:12px!important;padding:14px 18px!important;background:#f8fafc!important;box-sizing:border-box!important}
.bot-input{display:flex!important;gap:10px!important;padding:12px 18px 18px!important;background:#fff!important;box-sizing:border-box!important}
.bot-input input{min-width:0!important;flex:1!important}
</style>
<script id="final-real-fix-js">
(function(){
  'use strict';
  function qs(s,root){return (root||document).querySelector(s)}
  function qsa(s,root){return Array.prototype.slice.call((root||document).querySelectorAll(s))}
  function norm(s){return String(s||'').toLowerCase()}
  function byText(sel, words){
    return qsa(sel).filter(function(el){var t=norm(el.innerText||el.textContent);return words.some(function(w){return t.indexOf(norm(w))>-1})})
  }
  function getDeviceId(){
    var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/);
    if(m) return decodeURIComponent(m[1]);
    var id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=31536000; SameSite=Lax';
    return id;
  }
  window.getOrCreateDeviceId = window.getOrCreateDeviceId || getDeviceId;
  function planKeyFromText(text){
    text=norm(text);
    if(text.indexOf('1.959')>-1||text.indexOf('1959')>-1||text.indexOf('nhà bán')>-1||text.indexOf('seller')>-1) return 'sellerpro';
    if(text.indexOf('859')>-1||text.indexOf('1 năm')>-1) return 'yearly';
    if(text.indexOf('559')>-1||text.indexOf('6 tháng')>-1) return 'halfyear';
    if(text.indexOf('359')>-1||text.indexOf('3 tháng')>-1) return 'quarterly';
    return 'monthly';
  }
  function showPricing(){
    var premium=document.getElementById('premium') || qs('.premium-pricing-pro') || qs('.pricing-grid') || qs('.premium-grid-pro');
    if(premium) premium.scrollIntoView({behavior:'smooth', block:'start'});
  }
  function openPlan(planKey){
    planKey=planKey||'monthly';
    window.currentPremiumPlanKey=planKey;
    if(typeof window.openPayment==='function'){ window.openPayment(planKey); return false; }
    if(typeof window.openPremiumCheckout==='function'){ window.openPremiumCheckout(planKey); return false; }
    if(typeof window.openPremiumPopup==='function'){ window.openPremiumPopup(); return false; }
    showPricing(); return false;
  }
  window.scrollToPricing = function(){showPricing(); return false};
  window.openPremium = function(){showPricing(); return false};
  function isPremiumActive(){
    var txt=norm((qs('#sidebarPremiumStatus')||{}).innerText||'');
    return txt.indexOf('premium')>-1 && txt.indexOf('dùng thử')<0 && txt.indexOf('hết')<0;
  }
  function openLocked(feature){
    if(typeof window.openLockedFeature==='function') return window.openLockedFeature(feature||'Tính năng PRO');
    showPricing(); return false;
  }
  var realOpenModule = window.openModule;
  var freeModules = {dashboard:1, facebook_center:1, page_center_total:1, post:1, fanpage_manager:1, group_suite:1, comment_manager:1, premium:1};
  window.openModule = function(moduleId){
    moduleId = moduleId || 'dashboard';
    var proOnly = !freeModules[moduleId];
    if(proOnly && !isPremiumActive()) return openLocked(moduleId);
    if(typeof realOpenModule==='function'){
      try { return realOpenModule(moduleId); } catch(e) { console.warn('old openModule error',e); }
    }
    var target=document.getElementById(moduleId);
    if(target){ target.scrollIntoView({behavior:'smooth',block:'start'}); history.replaceState(null,'','#'+moduleId); return false; }
    if(moduleId==='premium') { showPricing(); return false; }
    return false;
  };
  function botReply(t){
    var l=norm(t);
    if(l.indexOf('premium')>-1||l.indexOf('nâng cấp')>-1||l.indexOf('gói')>-1||l.indexOf('giá')>-1)
      return 'Dạ hiện hệ thống có các gói:<br><br>• 1 tháng: <b>159K</b><br>• 3 tháng: <b>359K</b><br>• 6 tháng: <b>559K</b><br>• 1 năm: <b>859K</b><br>• Nhà bán hàng chuyên nghiệp: <b>1.959K</b><br><br>Anh/chị chọn gói phù hợp, nhập số điện thoại và Gmail để nhận hỗ trợ kích hoạt nhanh ạ.';
    if(l.indexOf('thanh toán')>-1||l.indexOf('qr')>-1||l.indexOf('chuyển khoản')>-1)
      return 'Dạ sau khi chuyển khoản, anh/chị gửi giúp em:<br><br>• ID thiết bị: <b>'+getDeviceId()+'</b><br>• Ảnh thanh toán<br>• Gói đã đăng ký<br><br>Nếu sau 5 phút chưa kích hoạt, vui lòng liên hệ Zalo <b>036 338 2629</b>.';
    if(l.indexOf('kích hoạt')>-1||l.indexOf('duyệt')>-1)
      return 'Dạ tài khoản sẽ được mở sau khi web admin duyệt đúng ID thiết bị.<br><br>ID thiết bị của anh/chị là: <b>'+getDeviceId()+'</b>';
    if(l.indexOf('liên hệ')>-1||l.indexOf('zalo')>-1||l.indexOf('hỗ trợ')>-1)
      return '<b>Zalo hỗ trợ:</b> 036 338 2629<br><b>Gmail hỗ trợ:</b> support@gptmini.pro';
    return 'Dạ em đã nhận thông tin. Anh/chị cần hỗ trợ Nâng cấp Premium, Kích hoạt tài khoản, Thanh toán hay Báo lỗi hệ thống ạ?';
  }
  window.toggleFloatingBot = function(){
    var p=qs('#floatingBotPanel'); if(!p) return false;
    var open=getComputedStyle(p).display!=='none';
    p.style.display=open?'none':'block';
    return false;
  };
  window.closeFloatingBot = function(){var p=qs('#floatingBotPanel'); if(p) p.style.display='none'; return false};
  window.botQuick = function(text){
    var body=qs('#floatingBotBody'); if(!body) return false;
    body.insertAdjacentHTML('beforeend','<div class="bot-msg"><b>Bạn:</b> '+String(text).replace(/[&<>]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;'}[c]})+'</div>');
    var id='typing_'+Date.now();
    body.insertAdjacentHTML('beforeend','<div class="bot-msg ai" id="'+id+'"><b>Bot hỗ trợ:</b><br>Đang nhập...</div>');
    body.scrollTop=body.scrollHeight;
    setTimeout(function(){var x=document.getElementById(id); if(x) x.outerHTML='<div class="bot-msg ai"><b>Bot hỗ trợ:</b><br>'+botReply(text)+'</div>'; body.scrollTop=body.scrollHeight;},500);
    if(window.fetch){try{fetch('/support_send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:getDeviceId(),message:String(text),sender:'customer'})}).catch(function(){});}catch(e){}}
    return false;
  };
  window.sendBotInput = function(){var i=qs('#botInputText'); if(!i||!i.value.trim()) return false; var v=i.value.trim(); i.value=''; return window.botQuick(v)};
  function bindAll(){
    // ID thiết bị
    var id=getDeviceId(); var side=qs('#sidebarDeviceId'); if(side) side.textContent=id; var pay=qs('#payDeviceId'); if(pay) pay.value=id;
    // Remove visible technical status if still rendered
    qsa('h2').forEach(function(h){ if(norm(h.innerText).indexOf('trạng thái hệ thống')>-1){ var p=h.nextElementSibling; if(p) p.remove(); if(h.previousElementSibling && h.previousElementSibling.tagName==='HR') h.previousElementSibling.remove(); h.remove(); }});
    qsa('p,div').forEach(function(el){ var t=norm(el.innerText); if(t.indexOf('thiếu gemini api')>-1 || t.indexOf('pages_json')>-1){ el.style.display='none'; }});
    // Sidebar PRO badges: 4 chức năng đầu Facebook Center
    byText('.v2-nav-link',['Facebook Center','Page Center Tổng','Đăng bài Facebook','Quản lý Fanpage']).forEach(function(a){a.classList.add('menu-pro')});
    qsa('.v2-nav-link').forEach(function(a){
      a.addEventListener('click',function(e){
        var oc=a.getAttribute('onclick')||''; var m=oc.match(/openModule\('([^']+)'\)/); if(m){ e.preventDefault(); return window.openModule(m[1]); }
      },true);
    });
    // Pricing cards and buttons, including Xem chi tiết gói
    qsa('.price-card,.premium-plan').forEach(function(card){
      card.style.cursor='pointer';
      card.onclick=function(e){ if(e.target && e.target.closest('button')) return; e.preventDefault(); return openPlan(planKeyFromText(card.innerText)); };
      qsa('button',card).forEach(function(btn){
        btn.onclick=function(e){ e.preventDefault(); e.stopPropagation(); return openPlan(planKeyFromText(card.innerText)); };
      });
    });
    byText('button,a',['Xem chi tiết gói','Nâng cấp Premium']).forEach(function(btn){
      if(btn.closest('.bot-actions')) return;
      btn.onclick=function(e){ e.preventDefault(); e.stopPropagation(); var card=btn.closest('.price-card,.premium-plan'); return openPlan(planKeyFromText(card?card.innerText:btn.innerText)); };
    });
    // Chat buttons
    var bubble=qs('.bot-bubble'); if(bubble) bubble.onclick=function(e){e.preventDefault();return window.toggleFloatingBot()};
    var close=qs('.bot-close'); if(close) close.onclick=function(e){e.preventDefault();return window.closeFloatingBot()};
    var input=qs('#botInputText'); if(input) input.onkeydown=function(e){if(e.key==='Enter'){e.preventDefault();window.sendBotInput();}};
    var send=qs('.bot-input button'); if(send) send.onclick=function(e){e.preventDefault();return window.sendBotInput()};
    byText('.bot-actions button',['Nâng cấp Premium']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Nâng cấp Premium')}});
    byText('.bot-actions button',['Hướng dẫn thanh toán']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Hướng dẫn thanh toán')}});
    byText('.bot-actions button',['Kích hoạt tài khoản']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Kích hoạt tài khoản')}});
    byText('.bot-actions button',['Liên hệ hỗ trợ']).forEach(function(b){b.onclick=function(e){e.preventDefault();return window.botQuick('Liên hệ hỗ trợ')}});
    // Counter auto jump
    var c=qs('#liveMemberCount'); if(c && !window.__liveCounterFinal){
      window.__liveCounterFinal=true;
      var n=parseInt((c.textContent||localStorage.getItem('mkt_live_count')||'231').replace(/\D/g,''),10)||231;
      setInterval(function(){n += Math.floor(Math.random()*3)+1; if(n>999)n=231; c.textContent=n; localStorage.setItem('mkt_live_count',String(n));},1800);
    }
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',bindAll); else bindAll();
  setTimeout(bindAll,400); setTimeout(bindAll,1200); setTimeout(bindAll,2500);
})();
</script>



<!-- FINAL PAYMENT UX OVERRIDE: no browser alert, professional checkout -->
<style id="mkt-payment-professional-css">
#mktPlanNote .pay-line,#mktPlanNote .pay-copy{display:block;margin:6px 0;color:#334155;font-weight:800}
#mktPlanNote .pay-copy{background:#f1f5f9;border:1px dashed #cbd5e1;border-radius:12px;padding:10px;color:#1e293b}
.payment-alert{background:#f8fafc!important;border:1px solid #dbeafe!important;color:#1e293b!important}
.bank-info{line-height:1.75!important}
</style>
<script id="mkt-payment-professional-js">
(function(){
  'use strict';
  function q(s){return document.querySelector(s)}
  function money(v){try{return Number(v||0).toLocaleString('vi-VN')+' VNĐ'}catch(e){return (v||'')+' VNĐ'}}
  function getDevice(){
    var id=localStorage.getItem('mkt_device_id');
    if(!id){
      var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/);
      id=m?decodeURIComponent(m[1]):'';
    }
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);}
    localStorage.setItem('mkt_device_id',id);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=157680000; SameSite=Lax';
    var side=q('#sidebarDeviceId'); if(side) side.textContent=id;
    var pay=q('#payDeviceId'); if(pay) pay.value=id;
    return id;
  }
  var plans=window.premiumPlans || {
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',package:'1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',package:'3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',package:'6THANG',desc:'Mở CRM và Sales Bot.'},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',package:'1NAM',desc:'Gói phổ biến nhất.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',package:'NHABANHANGCHUYENNGHIEP',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'},
    lifetime:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',package:'NHABANHANGCHUYENNGHIEP',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  plans.sellerpro=plans.sellerpro||plans.lifetime; plans.lifetime=plans.lifetime||plans.sellerpro;
  function plan(k){return plans[k]||plans.monthly;}
  function buildContent(p){
    var device=getDevice();
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    return device+' | '+phone+' | '+email+' | '+(p.package||p.title||'PREMIUM').toUpperCase();
  }
  window.refreshPaymentContent=function(){
    var p=plan(window.currentPremiumPlanKey||'monthly');
    var content=buildContent(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount||0)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
  };
  window.openPayment=function(planKey){
    var p=plan(planKey); window.currentPremiumPlanKey=planKey||'monthly';
    var modal=q('#paymentModal');
    if(!modal){return false;}
    var title=q('#payPlanTitle'); if(title) title.textContent=(p.title||'Gói Premium').replace(/^🚀 |^⭐ |^💎 |^🔥 |^👑 /,'');
    var desc=q('#payPlanDesc'); if(desc) desc.textContent='Nhập số điện thoại và Gmail để đăng ký kích hoạt. Nội dung chuyển khoản sẽ tự kèm ID thiết bị và gói đã chọn.';
    var price=q('#payPlanPrice'); if(price) price.textContent=money(p.amount);
    var device=q('#payDeviceId'); if(device) device.value=getDevice();
    var benefits=q('#payBenefits'); if(benefits && p.benefits){benefits.innerHTML=p.benefits.map(function(x){return '<div>'+x+'</div>';}).join('');}
    var locked=q('#payLocked'); if(locked){locked.innerHTML=''; var h=locked.previousElementSibling; if(h) h.style.display='none';}
    var alertBox=q('.payment-alert'); if(alertBox){alertBox.innerHTML='Sau khi chuyển khoản, bấm <b>Đã thanh toán</b> để gửi yêu cầu về web admin. Nội dung chuyển khoản gồm ID thiết bị, số điện thoại, Gmail và gói đã chọn. Zalo hỗ trợ: <b>036 338 2629</b>.';}
    var payBtn=q('.payment-actions .light'); if(payBtn){payBtn.textContent='Đã thanh toán'; payBtn.removeAttribute('href'); payBtn.onclick=function(e){e.preventDefault(); if(typeof window.submitPremiumRequest==='function') return window.submitPremiumRequest(); return false;};}
    window.refreshPaymentContent();
    modal.style.display='flex';
    var ph=q('#payPhone'); if(ph) setTimeout(function(){ph.focus();},150);
    return false;
  };
  window.mktOpenPlanDetail = (function(old){
    return function(k){
      if(typeof old==='function') old(k);
      setTimeout(function(){
        var p=plan(k||window.currentPremiumPlanKey||'monthly');
        var note=q('#mktPlanNote');
        if(note){
          var content=getDevice()+' | SỐ ĐIỆN THOẠI | GMAIL | '+(p.package||p.title||'PREMIUM').toUpperCase();
          note.innerHTML='<span class="pay-line">ID thiết bị: <b>'+getDevice()+'</b></span><span class="pay-line">Gói đang chọn: <b>'+p.title+'</b></span><span class="pay-copy">Nội dung thanh toán: '+content+'</span>';
        }
        var mini=q('.mktPlanMini'); if(mini) mini.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
        var btn=q('#mktPlanPay'); if(btn) btn.onclick=function(e){e.preventDefault(); if(q('#mktPlanOverlay')) q('#mktPlanOverlay').style.display='none'; return window.openPayment(k||'monthly');};
      },60);
      return false;
    };
  })(window.mktOpenPlanDetail);
  document.addEventListener('click',function(e){
    var t=e.target.closest('#mktPlanPay');
    if(t){e.preventDefault(); e.stopPropagation(); return window.openPayment(window.currentPremiumPlanKey||'monthly');}
  },true);
  getDevice();
})();
</script>



<!-- FINAL CLEAN PAYMENT PATCH V74 -->
<style id="mkt-final-clean-payment-css">
  #mktPlanOverlay #mktPlanNote{font-size:16px;line-height:1.55;color:#475569;font-weight:700}
  #mktPlanOverlay .mktPlanMini{background:#f8fafc!important;border:1px solid #dbeafe!important;color:#334155!important}
  #mktPlanOverlay .pay-copy,
  #paymentModal .pay-copy{display:block;margin-top:8px;background:#f1f5f9;border:1px dashed #cbd5e1;border-radius:12px;padding:10px;color:#1e293b;font-weight:800;word-break:break-word}
  #paymentModal .payment-alert{display:none!important}
  #paymentModal .payment-actions{display:grid!important;grid-template-columns:1fr 1fr;gap:10px}
  #paymentModal .payment-actions a{display:flex;align-items:center;justify-content:center;text-decoration:none}
  #paymentNotice{margin:10px 0;padding:12px 14px;border:1px solid #bbf7d0;background:#f0fdf4;color:#166534;border-radius:12px;font-weight:800;display:none}
</style>
<script id="mkt-final-clean-payment-js">
(function(){
  'use strict';
  var PLAN_MAP={
    monthly:{title:'Gói 1 tháng',amount:'159000',price:'159.000đ',code:'GOI1THANG'},
    quarterly:{title:'Gói 3 tháng',amount:'359000',price:'359.000đ',code:'GOI3THANG'},
    halfyear:{title:'Gói 6 tháng',amount:'559000',price:'559.000đ',code:'GOI6THANG'},
    yearly:{title:'Gói 1 năm',amount:'859000',price:'859.000đ',code:'GOI1NAM'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',amount:'1959000',price:'1.959.000đ',code:'NHABANHANGPRO'},
    lifetime:{title:'Gói nhà bán hàng chuyên nghiệp',amount:'1959000',price:'1.959.000đ',code:'NHABANHANGPRO'}
  };
  function q(s,root){return (root||document).querySelector(s)}
  function qa(s,root){return Array.prototype.slice.call((root||document).querySelectorAll(s))}
  function cleanText(){
    var bad=['Gói sẽ mở sau khi admin duyệt thanh toán.','Gói sẽ mở sau khi admin duyệt thanh toán','Sau khi thanh toán, web admin sẽ duyệt theo ID thiết bị. Duyệt xong khách mới sử dụng được tính năng PRO.','Sau khi thanh toán, web admin sẽ duyệt theo ID thiết bị.','Duyệt xong khách mới sử dụng được tính năng PRO.'];
    qa('body *').forEach(function(el){
      if(el.children.length>0) return;
      var t=el.textContent||''; var nt=t;
      bad.forEach(function(b){nt=nt.split(b).join('')});
      if(nt!==t) el.textContent=nt.trim();
    });
  }
  function deviceId(){
    var id=localStorage.getItem('mkt_device_id');
    if(!id){var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/); id=m?decodeURIComponent(m[1]):'';}
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-6);}
    localStorage.setItem('mkt_device_id',id);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=157680000; SameSite=Lax';
    var side=q('#sidebarDeviceId'); if(side) side.textContent=id;
    var pay=q('#payDeviceId'); if(pay) pay.value=id;
    return id;
  }
  function keyFromText(txt){
    txt=String(txt||'').toLowerCase();
    if(txt.indexOf('1.959')>-1||txt.indexOf('1959')>-1||txt.indexOf('nhà bán')>-1||txt.indexOf('seller')>-1) return 'sellerpro';
    if(txt.indexOf('859')>-1||txt.indexOf('1 năm')>-1) return 'yearly';
    if(txt.indexOf('559')>-1||txt.indexOf('6 tháng')>-1) return 'halfyear';
    if(txt.indexOf('359')>-1||txt.indexOf('3 tháng')>-1) return 'quarterly';
    return 'monthly';
  }
  function plan(k){return PLAN_MAP[k]||PLAN_MAP.monthly}
  function contentFor(p){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'SDT';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'GMAIL';
    return deviceId();
  }
  window.showPaymentNotice=function(msg){
    var n=q('#paymentNotice');
    if(!n){
      var btn=q('#paymentModal .primary');
      n=document.createElement('div'); n.id='paymentNotice';
      if(btn && btn.parentNode) btn.parentNode.insertBefore(n, btn.nextSibling);
    }
    n.textContent=msg||'Đã gửi yêu cầu về web admin.'; n.style.display='block';
  };
  window.refreshPaymentContent=function(){
    var p=plan(window.currentPremiumPlanKey||'monthly');
    var content=contentFor(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    var notice=q('#paymentNotice'); if(notice) notice.style.display='none';
  };
  window.openPayment=function(k){
    var p=plan(k); window.currentPremiumPlanKey=k||'monthly';
    var modal=q('#paymentModal'); if(!modal) return false;
    var title=q('#payPlanTitle'); if(title) title.textContent=p.title;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
    var price=q('#payPlanPrice'); if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
    var device=q('#payDeviceId'); if(device) device.value=deviceId();
    var locked=q('#payLocked'); if(locked) locked.innerHTML='';
    var h=locked && locked.previousElementSibling; if(h) h.style.display='none';
    var alertBox=q('#paymentModal .payment-alert'); if(alertBox) alertBox.style.display='none';
    var z=q('#paymentModal .payment-actions a:not(.light)'); if(z){z.textContent='Zalo hỗ trợ: 036 338 2629'; z.href='https://zalo.me/0363382629'; z.target='_blank';}
    var done=q('#paymentModal .payment-actions .light'); if(done){done.textContent='Gửi yêu cầu admin duyệt'; done.removeAttribute('href'); done.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    var btn=q('#paymentModal .primary'); if(btn) btn.textContent='Gửi yêu cầu kích hoạt';
    window.refreshPaymentContent();
    modal.style.display='flex';
    setTimeout(function(){var ph=q('#payPhone'); if(ph) ph.focus();},150);
    cleanText(); return false;
  };
  var oldSubmit=window.submitPremiumRequest;
  window.submitPremiumRequest=function(){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    if(!phone || !email){ window.showPaymentNotice('Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.'); return false; }
    if(typeof oldSubmit==='function'){
      try{
        var r=oldSubmit();
        setTimeout(function(){window.showPaymentNotice('Đã gửi yêu cầu về web admin. Nội dung chuyển khoản đã kèm ID thiết bị và gói đã chọn.');},250);
        return r;
      }catch(e){}
    }
    window.showPaymentNotice('Đã gửi yêu cầu về web admin. Nội dung chuyển khoản đã kèm ID thiết bị và gói đã chọn.');
    return false;
  };
  function patchPlanOverlay(){
    var note=q('#mktPlanNote');
    if(note){
      var p=plan(window.currentPremiumPlanKey||keyFromText((q('#mktPlanOverlay')||document.body).innerText));
      note.innerHTML='ID thiết bị: <b>'+deviceId()+'</b><br><span style="color:#64748b">Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</span>';
    }
    var mini=q('.mktPlanMini'); if(mini) mini.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
    var pay=q('#mktPlanPay'); if(pay){pay.onclick=function(e){e.preventDefault(); e.stopPropagation(); var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(window.currentPremiumPlanKey||keyFromText((ov||document.body).innerText));};}
    cleanText();
  }
  var oldPlan=window.mktOpenPlanDetail;
  window.mktOpenPlanDetail=function(k){
    window.currentPremiumPlanKey=k||window.currentPremiumPlanKey||'monthly';
    var r=false;
    if(typeof oldPlan==='function'){r=oldPlan(k);} 
    setTimeout(patchPlanOverlay,50); setTimeout(patchPlanOverlay,200);
    return r;
  };
  document.addEventListener('click',function(e){
    var pay=e.target.closest && e.target.closest('#mktPlanPay');
    if(pay){e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation(); var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(window.currentPremiumPlanKey||keyFromText((ov||document.body).innerText));}
    var detail=e.target.closest && e.target.closest('button,a');
    if(detail && /xem chi tiết gói/i.test(detail.innerText||detail.textContent||'')){setTimeout(patchPlanOverlay,120);}
  },true);
  function init(){deviceId(); cleanText(); patchPlanOverlay();}
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',init); else init();
  setTimeout(init,500); setTimeout(init,1500); setInterval(cleanText,2500);
})();
</script>



<!-- FINAL FIX V76: close package modals + clean payment text -->
<style id="mkt-v76-modal-close-css">
  #mktPlanClose,.mktPlanClose,.payment-close,[data-close-plan],[data-close-payment]{cursor:pointer!important}
</style>
<script id="mkt-v76-modal-close-js">
(function(){
  'use strict';
  var CLEAN_TEXT='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
  var BAD_TEXTS=[
    'Bấm nâng cấp để nhập số điện thoại, Gmail và nhận nội dung chuyển khoản tự động.',
    'Bấm nâng cấp gói này để nhập số điện thoại, Gmail và xem nội dung chuyển khoản tự động.',
    'Vui lòng nhập số điện thoại và Gmail để admin đối chiếu nhanh.',
    'Gói sẽ mở sau khi admin duyệt thanh toán.',
    'Sau khi thanh toán, web admin sẽ duyệt theo ID thiết bị. Duyệt xong khách mới sử dụng được tính năng PRO.',
    'Sau khi thanh toán, web admin sẽ duyệt theo ID thiết bị.',
    'Duyệt xong khách mới sử dụng được tính năng PRO.'
  ];
  function q(s,r){return (r||document).querySelector(s)}
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s))}
  function cleanNodeText(){
    qa('body *').forEach(function(el){
      if(el.children && el.children.length) return;
      var t=el.textContent||'';
      var nt=t;
      BAD_TEXTS.forEach(function(b){ nt=nt.split(b).join(''); });
      if(t.indexOf('admin đối chiếu nhanh')>-1) nt=CLEAN_TEXT;
      if(nt!==t) el.textContent=nt.trim();
    });
    var mini=q('.mktPlanMini'); if(mini) mini.textContent=CLEAN_TEXT;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent=CLEAN_TEXT;
    var note=q('#mktPlanNote');
    if(note && note.innerHTML.indexOf(CLEAN_TEXT)===-1){
      var device=(window.mktGetDeviceId&&window.mktGetDeviceId()) || localStorage.getItem('mkt_device_id') || (document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/)||[])[1] || 'MKT-DEVICE';
      try{ device=decodeURIComponent(device); }catch(e){}
      note.innerHTML='ID thiết bị: <b>'+device+'</b><br><span style="color:#64748b">'+CLEAN_TEXT+'</span>';
    }
  }
  function closePlan(){
    ['#mktPlanOverlay','#packageModal','#priceModal','#planModal','.package-modal','.price-modal'].forEach(function(sel){
      qa(sel).forEach(function(el){ el.style.display='none'; el.classList.remove('show','active','open'); });
    });
    document.body.classList.remove('modal-open');
    return false;
  }
  function closePayment(){
    ['#paymentModal','#premiumPaymentModal','#lockedFeatureModal'].forEach(function(sel){
      qa(sel).forEach(function(el){ el.style.display='none'; el.classList.remove('show','active','open'); });
    });
    document.body.classList.remove('modal-open');
    return false;
  }
  window.mktClosePlanDetail=closePlan;
  window.closePackageModal=closePlan;
  window.closePlanModal=closePlan;
  window.closePriceModal=closePlan;
  window.closePayment=closePayment;
  window.closeLockedFeature=closePayment;
  document.addEventListener('click',function(e){
    var t=e.target;
    var btn=t.closest && t.closest('#mktPlanClose,.mktPlanClose,[data-close-plan],.package-close,.price-close');
    if(btn || ((t.id==='mktPlanOverlay'||t.id==='packageModal'||t.id==='priceModal'||t.id==='planModal') && t===e.target)){
      e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation(); return closePlan();
    }
    var pbtn=t.closest && t.closest('.payment-close,[data-close-payment],#paymentClose,#premiumPaymentClose');
    if(pbtn || ((t.id==='paymentModal'||t.id==='premiumPaymentModal'||t.id==='lockedFeatureModal') && t===e.target)){
      e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation(); return closePayment();
    }
  },true);
  document.addEventListener('keydown',function(e){ if(e.key==='Escape'){ closePlan(); closePayment(); } });
  var oldOpen=window.mktOpenPlanDetail;
  if(typeof oldOpen==='function'){
    window.mktOpenPlanDetail=function(k){ var r=oldOpen(k); setTimeout(cleanNodeText,30); setTimeout(cleanNodeText,160); return r; };
  }
  var oldPayment=window.openPayment;
  if(typeof oldPayment==='function'){
    window.openPayment=function(k){ var r=oldPayment(k); setTimeout(cleanNodeText,30); setTimeout(cleanNodeText,160); return r; };
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',cleanNodeText); else cleanNodeText();
  setTimeout(cleanNodeText,500); setTimeout(cleanNodeText,1500); setInterval(cleanNodeText,2500);
})();
</script>






<!-- PATCH V85: Premium Glass badges + sellerpro payment amount fix -->
<style id="mkt-v85-glass-badges-payment-fix">
  /* Badge cao cấp kiểu SaaS/Glass, bỏ glow neon */
  .v2-nav-link.pro-feature::after,
  .v2-nav-link.real-pro::after{
    content:"PRO"!important;
    position:absolute!important;right:14px!important;top:14px!important;z-index:20!important;
    padding:6px 12px!important;border-radius:999px!important;
    font-size:11px!important;font-weight:900!important;letter-spacing:.7px!important;
    color:#BBF7D0!important;
    background:linear-gradient(180deg,rgba(22,163,74,.22),rgba(21,128,61,.14))!important;
    border:1px solid rgba(134,239,172,.45)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.25),0 8px 22px rgba(0,0,0,.18)!important;
    backdrop-filter:blur(12px)!important;-webkit-backdrop-filter:blur(12px)!important;
    animation:none!important;text-shadow:none!important;
  }
  .v2-nav-link.pro-feature::before,
  .v2-nav-link.real-pro::before{
    content:""!important;position:absolute!important;right:63px!important;top:23px!important;width:7px!important;height:7px!important;border-radius:50%!important;
    background:#22C55E!important;box-shadow:0 0 8px rgba(34,197,94,.7)!important;z-index:21!important;
  }
  .v2-nav-link.premium-locked::after,
  .v2-nav-link.real-premium-lock::after{
    content:"Premium"!important;
    position:absolute!important;right:12px!important;top:14px!important;z-index:20!important;
    padding:6px 11px!important;border-radius:999px!important;
    font-size:11px!important;font-weight:900!important;letter-spacing:.35px!important;
    color:#FDE68A!important;
    background:linear-gradient(180deg,rgba(245,158,11,.20),rgba(146,64,14,.12))!important;
    border:1px solid rgba(253,230,138,.45)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.22),0 8px 22px rgba(0,0,0,.18)!important;
    backdrop-filter:blur(12px)!important;-webkit-backdrop-filter:blur(12px)!important;
    animation:none!important;text-shadow:none!important;
  }
  .v2-nav-link.premium-locked::before,
  .v2-nav-link.real-premium-lock::before{
    content:"👑"!important;position:absolute!important;right:78px!important;top:18px!important;font-size:12px!important;z-index:21!important;
    filter:none!important;text-shadow:none!important;
  }
  .v2-nav-link{padding-right:112px!important;}
</style>
<script id="mkt-v85-sellerpro-payment-fix">
(function(){
  'use strict';
  var PLANS={
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',code:'GOI1THANG',package:'GOI1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',code:'GOI3THANG',package:'GOI3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',code:'GOI6THANG',package:'GOI6THANG',desc:'Mở CRM và Sales Bot.'},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',code:'GOI1NAM',package:'GOI1NAM',desc:'Gói phổ biến nhất.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',code:'NHABANHANGPRO',package:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'},
    lifetime:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',code:'NHABANHANGPRO',package:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  function q(s,r){return (r||document).querySelector(s)}
  function txt(el){return String((el&&el.innerText)||'').toLowerCase()}
  function normKey(k){
    k=String(k||'').toLowerCase().trim();
    if(k==='lifetime') return 'sellerpro';
    if(PLANS[k]) return k;
    return '';
  }
  function keyFromText(t){
    t=String(t||'').toLowerCase();
    if(t.indexOf('1.959')>-1||t.indexOf('1959')>-1||t.indexOf('nhà bán')>-1||t.indexOf('nha ban')>-1||t.indexOf('chuyên nghiệp')>-1||t.indexOf('chuyen nghiep')>-1||t.indexOf('seller')>-1||t.indexOf('trọn đời')>-1||t.indexOf('tron doi')>-1||t.indexOf('vĩnh')>-1) return 'sellerpro';
    if(t.indexOf('859')>-1||t.indexOf('1 năm')>-1||t.indexOf('1 nam')>-1) return 'yearly';
    if(t.indexOf('559')>-1||t.indexOf('6 tháng')>-1||t.indexOf('6 thang')>-1) return 'halfyear';
    if(t.indexOf('359')>-1||t.indexOf('3 tháng')>-1||t.indexOf('3 thang')>-1) return 'quarterly';
    if(t.indexOf('159')>-1||t.indexOf('1 tháng')>-1||t.indexOf('1 thang')>-1) return 'monthly';
    return '';
  }
  function keyFromNode(node){
    var card=node&&node.closest&&node.closest('.premium-plan,.v4-plan,.plan-card,.pricing-card,.mkt-plan-card,.v4-lifetime,[class*="plan"],[class*="price"]');
    return keyFromText(txt(card)||txt(node));
  }
  function getDevice(){
    var id=localStorage.getItem('mkt_device_id');
    if(!id){var m=document.cookie.match(/(?:^|; )mkt_device_id=([^;]+)/);id=m?decodeURIComponent(m[1]):'';}
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-6);}
    localStorage.setItem('mkt_device_id',id);
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age=157680000; SameSite=Lax';
    var side=q('#sidebarDeviceId'); if(side) side.textContent=id;
    var pay=q('#payDeviceId'); if(pay) pay.value=id;
    return id;
  }
  function selectedKey(fallback){
    return normKey(window.mktSelectedPlanKey)||normKey(window.currentPremiumPlanKey)||normKey(fallback)||keyFromText((q('#mktPlanOverlay')||{}).innerText)||'monthly';
  }
  function setSelected(k){
    k=normKey(k)||'monthly';
    window.mktSelectedPlanKey=k;
    window.currentPremiumPlanKey=k;
    return k;
  }
  function plan(k){return PLANS[normKey(k)||'monthly']||PLANS.monthly}
  function payContent(p){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'SDT';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'GMAIL';
    return getDevice();
  }
  window.refreshPaymentContent=function(){
    var p=plan(selectedKey());
    var content=payContent(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    var price=q('#payPlanPrice'); if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
  };
  window.openPayment=function(k){
    k=setSelected(k||selectedKey());
    var p=plan(k);
    var modal=q('#paymentModal'); if(!modal) return false;
    var title=q('#payPlanTitle'); if(title) title.textContent=p.title;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
    var price=q('#payPlanPrice'); if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
    var device=q('#payDeviceId'); if(device) device.value=getDevice();
    var locked=q('#payLocked'); if(locked) locked.innerHTML='';
    var alert=q('#paymentModal .payment-alert'); if(alert) alert.style.display='none';
    var z=q('#paymentModal .payment-actions a:not(.light)'); if(z){z.textContent='Zalo hỗ trợ: 036 338 2629'; z.href='https://zalo.me/0363382629'; z.target='_blank';}
    var done=q('#paymentModal .payment-actions .light'); if(done){done.textContent='Gửi yêu cầu admin duyệt'; done.removeAttribute('href'); done.onclick=function(e){e.preventDefault();return window.submitPremiumRequest();};}
    var primary=q('#paymentModal .primary'); if(primary){primary.textContent='Gửi yêu cầu kích hoạt'; primary.onclick=function(e){e.preventDefault();return window.submitPremiumRequest();};}
    window.refreshPaymentContent();
    modal.style.display='flex';
    setTimeout(function(){var ph=q('#payPhone'); if(ph) ph.focus();},120);
    return false;
  };
  window.submitPremiumRequest=function(){
    var k=selectedKey(); setSelected(k);
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    if(!phone||!email){
      if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.');
      else alert('Vui lòng nhập số điện thoại và Gmail.');
      return false;
    }
    var p=plan(k);
    window.refreshPaymentContent();
    fetch('/premium_request',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:getDevice(),phone:phone,email:email,package_key:k})})
      .then(function(r){return r.json();}).then(function(data){
        if(data&&data.ok){
          var pc=q('#payContent'); if(pc) pc.textContent=data.payment_note||payContent(p);
          if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Đã gửi yêu cầu '+p.title+' về web admin. Số tiền: '+p.price+'.');
        }else{
          if(typeof window.showPaymentNotice==='function') window.showPaymentNotice((data&&data.message)||'Chưa gửi được yêu cầu, vui lòng thử lại.');
          else alert((data&&data.message)||'Chưa gửi được yêu cầu, vui lòng thử lại.');
        }
      }).catch(function(){
        if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Kết nối chậm, vui lòng thử lại hoặc gửi Zalo hỗ trợ.');
      });
    return false;
  };
  var oldDetail=window.mktOpenPlanDetail;
  window.mktOpenPlanDetail=function(k){
    k=normKey(k)||selectedKey(); setSelected(k);
    var r=false;
    if(typeof oldDetail==='function') r=oldDetail(k);
    setTimeout(function(){
      var p=plan(selectedKey(k));
      var note=q('#mktPlanNote');
      if(note){note.innerHTML='ID thiết bị: <b>'+getDevice()+'</b><br>Gói đang chọn: <b>'+p.title+'</b><br><span style="color:#64748b">Số tiền: <b>'+p.price+'</b>. Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</span>';}
      var mini=q('.mktPlanMini'); if(mini) mini.textContent='Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.';
      var pay=q('#mktPlanPay'); if(pay){pay.onclick=function(e){e.preventDefault();e.stopPropagation();var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(selectedKey(k));};}
    },80);
    return r;
  };
  document.addEventListener('click',function(e){
    var b=e.target.closest&&e.target.closest('button,a');
    if(!b) return;
    var text=String(b.innerText||b.textContent||'');
    var k=keyFromNode(b);
    if(k) setSelected(k);
    if(/xem chi tiết gói/i.test(text)){
      setTimeout(function(){ if(k) setSelected(k); },30);
      setTimeout(function(){ if(window.mktOpenPlanDetail && k) window.mktOpenPlanDetail(k); },80);
    }
    if(b.id==='mktPlanPay'||/nâng cấp|mở khóa|thanh toán|kích hoạt/i.test(text)){
      if(k) setSelected(k);
      if(b.id==='mktPlanPay'){
        e.preventDefault();e.stopPropagation();if(e.stopImmediatePropagation)e.stopImmediatePropagation();
        var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none';
        return window.openPayment(selectedKey(k));
      }
    }
  },true);
  document.addEventListener('input',function(e){if(e.target&&(/payPhone|payEmail/.test(e.target.id||''))) window.refreshPaymentContent();},true);
  getDevice();
})();
<script id="mkt-price-detail-final-js">
(function(){
  'use strict';

  var plans={
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',desc:'Phù hợp người mới bắt đầu, shop nhỏ cần đăng bài, tạo content và quản lý Fanpage cơ bản.',benefits:['Đăng bài Facebook','Quản lý Fanpage','Quản lý Group','AI Comment','Tạo content cơ bản','Token Manager','Hỗ trợ kích hoạt theo ID thiết bị']},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',desc:'Tối ưu cho shop đang bán hàng cần dùng ổn định hơn và tiết kiệm hơn gói tháng.',benefits:['Toàn bộ gói 1 tháng','AI Messenger','CRM Kanban cơ bản','Kịch bản inbox','Lịch đăng nâng cao','Báo cáo cơ bản','Ưu tiên hỗ trợ']},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',desc:'Phù hợp shop cần CRM, chăm sóc khách và tối ưu quy trình bán hàng.',benefits:['Toàn bộ gói 3 tháng','CRM Pro','AI Sales Bot','Comment Manager','Auto Tag khách hàng','Quản lý khách hàng','Chuyển khách sang CRM']},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',desc:'Gói phổ biến nhất cho nhà bán hàng muốn dùng đầy đủ công cụ AI Marketing trong 1 năm.',benefits:['Toàn bộ gói 6 tháng','AI Marketing Director','AI Ads Chuyên Gia','Kho Content Premium','Automation Marketing','Export báo cáo','Ưu tiên xử lý']},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp, mở toàn bộ hệ thống sau khi admin duyệt.',benefits:['Toàn bộ tính năng Premium','AI Image Center','AI Video Center','AI Voice Studio','Dashboard Enterprise','Export PDF / Excel','Backup Database','Ưu tiên hỗ trợ VIP']}
  };

  plans.lifetime = plans.sellerpro;

  function norm(t){
    return String(t||'')
      .toLowerCase()
      .replace(/\s+/g,' ')
      .replace(/vnđ|vnd|đ/g,'')
      .trim();
  }

  function keyFromText(t){
    t = norm(t);

    if(t.includes('1959000') || t.includes('1.959') || t.includes('1959') || t.includes('nhà bán') || t.includes('seller') || t.includes('chuyên nghiệp') || t.includes('trọn đời')) return 'sellerpro';
    if(t.includes('859000') || t.includes('859') || t.includes('1 năm') || t.includes('12 tháng')) return 'yearly';
    if(t.includes('559000') || t.includes('559') || t.includes('6 tháng')) return 'halfyear';
    if(t.includes('359000') || t.includes('359') || t.includes('3 tháng')) return 'quarterly';
    if(t.includes('159000') || t.includes('159') || t.includes('1 tháng')) return 'monthly';

    return '';
  }

  function findPlanKey(btn){
    var key =
      btn.getAttribute('data-plan') ||
      btn.getAttribute('data-package') ||
      btn.getAttribute('data-key') ||
      btn.dataset.plan ||
      btn.dataset.package ||
      btn.dataset.key ||
      '';

    key = norm(key);

    if(plans[key]) return key;

    var card = btn.closest('.pricing-card,.price-card,.plan-card,.premium-card,.package-card,.mkt-plan-card,.v2-price-card,div');
    var text = (btn.innerText || '') + ' ' + (card ? card.innerText : '') + ' ' + document.body.innerText;

    return keyFromText(text) || 'monthly';
  }

  document.addEventListener('click', function(e){
    var btn = e.target.closest('button,a');
    if(!btn) return;

    var text = norm(btn.innerText || btn.textContent || '');

    if(
      text.includes('nâng cấp') ||
      text.includes('thanh toán') ||
      text.includes('chọn gói') ||
      text.includes('mua gói')
    ){
      var key = findPlanKey(btn);
      window.MKT_SELECTED_PLAN_KEY = key;
      localStorage.setItem('mkt_selected_plan_key', key);
    }
  }, true);

})();
</script>  function payContent(p){
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    return deviceId();
  }
  function fill(k){
    k=setPlan(k); var p=getPlan(k);
    var t=q('#payPlanTitle'); if(t) t.textContent=p.title;
    var d=q('#payPlanDesc'); if(d) d.textContent=p.desc;
    var pr=q('#payPlanPrice'); if(pr) pr.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
    var dev=q('#payDeviceId'); if(dev) dev.value=deviceId();
    var content=payContent(p);
    var pc=q('#payContent'); if(pc) pc.textContent=content;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(content)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    var lock=q('#payLocked'); if(lock) lock.innerHTML='';
    return p;
  }
  window.refreshPaymentContent=function(){fill(window.mktPaymentPlanKey||window.mktSelectedPlanKey||window.currentPremiumPlanKey||'monthly')};
  window.openPayment=function(planKey){
    var k=normKey(planKey)||normKey(window.mktPaymentPlanKey)||normKey(window.mktSelectedPlanKey)||normKey(window.currentPremiumPlanKey)||'monthly';
    fill(k);
    var modal=q('#paymentModal'); if(modal) modal.style.display='flex';
    var primary=q('#paymentModal .primary'); if(primary){primary.textContent='Gửi yêu cầu kích hoạt'; primary.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    var light=q('#paymentModal .payment-actions .light'); if(light){light.textContent='Gửi yêu cầu admin duyệt'; light.removeAttribute('href'); light.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    return false;
  };
  var previousDetail=window.mktOpenPlanDetail;
  window.mktOpenPlanDetail=function(k){
    k=setPlan(k||window.mktPaymentPlanKey||'monthly');
    var ret=false; if(typeof previousDetail==='function'){try{ret=previousDetail(k)}catch(e){}}
    setTimeout(function(){
      setPlan(k); var p=getPlan(k);
      var note=q('#mktPlanNote');
      if(note) note.innerHTML='ID thiết bị: <b>'+deviceId()+'</b><br>Gói đang chọn: <b>'+p.title+'</b><br><span style="color:#64748b">Số tiền: <b>'+p.price+'</b>. Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.</span>';
      var mini=q('.mktPlanMini'); if(mini) mini.textContent='Gói đang chọn: '+p.title+' - '+p.price;
      var pay=q('#mktPlanPay'); if(pay){pay.setAttribute('data-plan',k); pay.onclick=function(e){e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation(); var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none'; return window.openPayment(k);};}
    },100);
    return ret;
  };
  window.submitPremiumRequest=function(){
    var k=setPlan(window.mktPaymentPlanKey||window.mktSelectedPlanKey||window.currentPremiumPlanKey||'monthly');
    var p=fill(k);
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    if(!phone||!email){ if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Vui lòng nhập số điện thoại và Gmail để được duyệt nâng cấp tự động lên gói Premium sau khi thanh toán xong.'); else alert('Vui lòng nhập số điện thoại và Gmail.'); return false; }
    fetch('/premium_request',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:deviceId(),phone:phone,email:email,package_key:k})})
      .then(function(r){return r.json()}).then(function(data){
        if(data&&data.payment_note){var pc=q('#payContent'); if(pc) pc.textContent=data.payment_note;}
        var msg=(data&&data.ok)?('Đã gửi yêu cầu '+p.title+' về web admin. Số tiền: '+p.price+'.'):((data&&data.message)||'Chưa gửi được yêu cầu, vui lòng thử lại.');
        if(typeof window.showPaymentNotice==='function') window.showPaymentNotice(msg); else alert(msg);
      }).catch(function(){if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Kết nối chậm, vui lòng thử lại hoặc gửi Zalo hỗ trợ.');});
    return false;
  };
  function bindCards(){
    qa('.premium-plan,.v4-plan,.price-card,.pricing-card,.plan-card').forEach(function(card){
      var k=keyFromText(card.innerText||card.textContent||'');
      if(!k) return;
      card.setAttribute('data-plan',k);
      qa('button,a',card).forEach(function(b){b.setAttribute('data-plan',k);});
    });
  }
  document.addEventListener('click',function(e){
    var b=e.target.closest&&e.target.closest('button,a'); if(!b) return;
    var k=keyFromNode(b);
    if(k) setPlan(k);
    var text=norm(b.innerText||b.textContent||'');
    var isPay=b.id==='mktPlanPay'||/nang cap|mo khoa|dang ky|thanh toan|kich hoat|da thanh toan|gui yeu cau/.test(text)||/openPayment\(/i.test(String(b.getAttribute('onclick')||''));
    if(b.id==='mktPlanPay'){
      e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation();
      var ov=q('#mktPlanOverlay'); if(ov) ov.style.display='none';
      return window.openPayment(k||window.mktPaymentPlanKey||window.mktSelectedPlanKey||'monthly');
    }
    if(isPay && k){
      e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation();
      return window.openPayment(k);
    }
  },true);
  document.addEventListener('input',function(e){if(e.target&&(/payPhone|payEmail/.test(e.target.id||''))) window.refreshPaymentContent();},true);
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',bindCards); else bindCards();
  setTimeout(bindCards,300); setTimeout(bindCards,1000); setTimeout(bindCards,2500);
  deviceId();
})();
</script>

<script id="mkt-compact-pricing-final-fix">
(function(){
  'use strict';
  var PLANS={
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',code:'GOI1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',code:'GOI3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',code:'GOI6THANG',desc:'Mở thêm CRM, Sales Bot và công cụ nâng cao.'},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',code:'GOI1NAM',desc:'Gói phổ biến, tối ưu chi phí dài hạn.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',code:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  PLANS.lifetime=PLANS.sellerpro;
  function q(s,r){return (r||document).querySelector(s)}
  function norm(k){k=String(k||'').toLowerCase().replace(/[^a-z0-9]/g,''); if(k==='lifetime'||k==='seller'||k==='sellerpro'||k==='nhabanhangpro')return 'sellerpro'; if(k==='monthly'||k==='basic'||k==='goi1thang')return 'monthly'; if(k==='quarterly'||k==='pro'||k==='goi3thang')return 'quarterly'; if(k==='halfyear'||k==='business'||k==='goi6thang')return 'halfyear'; if(k==='yearly'||k==='goi1nam')return 'yearly'; return PLANS[k]?k:'monthly'}
  function device(){var id='';try{id=localStorage.getItem('mkt_device_id')||''}catch(e){} if(!id&&typeof window.getOrCreateDeviceId==='function')id=window.getOrCreateDeviceId(); if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);try{localStorage.setItem('mkt_device_id',id)}catch(e){}} document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age='+(60*60*24*365*5); return id}
  function setKey(k){k=norm(k); window.currentPremiumPlanKey=k; window.mktSelectedPlanKey=k; window.__mktClickedPlanKey=k; try{localStorage.setItem('mkt_selected_plan_key',k)}catch(e){} return k}
  function payNote(p){var phone=(q('#payPhone')&&q('#payPhone').value.trim())||''; var email=(q('#payEmail')&&q('#payEmail').value.trim())||''; return device();}
  window.refreshPaymentContent=function(){var k=norm(window.currentPremiumPlanKey||window.mktSelectedPlanKey); var p=PLANS[k]||PLANS.monthly; var note=payNote(p); var pc=q('#payContent'); if(pc)pc.textContent=note; var price=q('#payPlanPrice'); if(price)price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ'; var qr=q('#payQr'); if(qr)qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(note)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN')};
  window.openPayment=function(k){k=setKey(k); var p=PLANS[k]||PLANS.monthly; var modal=q('#paymentModal'); if(!modal)return false; var title=q('#payPlanTitle'); if(title)title.textContent=p.title; var desc=q('#payPlanDesc'); if(desc)desc.textContent=p.desc; var price=q('#payPlanPrice'); if(price)price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ'; var dev=q('#payDeviceId'); if(dev)dev.value=device(); window.refreshPaymentContent(); modal.style.display='flex'; return false};
  document.addEventListener('click',function(e){var btn=e.target.closest&&e.target.closest('[data-plan]'); if(!btn)return; var k=norm(btn.getAttribute('data-plan')||(btn.dataset&&btn.dataset.plan)); setKey(k); if(/button|a/i.test(btn.tagName)||btn.closest('button,a')){e.preventDefault();e.stopPropagation();if(e.stopImmediatePropagation)e.stopImmediatePropagation();return window.openPayment(k)}},true);
  document.addEventListener('input',function(e){if(e.target&&/payPhone|payEmail/.test(e.target.id||''))window.refreshPaymentContent()},true);
})();
</script>



<!-- FINAL MENU PRO UPGRADE: sang hơn, gọn hơn, không đổi logic menu -->
<style id="mkt-final-menu-pro-style">
  .sidebar{
    background:
      radial-gradient(circle at 20% 0%,rgba(59,130,246,.38),transparent 34%),
      radial-gradient(circle at 100% 20%,rgba(250,204,21,.20),transparent 28%),
      linear-gradient(180deg,#07111f 0%,#0f172a 52%,#111827 100%)!important;
    border:1px solid rgba(148,163,184,.24)!important;
    box-shadow:0 26px 80px rgba(2,6,23,.42), inset 0 1px 0 rgba(255,255,255,.08)!important;
    padding:18px!important;
  }
  .sidebar .logo{
    padding:16px 16px 14px!important;
    border-radius:22px!important;
    background:linear-gradient(135deg,rgba(255,255,255,.12),rgba(255,255,255,.04))!important;
    border:1px solid rgba(255,255,255,.14)!important;
    box-shadow:0 16px 36px rgba(0,0,0,.18)!important;
  }
  .sidebar .logo:after{
    width:100%!important;height:1px!important;margin-top:14px!important;
    background:linear-gradient(90deg,transparent,#60a5fa,#facc15,transparent)!important;
    border:0!important;
  }
  .v2-nav-title{
    margin:18px 0 9px!important;
    padding-left:10px!important;
    color:#93c5fd!important;
    letter-spacing:.12em!important;
    font-size:10px!important;
    position:relative!important;
  }
  .v2-nav-title:before{
    content:"";display:inline-block;width:7px;height:7px;border-radius:99px;
    background:#22c55e;box-shadow:0 0 16px rgba(34,197,94,.95);margin-right:8px;
  }
  .v2-nav-link{
    min-height:56px!important;
    padding:13px 14px 13px 16px!important;
    border-radius:18px!important;
    background:linear-gradient(135deg,rgba(255,255,255,.105),rgba(255,255,255,.035))!important;
    border:1px solid rgba(148,163,184,.16)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.06),0 8px 22px rgba(0,0,0,.12)!important;
    color:#f8fafc!important;
    position:relative!important;
    overflow:hidden!important;
  }
  .v2-nav-link:before{
    content:"";position:absolute;left:0;top:12px;bottom:12px;width:3px;border-radius:999px;
    background:linear-gradient(180deg,#38bdf8,#818cf8);opacity:.45;
  }
  .v2-nav-link:hover,.v2-nav-link.active{
    background:linear-gradient(135deg,rgba(37,99,235,.95),rgba(124,58,237,.88))!important;
    border-color:rgba(255,255,255,.24)!important;
    transform:translateX(4px)!important;
    box-shadow:0 16px 38px rgba(37,99,235,.28)!important;
  }
  .v2-nav-text{font-weight:900!important;letter-spacing:-.01em!important;line-height:1.22!important;}
  .v2-nav-tag{
    padding:5px 9px!important;
    border-radius:999px!important;
    font-size:10px!important;
    font-weight:1000!important;
    color:#0f172a!important;
    background:linear-gradient(135deg,#f8fafc,#c7d2fe)!important;
    border:1px solid rgba(255,255,255,.55)!important;
    box-shadow:0 0 20px rgba(129,140,248,.22)!important;
  }
  .v2-nav-link .v2-nav-tag:empty{display:none!important;}
  .v2-nav-link.premium-locked .v2-nav-tag,
  .v2-nav-link.real-premium-lock .v2-nav-tag,
  .v2-nav-link[href="#messenger_ai"] .v2-nav-tag,
  .v2-nav-link[href="#crm_sales"] .v2-nav-tag,
  .v2-nav-link[href="#marketing_director"] .v2-nav-tag{
    background:linear-gradient(135deg,#facc15,#fb923c)!important;
    color:#111827!important;
    box-shadow:0 0 24px rgba(250,204,21,.55)!important;
  }
  #sidebarDeviceId{
    padding:7px 10px!important;border-radius:12px!important;
    background:rgba(15,23,42,.52)!important;border:1px solid rgba(96,165,250,.22)!important;
    color:#bfdbfe!important;
  }
  #sidebarPremiumStatus{
    display:inline-flex!important;margin-top:8px!important;padding:7px 10px!important;border-radius:999px!important;
    background:rgba(34,197,94,.13)!important;border:1px solid rgba(34,197,94,.24)!important;color:#bbf7d0!important;
    font-weight:900!important;
  }
</style>


<!-- FINAL FIX V96: clean transfer note + auto open premium menus -->
<style id="mkt-v96-clean-pay-menu-pro">
  /* Ẩn mọi dòng hướng dẫn kỹ thuật dư trong khung giá nếu bản cũ còn sót */
  .pricing-note,
  .premium-pricing-compact .pricing-tip,
  .premium-pricing-compact .mkt-plan-footnote,
  .premium-note-small{
    display:none!important;
  }

  /* Nâng cấp menu trái cao cấp hơn, áp mạnh lên cả bản cũ */
  .sidebar,.v2-sidebar{
    background:
      radial-gradient(circle at 18% 0%,rgba(37,99,235,.42),transparent 34%),
      radial-gradient(circle at 100% 24%,rgba(250,204,21,.18),transparent 30%),
      linear-gradient(180deg,#050b18 0%,#0b1220 48%,#111827 100%)!important;
    border-right:1px solid rgba(148,163,184,.25)!important;
    box-shadow:18px 0 70px rgba(2,6,23,.38), inset -1px 0 0 rgba(255,255,255,.05)!important;
  }
  .v2-nav-link{
    margin:7px 10px!important;
    min-height:58px!important;
    border-radius:18px!important;
    color:#f8fafc!important;
    background:linear-gradient(135deg,rgba(255,255,255,.11),rgba(255,255,255,.035))!important;
    border:1px solid rgba(148,163,184,.18)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.08),0 10px 26px rgba(0,0,0,.16)!important;
    transition:.22s ease!important;
    position:relative!important;
    overflow:hidden!important;
  }
  .v2-nav-link:before{
    content:""!important;
    position:absolute!important;
    left:0!important;top:13px!important;bottom:13px!important;width:3px!important;
    border-radius:999px!important;
    background:linear-gradient(180deg,#38bdf8,#818cf8,#facc15)!important;
    opacity:.65!important;
  }
  .v2-nav-link:hover,.v2-nav-link.active{
    transform:translateX(5px)!important;
    background:linear-gradient(135deg,rgba(37,99,235,.92),rgba(124,58,237,.82))!important;
    border-color:rgba(255,255,255,.30)!important;
    box-shadow:0 16px 42px rgba(37,99,235,.28)!important;
  }
  .v2-nav-text{font-weight:900!important;letter-spacing:-.01em!important;line-height:1.22!important;}
  .v2-nav-tag,
  .v2-nav-link.premium-locked::after,
  .v2-nav-link.real-premium-lock::after{
    content:"PRO"!important;
    display:inline-flex!important;
    align-items:center!important;
    justify-content:center!important;
    min-width:48px!important;
    height:24px!important;
    padding:0 10px!important;
    border-radius:999px!important;
    background:linear-gradient(135deg,#facc15,#fb923c)!important;
    color:#111827!important;
    font-size:10px!important;
    font-weight:1000!important;
    letter-spacing:.03em!important;
    border:1px solid rgba(255,255,255,.45)!important;
    box-shadow:0 0 22px rgba(250,204,21,.55)!important;
    position:absolute!important;
    right:14px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    z-index:5!important;
  }
  .v2-nav-link{padding-right:82px!important;}
  .v2-nav-title{
    color:#93c5fd!important;
    font-weight:1000!important;
    letter-spacing:.13em!important;
    font-size:10px!important;
    margin:18px 14px 8px!important;
  }
</style>
<script id="mkt-v96-clean-pay-menu-pro">
(function(){
  'use strict';
  var PLANS={
    monthly:{title:'Gói 1 tháng',price:'159.000đ',amount:'159000',code:'GOI1THANG',desc:'Phù hợp người mới bắt đầu.'},
    quarterly:{title:'Gói 3 tháng',price:'359.000đ',amount:'359000',code:'GOI3THANG',desc:'Tối ưu hơn gói tháng.'},
    halfyear:{title:'Gói 6 tháng',price:'559.000đ',amount:'559000',code:'GOI6THANG',desc:'Mở thêm CRM, Sales Bot và công cụ nâng cao.'},
    yearly:{title:'Gói 1 năm',price:'859.000đ',amount:'859000',code:'GOI1NAM',desc:'Gói phổ biến, tối ưu chi phí dài hạn.'},
    sellerpro:{title:'Gói nhà bán hàng chuyên nghiệp',price:'1.959.000đ',amount:'1959000',code:'NHABANHANGPRO',desc:'Gói cao nhất cho nhà bán hàng chuyên nghiệp.'}
  };
  PLANS.lifetime=PLANS.sellerpro;
  function q(s,r){return (r||document).querySelector(s)}
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s))}
  function norm(k){
    k=String(k||'').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/đ/g,'d').replace(/[^a-z0-9]/g,'');
    if(k==='lifetime'||k==='seller'||k==='sellerpro'||k==='sellerpro'||k==='nhabanhangpro'||k==='nhabanhangchuyennghiep')return 'sellerpro';
    if(k==='monthly'||k==='basic'||k==='goi1thang'||k==='1thang')return 'monthly';
    if(k==='quarterly'||k==='pro'||k==='goi3thang'||k==='3thang')return 'quarterly';
    if(k==='halfyear'||k==='business'||k==='goi6thang'||k==='6thang')return 'halfyear';
    if(k==='yearly'||k==='goi1nam'||k==='1nam')return 'yearly';
    return PLANS[k]?k:'monthly';
  }
  function device(){
    var id='';
    try{id=localStorage.getItem('mkt_device_id')||''}catch(e){}
    if(!id&&typeof window.getOrCreateDeviceId==='function') id=window.getOrCreateDeviceId();
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);try{localStorage.setItem('mkt_device_id',id)}catch(e){}}
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age='+(60*60*24*365*5);
    return id;
  }
  function setKey(k){k=norm(k); window.currentPremiumPlanKey=k; window.mktSelectedPlanKey=k; window.__mktClickedPlanKey=k; try{localStorage.setItem('mkt_selected_plan_key',k)}catch(e){} return k}
  function cleanNote(){return device();}
  function fillPayment(k){
    k=setKey(k||window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly');
    var p=PLANS[k]||PLANS.monthly;
    var note=cleanNote();
    var title=q('#payPlanTitle'); if(title) title.textContent=p.title;
    var desc=q('#payPlanDesc'); if(desc) desc.textContent=p.desc;
    var price=q('#payPlanPrice'); if(price) price.textContent=Number(p.amount).toLocaleString('vi-VN')+' VNĐ';
    var dev=q('#payDeviceId'); if(dev) dev.value=device();
    var pc=q('#payContent'); if(pc) pc.textContent=note;
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(p.amount)+'&addInfo='+encodeURIComponent(note)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    return p;
  }
  window.refreshPaymentContent=function(){fillPayment(window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly')};
  window.openPayment=function(k){
    fillPayment(k);
    var modal=q('#paymentModal'); if(modal) modal.style.display='flex';
    var primary=q('#paymentModal .primary'); if(primary){primary.textContent='Gửi yêu cầu kích hoạt'; primary.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    var light=q('#paymentModal .payment-actions .light'); if(light){light.textContent='Gửi yêu cầu admin duyệt'; light.removeAttribute('href'); light.onclick=function(e){e.preventDefault(); return window.submitPremiumRequest();};}
    return false;
  };
  window.submitPremiumRequest=function(){
    var k=setKey(window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly');
    var p=fillPayment(k);
    var phone=(q('#payPhone')&&q('#payPhone').value.trim())||'';
    var email=(q('#payEmail')&&q('#payEmail').value.trim())||'';
    if(!phone||!email){ if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Vui lòng nhập số điện thoại và Gmail để admin duyệt kích hoạt.'); else alert('Vui lòng nhập số điện thoại và Gmail.'); return false; }
    fetch('/premium_request',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({device_id:device(),phone:phone,email:email,package_key:k})})
      .then(function(r){return r.json()}).then(function(data){
        var pc=q('#payContent'); if(pc) pc.textContent=device();
        var msg=(data&&data.ok)?('Đã gửi yêu cầu '+p.title+' về web admin. Số tiền: '+p.price+'.'):((data&&data.message)||'Chưa gửi được yêu cầu, vui lòng thử lại.');
        if(typeof window.showPaymentNotice==='function') window.showPaymentNotice(msg); else alert(msg);
      }).catch(function(){if(typeof window.showPaymentNotice==='function') window.showPaymentNotice('Kết nối chậm, vui lòng thử lại hoặc gửi Zalo hỗ trợ.');});
    return false;
  };
  document.addEventListener('click',function(e){
    var planBtn=e.target.closest&&e.target.closest('[data-plan]');
    if(planBtn){
      var k=setKey(planBtn.getAttribute('data-plan')||(planBtn.dataset&&planBtn.dataset.plan)||'monthly');
      var txt=(planBtn.innerText||planBtn.textContent||'').toLowerCase();
      if(/nâng cấp|thanh toán|kích hoạt|đã thanh toán|gui|gửi/.test(txt) || planBtn.closest('button,a')){
        e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation();
        return window.openPayment(k);
      }
    }
  },true);
  document.addEventListener('input',function(e){if(e.target&&/payPhone|payEmail/.test(e.target.id||''))fillPayment(window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly')},true);

  function showPricingFromMenu(){
    var target=q('.premium-pricing-compact')||q('#premium')||q('#pricing')||q('.pricing-head')||q('[data-section="premium"]');
    if(target){
      var wrap=target.closest('.section,.page-section,.content-section')||target;
      try{wrap.style.display='block'}catch(e){}
      target.scrollIntoView({behavior:'smooth',block:'center'});
    }else if(typeof window.openPremiumPopup==='function'){
      window.openPremiumPopup();
    }
    setTimeout(function(){
      var first=q('.premium-pricing-compact .price-card,.pricing-card,.price-card');
      if(first){first.style.boxShadow='0 0 0 4px rgba(250,204,21,.22),0 24px 60px rgba(37,99,235,.22)'; setTimeout(function(){first.style.boxShadow=''},1500)}
    },350);
    return false;
  }
  function isProMenu(a){
    if(!a || !a.classList) return false;
    if(a.classList.contains('premium-locked')||a.classList.contains('real-premium-lock')) return true;
    var tag=a.querySelector('.v2-nav-tag');
    var text=((tag&&tag.textContent)||a.textContent||'').toLowerCase();
    return /pro|premium/.test(text);
  }
  document.addEventListener('click',function(e){
    var a=e.target.closest&&e.target.closest('.v2-nav-link,a[href^="#"]');
    if(!a || !isProMenu(a)) return;
    e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation)e.stopImmediatePropagation();
    return showPricingFromMenu();
  },true);
  function upgradeMenuLook(){
    qa('.v2-nav-link').forEach(function(a){
      var text=(a.textContent||'').toLowerCase();
      if(/pro|premium/.test(text) || a.classList.contains('premium-locked') || a.classList.contains('real-premium-lock')){
        a.classList.add('premium-locked');
        if(!a.querySelector('.v2-nav-tag')){var sp=document.createElement('span'); sp.className='v2-nav-tag'; sp.textContent='PRO'; a.appendChild(sp);}
      }
    });
  }
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',upgradeMenuLook); else upgradeMenuLook();
  setTimeout(upgradeMenuLook,500); setTimeout(upgradeMenuLook,1500);
  fillPayment(window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly');
})();
</script>


<!-- FINAL FIX V97: remove transfer placeholders + premium menu micro badge -->
<style id="mkt-v97-menu-micro-pro-css">
  #payContent{white-space:normal!important;word-break:break-word!important;}
  .v2-nav-link.premium-locked,
  .v2-nav-link.real-premium-lock{
    position:relative!important;
    overflow:hidden!important;
    background:linear-gradient(135deg,rgba(17,24,39,.94),rgba(49,46,129,.72))!important;
    border:1px solid rgba(129,140,248,.28)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.08),0 12px 30px rgba(15,23,42,.28)!important;
    padding-right:76px!important;
  }
  .v2-nav-link.premium-locked:hover,
  .v2-nav-link.real-premium-lock:hover{
    background:linear-gradient(135deg,rgba(37,99,235,.95),rgba(124,58,237,.85))!important;
    transform:translateX(5px)!important;
    box-shadow:0 18px 45px rgba(37,99,235,.32)!important;
  }
  .v2-nav-link.premium-locked .v2-nav-tag,
  .v2-nav-link.real-premium-lock .v2-nav-tag,
  .v2-nav-link.premium-locked::after,
  .v2-nav-link.real-premium-lock::after{
    content:"PRO"!important;
    min-width:34px!important;
    height:18px!important;
    padding:0 7px 0 17px!important;
    border-radius:999px!important;
    font-size:8px!important;
    line-height:18px!important;
    font-weight:1000!important;
    letter-spacing:.04em!important;
    color:#111827!important;
    background:linear-gradient(135deg,#fde047,#fb923c)!important;
    border:1px solid rgba(255,255,255,.55)!important;
    box-shadow:0 0 16px rgba(250,204,21,.55)!important;
    position:absolute!important;
    right:12px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    z-index:9!important;
  }
  .v2-nav-link.premium-locked .v2-nav-tag::before,
  .v2-nav-link.real-premium-lock .v2-nav-tag::before{
    content:""!important;
    width:7px!important;
    height:7px!important;
    border-radius:50%!important;
    background:#22c55e!important;
    box-shadow:0 0 0 4px rgba(34,197,94,.20),0 0 15px rgba(34,197,94,.95)!important;
    position:absolute!important;
    left:6px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
  }
</style>
<script id="mkt-v97-clean-transfer-note-final">
(function(){
  'use strict';
  function q(s,r){return (r||document).querySelector(s)}
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s))}
  function device(){
    var id='';
    try{id=localStorage.getItem('mkt_device_id')||''}catch(e){}
    if(!id && typeof window.getOrCreateDeviceId==='function') id=window.getOrCreateDeviceId();
    if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);try{localStorage.setItem('mkt_device_id',id)}catch(e){}}
    document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age='+(60*60*24*365*5);
    return id;
  }
  function amountOf(k){
    k=String(k||window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly').toLowerCase();
    if(/seller|lifetime|nhabanhang/.test(k)) return '1959000';
    if(/year|1nam/.test(k)) return '859000';
    if(/half|6thang/.test(k)) return '559000';
    if(/quarter|3thang/.test(k)) return '359000';
    return '159000';
  }
  function cleanTransfer(){
    var id=device();
    var pc=q('#payContent');
    if(pc && pc.textContent!==id) pc.textContent=id;
    var dev=q('#payDeviceId'); if(dev) dev.value=id;
    var qr=q('#payQr');
    if(qr){
      var k=window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly';
      qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(amountOf(k))+'&addInfo='+encodeURIComponent(id)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
    }
    qa('span,div,p,b').forEach(function(el){
      if(el.id==='payContent') return;
      var t=el.childNodes&&el.childNodes.length===1?el.textContent:'';
      if(/CHUA_SDT|CHUA_GMAIL|GOI1THANG|GOI3THANG|GOI6THANG|GOI1NAM|NHABANHANGPRO/.test(t||'')){
        el.textContent=(t||'').replace(/\s*\|\s*CHUA_SDT\s*\|\s*CHUA_GMAIL\s*\|\s*(GOI1THANG|GOI3THANG|GOI6THANG|GOI1NAM|NHABANHANGPRO)/g,'').replace(/CHUA_SDT|CHUA_GMAIL|GOI1THANG|GOI3THANG|GOI6THANG|GOI1NAM|NHABANHANGPRO/g,'').trim();
      }
    });
  }
  var oldRefresh=window.refreshPaymentContent;
  window.refreshPaymentContent=function(){try{if(typeof oldRefresh==='function') oldRefresh.apply(this,arguments)}catch(e){} cleanTransfer();};
  var oldOpen=window.openPayment;
  window.openPayment=function(k){var r; try{if(typeof oldOpen==='function') r=oldOpen.apply(this,arguments)}catch(e){} setTimeout(cleanTransfer,0); setTimeout(cleanTransfer,80); setTimeout(cleanTransfer,250); return r===undefined?false:r;};
  function upgradeMenu(){
    qa('.v2-nav-link').forEach(function(a){
      var txt=(a.textContent||'').toLowerCase();
      if(/pro|premium/.test(txt)||a.classList.contains('premium-locked')||a.classList.contains('real-premium-lock')){
        a.classList.add('premium-locked');
        var tag=a.querySelector('.v2-nav-tag');
        if(!tag){tag=document.createElement('span');tag.className='v2-nav-tag';tag.textContent='PRO';a.appendChild(tag)}
        tag.textContent='PRO';
      }
    });
  }
  document.addEventListener('click',function(e){setTimeout(cleanTransfer,0);setTimeout(cleanTransfer,150);},true);
  document.addEventListener('input',function(e){if(e.target&&/payPhone|payEmail/.test(e.target.id||''))setTimeout(cleanTransfer,0);},true);
  if(window.MutationObserver){new MutationObserver(function(){cleanTransfer();upgradeMenu();}).observe(document.documentElement,{childList:true,subtree:true,characterData:true});}
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',function(){cleanTransfer();upgradeMenu();}); else {cleanTransfer();upgradeMenu();}
  setInterval(cleanTransfer,700);
  setInterval(upgradeMenu,1200);
})();
</script>



<!-- FINAL UPGRADE V98: SaaS premium polish + smart locked menu routing -->
<style id="mkt-v98-saas-menu-premium-polish">
  :root{--mkt-blue:#3b82f6;--mkt-violet:#7c3aed;--mkt-gold:#facc15;--mkt-green:#22c55e;}
  .sidebar,.v2-sidebar{
    background:
      radial-gradient(circle at 10% -10%,rgba(96,165,250,.42),transparent 33%),
      radial-gradient(circle at 102% 10%,rgba(250,204,21,.20),transparent 28%),
      radial-gradient(circle at 50% 110%,rgba(124,58,237,.30),transparent 38%),
      linear-gradient(180deg,#06101f 0%,#0b1220 46%,#111827 100%)!important;
    border-right:1px solid rgba(148,163,184,.26)!important;
    box-shadow:22px 0 80px rgba(2,6,23,.46),inset -1px 0 0 rgba(255,255,255,.06)!important;
  }
  .sidebar .logo,.mkt-sidebar-brand{
    position:relative!important;
    padding:16px!important;
    margin-bottom:14px!important;
    border-radius:24px!important;
    background:linear-gradient(135deg,rgba(255,255,255,.13),rgba(255,255,255,.04))!important;
    border:1px solid rgba(255,255,255,.16)!important;
    box-shadow:0 18px 44px rgba(0,0,0,.26),inset 0 1px 0 rgba(255,255,255,.10)!important;
    overflow:hidden!important;
  }
  .sidebar .logo:before,.mkt-sidebar-brand:before{
    content:""!important;position:absolute!important;inset:-60% -20% auto auto!important;width:160px!important;height:160px!important;
    background:radial-gradient(circle,rgba(96,165,250,.38),transparent 62%)!important;pointer-events:none!important;
  }
  .mkt-sidebar-brand .brand-title{font-size:15px!important;font-weight:1000!important;color:#f8fafc!important;letter-spacing:-.02em!important;line-height:1.15!important;position:relative!important;z-index:2!important;}
  .mkt-sidebar-brand .brand-sub{display:inline-flex!important;align-items:center!important;gap:7px!important;margin-top:9px!important;padding:7px 10px!important;border-radius:999px!important;background:rgba(34,197,94,.13)!important;border:1px solid rgba(34,197,94,.25)!important;color:#bbf7d0!important;font-size:11px!important;font-weight:900!important;position:relative!important;z-index:2!important;}
  .mkt-sidebar-brand .brand-sub:before,.mkt-online-dot{
    content:""!important;width:8px!important;height:8px!important;border-radius:50%!important;background:#00ff7f!important;
    box-shadow:0 0 8px #00ff7f,0 0 18px #00ff7f,0 0 30px rgba(0,255,127,.8)!important;display:inline-block!important;flex:0 0 auto!important;
  }
  .v2-nav-title{color:#a5b4fc!important;font-size:10px!important;letter-spacing:.16em!important;font-weight:1000!important;margin:18px 12px 8px!important;text-transform:uppercase!important;}
  .v2-nav-link{
    margin:7px 10px!important;min-height:56px!important;padding:13px 72px 13px 16px!important;border-radius:18px!important;
    color:#f8fafc!important;background:linear-gradient(135deg,rgba(255,255,255,.105),rgba(255,255,255,.035))!important;
    border:1px solid rgba(148,163,184,.18)!important;box-shadow:inset 0 1px 0 rgba(255,255,255,.075),0 10px 26px rgba(0,0,0,.16)!important;
    transition:transform .2s ease,box-shadow .2s ease,border-color .2s ease,background .2s ease!important;position:relative!important;overflow:hidden!important;
  }
  .v2-nav-link:before{content:""!important;position:absolute!important;left:0!important;top:13px!important;bottom:13px!important;width:3px!important;border-radius:999px!important;background:linear-gradient(180deg,#38bdf8,#818cf8,#facc15)!important;opacity:.62!important;}
  .v2-nav-link:hover,.v2-nav-link.active{transform:translateX(6px)!important;border-color:rgba(96,165,250,.46)!important;background:linear-gradient(135deg,rgba(37,99,235,.94),rgba(124,58,237,.84))!important;box-shadow:0 18px 48px rgba(37,99,235,.32)!important;}
  .v2-nav-text{font-weight:950!important;letter-spacing:-.015em!important;line-height:1.24!important;text-shadow:0 1px 0 rgba(0,0,0,.14)!important;}
  .v2-nav-link.premium-locked,.v2-nav-link.real-premium-lock{padding-right:74px!important;}
  .v2-nav-link .v2-nav-tag,.v2-nav-link.premium-locked::after,.v2-nav-link.real-premium-lock::after{
    content:"PRO"!important;position:absolute!important;right:12px!important;top:50%!important;transform:translateY(-50%)!important;
    min-width:31px!important;height:17px!important;padding:0 7px 0 16px!important;border-radius:999px!important;font-size:7.5px!important;line-height:17px!important;font-weight:1000!important;letter-spacing:.045em!important;
    color:#111827!important;background:linear-gradient(135deg,#fde047,#fb923c)!important;border:1px solid rgba(255,255,255,.58)!important;box-shadow:0 0 16px rgba(250,204,21,.52)!important;z-index:10!important;text-align:center!important;
  }
  .v2-nav-link .v2-nav-tag:before{content:""!important;width:6px!important;height:6px!important;border-radius:50%!important;background:#00ff7f!important;box-shadow:0 0 8px #00ff7f,0 0 16px rgba(0,255,127,.9)!important;position:absolute!important;left:6px!important;top:50%!important;transform:translateY(-50%)!important;}
  .v2-nav-link[data-auto-plan="quarterly"] .v2-nav-tag,.v2-nav-link[data-auto-plan="quarterly"]::after{content:"PRO"!important;}
  .v2-nav-link[data-auto-plan="halfyear"] .v2-nav-tag,.v2-nav-link[data-auto-plan="halfyear"]::after{content:"CRM"!important;}
  .v2-nav-link[data-auto-plan="yearly"] .v2-nav-tag,.v2-nav-link[data-auto-plan="yearly"]::after{content:"AI"!important;}
  .v2-nav-link[data-auto-plan="sellerpro"] .v2-nav-tag,.v2-nav-link[data-auto-plan="sellerpro"]::after{content:"VIP"!important;background:linear-gradient(135deg,#facc15,#f97316)!important;}
  .tool-card,.feature-card,.v3-feature-card,.dashboard-card,.mini-card,.stat-card,.card:not(.compact-price-card){background:rgba(255,255,255,.96)!important;backdrop-filter:blur(10px)!important;border:1px solid rgba(99,102,241,.13)!important;box-shadow:0 12px 30px rgba(0,0,0,.08)!important;}
  .tool-card:hover,.feature-card:hover,.v3-feature-card:hover,.dashboard-card:hover,.mini-card:hover{transform:translateY(-3px)!important;box-shadow:0 18px 44px rgba(37,99,235,.13)!important;border-color:rgba(99,102,241,.24)!important;}
  .compact-price-card{box-shadow:0 18px 45px rgba(15,23,42,.10)!important;border:1px solid rgba(99,102,241,.16)!important;}
  .compact-price-card:hover{transform:translateY(-6px)!important;box-shadow:0 24px 60px rgba(37,99,235,.18)!important;}
  .mkt-shake-premium{animation:mktPremiumShake .45s ease both!important;}
  @keyframes mktPremiumShake{0%,100%{transform:translateX(0)}20%{transform:translateX(-4px)}40%{transform:translateX(5px)}60%{transform:translateX(-3px)}80%{transform:translateX(2px)}}
</style>
<script id="mkt-v98-saas-menu-smart-routing">
(function(){
  'use strict';
  function q(s,r){return (r||document).querySelector(s)}
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s))}
  function device(){var id='';try{id=localStorage.getItem('mkt_device_id')||''}catch(e){};if(!id&&typeof window.getOrCreateDeviceId==='function')id=window.getOrCreateDeviceId();if(!id){id='MKT-'+Math.random().toString(36).slice(2,8).toUpperCase()+Date.now().toString().slice(-4);try{localStorage.setItem('mkt_device_id',id)}catch(e){}}document.cookie='mkt_device_id='+encodeURIComponent(id)+'; path=/; max-age='+(60*60*24*365*5);return id;}
  function planForText(t){t=String(t||'').toLowerCase();if(/studio|image|video|voice|giọng|livestream|seller|vip|enterprise/.test(t))return 'sellerpro';if(/marketing director|ads|phễu|funnel|kpi|content brain/.test(t))return 'yearly';if(/crm|kanban|sales|khách hàng|chốt/.test(t))return 'halfyear';if(/messenger|inbox|kịch bản/.test(t))return 'quarterly';return 'monthly';}
  function cleanPay(){
    var id=device();
    var pc=q('#payContent'); if(pc) pc.textContent=id;
    var dev=q('#payDeviceId'); if(dev) dev.value=id;
    var bank=q('.bank-info'); if(bank){
      var html=bank.innerHTML;
      if(/Nội dung CK/i.test(bank.textContent||'')){
        bank.innerHTML='<b>Ngân hàng:</b> Agribank<br><b>STK:</b> 8888363382629<br><b>Chủ TK:</b> NGUYEN DANG THI XUAN<br><b>Nội dung CK:</b> <span id="payContent">'+id+'</span>';
      }
    }
    var k=window.currentPremiumPlanKey||window.mktSelectedPlanKey||'monthly';
    var amounts={monthly:'159000',quarterly:'359000',halfyear:'559000',yearly:'859000',sellerpro:'1959000',lifetime:'1959000'};
    var qr=q('#payQr'); if(qr) qr.src='https://img.vietqr.io/image/970405-8888363382629-compact2.png?amount='+encodeURIComponent(amounts[k]||amounts.monthly)+'&addInfo='+encodeURIComponent(id)+'&accountName='+encodeURIComponent('NGUYEN DANG THI XUAN');
  }
  function addBrand(){
    var side=q('.sidebar')||q('.v2-sidebar'); if(!side||q('.mkt-sidebar-brand',side))return;
    var logo=q('.logo',side);
    var box=document.createElement('div');box.className='mkt-sidebar-brand';
    box.innerHTML='<div class="brand-title">Marketing Automation Pro</div><div class="brand-sub">Premium AI Suite</div>';
    if(logo&&logo.parentNode){logo.parentNode.insertBefore(box,logo.nextSibling)}else side.insertBefore(box,side.firstChild);
  }
  function upgradeMenu(){
    qa('.v2-nav-link').forEach(function(a){
      var txt=(a.textContent||'');
      var isLocked=/pro|premium/i.test(txt)||a.classList.contains('premium-locked')||a.classList.contains('real-premium-lock');
      if(isLocked){
        a.classList.add('premium-locked');
        a.setAttribute('data-auto-plan',planForText(txt));
        var tag=a.querySelector('.v2-nav-tag');
        if(!tag){tag=document.createElement('span');tag.className='v2-nav-tag';a.appendChild(tag)}
        var p=a.getAttribute('data-auto-plan'); tag.textContent=p==='sellerpro'?'VIP':(p==='halfyear'?'CRM':(p==='yearly'?'AI':'PRO'));
      }
    });
  }
  function focusPricing(plan){
    var target=q('[data-plan="'+plan+'"].compact-price-card')||q('.compact-price-card[data-plan="'+plan+'"]')||q('[data-plan="'+plan+'"]')||q('.premium-pricing-compact')||q('#premium');
    if(target){target.scrollIntoView({behavior:'smooth',block:'center'});target.classList.add('mkt-shake-premium');setTimeout(function(){target.classList.remove('mkt-shake-premium')},650);}
    if(window.openPlanBenefit){setTimeout(function(){window.openPlanBenefit(plan)},450)}
  }
  document.addEventListener('click',function(e){
    var a=e.target.closest&&e.target.closest('.v2-nav-link');
    if(!a||!a.classList.contains('premium-locked'))return;
    var plan=a.getAttribute('data-auto-plan')||planForText(a.textContent||'');
    e.preventDefault();e.stopPropagation();if(e.stopImmediatePropagation)e.stopImmediatePropagation();
    focusPricing(plan);return false;
  },true);
  var oldRefresh=window.refreshPaymentContent;window.refreshPaymentContent=function(){try{if(typeof oldRefresh==='function')oldRefresh.apply(this,arguments)}catch(e){}cleanPay();};
  var oldOpen=window.openPayment;window.openPayment=function(){var r;try{if(typeof oldOpen==='function')r=oldOpen.apply(this,arguments)}catch(e){}setTimeout(cleanPay,0);setTimeout(cleanPay,90);setTimeout(cleanPay,260);return r===undefined?false:r;};
  if(window.MutationObserver){new MutationObserver(function(){addBrand();upgradeMenu();cleanPay();}).observe(document.documentElement,{childList:true,subtree:true,characterData:true});}
  if(document.readyState==='loading')document.addEventListener('DOMContentLoaded',function(){addBrand();upgradeMenu();cleanPay();});else{addBrand();upgradeMenu();cleanPay();}
  setTimeout(function(){addBrand();upgradeMenu();cleanPay();},500);setTimeout(function(){addBrand();upgradeMenu();cleanPay();},1500);
})();
</script>


<!-- FINAL UPGRADE V99: Luxury sidebar + compact PRO badges + smart Premium routing -->
<style id="mkt-v99-luxury-sidebar-fix">
  :root{
    --lux-bg:#07111f;
    --lux-card:rgba(255,255,255,.075);
    --lux-card2:rgba(255,255,255,.045);
    --lux-line:rgba(148,163,184,.18);
    --lux-blue:#38bdf8;
    --lux-violet:#8b5cf6;
    --lux-green:#22f28a;
    --lux-gold:#facc15;
  }

  body .sidebar,
  body .v2-sidebar,
  body aside,
  body [class*="sidebar"]{
    background:
      radial-gradient(circle at 8% 0%,rgba(56,189,248,.26),transparent 28%),
      radial-gradient(circle at 95% 8%,rgba(250,204,21,.18),transparent 30%),
      radial-gradient(circle at 50% 105%,rgba(139,92,246,.28),transparent 38%),
      linear-gradient(180deg,#060b16 0%,#091426 45%,#0f172a 100%)!important;
    border-right:1px solid rgba(148,163,184,.22)!important;
    box-shadow:26px 0 80px rgba(0,0,0,.50), inset -1px 0 0 rgba(255,255,255,.05)!important;
  }

  body .v2-sidebar::before,
  body .sidebar::before{
    content:"Marketing Automation Pro\A Premium AI Suite";
    white-space:pre-line;
    display:block;
    margin:14px 14px 12px!important;
    padding:18px 16px!important;
    border-radius:24px!important;
    color:#fff!important;
    font-weight:1000!important;
    font-size:15px!important;
    line-height:1.28!important;
    letter-spacing:-.02em!important;
    background:
      linear-gradient(135deg,rgba(255,255,255,.16),rgba(255,255,255,.045)),
      radial-gradient(circle at 85% 20%,rgba(56,189,248,.28),transparent 35%)!important;
    border:1px solid rgba(255,255,255,.14)!important;
    box-shadow:0 18px 48px rgba(0,0,0,.32), inset 0 1px 0 rgba(255,255,255,.14)!important;
  }

  body .v2-nav-title,
  body [class*="nav-title"]{
    margin:18px 18px 8px!important;
    color:#a5b4fc!important;
    font-size:10px!important;
    line-height:1!important;
    font-weight:1000!important;
    letter-spacing:.18em!important;
    text-transform:uppercase!important;
    opacity:.96!important;
  }

  body .v2-nav-link,
  body .sidebar a,
  body .sidebar button,
  body [class*="sidebar"] a,
  body [class*="sidebar"] button{
    position:relative!important;
    min-height:62px!important;
    margin:8px 14px!important;
    padding:13px 62px 13px 18px!important;
    border-radius:20px!important;
    color:#f8fafc!important;
    background:
      linear-gradient(135deg,rgba(255,255,255,.11),rgba(255,255,255,.045))!important;
    border:1px solid rgba(148,163,184,.18)!important;
    box-shadow:
      0 12px 30px rgba(0,0,0,.20),
      inset 0 1px 0 rgba(255,255,255,.09)!important;
    transition:transform .20s ease, box-shadow .20s ease, border-color .20s ease, background .20s ease!important;
    overflow:hidden!important;
  }

  body .v2-nav-link::before,
  body [class*="sidebar"] a::before{
    content:""!important;
    position:absolute!important;
    left:0!important;
    top:16px!important;
    bottom:16px!important;
    width:3px!important;
    border-radius:999px!important;
    background:linear-gradient(180deg,#38bdf8,#8b5cf6,#facc15)!important;
    opacity:.7!important;
  }

  body .v2-nav-link:hover,
  body .v2-nav-link.active,
  body .sidebar a:hover,
  body [class*="sidebar"] a:hover{
    transform:translateX(7px)!important;
    border-color:rgba(96,165,250,.55)!important;
    background:
      radial-gradient(circle at 90% 20%,rgba(250,204,21,.22),transparent 30%),
      linear-gradient(135deg,rgba(37,99,235,.92),rgba(124,58,237,.84))!important;
    box-shadow:
      0 20px 54px rgba(37,99,235,.36),
      inset 0 1px 0 rgba(255,255,255,.16)!important;
  }

  body .v2-nav-text,
  body .v2-nav-link span:first-child,
  body [class*="sidebar"] a span:first-child{
    color:#f8fafc!important;
    font-size:16px!important;
    line-height:1.22!important;
    font-weight:1000!important;
    letter-spacing:-.015em!important;
    text-shadow:0 1px 0 rgba(0,0,0,.18)!important;
  }

  /* compact PRO/PREMIUM badge - small, luxury, with green online dot */
  body .v2-nav-link .v2-nav-tag,
  body .v2-nav-link .premium-badge,
  body .v2-nav-link .pro-badge,
  body .v2-nav-link [class*="tag"],
  body .v2-nav-link [class*="badge"],
  body [class*="sidebar"] a .v2-nav-tag,
  body [class*="sidebar"] a [class*="badge"],
  body [class*="sidebar"] a [class*="tag"],
  body .mkt-pro-mini{
    position:absolute!important;
    right:14px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    width:auto!important;
    min-width:38px!important;
    max-width:68px!important;
    height:20px!important;
    padding:0 9px 0 19px!important;
    border-radius:999px!important;
    font-size:8px!important;
    line-height:20px!important;
    font-weight:1000!important;
    letter-spacing:.05em!important;
    text-align:center!important;
    color:#06220f!important;
    background:linear-gradient(135deg,#bbf7d0,#22c55e)!important;
    border:1px solid rgba(255,255,255,.65)!important;
    box-shadow:0 0 12px rgba(34,242,138,.50),0 0 24px rgba(34,242,138,.25)!important;
    text-shadow:none!important;
    z-index:30!important;
  }

  body .v2-nav-link .v2-nav-tag::before,
  body .v2-nav-link .premium-badge::before,
  body .v2-nav-link .pro-badge::before,
  body .v2-nav-link [class*="tag"]::before,
  body .v2-nav-link [class*="badge"]::before,
  body .mkt-pro-mini::before{
    content:""!important;
    position:absolute!important;
    left:7px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    width:7px!important;
    height:7px!important;
    border-radius:50%!important;
    background:#00ff7f!important;
    box-shadow:0 0 8px #00ff7f,0 0 18px rgba(0,255,127,.95),0 0 30px rgba(0,255,127,.65)!important;
  }

  body .v2-nav-link.premium-locked .v2-nav-tag,
  body .v2-nav-link.real-premium-lock .v2-nav-tag,
  body [data-auto-plan="sellerpro"] .v2-nav-tag{
    color:#241400!important;
    background:linear-gradient(135deg,#fde047,#fb923c)!important;
    box-shadow:0 0 14px rgba(250,204,21,.55),0 0 28px rgba(251,146,60,.25)!important;
  }

  body .v2-nav-link.premium-locked::after,
  body .v2-nav-link.real-premium-lock::after{
    display:none!important;
    content:""!important;
  }

  body .v2-nav-link.premium-locked::before,
  body .v2-nav-link.real-premium-lock::before{
    content:""!important;
    left:0!important;
    right:auto!important;
    top:16px!important;
    bottom:16px!important;
    width:3px!important;
    height:auto!important;
    font-size:0!important;
    background:linear-gradient(180deg,#facc15,#fb923c)!important;
    filter:none!important;
    text-shadow:none!important;
  }

  /* feature cards softer and premium */
  body .feature-card,
  body .mkt-feature-card,
  body .dashboard-card,
  body .tool-card,
  body .v2-card,
  body [class*="feature"] [class*="card"]{
    background:rgba(255,255,255,.96)!important;
    backdrop-filter:blur(12px)!important;
    border:1px solid rgba(99,102,241,.13)!important;
    box-shadow:0 16px 38px rgba(2,6,23,.10), inset 0 1px 0 rgba(255,255,255,.7)!important;
    border-radius:24px!important;
    transition:transform .2s ease, box-shadow .2s ease, border-color .2s ease!important;
  }

  body .feature-card:hover,
  body .mkt-feature-card:hover,
  body .dashboard-card:hover,
  body .tool-card:hover,
  body .v2-card:hover{
    transform:translateY(-5px)!important;
    border-color:rgba(96,165,250,.34)!important;
    box-shadow:0 24px 62px rgba(37,99,235,.16), inset 0 1px 0 rgba(255,255,255,.75)!important;
  }

  .mkt-menu-shake{animation:mktMenuShake .32s ease both!important;}
  @keyframes mktMenuShake{
    0%,100%{transform:translateX(0)}
    25%{transform:translateX(7px)}
    55%{transform:translateX(2px)}
    75%{transform:translateX(6px)}
  }

  .mkt-highlight-plan{
    animation:mktPlanPulse 1.25s ease 0s 2!important;
    box-shadow:0 0 0 4px rgba(250,204,21,.38),0 26px 70px rgba(250,204,21,.30)!important;
    border-color:#f59e0b!important;
  }
  @keyframes mktPlanPulse{
    0%,100%{transform:translateY(0)}
    50%{transform:translateY(-8px)}
  }
</style>

<script id="mkt-v99-luxury-sidebar-smart-routing">
(function(){
  'use strict';
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s));}
  function txt(el){return String((el&&el.textContent)||'').toLowerCase();}
  function planForMenu(el){
    var t=txt(el);
    if(/ai studio|image|video|voice|giọng|livestream|seller|vip|business/.test(t)) return 'sellerpro';
    if(/marketing director|ads|funnel|kpi|director|content calendar/.test(t)) return 'yearly';
    if(/crm|kanban|sales bot|khách hàng/.test(t)) return 'halfyear';
    if(/messenger|inbox|chat/.test(t)) return 'quarterly';
    if(/premium|pro/.test(t)) return 'yearly';
    return '';
  }
  function ensureBadge(a, plan){
    if(!a || !plan) return;
    a.classList.add('premium-locked','mkt-lux-menu');
    a.setAttribute('data-auto-plan', plan);
    var tag=a.querySelector('.v2-nav-tag,.premium-badge,.pro-badge,.mkt-pro-mini,[class*="badge"],[class*="tag"]');
    if(!tag){
      tag=document.createElement('span');
      tag.className='mkt-pro-mini';
      a.appendChild(tag);
    }
    tag.className=(tag.className||'')+' mkt-pro-mini';
    tag.textContent = plan==='sellerpro' ? 'VIP' : (plan==='halfyear' ? 'CRM' : (plan==='yearly' ? 'AI' : 'PRO'));
  }
  function polishMenus(){
    qa('.v2-nav-link,.sidebar a,[class*="sidebar"] a').forEach(function(a){
      var plan=planForMenu(a);
      if(plan) ensureBadge(a, plan);
    });
  }
  function scrollPremium(plan){
    var target=document.querySelector('#premium,.premium-pricing-compact,.premium-pricing-pro,.pricing-grid-5,[data-section="premium"],[id*="premium"]');
    if(target && target.scrollIntoView) target.scrollIntoView({behavior:'smooth',block:'center'});
    try{history.replaceState(null,'','#premium');}catch(e){}
    setTimeout(function(){
      var cards=qa('[data-plan="'+plan+'"],.price-card,.premium-card,.plan-card');
      var card=cards.find(function(c){return c.getAttribute('data-plan')===plan || txt(c).indexOf(plan)>-1;}) || document.querySelector('[data-plan="'+plan+'"]');
      if(card){
        card.classList.add('mkt-highlight-plan');
        setTimeout(function(){card.classList.remove('mkt-highlight-plan');},2600);
      }
      if(typeof window.mktOpenPlanDetail==='function'){
        try{ window.mktOpenPlanDetail(plan); }catch(e){}
      }
    },420);
  }
  document.addEventListener('click',function(e){
    var a=e.target.closest && e.target.closest('.v2-nav-link,.sidebar a,[class*="sidebar"] a');
    if(!a) return;
    var plan=a.getAttribute('data-auto-plan') || planForMenu(a);
    if(!plan) return;
    if(/pro|premium|ai messenger|crm|marketing director|ai studio|image|video|voice|seller/i.test(a.textContent||'')){
      e.preventDefault();
      e.stopPropagation();
      if(e.stopImmediatePropagation) e.stopImmediatePropagation();
      a.classList.add('mkt-menu-shake');
      setTimeout(function(){a.classList.remove('mkt-menu-shake');},380);
      scrollPremium(plan);
      return false;
    }
  },true);
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',polishMenus); else polishMenus();
  if(window.MutationObserver){
    new MutationObserver(function(){polishMenus();}).observe(document.documentElement,{childList:true,subtree:true});
  }
  setInterval(polishMenus,1200);
})();
</script>



<!-- FINAL PATCH V100: only compact PRO/PREMIUM badges + upgraded mobile install link -->
<style id="mkt-v100-badge-pwa-only">
  /* Chỉ tối ưu badge PRO/PREMIUM, không đổi nội dung giao diện */
  body .v2-nav-link.mkt-menu-pro,
  body .v2-nav-link.mkt-menu-premium,
  body .sidebar a.mkt-menu-pro,
  body .sidebar a.mkt-menu-premium{
    padding-right:78px!important;
  }
  body .v2-nav-link.mkt-menu-pro .v2-nav-tag,
  body .v2-nav-link.mkt-menu-premium .v2-nav-tag,
  body .v2-nav-link.mkt-menu-pro .mkt-pro-mini,
  body .v2-nav-link.mkt-menu-premium .mkt-pro-mini{
    position:absolute!important;
    right:13px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    display:inline-flex!important;
    align-items:center!important;
    justify-content:center!important;
    width:auto!important;
    min-width:42px!important;
    max-width:74px!important;
    height:19px!important;
    padding:0 8px 0 18px!important;
    border-radius:999px!important;
    font-size:7.8px!important;
    line-height:19px!important;
    font-weight:1000!important;
    letter-spacing:.045em!important;
    text-shadow:none!important;
    z-index:60!important;
    overflow:visible!important;
    white-space:nowrap!important;
  }
  body .v2-nav-link.mkt-menu-pro .v2-nav-tag,
  body .v2-nav-link.mkt-menu-pro .mkt-pro-mini{
    color:#052e16!important;
    background:linear-gradient(135deg,#bbf7d0,#22c55e)!important;
    border:1px solid rgba(255,255,255,.62)!important;
    box-shadow:0 0 10px rgba(34,197,94,.44),0 0 22px rgba(34,197,94,.22)!important;
  }
  body .v2-nav-link.mkt-menu-premium .v2-nav-tag,
  body .v2-nav-link.mkt-menu-premium .mkt-pro-mini{
    color:#231400!important;
    background:linear-gradient(135deg,#fde047,#fb923c)!important;
    border:1px solid rgba(255,255,255,.62)!important;
    box-shadow:0 0 10px rgba(250,204,21,.46),0 0 22px rgba(251,146,60,.24)!important;
  }
  body .v2-nav-link.mkt-menu-pro .v2-nav-tag:before,
  body .v2-nav-link.mkt-menu-premium .v2-nav-tag:before,
  body .v2-nav-link.mkt-menu-pro .mkt-pro-mini:before,
  body .v2-nav-link.mkt-menu-premium .mkt-pro-mini:before{
    content:""!important;
    position:absolute!important;
    left:6px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    width:6px!important;
    height:6px!important;
    border-radius:50%!important;
    background:#00ff7f!important;
    box-shadow:0 0 7px #00ff7f,0 0 15px rgba(0,255,127,.92)!important;
  }
  body .v2-nav-link.mkt-menu-pro::after,
  body .v2-nav-link.mkt-menu-premium::after,
  body .v2-nav-link.real-premium-lock::after,
  body .v2-nav-link.premium-locked::after{
    display:none!important;
    content:""!important;
  }
  body .app-install-banner{
    border:1px solid rgba(255,255,255,.16)!important;
    background:radial-gradient(circle at 90% 10%,rgba(34,197,94,.30),transparent 28%),linear-gradient(135deg,#0f172a,#312e81)!important;
    box-shadow:0 18px 44px rgba(15,23,42,.28),inset 0 1px 0 rgba(255,255,255,.10)!important;
  }
  body .app-install-banner button{
    border:0!important;
    border-radius:999px!important;
    padding:12px 16px!important;
    font-weight:1000!important;
    color:#07111f!important;
    background:linear-gradient(135deg,#bbf7d0,#38bdf8,#c7d2fe)!important;
    box-shadow:0 0 22px rgba(56,189,248,.35)!important;
  }
  body.mkt-installed-mode .app-install-banner{display:none!important;}
</style>
<script id="mkt-v100-badge-pwa-only-js">
(function(){
  'use strict';
  var deferredInstallPrompt=null;
  function qa(s,r){return Array.prototype.slice.call((r||document).querySelectorAll(s));}
  function cleanText(el){return String((el&&el.textContent)||'').toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/đ/g,'d');}
  function ensureTag(a,label){
    if(!a) return;
    var tag=a.querySelector('.v2-nav-tag,.mkt-pro-mini,.premium-badge,.pro-badge,[class*="badge"],[class*="tag"]');
    if(!tag){tag=document.createElement('span');tag.className='v2-nav-tag';a.appendChild(tag);}
    tag.className='v2-nav-tag';
    tag.textContent=label;
  }
  function refreshBadges(){
    qa('.v2-nav-link,.sidebar a,[class*="sidebar"] a').forEach(function(a){
      var t=cleanText(a);
      if(/ai messenger|crm|kanban|marketing director|ai studio|creative|analytics|automation|premium|video|image|giong noi|livestream|seller|vip/.test(t)){
        a.classList.remove('mkt-menu-pro');
        a.classList.add('mkt-menu-premium');
        ensureTag(a,'PREMIUM');
      }else if(/group center tong|group center|quan ly group|quan ly fanpage|ai comment|facebook center|fanpage|comment manager|dang bai facebook|token fanpage/.test(t)){
        a.classList.remove('mkt-menu-premium');
        a.classList.add('mkt-menu-pro');
        ensureTag(a,'PRO');
      }
    });
  }
  function isStandalone(){
    return window.matchMedia('(display-mode: standalone)').matches || window.navigator.standalone===true;
  }
  function showGuide(){
    alert('Cách cài App Mini:\n\nAndroid Chrome: bấm nút Cài app nếu hiện thông báo. Nếu chưa hiện, bấm dấu 3 chấm → Thêm vào màn hình chính.\n\niPhone Safari: bấm Chia sẻ → Thêm vào màn hình chính.\n\nLưu ý: cần mở bằng domain HTTPS thật, không dùng file:// hoặc localhost.');
  }
  window.addEventListener('beforeinstallprompt',function(e){
    e.preventDefault();
    deferredInstallPrompt=e;
    qa('.app-install-banner button').forEach(function(btn){btn.textContent='Cài app vào điện thoại';});
  });
  window.addEventListener('appinstalled',function(){document.body.classList.add('mkt-installed-mode');deferredInstallPrompt=null;});
  window.showInstallGuide=function(){
    if(isStandalone()){document.body.classList.add('mkt-installed-mode');return false;}
    if(deferredInstallPrompt){
      deferredInstallPrompt.prompt();
      deferredInstallPrompt.userChoice.finally(function(){deferredInstallPrompt=null;});
      return false;
    }
    showGuide();
    return false;
  };
  document.addEventListener('click',function(e){
    var btn=e.target.closest&&e.target.closest('.app-install-banner button,[data-install-app],#installAppBtn');
    if(!btn) return;
    e.preventDefault();e.stopPropagation();if(e.stopImmediatePropagation)e.stopImmediatePropagation();
    return window.showInstallGuide();
  },true);
  function init(){if(isStandalone()) document.body.classList.add('mkt-installed-mode');refreshBadges();}
  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',init); else init();
  if(window.MutationObserver){new MutationObserver(refreshBadges).observe(document.documentElement,{childList:true,subtree:true,characterData:true});}
  setTimeout(refreshBadges,400);setTimeout(refreshBadges,1200);setInterval(refreshBadges,1800);
})();
</script>



<!-- FINAL PATCH V101: matte luxury PRO/PREMIUM badges - giảm chói, nhỏ gọn, không đổi nội dung -->
<style id="mkt-v101-matte-luxury-badges">
  body .v2-nav-link.mkt-menu-pro,
  body .v2-nav-link.mkt-menu-premium,
  body .sidebar a.mkt-menu-pro,
  body .sidebar a.mkt-menu-premium{
    padding-right:82px!important;
  }

  body .v2-nav-link.mkt-menu-pro .v2-nav-tag,
  body .v2-nav-link.mkt-menu-premium .v2-nav-tag,
  body .v2-nav-link.mkt-menu-pro .mkt-pro-mini,
  body .v2-nav-link.mkt-menu-premium .mkt-pro-mini,
  body .sidebar a.mkt-menu-pro .v2-nav-tag,
  body .sidebar a.mkt-menu-premium .v2-nav-tag{
    position:absolute!important;
    right:16px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    display:inline-flex!important;
    align-items:center!important;
    justify-content:center!important;
    height:18px!important;
    min-width:42px!important;
    max-width:68px!important;
    padding:0 8px 0 14px!important;
    border-radius:999px!important;
    font-size:7.6px!important;
    line-height:18px!important;
    font-weight:900!important;
    letter-spacing:.035em!important;
    text-transform:uppercase!important;
    text-shadow:none!important;
    white-space:nowrap!important;
    overflow:visible!important;
    z-index:80!important;
    filter:none!important;
  }

  body .v2-nav-link.mkt-menu-pro .v2-nav-tag,
  body .v2-nav-link.mkt-menu-pro .mkt-pro-mini,
  body .sidebar a.mkt-menu-pro .v2-nav-tag{
    color:#dcfff0!important;
    background:linear-gradient(135deg,rgba(21,128,61,.96),rgba(16,185,129,.92))!important;
    border:1px solid rgba(134,239,172,.38)!important;
    box-shadow:0 3px 10px rgba(0,0,0,.22),0 0 8px rgba(34,197,94,.20)!important;
  }

  body .v2-nav-link.mkt-menu-premium .v2-nav-tag,
  body .v2-nav-link.mkt-menu-premium .mkt-pro-mini,
  body .sidebar a.mkt-menu-premium .v2-nav-tag{
    color:#fff7d6!important;
    background:linear-gradient(135deg,rgba(180,83,9,.96),rgba(245,158,11,.92))!important;
    border:1px solid rgba(253,230,138,.36)!important;
    box-shadow:0 3px 10px rgba(0,0,0,.22),0 0 8px rgba(245,158,11,.18)!important;
  }

  body .v2-nav-link.mkt-menu-pro .v2-nav-tag:before,
  body .v2-nav-link.mkt-menu-premium .v2-nav-tag:before,
  body .v2-nav-link.mkt-menu-pro .mkt-pro-mini:before,
  body .v2-nav-link.mkt-menu-premium .mkt-pro-mini:before{
    content:""!important;
    position:absolute!important;
    left:6px!important;
    top:50%!important;
    transform:translateY(-50%)!important;
    width:4px!important;
    height:4px!important;
    border-radius:50%!important;
    background:rgba(255,255,255,.88)!important;
    box-shadow:0 0 4px rgba(255,255,255,.36)!important;
  }

  body .v2-nav-link.mkt-menu-pro::before,
  body .v2-nav-link.mkt-menu-premium::before,
  body .v2-nav-link.real-premium-lock::before,
  body .v2-nav-link.premium-locked::before,
  body .v2-nav-link.mkt-menu-pro::after,
  body .v2-nav-link.mkt-menu-premium::after,
  body .v2-nav-link.real-premium-lock::after,
  body .v2-nav-link.premium-locked::after{
    display:none!important;
    opacity:0!important;
    content:""!important;
    box-shadow:none!important;
    filter:none!important;
  }

  body .v2-nav-link,
  body .sidebar a{
    overflow:hidden!important;
  }

  body .app-install-banner{
    border:1px solid rgba(255,255,255,.14)!important;
    background:linear-gradient(135deg,rgba(15,23,42,.96),rgba(30,41,99,.96))!important;
    box-shadow:0 12px 30px rgba(0,0,0,.26),inset 0 1px 0 rgba(255,255,255,.08)!important;
  }
  body .app-install-banner button{
    min-height:38px!important;
    padding:10px 15px!important;
    border-radius:999px!important;
    color:#eff6ff!important;
    background:linear-gradient(135deg,#2563eb,#7c3aed)!important;
    box-shadow:0 6px 16px rgba(37,99,235,.25)!important;
  }
</style>



<!-- FINAL OVERRIDE 20260610: SaaS left menu dot Pro/Premium + remove browser alert lock -->
<style id="mkt-final-saas-sidebar-dot-override">
  body .v2-sidebar,
  body .sidebar,
  body aside{
    background:linear-gradient(180deg,#070b18 0%,#0b1024 52%,#0a0b18 100%)!important;
    border-right:1px solid rgba(148,163,184,.16)!important;
  }

  body .v2-nav-link,
  body .sidebar a.v2-nav-link{
    min-height:52px!important;
    height:auto!important;
    padding:10px 14px!important;
    margin:0 0 8px 0!important;
    border-radius:14px!important;
    display:flex!important;
    align-items:center!important;
    justify-content:space-between!important;
    gap:12px!important;
    background:linear-gradient(135deg,rgba(15,23,42,.82),rgba(30,41,59,.52))!important;
    border:1px solid rgba(148,163,184,.12)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.045),0 8px 22px rgba(0,0,0,.20)!important;
    transform:none!important;
    overflow:hidden!important;
    cursor:pointer!important;
    pointer-events:auto!important;
  }

  body .v2-nav-link:hover,
  body .sidebar a.v2-nav-link:hover{
    transform:translateX(3px)!important;
    background:linear-gradient(135deg,rgba(37,99,235,.20),rgba(124,58,237,.13))!important;
    border-color:rgba(96,165,250,.24)!important;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.06),0 12px 28px rgba(0,0,0,.28)!important;
  }

  body .v2-nav-link.active,
  body .sidebar a.v2-nav-link.active{
    background:linear-gradient(135deg,rgba(37,99,235,.34),rgba(124,58,237,.25))!important;
    border-color:rgba(129,140,248,.38)!important;
  }

  body .v2-nav-link .v2-nav-ico{display:none!important;width:0!important;min-width:0!important;margin:0!important;padding:0!important;}
  body .v2-nav-link .v2-nav-text{
    flex:1 1 auto!important;
    min-width:0!important;
    color:#f8fafc!important;
    font-size:14.5px!important;
    line-height:1.24!important;
    font-weight:800!important;
    letter-spacing:-.02em!important;
    text-shadow:none!important;
  }

  body .v2-nav-link .v2-nav-tag,
  body .v2-nav-link .mkt-pro-mini{display:none!important;}

  body .v2-nav-link::before,
  body .v2-nav-link::after,
  body .v2-nav-link.pro-feature::after,
  body .v2-nav-link.premium-locked::after,
  body .v2-nav-link.real-pro::after,
  body .v2-nav-link.real-premium-lock::after,
  body .v2-nav-link.mkt-menu-pro::after,
  body .v2-nav-link.mkt-menu-premium::after{
    display:none!important;
    content:""!important;
    opacity:0!important;
    visibility:hidden!important;
    width:0!important;
    height:0!important;
    padding:0!important;
    margin:0!important;
    box-shadow:none!important;
    background:none!important;
    animation:none!important;
    filter:none!important;
  }

  body .mkt-level-badge{
    flex:0 0 auto!important;
    display:inline-flex!important;
    align-items:center!important;
    gap:6px!important;
    height:20px!important;
    padding:0 3px!important;
    border-radius:999px!important;
    background:transparent!important;
    border:none!important;
    box-shadow:none!important;
    color:#dbeafe!important;
    font-size:11.5px!important;
    line-height:20px!important;
    font-weight:800!important;
    letter-spacing:-.01em!important;
    text-transform:none!important;
    text-shadow:none!important;
    white-space:nowrap!important;
  }

  body .mkt-level-badge::before{
    content:""!important;
    display:inline-block!important;
    width:8px!important;
    height:8px!important;
    border-radius:999px!important;
    background:#22c55e!important;
    box-shadow:0 0 0 3px rgba(34,197,94,.10),0 0 10px rgba(34,197,94,.45)!important;
  }
  body .mkt-level-badge.premium::before{
    background:#facc15!important;
    box-shadow:0 0 0 3px rgba(250,204,21,.10),0 0 10px rgba(250,204,21,.42)!important;
  }

  body .mkt-level-badge.pro{color:#d1fae5!important;}
  body .mkt-level-badge.premium{color:#fef3c7!important;}

  body .v2-sidebar-title,
  body .sidebar-title{
    font-size:11px!important;
    letter-spacing:.13em!important;
    color:#a5b4fc!important;
    margin:16px 6px 8px!important;
    opacity:.88!important;
  }

  body .app-install-banner,
  body #installAppBtn,
  body #pwaInstallBtn{
    box-shadow:none!important;
    filter:none!important;
  }
</style>

<script id="mkt-final-saas-sidebar-dot-no-alert-js">
(function(){
  var IS_PREMIUM = {{ 'true' if is_device_premium else 'false' }};
  var TRIAL_EXPIRED = {{ 'true' if free_status.is_expired else 'false' }};
  var CORE_PRO = {
    facebook_center:1,page_center_total:1,post:1,fanpage_manager:1,group_suite:1,comment_manager:1
  };
  var PREMIUM_ONLY = {
    messenger_ai:1,crm_sales:1,marketing_director:1,ai_studio:1,creative_center:1,analytics:1,analytics_center:1,automation_center:1
  };
  var ALIAS = {
    page_center:'page_center_total',page_comment_pro:'page_center_total',page_comment_queue:'page_center_total',
    group_marketing:'group_suite',group_finder:'group_suite',group_uid_splitter:'group_suite',group_join_queue:'group_suite',group_post_filter:'group_suite'
  };
  function norm(id){return ALIAS[id] || id || 'dashboard';}
  function isPremium(id){id=norm(id); return !!PREMIUM_ONLY[id];}
  function isPro(id){id=norm(id); return !!CORE_PRO[id];}
  function isLocked(id){id=norm(id); if(IS_PREMIUM) return false; if(PREMIUM_ONLY[id]) return true; if(TRIAL_EXPIRED && CORE_PRO[id]) return true; return false;}
  function labelName(id){id=norm(id); var m={messenger_ai:'AI Messenger',crm_sales:'CRM Kanban',marketing_director:'AI Marketing Director',ai_studio:'AI Studio',creative_center:'Image / Video / Voice',analytics:'Analytics Center',automation_center:'Cài đặt Automation',facebook_center:'Facebook Center',page_center_total:'Page Center Tổng',fanpage_manager:'Quản lý Fanpage',group_suite:'Group Center Tổng',comment_manager:'AI Comment'}; return m[id] || 'Tính năng';}

  function pricingTarget(){
    return document.querySelector('#premium') || document.querySelector('.premium-pricing-compact') || document.querySelector('.price-card') || document.querySelector('.pricing-card');
  }
  function showUpgrade(id){
    var t = pricingTarget();
    if(t){
      try{t.closest('.module-section,.section,.page-section,.content-section').style.display='block';}catch(e){}
      setTimeout(function(){t.scrollIntoView({behavior:'smooth',block:'start'});},30);
    }
    if(typeof window.showPaymentNotice==='function'){
      window.showPaymentNotice(labelName(id)+' cần nâng cấp Pro/Premium. Chọn gói phù hợp ở bảng giá để kích hoạt theo ID thiết bị.');
    }
    return false;
  }

  window.openLockedFeature = function(name){ return showUpgrade(name || 'premium'); };

  window.openModule = function(id){
    id = norm(id);
    if(isLocked(id)) return showUpgrade(id);
    document.querySelectorAll('.module-section').forEach(function(el){el.classList.remove('active-module');});
    var target=document.getElementById(id)||document.getElementById('dashboard');
    if(target){target.classList.add('active-module'); setTimeout(function(){target.scrollIntoView({behavior:'smooth',block:'start'});},30);}
    document.querySelectorAll('.v2-nav-link').forEach(function(a){a.classList.remove('active');});
    var active=document.querySelector('.v2-nav-link[href="#'+id+'"]'); if(active) active.classList.add('active');
    return false;
  };

  function decorateMenu(){
    document.querySelectorAll('.v2-nav-link').forEach(function(a){
      var raw=(a.getAttribute('href')||'').replace('#','');
      var id=norm(raw);
      a.querySelectorAll('.v2-nav-ico,.v2-nav-tag,.mkt-pro-mini,.mkt-level-badge').forEach(function(x){x.remove();});
      a.classList.remove('premium-locked','real-premium-lock','pro-feature','real-pro','mkt-menu-pro','mkt-menu-premium');
      if(isPremium(id)){
        var p=document.createElement('span'); p.className='mkt-level-badge premium'; p.textContent='Premium'; a.appendChild(p);
      }else if(isPro(id)){
        var g=document.createElement('span'); g.className='mkt-level-badge pro'; g.textContent='Pro'; a.appendChild(g);
      }
    });
  }

  document.addEventListener('click',function(e){
    var a=e.target.closest && e.target.closest('.v2-nav-link[href^="#"]');
    if(!a) return;
    var id=(a.getAttribute('href')||'').replace('#','');
    if(!id) return;
    e.preventDefault(); e.stopPropagation(); if(e.stopImmediatePropagation) e.stopImmediatePropagation();
    return window.openModule(id);
  },true);

  if(document.readyState==='loading') document.addEventListener('DOMContentLoaded',decorateMenu); else decorateMenu();
  setTimeout(decorateMenu,300); setTimeout(decorateMenu,1000); setTimeout(decorateMenu,2200);
})();
</script>

</body></html>
"""

@app.route("/admin")
def admin_home():
    return render_template_string(ADMIN_HTML, rows=get_premium_requests(), support_rows=get_support_messages(limit=200), admin_stats=admin_ceo_stats())


@app.route('/api/admin/ceo_dashboard')
def api_admin_ceo_dashboard():
    return jsonify(admin_ceo_stats())

@app.route("/admin/premium_action", methods=["POST"])
def admin_premium_action():
    request_id = request.form.get("request_id")
    status = request.form.get("status", "Đã duyệt")
    approve_premium_request(request_id, status=status, admin_note="Duyệt từ Web Admin")
    return admin_home()

@app.route("/api/device_status")
def api_device_status():
    device_id = request.args.get("device_id") or get_device_id()
    sub = get_device_subscription(device_id)
    notice = get_renewal_notice(device_id)
    if not sub:
        return jsonify({"ok": True, "device_id": device_id, "premium": False, "notice": notice})
    return jsonify({"ok": True, "device_id": device_id, "premium": True, "package_name": sub[4], "end_date": sub[6], "status": sub[7], "notice": notice})

@app.route("/healthz")
def healthz_route():
    return jsonify({"ok": True, "app": APP_TITLE, "pages": len(get_pages_dynamic())})

@app.route("/support_poll")
def support_poll_route():
    device_id = (request.args.get("device_id") or get_device_id()).strip().upper()
    after_id = request.args.get("after_id", 0)
    rows = get_support_messages(device_id=device_id, after_id=after_id, limit=100)
    messages = [{"id": r[0], "device_id": r[1], "sender": r[2], "message": r[3], "created_at": r[4]} for r in rows]
    last_id = messages[-1]["id"] if messages else int(after_id or 0)
    return jsonify({"success": True, "messages": messages, "last_id": last_id})

@app.route("/support_send", methods=["POST"])
def support_send_route():
    data = request.get_json(silent=True) or request.form
    device_id = (data.get("device_id") or get_device_id()).strip().upper()
    message = (data.get("message") or "").strip()
    sender = (data.get("sender") or "user").strip().lower()
    msg_id = save_support_message(device_id, sender, message)
    if not msg_id:
        return jsonify({"success": False, "message": "Tin nhắn trống."}), 400
    return jsonify({"success": True, "id": msg_id, "message": "Đã gửi tin nhắn hỗ trợ."})

@app.route("/admin/support_reply", methods=["POST"])
def admin_support_reply_route():
    device_id = (request.form.get("device_id") or "").strip().upper()
    message = (request.form.get("message") or "").strip()
    if device_id and message:
        save_support_message(device_id, "admin", message)
    return admin_home()

@app.route("/page_comment_queue_add", methods=["POST"])
def page_comment_queue_add_route():
    comment_text = request.form.get("comment_text", "").strip()
    raw_targets = request.form.get("raw_targets", "").strip()
    target_type = request.form.get("target_type", "post")
    page_indexes = request.form.getlist("page_indexes")
    page_index = request.form.get("page_index", "")
    if not page_indexes and page_index != "":
        page_indexes = [page_index]
    min_delay = request.form.get("min_delay", "45")
    max_delay = request.form.get("max_delay", "60")
    single_post_uid = request.form.get("single_post_uid", "").strip()
    single_group_uid = request.form.get("single_group_uid", "").strip()
    single_user_uid = request.form.get("single_user_uid", "").strip()
    comments = split_comment_lines(comment_text)
    if not comments:
        return render(message="Vui lòng nhập nội dung bình luận. Mỗi dòng là 1 bình luận riêng.", ok=False)
    if not page_indexes:
        return render(message="Vui lòng chọn ít nhất 1 Page để tạo hàng chờ bình luận.", ok=False)
    total_added = total_skipped = 0
    for pi in page_indexes:
        if raw_targets:
            result = bulk_import_page_comment_queue(pi, raw_targets, comment_text, min_delay, max_delay, target_type)
            total_added += result['added']; total_skipped += result['skipped']
        else:
            first_comment = comments[0]
            qid = add_page_comment_queue(pi, target_type, single_user_uid, single_post_uid, single_group_uid, first_comment, min_delay, max_delay)
            if qid: total_added += 1
            else: total_skipped += 1
    if total_added <= 0:
        return render(message="Vui lòng nhập ít nhất một UID bài viết, UID Group hoặc UID người dùng hợp lệ.", ok=False)
    return render(message=f"Đã tạo {total_added} hàng chờ bình luận cho {len(page_indexes)} Page. Tổng bình luận nhập: {len(comments)}. Mỗi UID nhận 1 bình luận theo từng dòng và tự quay vòng nếu thiếu nội dung. Bỏ qua {total_skipped}. Giãn cách {normalize_comment_delay_seconds(min_delay,45)}-{normalize_comment_delay_seconds(max_delay,60)} giây.", ok=True)

@app.route("/page_comment_queue_action", methods=["POST"])
def page_comment_queue_action_route():
    queue_id = request.form.get("queue_id", "")
    action = request.form.get("action", "")
    if not queue_id:
        return render(message="Thiếu ID hàng chờ.", ok=False)
    if action == "approve":
        approve_page_comment_queue(queue_id)
        return render(message=f"Đã duyệt hàng chờ bình luận ID {queue_id}.", ok=True)
    if action == "done":
        mark_page_comment_done(queue_id, "Hoàn thành", "Đã xử lý hợp lệ và đã ghi log.")
        return render(message=f"Đã đánh dấu hoàn thành ID {queue_id}.", ok=True)
    mark_page_comment_done(queue_id, "Lỗi quyền", "Không xử lý vì thiếu quyền hoặc điều kiện truy cập không hợp lệ.")
    return render(message=f"Đã đánh dấu lỗi quyền ID {queue_id}.", ok=True)

@app.route("/export_page_comment_queue")
def export_page_comment_queue_route():
    path = export_page_comment_queue_csv()
    return send_file(path, as_attachment=True)

@app.route("/comment_ai", methods=["POST"])
def comment_ai_route():
    customer_name = request.form.get("customer_name", "").strip()
    phone = request.form.get("phone", "").strip()
    comment_text = request.form.get("comment_text", "").strip()
    label = request.form.get("label", "Khách nóng").strip()
    if not comment_text:
        return render(message="Vui lòng nhập comment khách hàng.", ok=False)
    clean_comment = hide_phone_preview(comment_text)
    prompt = v3_ai_tool_prompt('comment_reply', clean_comment, f"Nhãn khách: {label}")
    fallback = f"Dạ em đã nhận thông tin. Anh/chị vui lòng inbox để bên em tư vấn chi tiết và hỗ trợ nhanh hơn.\n\nNhãn gợi ý: {label}\nHành động: Chuyển CRM"
    ai_reply = safe_ai_generate(prompt, fallback=fallback)
    add_comment_lead(customer_name, phone, clean_comment, ai_reply, label)
    if request.form.get("to_crm", "1") == "1":
        add_pipeline_lead(customer_name or "Khách từ Comment", phone, "", "Comment Facebook", "Khách mới", 0, clean_comment)
    return render(content=ai_reply, message="Đã xử lý comment, ẩn SĐT và chuyển CRM nếu được chọn.", ok=True)

@app.route("/messenger_ai_script", methods=["POST"])
def messenger_ai_script_route():
    product = request.form.get("product", "").strip()
    script_type = request.form.get("script_type", "Kịch bản Inbox").strip()
    extra = request.form.get("extra", "").strip()
    if not product:
        return render(message="Vui lòng nhập sản phẩm/dịch vụ để tạo kịch bản Messenger.", ok=False)
    prompt = v3_ai_tool_prompt('messenger_ai', product, f"Loại kịch bản: {script_type}. {extra}")
    fallback = f"{script_type} cho {product}:\n1. Chào khách tự nhiên.\n2. Hỏi nhu cầu chính.\n3. Tư vấn gói phù hợp.\n4. Xử lý băn khoăn về giá.\n5. Chốt hành động: inbox/Zalo/thanh toán."
    content = safe_ai_generate(prompt, fallback=fallback)
    add_messenger_script(product, script_type, content)
    return render(content=content, message="Đã tạo kịch bản Messenger AI.", ok=True)

@app.route("/export")
def export_route():
    path = export_csv()
    return send_file(path, as_attachment=True)


@app.route("/backup")
def backup_route():
    path = backup_database()
    if not path:
        return render(message="Không tìm thấy database để backup.", ok=False)
    return send_file(path, as_attachment=True)

@app.route("/export_pdf")
def export_pdf_route():
    path = export_pdf_report()
    return send_file(path, as_attachment=True)


@app.route("/api/templates")
def api_templates():
    industry = request.args.get("industry", "spa")
    return jsonify(current_library(industry))


@app.get("/manifest.json")
def pwa_manifest():
    return jsonify({
        "name": "GPT MKT Pro",
        "short_name": "GPTMKT",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#2563eb",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    })

@app.get("/sw.js")
def pwa_sw():
    js = """
const CACHE_NAME = "gptmkt-v1";

self.addEventListener("install", (event) => {
  event.waitUntil(
    caches.open(CACHE_NAME).then((cache) => {
      return cache.addAll([
        "/",
        "/static/css/style.css",
        "/static/js/app.js"
      ]).catch(() => Promise.resolve());
    })
  );
  self.skipWaiting();
});

self.addEventListener("activate", (event) => {
  event.waitUntil(self.clients.claim());
});

self.addEventListener("fetch", (event) => {
  event.respondWith(
    caches.match(event.request).then((response) => {
      return response || fetch(event.request);
    })
  );
});
"""
    return app.response_class(js, mimetype="application/javascript")

@app.get("/service-worker.js")
def pwa_service_worker():
    return pwa_sw()

@app.get("/static/icon-192.png")
def static_icon_192():
    if os.path.exists(os.path.join("static", "icon-192.png")):
        return send_file(os.path.join("static", "icon-192.png"), mimetype="image/png")
    if os.path.exists("pwa-icon-192.png"):
        return send_file("pwa-icon-192.png", mimetype="image/png")
    import base64
    png = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII=")
    return app.response_class(png, mimetype="image/png")

@app.get("/static/icon-512.png")
def static_icon_512():
    if os.path.exists(os.path.join("static", "icon-512.png")):
        return send_file(os.path.join("static", "icon-512.png"), mimetype="image/png")
    if os.path.exists("pwa-icon-512.png"):
        return send_file("pwa-icon-512.png", mimetype="image/png")
    import base64
    png = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII=")
    return app.response_class(png, mimetype="image/png")

@app.get("/pwa-icon-192.png")
def pwa_icon_192():
    if os.path.exists("pwa-icon-192.png"):
        return send_file("pwa-icon-192.png", mimetype="image/png")
    import base64
    png = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII=")
    return app.response_class(png, mimetype="image/png")

@app.get("/pwa-icon-512.png")
def pwa_icon_512():
    if os.path.exists("pwa-icon-512.png"):
        return send_file("pwa-icon-512.png", mimetype="image/png")
    import base64
    png = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII=")
    return app.response_class(png, mimetype="image/png")


# =========================
# CTV / AFFILIATE CENTER V1
# =========================
def ensure_affiliate_tables():
    conn = db(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS affiliate_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT UNIQUE,
            full_name TEXT,
            phone TEXT,
            email TEXT,
            affiliate_code TEXT UNIQUE,
            custom_slug TEXT UNIQUE,
            level_name TEXT DEFAULT 'CTV thường',
            commission_percent REAL DEFAULT 20,
            status TEXT DEFAULT 'active',
            created_at TEXT,
            updated_at TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS affiliate_clicks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            affiliate_code TEXT,
            visitor_device_id TEXT,
            ip TEXT,
            user_agent TEXT,
            created_at TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS affiliate_referrals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            affiliate_code TEXT,
            customer_device_id TEXT,
            customer_phone TEXT,
            customer_email TEXT,
            package_key TEXT,
            package_name TEXT,
            amount INTEGER DEFAULT 0,
            commission_percent REAL DEFAULT 20,
            commission_amount INTEGER DEFAULT 0,
            status TEXT DEFAULT 'Chờ duyệt',
            premium_request_id INTEGER,
            created_at TEXT,
            approved_at TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS affiliate_withdrawals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            affiliate_code TEXT,
            amount INTEGER DEFAULT 0,
            bank_name TEXT,
            bank_number TEXT,
            account_name TEXT,
            status TEXT DEFAULT 'Chờ duyệt',
            admin_note TEXT,
            created_at TEXT,
            paid_at TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS affiliate_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level_name TEXT UNIQUE,
            min_orders INTEGER DEFAULT 0,
            commission_percent REAL DEFAULT 20,
            note TEXT,
            updated_at TEXT
        )
    """)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    defaults = [
        ('CTV thường', 0, 20, '0-10 đơn'),
        ('CTV Bạc', 11, 25, '11-30 đơn'),
        ('CTV Vàng', 31, 30, '31-70 đơn'),
        ('CTV Kim Cương', 71, 35, '71+ đơn'),
    ]
    for name, min_orders, pct, note in defaults:
        c.execute("""
            INSERT OR IGNORE INTO affiliate_settings(level_name,min_orders,commission_percent,note,updated_at)
            VALUES(?,?,?,?,?)
        """, (name, min_orders, pct, note, now))
    try:
        c.execute("ALTER TABLE premium_upgrade_requests ADD COLUMN affiliate_code TEXT")
    except Exception:
        pass
    conn.commit(); conn.close()


def current_affiliate_code():
    ref = (request.args.get('ref') or request.cookies.get('affiliate_code') or '').strip().upper()
    ref = ref.replace(' ', '').replace('_', '-')
    if ref and not ref.startswith('CTV-') and ref.startswith('CTV'):
        ref = 'CTV-' + ref[3:]
    return ref[:40]


def generate_affiliate_code():
    ensure_affiliate_tables()
    conn = db(); c = conn.cursor()
    while True:
        code = 'CTV-' + ''.join(random.choice('0123456789') for _ in range(6))
        c.execute('SELECT id FROM affiliate_users WHERE affiliate_code=?', (code,))
        if not c.fetchone():
            conn.close(); return code


def get_or_create_affiliate(device_id=None, full_name='', phone='', email=''):
    ensure_affiliate_tables()
    device_id = (device_id or get_device_id()).strip().upper()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute('SELECT id,device_id,full_name,phone,email,affiliate_code,custom_slug,level_name,commission_percent,status FROM affiliate_users WHERE device_id=?', (device_id,))
    row = c.fetchone()
    if row:
        if full_name or phone or email:
            c.execute("""
                UPDATE affiliate_users SET
                    full_name=COALESCE(NULLIF(?,''),full_name),
                    phone=COALESCE(NULLIF(?,''),phone),
                    email=COALESCE(NULLIF(?,''),email),
                    updated_at=?
                WHERE device_id=?
            """, (full_name.strip(), phone.strip(), email.strip(), now, device_id))
            conn.commit()
        c.execute('SELECT id,device_id,full_name,phone,email,affiliate_code,custom_slug,level_name,commission_percent,status FROM affiliate_users WHERE device_id=?', (device_id,))
        row = c.fetchone(); conn.close(); return row
    code = generate_affiliate_code()
    slug = code.lower().replace('-', '')
    c.execute("""
        INSERT INTO affiliate_users(device_id,full_name,phone,email,affiliate_code,custom_slug,level_name,commission_percent,status,created_at,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?)
    """, (device_id, full_name.strip(), phone.strip(), email.strip(), code, slug, 'CTV thường', 20, 'active', now, now))
    conn.commit()
    c.execute('SELECT id,device_id,full_name,phone,email,affiliate_code,custom_slug,level_name,commission_percent,status FROM affiliate_users WHERE device_id=?', (device_id,))
    row = c.fetchone(); conn.close(); return row


def affiliate_summary(affiliate_code):
    ensure_affiliate_tables()
    affiliate_code = (affiliate_code or '').strip().upper()
    conn = db(); c = conn.cursor()
    def one(sql, params=()):
        c.execute(sql, params); r = c.fetchone(); return r[0] if r and r[0] is not None else 0
    clicks = one('SELECT COUNT(*) FROM affiliate_clicks WHERE affiliate_code=?', (affiliate_code,))
    referrals = one('SELECT COUNT(*) FROM affiliate_referrals WHERE affiliate_code=?', (affiliate_code,))
    premium = one("SELECT COUNT(*) FROM affiliate_referrals WHERE affiliate_code=? AND status IN ('Đã duyệt','Đã thanh toán')", (affiliate_code,))
    revenue = one("SELECT COALESCE(SUM(amount),0) FROM affiliate_referrals WHERE affiliate_code=? AND status IN ('Đã duyệt','Đã thanh toán')", (affiliate_code,))
    commission = one("SELECT COALESCE(SUM(commission_amount),0) FROM affiliate_referrals WHERE affiliate_code=? AND status IN ('Đã duyệt','Đã thanh toán')", (affiliate_code,))
    paid = one("SELECT COALESCE(SUM(amount),0) FROM affiliate_withdrawals WHERE affiliate_code=? AND status='Đã thanh toán'", (affiliate_code,))
    pending_withdraw = one("SELECT COALESCE(SUM(amount),0) FROM affiliate_withdrawals WHERE affiliate_code=? AND status='Chờ duyệt'", (affiliate_code,))
    c.execute("""SELECT customer_device_id,customer_phone,customer_email,package_name,amount,commission_amount,status,created_at
                 FROM affiliate_referrals WHERE affiliate_code=? ORDER BY id DESC LIMIT 20""", (affiliate_code,))
    recent = c.fetchall()
    c.execute("SELECT amount,bank_name,bank_number,account_name,status,created_at,paid_at FROM affiliate_withdrawals WHERE affiliate_code=? ORDER BY id DESC LIMIT 20", (affiliate_code,))
    withdrawals = c.fetchall()
    conn.close()
    return {
        'clicks': clicks, 'referrals': referrals, 'premium': premium, 'revenue': revenue,
        'commission': commission, 'paid': paid, 'balance': max(0, int(commission or 0) - int(paid or 0) - int(pending_withdraw or 0)),
        'pending_withdraw': pending_withdraw, 'recent': recent, 'withdrawals': withdrawals
    }


def record_affiliate_click(affiliate_code):
    ensure_affiliate_tables()
    affiliate_code = (affiliate_code or '').strip().upper()
    if not affiliate_code:
        return
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute('SELECT id FROM affiliate_users WHERE affiliate_code=? AND status="active"', (affiliate_code,))
    if c.fetchone():
        c.execute("""INSERT INTO affiliate_clicks(affiliate_code,visitor_device_id,ip,user_agent,created_at)
                     VALUES(?,?,?,?,?)""", (affiliate_code, get_device_id(), request.remote_addr or '', request.headers.get('User-Agent','')[:250], now))
        conn.commit()
    conn.close()


def create_affiliate_commission_for_request(cursor, request_id):
    try:
        cursor.execute("SELECT device_id,phone,email,package_key,package_name,amount,affiliate_code,status FROM premium_upgrade_requests WHERE id=?", (request_id,))
    except Exception:
        return False
    row = cursor.fetchone()
    if not row:
        return False
    device_id, phone, email, package_key, package_name, amount, affiliate_code, req_status = row
    affiliate_code = (affiliate_code or '').strip().upper()
    if not affiliate_code:
        return False
    cursor.execute('SELECT commission_percent FROM affiliate_users WHERE affiliate_code=? AND status="active"', (affiliate_code,))
    ar = cursor.fetchone()
    if not ar:
        return False
    pct = float(ar[0] or 20)
    amount = int(amount or 0)
    commission = int(round(amount * pct / 100))
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute('SELECT id FROM affiliate_referrals WHERE premium_request_id=?', (request_id,))
    if cursor.fetchone():
        return False
    cursor.execute("""
        INSERT INTO affiliate_referrals(affiliate_code,customer_device_id,customer_phone,customer_email,package_key,package_name,amount,commission_percent,commission_amount,status,premium_request_id,created_at,approved_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (affiliate_code, device_id, phone, email, package_key, package_name, amount, pct, commission, 'Đã duyệt', request_id, now, now))
    cursor.execute("""INSERT INTO notifications(title,detail,level,status,created_at) VALUES(?,?,?,?,?)""",
                   ('Hoa hồng CTV mới', f'{affiliate_code} nhận {commission:,}đ từ {package_name}'.replace(',', '.'), 'success', 'new', now))
    return True


@app.route('/api/affiliate/me', methods=['GET', 'POST'])
def api_affiliate_me():
    ensure_affiliate_tables()
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form
        row = get_or_create_affiliate(full_name=data.get('full_name',''), phone=data.get('phone',''), email=data.get('email',''))
    else:
        row = get_or_create_affiliate()
    _, device_id, full_name, phone, email, code, slug, level, pct, status = row
    base = request.host_url.rstrip('/')
    summary = affiliate_summary(code)
    return jsonify({
        'ok': True,
        'device_id': device_id,
        'full_name': full_name or '',
        'phone': phone or '',
        'email': email or '',
        'affiliate_code': code,
        'custom_slug': slug or '',
        'level_name': level or 'CTV thường',
        'commission_percent': pct or 20,
        'status': status or 'active',
        'ref_link': f'{base}/?ref={code}',
        'pretty_link': f'{base}/ctv/{slug or code.lower().replace("-","")}',
        'summary': summary
    })


@app.route('/api/affiliate/withdraw', methods=['POST'])
def api_affiliate_withdraw():
    ensure_affiliate_tables()
    data = request.get_json(silent=True) or request.form
    row = get_or_create_affiliate()
    code = row[5]
    summary = affiliate_summary(code)
    amount = int(str(data.get('amount','0')).replace('.','').replace(',','') or 0)
    if amount <= 0 or amount > int(summary.get('balance',0)):
        return jsonify({'ok': False, 'message': 'Số tiền rút không hợp lệ hoặc vượt số dư khả dụng.'}), 400
    bank = (data.get('bank_name','') or '').strip()
    number = (data.get('bank_number','') or '').strip()
    account = (data.get('account_name','') or '').strip()
    if not bank or not number or not account:
        return jsonify({'ok': False, 'message': 'Vui lòng nhập đủ ngân hàng, số tài khoản và chủ tài khoản.'}), 400
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    c.execute("""INSERT INTO affiliate_withdrawals(affiliate_code,amount,bank_name,bank_number,account_name,status,admin_note,created_at,paid_at)
                 VALUES(?,?,?,?,?,?,?,?,?)""", (code, amount, bank, number, account, 'Chờ duyệt', '', now, ''))
    c.execute("""INSERT INTO notifications(title,detail,level,status,created_at) VALUES(?,?,?,?,?)""",
              ('CTV yêu cầu rút hoa hồng', f'{code} yêu cầu rút {amount:,}đ'.replace(',', '.'), 'warning', 'new', now))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'message': 'Đã gửi yêu cầu rút hoa hồng về web admin.'})


@app.route('/api/admin/affiliate')
def api_admin_affiliate():
    ensure_affiliate_tables()
    conn = db(); c = conn.cursor()
    c.execute("""SELECT u.id,u.device_id,u.full_name,u.phone,u.email,u.affiliate_code,u.level_name,u.commission_percent,u.status,u.created_at,
                        COALESCE(COUNT(r.id),0) AS orders,
                        COALESCE(SUM(CASE WHEN r.status IN ('Đã duyệt','Đã thanh toán') THEN r.amount ELSE 0 END),0) AS revenue,
                        COALESCE(SUM(CASE WHEN r.status IN ('Đã duyệt','Đã thanh toán') THEN r.commission_amount ELSE 0 END),0) AS commission
                 FROM affiliate_users u
                 LEFT JOIN affiliate_referrals r ON r.affiliate_code=u.affiliate_code
                 GROUP BY u.id ORDER BY u.id DESC LIMIT 200""")
    users = c.fetchall()
    c.execute('SELECT id,level_name,min_orders,commission_percent,note,updated_at FROM affiliate_settings ORDER BY min_orders ASC')
    settings = c.fetchall()
    c.execute("SELECT id,affiliate_code,amount,bank_name,bank_number,account_name,status,created_at,paid_at FROM affiliate_withdrawals ORDER BY id DESC LIMIT 100")
    withdrawals = c.fetchall()
    conn.close()
    return jsonify({'ok': True, 'users': users, 'settings': settings, 'withdrawals': withdrawals})


@app.route('/api/admin/affiliate/settings', methods=['POST'])
def api_admin_affiliate_settings():
    ensure_affiliate_tables()
    data = request.get_json(silent=True) or request.form
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    for key, default in [('normal',20),('silver',25),('gold',30),('diamond',35)]:
        pct = data.get(key)
        if pct is None: continue
        try: pct = float(pct)
        except Exception: pct = default
        name = {'normal':'CTV thường','silver':'CTV Bạc','gold':'CTV Vàng','diamond':'CTV Kim Cương'}[key]
        c.execute('UPDATE affiliate_settings SET commission_percent=?, updated_at=? WHERE level_name=?', (pct, now, name))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'message': 'Đã cập nhật phần trăm hoa hồng.'})



@app.route('/api/admin/affiliate/user/<int:user_id>', methods=['POST'])
def api_admin_affiliate_user_update(user_id):
    ensure_affiliate_tables()
    data = request.get_json(silent=True) or request.form
    status = (data.get('status') or 'active').strip()
    if status not in ['active','paused','blocked']:
        status = 'active'
    try:
        pct = float(data.get('commission_percent', 20))
    except Exception:
        pct = 20
    pct = max(0, min(80, pct))
    level_name = (data.get('level_name') or '').strip() or None
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = db(); c = conn.cursor()
    if level_name:
        c.execute('UPDATE affiliate_users SET commission_percent=?, status=?, level_name=?, updated_at=? WHERE id=?', (pct, status, level_name, now, user_id))
    else:
        c.execute('UPDATE affiliate_users SET commission_percent=?, status=?, updated_at=? WHERE id=?', (pct, status, now, user_id))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'message': 'Đã cập nhật hoa hồng riêng cho CTV.'})

@app.route('/api/admin/affiliate/withdraw/<int:withdraw_id>', methods=['POST'])
def api_admin_affiliate_withdraw_action(withdraw_id):
    ensure_affiliate_tables()
    data = request.get_json(silent=True) or request.form
    status = data.get('status','Đã thanh toán')
    if status not in ['Đã thanh toán','Từ chối','Chờ duyệt']:
        status = 'Đã thanh toán'
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') if status == 'Đã thanh toán' else ''
    conn = db(); c = conn.cursor()
    c.execute('UPDATE affiliate_withdrawals SET status=?, paid_at=? WHERE id=?', (status, now, withdraw_id))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route('/ctv/<slug>')
def affiliate_pretty_link(slug):
    ensure_affiliate_tables()
    slug = (slug or '').strip().lower()
    conn = db(); c = conn.cursor()
    c.execute('SELECT affiliate_code FROM affiliate_users WHERE lower(custom_slug)=? OR lower(replace(affiliate_code,"-",""))=? LIMIT 1', (slug, slug))
    row = c.fetchone(); conn.close()
    code = row[0] if row else ''
    resp = home()
    if code:
        record_affiliate_click(code)
        try:
            resp.set_cookie('affiliate_code', code, max_age=60*60*24*90)
        except Exception:
            pass
    return resp


if __name__ == "__main__":
    init_db()
    threading.Thread(target=scheduler_loop, daemon=True).start()
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=False)
